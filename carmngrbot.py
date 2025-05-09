# (1) --------------- ИМПОРТ МОДУЛЕЙ ---------------

import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telebot import types
import datetime
from datetime import datetime
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import os
import json
import locale
import re
import requests
from functools import partial
from urllib.parse import quote
import traceback
from bs4 import BeautifulSoup
import time
import logging
import calendar
from telegram_bot_calendar import DetailedTelegramCalendar
from datetime import date
from requests.exceptions import ReadTimeout, ConnectionError
from scipy.spatial import cKDTree
import threading
import csv
import shutil
import hashlib
from statistics import mean
from functools import wraps
import schedule
import threading
from datetime import timedelta

# (2) --------------- ТОКЕН БОТА ---------------

bot = telebot.TeleBot("7519948621:AAGPoPBJrnL8-vZepAYvTmm18TipvvmLUoE")


import pytz

# Создаем объект часового пояса для GMT+3
moscow_tz = pytz.timezone('Europe/Moscow')

# Пример использования часового пояса
current_time = datetime.now(moscow_tz)
formatted_time = current_time.strftime('%d.%m.%Y в %H:%M:%S')


# (4) --------------- ЗАГРУЗКА ТЕКСТОВОГО ФАЙЛА С РЕГИОНАМИ ---------------

regions = {}
try:
    with open('files/regions.txt', 'r', encoding='utf-8') as file:
        for line in file:
            parts = line.strip().split(' — ')
            if len(parts) == 2:
                code, name = parts
                regions[code.strip()] = name.strip()
except FileNotFoundError:
    pass

# (5) --------------- ОБРАБОТЧИК КОМАНДЫ /START ---------------

# Обработчик команды /start
def restricted(func):
    """Декоратор для ограничения доступа заблокированным пользователям."""
    def wrapper(message, *args, **kwargs):
        user_id = message.from_user.id
        username = message.from_user.username

        # Проверка по ID
        if is_user_blocked(user_id):
            bot.send_message(message.chat.id, "🚫 Вы *заблокированы* и не можете выполнять это действие!", parse_mode="Markdown")
            return

        # Проверка по username
        if username and is_user_blocked(get_user_id_by_username(username)):
            bot.send_message(message.chat.id, "🚫 Вы *заблокированы* и не можете выполнять это действие!", parse_mode="Markdown")
            return
        
        return func(message, *args, **kwargs)
    return wrapper

def track_user_activity(func):
    def wrapper(message, *args, **kwargs):
        user_id = message.from_user.id
        username = message.from_user.username if message.from_user.username else "Неизвестный пользователь"
        first_name = message.from_user.first_name if message.from_user.first_name else ""
        last_name = message.from_user.last_name if message.from_user.last_name else ""
        update_user_activity(user_id, username, first_name, last_name)  # Обновляем активность пользователя
        return func(message, *args, **kwargs)
    return wrapper

# Декоратор для проверки состояния чата

# Глобальная переменная для хранения истории сообщений
message_history = {}

# Декоратор для проверки состояния чата
def check_chat_state(func):
    @wraps(func)
    def wrapper(message, *args, **kwargs):
        global active_chats  # Убедимся, что используем глобальную переменную
        user_id = message.from_user.id

        # Проверяем, есть ли у пользователя активный запрос на чат
        if user_id in active_chats and active_chats[user_id].get("awaiting_response", False):
            if message.text.strip().lower() not in ["принять", "отклонить"]:
                bot.send_message(user_id, "Пожалуйста, выберите *ПРИНЯТЬ* или *ОТКЛОНИТЬ*!", parse_mode="Markdown")
                return

        return func(message, *args, **kwargs)
    return wrapper


# Функция для сохранения последнего сообщения от бота
def save_last_bot_message(user_id, message_text):
    message_history[user_id] = {"last_bot_message": message_text}

# Декоратор для проверки состояния функции

def check_function_state(function_name):
    if function_name in function_states:
        return function_states[function_name]['state']
    return False

def check_function_state_decorator(function_name):
    def decorator(func):
        @wraps(func)
        def wrapped(message, *args, **kwargs):
            if check_function_state(function_name):
                return func(message, *args, **kwargs)
            else:
                bot.send_message(message.chat.id, f"Функция *{function_name.lower()}* временно недоступна!", parse_mode="Markdown")
        return wrapped
    return decorator

# Декоратор для отслеживания вызовов функций
def track_usage(func_name):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            statistics = load_statistics()
            current_date = datetime.now().strftime('%d.%m.%Y')

            if current_date not in statistics:
                statistics[current_date] = {'users': set(), 'functions': {}}
            if 'functions' not in statistics[current_date]:
                statistics[current_date]['functions'] = {}
            if func_name not in statistics[current_date]['functions']:
                statistics[current_date]['functions'][func_name] = 0
            statistics[current_date]['functions'][func_name] += 1

            save_statistics(statistics)
            return func(*args, **kwargs)
        return wrapper
    return decorator


# Настройка логирования для файла
file_logger = logging.getLogger('fileLogger')
file_handler = logging.FileHandler('data base/log/bot_logs.log', encoding='utf-8')
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
file_logger.addHandler(file_handler)

# Настройка логирования для консоли
console_logger = logging.getLogger('consoleLogger')
console_logger.setLevel(logging.ERROR)  # Устанавливаем уровень ERROR, чтобы выводить только ошибки в консоль
console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
console_logger.addHandler(console_handler)

# Функция для записи логов в JSON файл
def log_to_json(user_id, log_entry):
    log_dir = "data base/log"
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"{user_id}_log.json")

    # Если файл существует, пытаемся загрузить данные
    if os.path.exists(log_file):
        try:
            with open(log_file, 'r', encoding='utf-8') as file:
                logs = json.load(file)
        except json.decoder.JSONDecodeError:  # Если файл поврежден
            print(f"Ошибка при чтении файла {log_file}, файл поврежден! Создаем новый файл...")
            logs = []  # Создаем новый список
    else:
        logs = []  # Если файл не существует, создаем новый список

    # Добавляем новую запись в логи
    logs.append(log_entry)

    # Сохраняем обновленные логи обратно в файл
    with open(log_file, 'w', encoding='utf-8') as file:
        json.dump(logs, file, ensure_ascii=False, indent=4)

# Функция для очистки логов и переноса ошибок
def clear_logs_and_transfer_errors():
    log_dir = "data base/log"
    error_log_file = os.path.join(log_dir, "errors_log.json")

    # Загружаем существующие ошибки из errors_log.json
    if os.path.exists(error_log_file):
        with open(error_log_file, 'r', encoding='utf-8') as file:
            errors = json.load(file)
    else:
        errors = []

    for filename in os.listdir(log_dir):
        if filename.endswith("_log.json") and filename != "errors_log.json":
            file_path = os.path.join(log_dir, filename)
            with open(file_path, 'r', encoding='utf-8') as file:
                logs = json.load(file)

            # Переносим ошибки в errors_log.json
            for log in logs:
                if log.get("level") == "ERROR":
                    errors.append(log)

            # Очищаем содержимое файла
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump([], file, ensure_ascii=False, indent=4)

    # Сохраняем ошибки в errors_log.json
    with open(error_log_file, 'w', encoding='utf-8') as file:
        json.dump(errors, file, ensure_ascii=False, indent=4)

# Функция для удаления старых ошибок из errors_log.json
def remove_old_errors():
    log_dir = "data base/log"
    error_log_file = os.path.join(log_dir, "errors_log.json")

    # Загружаем существующие ошибки из errors_log.json
    if os.path.exists(error_log_file):
        with open(error_log_file, 'r', encoding='utf-8') as file:
            errors = json.load(file)
    else:
        errors = []

    # Удаляем ошибки, которые не повторялись в течение 7 дней
    current_time = datetime.now()
    errors = [error for error in errors if (current_time - datetime.strptime(error["timestamp"], '%d.%m.%Y в %H:%M:%S')) <= timedelta(days=7)]

    # Сохраняем обновленные ошибки в errors_log.json
    with open(error_log_file, 'w', encoding='utf-8') as file:
        json.dump(errors, file, ensure_ascii=False, indent=4)

# Продвинутый декоратор для логирования действий пользователя и бота
def log_user_actions(func):
    @wraps(func)
    def wrapper(message, *args, **kwargs):
        # Проверяем корректность объекта message
        if not hasattr(message, 'from_user') or not hasattr(message, 'chat'):
            file_logger.error(f"Invalid message type: {type(message)}. Message: {message}")
            console_logger.error(f"Invalid message type: {type(message)}. Message: {message}")
            return  # Выходим, если сообщение некорректно

        try:
            user_id = message.from_user.id
            username = message.from_user.username if message.from_user.username else "Неизвестный пользователь"
            first_name = message.from_user.first_name if message.from_user.first_name else ""
            last_name = message.from_user.last_name if message.from_user.last_name else ""
            message_text = message.text if hasattr(message, 'text') and message.text else "Нет текста"
            command = func.__name__
            chat_id = message.chat.id
            message_id = message.message_id
            message_type = "text" if hasattr(message, 'text') and message.text else "media"
            context = message_text
            language_code = message.from_user.language_code if hasattr(message.from_user, 'language_code') else "Неизвестный язык"
            chat_type = message.chat.type if hasattr(message.chat, 'type') else "Неизвестный тип"
            location = message.location if hasattr(message, 'location') and message.location else "Нет геолокации"

            start_time = time.time()

            # Выполняем основную функцию
            result = func(message, *args, **kwargs)
            execution_time = time.time() - start_time

            log_entry = {
                "timestamp": datetime.now().strftime('%d.%m.%Y в %H:%M:%S'),
                "level": "INFO",
                "event_type": "User Action",
                "user_id": user_id,
                "username": username,
                "first_name": first_name,
                "last_name": last_name,
                "message_text": message_text,
                "command": command,
                "result": "Success",
                "execution_time": f"{execution_time:.2f} sec",
                "chat_id": chat_id,
                "message_id": message_id,
                "message_type": message_type,
                "context": context,
                "language_code": language_code,
                "chat_type": chat_type,
                "location": location,
                "bot_info": {
                    "version": "1.0",
                    "bot_id": "7519948621:AAGPoPBJrnL8-vZepAYvTmm18TipvvmLUoE"
                }
            }
            file_logger.info(f"User {user_id} executed command {command} successfully in {execution_time:.2f} sec.")
        except Exception as e:
            execution_time = time.time() - start_time
            error_details_more = traceback.format_exc().splitlines()
            log_entry = {
                "timestamp": datetime.now().strftime('%d.%m.%Y в %H:%M:%S'),
                "level": "ERROR",
                "event_type": "User Action",
                "user_id": "Unknown" if 'user_id' not in locals() else user_id,
                "username": "Unknown" if 'username' not in locals() else username,
                "first_name": "Unknown" if 'first_name' not in locals() else first_name,
                "last_name": "Unknown" if 'last_name' not in locals() else last_name,
                "message_text": "Unknown" if 'message_text' not in locals() else message_text,
                "command": command,
                "result": "Error",
                "execution_time": f"{execution_time:.2f} sec",
                "error_details": str(e),
                "error_details_more": error_details_more,
                "chat_id": "Unknown" if 'chat_id' not in locals() else chat_id,
                "message_id": "Unknown" if 'message_id' not in locals() else message_id,
                "message_type": "Unknown" if 'message_type' not in locals() else message_type,
                "context": "Unknown" if 'context' not in locals() else context,
                "language_code": "Unknown" if 'language_code' not in locals() else language_code,
                "chat_type": "Unknown" if 'chat_type' not in locals() else chat_type,
                "location": "Unknown" if 'location' not in locals() else location,
                "bot_info": {
                    "version": "1.0",
                    "bot_id": "7519948621:AAGPoPBJrnL8-vZepAYvTmm18TipvvmLUoE"
                }
            }
            file_logger.error(f"Error while executing command {command}: {e}")
            console_logger.error(f"Error while executing command {command}: {e}")
            raise
        finally:
            # Сохраняем логи в JSON
            log_to_json(user_id if 'user_id' in locals() else "Unknown", log_entry)

        return result
    return wrapper

# Функция для запуска задачи удаления старых ошибок каждую неделю
def run_weekly_task():
    while True:
        remove_old_errors()
        time.sleep(7 * 24 * 60 * 60)  # Спим 7 дней

# Запуск задачи очистки логов и переноса ошибок каждый день в 00:00
schedule.every().day.at("00:00").do(clear_logs_and_transfer_errors)

# Функция для запуска запланированных задач
def run_scheduled_tasks():
    while True:
        schedule.run_pending()
        time.sleep(1)

# Запуск запланированных задач в отдельном потоке
scheduler_thread = threading.Thread(target=run_scheduled_tasks)
scheduler_thread.start()

# Запуск задачи удаления старых ошибок каждую неделю в отдельном потоке
weekly_task_thread = threading.Thread(target=run_weekly_task)
weekly_task_thread.start()


from functools import wraps
from telebot.apihelper import ApiTelegramException

# Путь к файлу с заблокированными пользователями
BLOCKED_USERS_FILE = 'data base/admin/blocked_bot_users.json'

# Загрузка списка заблокированных пользователей из файла
def load_blocked_users():
    if os.path.exists(BLOCKED_USERS_FILE):
        with open(BLOCKED_USERS_FILE, 'r') as file:
            return json.load(file)
    return []

# Сохранение списка заблокированных пользователей в файл
def save_blocked_users(blocked_users):
    with open(BLOCKED_USERS_FILE, 'w') as file:
        json.dump(blocked_users, file, indent=4)

# Декоратор для проверки, заблокировал ли пользователь бота
def check_user_blocked(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except ApiTelegramException as e:
            if e.result_json['error_code'] == 403 and 'bot was blocked by the user' in e.result_json['description']:
                user_id = kwargs.get('user_id') or args[0].chat.id
                print(f"User blocked the bot: {user_id}")

                # Загрузка текущего списка заблокированных пользователей
                blocked_users = load_blocked_users()

                # Добавление нового пользователя в список, если его там еще нет
                if user_id not in blocked_users:
                    blocked_users.append(user_id)
                    save_blocked_users(blocked_users)
            else:
                raise e
    return wrapper


# -------- Уведомление о неактивности ---------

# Путь к файлу базы данных
DB_PATH = 'data base/admin/users.json'

# Интервал в секундах для проверки неактивности (3 дня)
INACTIVITY_THRESHOLD = 3 * 24 * 60 * 60

# Интервал в секундах для проверки (12 часов)
CHECK_INTERVAL = 12 * 60 * 60

def load_users():
    with open(DB_PATH, 'r', encoding='utf-8') as file:
        return json.load(file)

def save_users(users):
    with open(DB_PATH, 'w', encoding='utf-8') as file:
        json.dump(users, file, ensure_ascii=False, indent=4)

def check_inactivity():
    while True:
        users = load_users()
        current_time = datetime.now()

        for user_id, user_data in users.items():
            last_active_str = user_data.get('last_active')
            if last_active_str:
                last_active = datetime.strptime(last_active_str, '%d.%m.%Y в %H:%M:%S')
                if current_time - last_active > timedelta(seconds=INACTIVITY_THRESHOLD):
                    username = user_data.get('username', 'Неизвестный пользователь')
                    message = f"Уважаемый пользователь, @{escape_markdown(username)}, от вас давно не было активности! Используйте бота или ваши данные будут удалены!"
                    bot.send_message(user_id, message, parse_mode="Markdown")

        # Ждем 12 часов перед следующей проверкой
        time.sleep(CHECK_INTERVAL)

# Запуск потока для проверки неактивности
inactivity_thread = threading.Thread(target=check_inactivity)
inactivity_thread.daemon = True
inactivity_thread.start()


# -------- Экстренная остановка ---------

import signal

def emergency_stop(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Экстренная остановка'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Подтвердить остановку", "Отмена остановки")
    markup.add("В меню админ-панели")
    bot.send_message(message.chat.id, "Вы уверены, что хотите остановить бота?", reply_markup=markup)
    bot.register_next_step_handler(message, confirm_emergency_stop)

# Обработчик команды для экстренной остановки бота
@bot.message_handler(func=lambda message: message.text == 'Экстренная остановка' and check_admin_access(message))
def handle_emergency_stop(message):
    emergency_stop(message)

# Функция для остановки бота через 5 секунд
def stop_bot_after_delay():
    threading.Timer(5.0, stop_bot).start()

# Функция для остановки бота
def stop_bot():
    bot.stop_polling()
    os.kill(os.getpid(), signal.SIGINT)

# Обработчик подтверждения остановки бота
def confirm_emergency_stop(message):
    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.text == "Подтвердить остановку":
        bot.send_message(message.chat.id, "🛑 Бот будет остановлен через 5 секунд...")
        stop_bot_after_delay()
        show_admin_panel(message)
    elif message.text == "Отмена остановки":
        bot.send_message(message.chat.id, "Остановка бота отменена!")
        show_admin_panel(message)
    else:
        bot.send_message(message.chat.id, "Неверная команда! Пожалуйста, выберите действие")
        bot.register_next_step_handler(message, confirm_emergency_stop)


# (3) --------------- СОХРАНЕНИЯ И ЗАГРУЗКА ДАННЫХ ПОЛЬЗОВАТЕЛЯ ---------------

# (3.1) --------------- СОХРАНЕНИЯ И ЗАГРУЗКА ДАННЫХ ПОЛЬЗОВАТЕЛЯ (ПОЕЗДКИ) ---------------

def save_data(user_id): 
    if user_id in user_trip_data:
        folder_path = "data base\\trip"  # или используйте прямой слэш "data base/trip"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        with open(os.path.join(folder_path, f"{user_id}_trip_data.json"), "w") as json_file:
            json.dump(user_trip_data[user_id], json_file)

# Псевдоданные - просто для примера
# В реальном случае данные приходят из пользовательского ввода или другого источника
user_trip_data = {}

# Функция для добавления поездки
#@log_user_actions
def add_trip(user_id, trip):
    if user_id not in user_trip_data:
        user_trip_data[user_id] = []
    user_trip_data[user_id].append(trip)

# Функция для сохранения данных
def save_trip_data(user_id):
    folder_path = "data base\\trip"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Сохраняем данные поездок для конкретного пользователя
    file_path = os.path.join(folder_path, f"{user_id}_trip_data.json")
    with open(file_path, "w", encoding='utf-8') as json_file:
        json.dump(user_trip_data.get(user_id, []), json_file, ensure_ascii=False, indent=4)


# Данные поездок для всех пользователей
user_trip_data = {}

# Путь к папке с данными
folder_path = "data base\\trip"

# Функция для загрузки данных для одного пользователя
def load_trip_data(user_id):
    folder_path = "data base\\trip"
    file_path = os.path.join(folder_path, f"{user_id}_trip_data.json")
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding='utf-8') as json_file:
                return json.load(json_file)
        except UnicodeDecodeError:
            with open(file_path, "r", encoding='windows-1251') as json_file:
                data = json.load(json_file)
            with open(file_path, "w", encoding='utf-8') as json_file:
                json.dump(data, json_file, ensure_ascii=False, indent=4)
            return data
    else:
        return []  # Если данных нет, возвращаем пустой список


# Функция для загрузки данных для всех пользователей
def load_all_user_data():
    folder_path = "data base\\trip"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)  # Если папка не существует, создаем её
    for filename in os.listdir(folder_path):
        if filename.endswith("_trip_data.json"):
            user_id = filename.split("_")[0]  # Извлекаем user_id из имени файла
            user_trip_data[user_id] = load_trip_data(user_id)  # Загружаем данные для пользователя

# Функция для добавления поездки
#@log_user_actions
def add_trip(user_id, trip):
    if user_id not in user_trip_data:
        user_trip_data[user_id] = []
    user_trip_data[user_id].append(trip)

# Функция для сохранения данных всех пользователей при выходе
def save_all_trip_data():
    for user_id in user_trip_data:
        save_trip_data(user_id)

# Загружаем все данные при старте бота
load_all_user_data()


# Старт
@bot.message_handler(commands=['start'])
@restricted
@track_user_activity
@check_chat_state
#@log_user_actions
def start(message):
    chat_id = message.chat.id
    user_id = message.from_user.id
    username = message.from_user.username if message.from_user.username else "Неизвестный пользователь"

    # Сохраняем информацию о пользователе
    update_user_activity(user_id, username)

    # Создание кнопок меню
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Расход топлива")
    item2 = types.KeyboardButton("Траты и ремонты")
    item3 = types.KeyboardButton("Найти транспорт")
    item4 = types.KeyboardButton("Поиск мест")
    item5 = types.KeyboardButton("Погода")
    item6 = types.KeyboardButton("Код региона")
    item7 = types.KeyboardButton("Цены на топливо")
    item8 = types.KeyboardButton("Анти-радар")
    item9 = types.KeyboardButton("Напоминания")
    item10 = types.KeyboardButton("Коды OBD2")
    item11 = types.KeyboardButton("Прочее")

    markup.add(item1, item2)
    markup.add(item3, item4)
    markup.add(item5, item7)
    markup.add(item6, item8)
    markup.add(item9, item10)
    markup.add(item11)

    welcome_message = f"Добро пожаловать, @{escape_markdown(username)}!\nВыберите действие из меню:"
    bot.send_message(chat_id, welcome_message, parse_mode="Markdown", reply_markup=markup)

# (6) --------------- ОБРАБОТЧИК КОМАНДЫ /MAINMENU ---------------
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('В главное меню')
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == "В главное меню")
@bot.message_handler(commands=['mainmenu'])
def return_to_menu(message):
    start(message)

# (7) --------------- ДОПОЛНИТЕЛЬНОЙ ИНФОРМАЦИИ ДЛЯ ПОЛЬЗОВАТЕЛЯ ---------------

@bot.message_handler(func=lambda message: message.text == "Сайт")
@bot.message_handler(commands=['website'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Сайт')
@track_usage('Сайт')  # Добавление отслеживания статистики
#@log_user_actions
def send_website_file(message):
    # Отправка гиперссылки на сайт
    bot.send_message(message.chat.id, "[Сайт CAR MANAGER](carmngrbot.com.swtest.ru)", parse_mode="Markdown")  # http://carmngrbot.com.swtest.ru/ # https://goo.su/5htqWmk
    
    # Закомментированная часть для отправки файла
    # try:
    #     with open('files/website.html', 'rb') as website_file:
    #         bot.send_document(chat_id, website_file)
    # except FileNotFoundError:
    #     bot.send_message(chat_id, "Извините, файл website не найден")

# (8) --------------- ОБРАБОТЧИК КОМАНДЫ "РАСХОД ТОПЛИВА"---------------

@bot.message_handler(func=lambda message: message.text == "Расход топлива")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Расход топлива')
@track_usage('Расход топлива')  # Добавление отслеживания статистики
#@log_user_actions
def handle_fuel_expense(message):

    user_id = message.from_user.id

    if user_id not in user_trip_data:
        user_trip_data[user_id] = load_trip_data(user_id)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Рассчитать расход топлива")
    item2 = types.KeyboardButton("Посмотреть поездки")
    item3 = types.KeyboardButton("Удалить поездку")
    item4 = types.KeyboardButton("В главное меню")

    markup.add(item1)
    markup.add(item2, item3)
    markup.add(item4)

    bot.clear_step_handler_by_chat_id(user_id)
    bot.send_message(user_id, "Меню для учета расхода топлива. Выберите действие:", reply_markup=markup)

# (9) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" ---------------

# (9.1) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (СЛОВАРИ, СПИСОК, РЕГУЛЯРНОЕ ВЫРАЖЕНИЕ) ---------------

user_trip_data = {}

trip_data = {}

temporary_trip_data = {}

fuel_types = ["АИ-92", "АИ-95", "АИ-98", "АИ-100", "ДТ", "ГАЗ"]

date_pattern = r"^\d{2}.\d{2}.\d{4}$"

# (9.2) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (КОМАНДЫ /restart1 ) ---------------

# Пример функции, которая вызывается при выходе в меню расчета топлива
@bot.message_handler(func=lambda message: message.text == "Вернуться в меню расчета топлива")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Вернуться в меню расчета топлива')
#@log_user_actions
def restart_handler(message):
    user_id = message.chat.id

    # Сохраняем данные перед выходом в меню
    save_trip_data(user_id)

    # Загружаем данные для пользователя при возвращении в меню
    user_trip_data[user_id] = load_trip_data(user_id)

    # Возвращаемся в меню
    reset_and_start_over(message.chat.id)

#@log_user_actions
def reset_and_start_over(chat_id):
    # Отправляем новое меню пользователю
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Рассчитать расход топлива")
    item2 = types.KeyboardButton("Посмотреть поездки")
    item3 = types.KeyboardButton("Удалить поездку")
    item4 = types.KeyboardButton("В главное меню")

    markup.add(item1)
    markup.add(item2, item3)
    markup.add(item4)

    bot.send_message(chat_id, "Вы вернулись в меню расчета топлива. Выберите действие:", reply_markup=markup)

# Когда бот завершает работу или пользователь выходит из меню, сохраняем все данные
# Например, при перезапуске или выходе:
save_all_trip_data()

# (9.3) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (ЗАГРУЗКА ДАННЫХ ПРИ /restart1) ---------------

#@log_user_actions
def reset_user_data(user_id):
    if user_id not in user_trip_data:
        user_trip_data[user_id] = load_trip_data(user_id)
    bot.clear_step_handler_by_chat_id(user_id) 

# (9.4) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (ФУНКЦИЯ НАЧАЛЬНОГО МЕСТОПОЛОЖЕНИЯ) ---------------


@bot.message_handler(func=lambda message: message.text == "Рассчитать расход топлива")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Рассчитать расход топлива')
@track_usage('Рассчитать расход топлива')  # Добавление отслеживания статистики
#@log_user_actions
def calculate_fuel_cost_handler(message):
    chat_id = message.chat.id

    bot.clear_step_handler_by_chat_id(chat_id)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Отправить геолокацию", request_location=True)
    item2 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item3 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)
    markup.add(item3)
    sent = bot.send_message(chat_id, "Введите начальное местоположение или отправьте геолокацию:", reply_markup=markup)
    reset_user_data(chat_id)  

    bot.register_next_step_handler(sent, process_start_location_step)
	
#@log_user_actions
def process_start_location_step(message):
    chat_id = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Отправить геолокацию", request_location=True)
    item2 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item3 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)
    markup.add(item3)

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return  
    if message.text == "В главное меню":
        return_to_menu(message) 
        return    

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_start_location_step)
        return  

    if message.location:  # Если пользователь отправил геолокацию
        location = message.location
        try:
            start_address = geolocator.reverse((location.latitude, location.longitude), timeout=10).address
        except GeocoderUnavailable: # type: ignore
            bot.send_message(chat_id, "Сервис геолокации временно недоступен. Попробуйте позже.")
            return
        trip_data[chat_id] = {
            "start_location": {
                "address": start_address,
                "latitude": location.latitude,
                "longitude": location.longitude
            }
        }
        bot.send_message(chat_id, f"Ваше начальное местоположение:\n\n{start_address}")
    else:  # Если пользователь ввел текстовое местоположение
        start_location = message.text
        try:
            location = geolocator.geocode(start_location, timeout=10)
        except GeocoderUnavailable: # type: ignore
            bot.send_message(chat_id, "Сервис геолокации временно недоступен. Попробуйте позже.")
            return
        if location:
            trip_data[chat_id] = {
                "start_location": {
                    "address": start_location,
                    "latitude": location.latitude,
                    "longitude": location.longitude
                }
            }
            bot.send_message(chat_id, f"Ваше начальное местоположение:\n\n{start_location}")
        else:
            sent = bot.send_message(chat_id, "Не удалось найти местоположение! Пожалуйста, введите корректный адрес")
            bot.register_next_step_handler(sent, process_start_location_step)
            return
    
    sent = bot.send_message(chat_id, "Введите конечное местоположение или отправьте геолокацию:", reply_markup=markup)
    bot.register_next_step_handler(sent, process_end_location_step)

#@log_user_actions
def process_start_location_step(message):
    chat_id = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Отправить геолокацию", request_location=True)
    item2 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item3 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)
    markup.add(item3)

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return  
    if message.text == "В главное меню":
        return_to_menu(message) 
        return    

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_start_location_step)
        return  

    if message.location:  # Если пользователь отправил геолокацию
        location = message.location
        try:
            start_address = geolocator.reverse((location.latitude, location.longitude), timeout=10).address
        except GeocoderUnavailable: # type: ignore
            bot.send_message(chat_id, "Сервис геолокации временно недоступен. Попробуйте позже.")
            return
        trip_data[chat_id] = {
            "start_location": {
                "address": start_address,
                "latitude": location.latitude,
                "longitude": location.longitude
            }
        }
        bot.send_message(chat_id, f"Ваше начальное местоположение:\n\n{start_address}")
    else:  # Если пользователь ввел текстовое местоположение
        start_location = message.text
        try:
            location = geolocator.geocode(start_location, timeout=10)
        except GeocoderUnavailable: # type: ignore
            bot.send_message(chat_id, "Сервис геолокации временно недоступен. Попробуйте позже.")
            return
        if location:
            trip_data[chat_id] = {
                "start_location": {
                    "address": start_location,
                    "latitude": location.latitude,
                    "longitude": location.longitude
                }
            }
            bot.send_message(chat_id, f"Ваше начальное местоположение:\n\n{start_location}")
        else:
            sent = bot.send_message(chat_id, "Не удалось найти местоположение! Пожалуйста, введите корректный адрес")
            bot.register_next_step_handler(sent, process_start_location_step)
            return
    
    sent = bot.send_message(chat_id, "Введите конечное местоположение или отправьте геолокацию:", reply_markup=markup)
    bot.register_next_step_handler(sent, process_end_location_step)

#@log_user_actions
def process_end_location_step(message):
    chat_id = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return  
    if message.text == "В главное меню":
        return_to_menu(message) 
        return    

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_end_location_step)
        return

    if message.location:  # Если пользователь отправил геолокацию
        location = message.location
        try:
            end_address = geolocator.reverse((location.latitude, location.longitude), timeout=10).address
        except GeocoderUnavailable: # type: ignore
            bot.send_message(chat_id, "Сервис геолокации временно недоступен. Попробуйте позже.")
            return
        trip_data[chat_id]["end_location"] = {
            "address": end_address,
            "latitude": location.latitude,
            "longitude": location.longitude
        }
        bot.send_message(chat_id, f"Ваше конечное местоположение:\n\n{end_address}")
    else:  # Если пользователь ввел текстовое местоположение
        end_location = message.text
        try:
            location = geolocator.geocode(end_location, timeout=10)
        except GeocoderUnavailable: # type: ignore
            bot.send_message(chat_id, "Сервис геолокации временно недоступен. Попробуйте позже.")
            return
        if location:
            trip_data[chat_id]["end_location"] = {
                "address": end_location,
                "latitude": location.latitude,
                "longitude": location.longitude
            }
            bot.send_message(chat_id, f"Ваше конечное местоположение:\n\n{end_location}")
        else:
            sent = bot.send_message(chat_id, "Не удалось найти местоположение! Пожалуйста, введите корректный адрес")
            bot.register_next_step_handler(sent, process_end_location_step)
            return

    start_coords = (trip_data[chat_id]["start_location"]["latitude"], trip_data[chat_id]["start_location"]["longitude"])
    end_coords = (trip_data[chat_id]["end_location"]["latitude"], trip_data[chat_id]["end_location"]["longitude"])
    distance_km = geodesic(start_coords, end_coords).kilometers

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_auto = types.KeyboardButton("Использовать автоматическое расстояние")
    item_input = types.KeyboardButton("Ввести свое расстояние")
    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item2 = types.KeyboardButton("В главное меню")

    markup.add(item_auto, item_input)
    markup.add(item1)
    markup.add(item2)
    
    sent = bot.send_message(chat_id, "Выберите вариант ввода расстояния:", reply_markup=markup)
    bot.register_next_step_handler(sent, process_distance_choice_step, distance_km)

#@log_user_actions
def process_custom_distance_step(message):
    chat_id = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.", reply_markup=markup)
        bot.register_next_step_handler(sent, process_custom_distance_step)
        return

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    try:
        custom_distance = float(message.text)
        bot.send_message(chat_id, f"Вы ввели свое расстояние: {custom_distance:.2f} км.", reply_markup=markup)
        
        # Переход к выбору даты
        markup_date = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_calendar = types.KeyboardButton("Из календаря")
        item_manual = types.KeyboardButton("Ввести дату вручную")
        item_skip = types.KeyboardButton("Пропустить ввод даты")
        markup_date.add(item_calendar, item_manual, item_skip)
        markup_date.add(item1)
        markup_date.add(item2)

        sent = bot.send_message(chat_id, "Выберите способ ввода даты:", reply_markup=markup_date)
        bot.register_next_step_handler(sent, process_date_step, custom_distance)
    except ValueError:
        sent = bot.send_message(chat_id, "Пожалуйста, введите корректное число для расстояния.", reply_markup=markup)
        bot.register_next_step_handler(sent, process_custom_distance_step)

#@log_user_actions
def process_distance_choice_step(message, distance_km):
    chat_id = message.chat.id
    trip_data[chat_id]["distance"] = distance_km

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return
    
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_auto = types.KeyboardButton("Использовать автоматическое расстояние")
    item_input = types.KeyboardButton("Ввести свое расстояние")
    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item2 = types.KeyboardButton("В главное меню")

    markup.add(item_auto, item_input)
    markup.add(item1)
    markup.add(item2)

    if message.text == "Использовать автоматическое расстояние":
        bot.send_message(chat_id, f"Расстояние между точками: {distance_km:.2f} км.")
        process_date_step(message, distance_km)  # Переходим к выбору даты

    elif message.text == "Ввести свое расстояние":
        custom_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        custom_markup.add(item1)
        custom_markup.add(item2)
        
        sent = bot.send_message(chat_id, "Пожалуйста, введите ваше расстояние в километрах:", reply_markup=custom_markup)
        bot.register_next_step_handler(sent, process_custom_distance_step)
    else:
        sent = bot.send_message(chat_id, "Пожалуйста, выберите один из вариантов.", reply_markup=markup)
        bot.register_next_step_handler(sent, process_distance_choice_step, distance_km)

#@log_user_actions
def process_date_step(message, distance):
    chat_id = message.chat.id
    user_code = trip_data[chat_id].get("user_code", "ru")  # Задаем код по умолчанию

    if message.text == "Пропустить ввод даты":
        selected_date = "Без даты"
        process_selected_date(message, selected_date)
        return

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Из календаря":
        show_calendar(chat_id, user_code)
    elif message.text == "Ввести дату вручную":
        # Создаем разметку с двумя кнопками
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        markup.add(item2)

        # Отправляем сообщение с кнопками и текстом для ввода даты вручную
        sent = bot.send_message(chat_id, "Введите дату поездки:", reply_markup=markup)
        bot.register_next_step_handler(sent, process_manual_date_step, distance)
    else:
        # Повторный запрос способа ввода даты, если ввод неверный
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_calendar = types.KeyboardButton("Из календаря")
        item_manual = types.KeyboardButton("Ввести дату вручную")
        item_skip = types.KeyboardButton("Пропустить ввод даты")
        markup.add(item_calendar, item_manual, item_skip)

        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        markup.add(item2)

        sent = bot.send_message(chat_id, "Выберите способ ввода даты:", reply_markup=markup)
        bot.register_next_step_handler(sent, process_date_step, distance)

#@log_user_actions
def process_date_input_step(message, distance):
    chat_id = message.chat.id
    date_input = message.text.strip()  # Получаем введенную дату

    # Проверка формата даты (ДД.ММ.ГГГГ)
    date_pattern = r"^(0[1-9]|[12][0-9]|3[01])\.(0[1-9]|1[0-2])\.\d{4}$"
    if re.match(date_pattern, date_input):
        selected_date = date_input
        process_selected_date(message, selected_date)  # Переход к следующему шагу с двумя аргументами
    else:
        # Если формат неверный, сообщаем пользователю и запрашиваем ввод снова
        bot.send_message(chat_id, "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, process_date_input_step, distance)

#@log_user_actions
def handle_date_selection(message, distance):
    chat_id = message.chat.id
    user_code = trip_data[chat_id].get("user_code", "ru")  # Задаем код по умолчанию

    if message.text == "Пропустить ввод даты":
        selected_date = "Без даты"
        process_selected_date(message, selected_date)

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Из календаря":
        show_calendar(chat_id, user_code)  # Передаем user_code
    elif message.text == "Ввести дату вручную":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        markup.add(item2)

        sent = bot.send_message(chat_id, "Введите дату поездки:", reply_markup=markup)
        bot.register_next_step_handler(sent, process_manual_date_step, distance)
    elif message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
    elif message.text == "В главное меню":
        return_to_menu(message)
    else:
        sent = bot.send_message(chat_id, "Пожалуйста, выберите корректный вариант.")
        bot.register_next_step_handler(sent, handle_date_selection, distance)

#@log_user_actions
def show_calendar(chat_id, user_code):
    # Inline-календарь с использованием кода пользователя
    calendar, _ = DetailedTelegramCalendar(min_date=date(2000, 1, 1), max_date=date(3000, 12, 31), locale=user_code).build()

    # Обычная клавиатура для кнопок навигации
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    # Отправляем сообщение с навигацией и календарем
    bot.send_message(chat_id, "Календарь:", reply_markup=markup)
    bot.send_message(chat_id, "Выберите дату", reply_markup=calendar)

#@log_user_actions
@bot.callback_query_handler(func=DetailedTelegramCalendar.func())
@restricted
@track_user_activity
def handle_calendar(call):
    result, key, step = DetailedTelegramCalendar(
        min_date=date(2000, 1, 1), max_date=date(3000, 12, 31), unique_key=call.data
    ).process(call.data)

    if not result and key:
        bot.edit_message_text(f"Выберите {step}",
                              call.message.chat.id,
                              call.message.message_id,
                              reply_markup=key)
    elif result:
        selected_date = result.strftime('%d.%m.%Y')
        bot.edit_message_text(f"Вы выбрали дату: {selected_date}",
                              call.message.chat.id,
                              call.message.message_id)

        # Переходим к следующему шагу
        process_selected_date(call.message, selected_date)

#@log_user_actions
def process_selected_date(message, selected_date):
    chat_id = message.chat.id
    distance_km = trip_data[chat_id].get("distance")  # Получаем расстояние

    if distance_km is None:
        bot.send_message(chat_id, "Расстояние не было задано. Пожалуйста, попробуйте снова.")
        return

    show_fuel_types(chat_id, selected_date, distance_km)

#@log_user_actions
def process_manual_date_step(message, distance):
    chat_id = message.chat.id
    date_pattern = r"\d{2}\.\d{2}\.\d{4}"  # Формат ДД.ММ.ГГГГ
    
    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return
    
    # Разметка только с двумя кнопками
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.", reply_markup=markup)
        bot.register_next_step_handler(sent, process_manual_date_step, distance)
        return

    if re.match(date_pattern, message.text):
        day, month, year = map(int, message.text.split('.'))
        if 2000 <= year <= 3000:  # Проверка корректности года
            try:
                datetime(year, month, day)  # Проверка правильности даты
                bot.send_message(chat_id, f"Вы выбрали дату: {message.text}", reply_markup=markup)  # Отправляем сообщение с двумя кнопками
                show_fuel_types(chat_id, message.text, distance)
            except ValueError:
                sent = bot.send_message(chat_id, "Неправильная дата. Пожалуйста, введите корректную дату.", reply_markup=markup)
                bot.register_next_step_handler(sent, process_manual_date_step, distance)
        else:
            sent = bot.send_message(chat_id, "Год должен быть в диапазоне от 2000 г. до 3000 г.", reply_markup=markup)
            bot.register_next_step_handler(sent, process_manual_date_step, distance)
    else:
        sent = bot.send_message(chat_id, "Неправильный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ", reply_markup=markup)
        bot.register_next_step_handler(sent, process_manual_date_step, distance)

#@log_user_actions
def show_fuel_types(chat_id, date, distance):
    # Создаём клавиатуру с типами топлива и дополнительными кнопками
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    
    # Добавляем кнопки с типами топлива
    row1 = [KeyboardButton(fuel_type) for fuel_type in fuel_types[:3]] 
    row2 = [KeyboardButton(fuel_type) for fuel_type in fuel_types[3:]] 
    
    # Добавляем кнопки для возврата в меню и главное меню
    row3 = [KeyboardButton("Вернуться в меню расчета топлива")]
    row4 = [KeyboardButton("В главное меню")]

    markup.add(*row1, *row2, *row3)
    markup.add(*row4)
    
    sent = bot.send_message(chat_id, "Выберите тип топлива:", reply_markup=markup)
    bot.register_next_step_handler(sent, process_fuel_type, date, distance)

#@log_user_actions
def clean_price(price):
    # Удаляем все символы, кроме цифр и точек
    cleaned_price = re.sub(r'[^\d.]', '', price)
    if cleaned_price.count('.') > 1:
        cleaned_price = cleaned_price[:cleaned_price.find('.') + 1] + cleaned_price[cleaned_price.find('.') + 1:].replace('.', '')
    return cleaned_price

fuel_type_mapping = {
    "аи-92": "Аи-92",
    "аи-95": "Аи-95",
    "аи-98": "Аи-98",
    "аи-100": "Аи-100",
    "дт": "ДТ",
    "газ": "Газ СПБТ",
}

#@log_user_actions
def get_average_fuel_price_from_files(fuel_type, directory="data base/azs"):
    fuel_prices = []

    # Приводим fuel_type к нижнему регистру для унификации
    fuel_type = fuel_type.lower()

    # Проверяем, существует ли папка
    if not os.path.exists(directory):
        return None

    # Перебираем все файлы в папке
    for filename in os.listdir(directory):
        if filename.endswith(".json"):
            file_path = os.path.join(directory, filename)

            # Открываем и читаем файл JSON
            try:
                with open(file_path, "r", encoding="utf-8") as file:
                    data = json.load(file)

                    # Перебираем записи в файле и находим нужные по типу топлива
                    for entry in data:
                        if len(entry) == 3:  # Проверка структуры данных
                            company, fuel, price = entry
                            # Приводим fuel к нижнему регистру и сравниваем с введенным пользователем
                            if fuel.lower() == fuel_type:
                                try:
                                    price = float(price)
                                    fuel_prices.append(price)
                                except ValueError:
                                    continue  # Если цена не может быть преобразована в число, пропускаем
            except Exception as e:
                pass  # Игнорируем ошибку при чтении файла

    # Если есть собранные цены, возвращаем среднее значение
    if fuel_prices:
        average_price = sum(fuel_prices) / len(fuel_prices)
        return average_price
    else:
        return None

#@log_user_actions
def get_average_fuel_prices(city_code='default_city_code'):
    url = f'https://azsprice.ru/benzin-{city_code}'  # Замените на подходящий URL с использованием city_code
    
    try:
        response = requests.get(url)
        response.raise_for_status()  # Проверка на успешный статус ответа
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        table = soup.find('table')
        if not table:
            raise ValueError("Не найдена таблица с ценами")
        
        fuel_prices = {}
        rows = table.find_all('tr')
        
        for row in rows[1:]:  # Пропускаем заголовок таблицы
            columns = row.find_all('td')
            if len(columns) < 5:
                continue
            
            fuel_type = columns[2].text.strip()  # Марка топлива
            today_price = clean_price(columns[3].text.strip())  # Цена на топливо
            
            if today_price:
                try:
                    price = float(today_price)
                    if fuel_type not in fuel_prices:
                        fuel_prices[fuel_type] = []
                    fuel_prices[fuel_type].append(price)
                except ValueError:
                    pass  # Игнорируем ошибки преобразования цены

        average_prices = {fuel: sum(prices) / len(prices) for fuel, prices in fuel_prices.items()}
        return average_prices
    
    except (requests.RequestException, ValueError) as e:
        return None  # Вернем None в случае ошибки

#@log_user_actions
def process_fuel_type(message, date, distance):
    if message is None:
        return  # Проверка на None

    chat_id = message.chat.id

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверка на наличие мультимедийных элементов
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.voice or message.video_note:
        bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        # Зарегистрируем следующий шаг для ожидания текстового сообщения
        sent = bot.send_message(chat_id, "Выберите тип топлива:")
        bot.register_next_step_handler(sent, process_fuel_type, date, distance)
        return

    fuel_type = message.text.strip().lower() if message.text else ""

    # Проверка на допустимые типы топлива
    fuel_type_mapping = {
        "аи-92": "аи-92",
        "аи-95": "аи-95",
        "аи-98": "аи-98",
        "аи-100": "аи-100",
        "дт": "дт",
        "газ": "газ спбт",
    }

    # Проверка на допустимые значения
    if fuel_type not in fuel_type_mapping:
        sent = bot.send_message(chat_id, "Пожалуйста, выберите тип топлива только из предложенных вариантов.")
        bot.register_next_step_handler(sent, process_fuel_type, date, distance)
        return

    actual_fuel_type = fuel_type_mapping[fuel_type]

    # Дальнейшая логика обработки
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Использовать актуальную цену")
    item2 = types.KeyboardButton("Ввести свою цену")
    item3 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item4 = types.KeyboardButton("В главное меню")
    markup.add(item1, item2)
    markup.add(item3)
    markup.add(item4)

    sent = bot.send_message(chat_id, "Выберите вариант ввода цены топлива:", reply_markup=markup)
    bot.register_next_step_handler(sent, handle_price_input_choice, date, distance, actual_fuel_type)

#@log_user_actions
def handle_price_input_choice(message, date, distance, fuel_type):
    chat_id = message.chat.id

    if message.text == "Ввести свою цену":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        markup.add(item2)

        sent = bot.send_message(chat_id, "Пожалуйста, введите цену за литр топлива:", reply_markup=markup)
        bot.register_next_step_handler(sent, process_price_per_liter_step, date, distance, fuel_type)
    
    elif message.text == "Использовать актуальную цену":
        # Пытаемся получить цену с сайта
        fuel_prices = get_average_fuel_prices(city_code="cheboksary")
        
        if fuel_prices:  # Если сайт доступен и данные получены
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
            item2 = types.KeyboardButton("В главное меню")
            markup.add(item1)
            markup.add(item2)
            
            if fuel_type.lower() in fuel_prices:  # Если выбранный тип топлива существует
                price = fuel_prices[fuel_type.lower()]
                bot.send_message(chat_id, f"Актуальная средняя цена на {fuel_type.upper()} по РФ: {price:.2f} руб./л.", reply_markup=markup)
                sent = bot.send_message(chat_id, "Введите расход топлива на 100 км:", reply_markup=markup)
                bot.register_next_step_handler(sent, process_fuel_consumption_step, date, distance, fuel_type, price)
            else:
                # Если для выбранного топлива нет данных на сайте, сразу переходим к проверке файлов
                price_from_files = get_average_fuel_price_from_files(fuel_type, directory="data base/azs")
                
                if price_from_files:  # Если цена найдена в файлах
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
                    item2 = types.KeyboardButton("В главное меню")
                    markup.add(item1)
                    markup.add(item2)
                    
                    bot.send_message(chat_id, f"Актуальная средняя цена на {fuel_type.upper()} по РФ: {price_from_files:.2f} руб./л.", reply_markup=markup)
                    sent = bot.send_message(chat_id, "Введите расход топлива на 100 км:", reply_markup=markup)
                    bot.register_next_step_handler(sent, process_fuel_consumption_step, date, distance, fuel_type, price_from_files)
                else:
                    # Если нет данных в файлах, запрашиваем цену вручную
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
                    item2 = types.KeyboardButton("В главное меню")
                    markup.add(item1)
                    markup.add(item2)
                    sent = bot.send_message(chat_id, f"Для выбранного топлива '{fuel_type}' данных нет. Пожалуйста, введите цену", reply_markup=markup)
                    bot.register_next_step_handler(sent, process_price_per_liter_step, date, distance, fuel_type)
        
        else:  # Если сайт недоступен
            # Получаем цены из файлов
            price_from_files = get_average_fuel_price_from_files(fuel_type, directory="data base/azs")
            
            if price_from_files:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
                item2 = types.KeyboardButton("В главное меню")
                markup.add(item1)
                markup.add(item2)
                
                bot.send_message(chat_id, f"Актуальная средняя цена на {fuel_type.upper()} по РФ: {price_from_files:.2f} руб./л.", reply_markup=markup)
                sent = bot.send_message(chat_id, "Введите расход топлива на 100 км:", reply_markup=markup)
                bot.register_next_step_handler(sent, process_fuel_consumption_step, date, distance, fuel_type, price_from_files)
            else:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
                item2 = types.KeyboardButton("В главное меню")
                markup.add(item1)
                markup.add(item2)
                sent = bot.send_message(chat_id, f"Для выбранного топлива '{fuel_type}' данных нет. Пожалуйста, введите цену", reply_markup=markup)
                bot.register_next_step_handler(sent, process_price_per_liter_step, date, distance, fuel_type)
    
    elif message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
    elif message.text == "В главное меню":
        return_to_menu(message)
    else:
        sent = bot.send_message(chat_id, "Пожалуйста, выберите один из предложенных вариантов")
        bot.register_next_step_handler(sent, handle_price_input_choice, date, distance, fuel_type)

#@log_user_actions
def process_price_per_liter_step(message, date, distance, fuel_type):
    chat_id = message.chat.id

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return    
    if message.text == "В главное меню":
        return_to_menu(message) 
        return    

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_price_per_liter_step, date, distance, fuel_type)
        return

    input_text = message.text.replace(',', '.')
    
    try:
        price_per_liter = float(input_text)
        if price_per_liter <= 0:
            raise ValueError
        
        # Создаем новую клавиатуру с двумя кнопками
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        markup.add(item2)

        # Запрашиваем расход топлива
        sent = bot.send_message(chat_id, "Введите расход топлива на 100 км:", reply_markup=markup)
        bot.clear_step_handler_by_chat_id(chat_id)  # Очистка предыдущего хендлера
        bot.register_next_step_handler(sent, process_fuel_consumption_step, date, distance, fuel_type, price_per_liter)
        
    except ValueError:
        sent = bot.send_message(chat_id, "Пожалуйста, введите положительное число для цены топлива за литр:")
        bot.register_next_step_handler(sent, process_price_per_liter_step, date, distance, fuel_type)

# (9.11) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (ФУНКЦИЯ ДЛЯ ПАССАЖИРОВ ) ---------------

#@log_user_actions
def process_fuel_consumption_step(message, date, distance, fuel_type, price_per_liter):
    chat_id = message.chat.id
    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return
    if message.text == "В главное меню":
        return_to_menu(message) 
        return
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_fuel_consumption_step, date, distance, fuel_type, price_per_liter)
        return
    input_text = message.text.replace(',', '.') 
    try:
        fuel_consumption = float(input_text)
        if fuel_consumption <= 0:
            raise ValueError
        sent = bot.send_message(chat_id, "Введите количество пассажиров в машине:")
        bot.clear_step_handler_by_chat_id(chat_id) 
        bot.register_next_step_handler(sent, process_passengers_step, date, distance, fuel_type, price_per_liter, fuel_consumption)
    except ValueError:
        sent = bot.send_message(chat_id, "Пожалуйста, введите положительное число для расхода топлива на 100 км:")
        bot.register_next_step_handler(sent, process_fuel_consumption_step, date, distance, fuel_type, price_per_liter)

# (9.12) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (ФУНКЦИЯ ДЛЯ РАСЧЕТА) ---------------

#@log_user_actions
def process_passengers_step(message, date, distance, fuel_type, price_per_liter, fuel_consumption):
    chat_id = message.chat.id
    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(chat_id)
        return
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(chat_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_passengers_step, date, distance, fuel_type, price_per_liter, fuel_consumption)
        return
    try:
        passengers = int(message.text)
        if passengers <= 0:
            raise ValueError

        # Считаем стоимость топлива
        fuel_cost = (distance / 100) * fuel_consumption * price_per_liter
        fuel_cost_per_person = fuel_cost / passengers

        # Получаем координаты начального и конечного местоположений
        start_location = trip_data[chat_id]['start_location']
        end_location = trip_data[chat_id]['end_location']
        
        # Сформируем ссылку на Яндекс.Карты для автомобилиста
        yandex_maps_url = f"https://yandex.ru/maps/?rtext={start_location['latitude']},{start_location['longitude']}~{end_location['latitude']},{end_location['longitude']}&rtt=auto"
        
        # Отправляем запрос на clck.ru для сокращения ссылки
        try:
            response = requests.get(f'https://clck.ru/--?url={yandex_maps_url}')
            short_url = response.text
        except Exception as e:
            bot.send_message(chat_id, f"Не удалось сократить ссылку: {str(e)}")
            short_url = yandex_maps_url  # Используем оригинальную ссылку, если сокращение не удалось
        
        if chat_id not in temporary_trip_data:
            temporary_trip_data[chat_id] = []
        
        # Добавляем временные данные поездки
        temporary_trip_data[chat_id].append({
            "start_location": start_location,
            "end_location": end_location,
            "date": date,
            "distance": distance,
            "fuel_type": fuel_type,
            "price_per_liter": price_per_liter,
            "fuel_consumption": fuel_consumption,
            "passengers": passengers,
            "fuel_spent": (distance / 100) * fuel_consumption,
            "fuel_cost": fuel_cost,
            "fuel_cost_per_person": fuel_cost_per_person,
            "route_link": short_url  # Сохраняем сокращенную ссылку в временные данные
        })

        display_summary(chat_id, fuel_cost, fuel_cost_per_person, fuel_type, date, distance, price_per_liter, fuel_consumption, passengers)
    except ValueError:
        sent = bot.send_message(chat_id, "Пожалуйста, введите положительное целое число для количества пассажиров:")
        bot.register_next_step_handler(sent, process_passengers_step, date, distance, fuel_type, price_per_liter, fuel_consumption)

# (9.13) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (ИТОГОВАЯ ИНФОРМАЦИЯ) ---------------

#@log_user_actions
def display_summary(chat_id, fuel_cost, fuel_cost_per_person, fuel_type, date, distance, price_per_liter, fuel_consumption, passengers):
    fuel_spent = (distance / 100) * fuel_consumption 

    # Получаем координаты начального и конечного местоположений
    start_location = trip_data[chat_id]['start_location']
    end_location = trip_data[chat_id]['end_location']
    
    # Сформируем ссылку на Яндекс.Карты для автомобилиста
    yandex_maps_url = f"https://yandex.ru/maps/?rtext={start_location['latitude']},{start_location['longitude']}~{end_location['latitude']},{end_location['longitude']}&rtt=auto"
    
    # Отправляем запрос на clck.ru для сокращения ссылки
    try:
        response = requests.get(f'https://clck.ru/--?url={yandex_maps_url}')
        short_url = response.text
    except Exception as e:
        bot.send_message(chat_id, f"Не удалось сократить ссылку: {str(e)}")
        short_url = yandex_maps_url  # Используем оригинальную ссылку, если сокращение не удалось

    # Формируем итоговое сообщение
    summary_message = "🚗 *ИНФОРМАЦИЯ О ПОЕЗДКЕ* 🚗\n"
    summary_message += "-------------------------------------------------------------\n"
    summary_message += f"📍 *Начальное местоположение:*\n{start_location['address']}\n"
    summary_message += f"🏁 *Конечное местоположение:*\n{end_location['address']}\n"
    summary_message += f"🗓️ *Дата поездки:* {date}\n"
    summary_message += f"📏 *Расстояние:* {distance:.2f} км\n"
    summary_message += f"⛽ *Тип топлива:* {fuel_type}\n"
    summary_message += f"💵 *Цена топлива за литр:* {price_per_liter:.2f} руб.\n"
    summary_message += f"⚙️ *Расход топлива на 100 км:* {fuel_consumption} л.\n"
    summary_message += f"👥 *Количество пассажиров:* {passengers}\n"
    summary_message += "-------------------------------------------------------------\n"
    summary_message += f"🛢️ *ПОТРАЧЕНО ЛИТРОВ ТОПЛИВА:* {fuel_spent:.2f} л.\n"
    summary_message += f"💰 *СТОИМОСТЬ ТОПЛИВА ДЛЯ ПОЕЗДКИ:* {fuel_cost:.2f} руб.\n"
    summary_message += f"👤 *СТОИМОСТЬ ТОПЛИВА НА ЧЕЛОВЕКА:* {fuel_cost_per_person:.2f} руб.\n"
    summary_message += f"[ССЫЛКА НА МАРШРУТ]({short_url})\n"

    summary_message = summary_message.replace('\n', '\n\n')
    bot.clear_step_handler_by_chat_id(chat_id)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Сохранить поездку")
    item2 = types.KeyboardButton("Вернуться в меню расчета топлива")
    item3 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)
    markup.add(item3)

    bot.send_message(chat_id, summary_message, reply_markup=markup, parse_mode="Markdown")

#@log_user_actions
def update_excel_file(user_id):
    # Путь к папке с файлами Excel
    folder_path = "data base/trip/excel"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    file_path = os.path.join(folder_path, f"{user_id}_trips.xlsx")

    # Проверяем и создаем файл с заголовками
    if not os.path.exists(file_path):
        df = pd.DataFrame(columns=[
            "Дата", "Начальное местоположение", "Конечное местоположение",
            "Расстояние (км)", "Тип топлива", "Цена топлива (руб/л)",
            "Расход топлива (л/100 км)", "Количество пассажиров",
            "Потрачено литров", "Стоимость топлива (руб)", 
            "Стоимость на человека (руб)", "Ссылка на маршрут"
        ])
        df.to_excel(file_path, index=False)
    
    # Обновляем данные файла Excel
    df = pd.read_excel(file_path)
    trips = user_trip_data.get(user_id, [])
    trip_records = [
        [
            trip['start_location']['address'], trip['end_location']['address'], trip['date'],
            trip['distance'], trip['fuel_type'], trip['price_per_liter'], trip['fuel_consumption'], 
            trip['passengers'], trip['fuel_spent'], trip['fuel_cost'], trip['fuel_cost_per_person'], 
            trip.get('route_link', "Нет ссылки")
        ] for trip in trips
    ]
    df = pd.DataFrame(trip_records, columns=df.columns)
    df.to_excel(file_path, index=False)

    # Устанавливаем стилизацию Excel
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
        worksheet.column_dimensions[column[0].column_letter].width = max_length
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), 
                          top=Side(style='thick'), bottom=Side(style='thick'))
    for row in worksheet.iter_rows(min_row=2, min_col=9, max_col=12):
        for cell in row:
            cell.border = thick_border
    workbook.save(file_path)


import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Alignment, Border, Side

import pandas as pd
import os
from openpyxl import load_workbook

def save_trip_to_excel(user_id, trip):
    # Путь к папке с файлами Excel
    directory = "data base/trip/excel"
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_path = os.path.join(directory, f"{user_id}_trips.xlsx")

    # Создаем и обновляем файл Excel для конкретной поездки
    new_trip_data = {
        "Начальное местоположение": trip['start_location']['address'],
        "Конечное местоположение": trip['end_location']['address'],
        "Дата": trip['date'],
        "Расстояние (км)": round(trip.get('distance', None), 2),
        "Тип топлива": trip.get('fuel_type', None),
        "Цена топлива (руб/л)": round(trip.get('price_per_liter', None), 2),
        "Расход топлива (л/100 км)": round(trip.get('fuel_consumption', None), 2),
        "Количество пассажиров": trip.get('passengers', None),
        "Потрачено литров": round(trip.get('fuel_spent', None), 2),
        "Стоимость топлива (руб)": round(trip.get('fuel_cost', None), 2),
        "Стоимость на человека (руб)": round(trip.get('fuel_cost_per_person', None), 2),
        "Ссылка на маршрут": trip.get('route_link', None)
    }
    new_trip_df = pd.DataFrame([new_trip_data])

    if os.path.exists(file_path):
        existing_data = pd.read_excel(file_path).dropna(axis=1, how='all')
        updated_data = pd.concat([existing_data, new_trip_df], ignore_index=True)
    else:
        updated_data = new_trip_df

    updated_data.to_excel(file_path, index=False)
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    for column in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value) + 2
        worksheet.column_dimensions[column[0].column_letter].width = max_length
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                          top=Side(style='thick'), bottom=Side(style='thick'))
    for row in worksheet.iter_rows(min_col=worksheet.max_column-3, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thick_border
    workbook.save(file_path)



# (9.14) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (КОМАНДА "СОХРАНИТЬ ПОЕЗДКУ") ---------------

@bot.message_handler(func=lambda message: message.text == "Сохранить поездку")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Сохранить поездку')
@track_usage('Сохранить поездку')  # Добавление отслеживания статистики
def save_data_handler(message):
    user_id = message.chat.id
    if user_id in temporary_trip_data and temporary_trip_data[user_id]:
        if user_id not in user_trip_data:
            user_trip_data[user_id] = []

        # Переносим данные из временных данных в постоянные
        user_trip_data[user_id].extend(temporary_trip_data[user_id])

        # Получаем последнюю поездку
        last_trip = user_trip_data[user_id][-1]

        # Сохраняем данные поездки в базу данных
        save_trip_data(user_id) 
        
        # Сохраняем последнюю поездку в Excel
        save_trip_to_excel(user_id, last_trip)

        # Очищаем временные данные
        temporary_trip_data[user_id] = []
        
        # Создаем клавиатуру для возврата в меню
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Вернуться в меню расчета топлива")
        item2 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        markup.add(item2) 
        
        bot.send_message(user_id, "Данные поездки успешно сохранены!", reply_markup=markup)

# (9.15) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (КОМАНДА "В ГЛАВНОЕ МЕНЮ  ВРЕМЕННЫХ ДАННЫХ") ---------------

@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('В главное меню')
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == "В главное меню")
@bot.message_handler(commands=['mainmenu'])
def return_to_menu(message):
    user_id = message.chat.id
    if user_id in temporary_trip_data:
        temporary_trip_data[user_id] = []
    start(message)

# # (9.16) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (КОМАНДА "ВЕРНУТЬСЯ В МЕНЮ РАСЧЕТА ТОПЛИВА   ВРЕМЕННЫХ ДАННЫХ") ---------------

@bot.message_handler(func=lambda message: message.text == "Вернуться в меню расчета топлива")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Вернуться в меню расчета топлива')
#@log_user_actions
def restart_handler(message):
    user_id = message.chat.id

    # Убедитесь, что при возвращении в меню расчета топлива, данные не удаляются
    # Например, сохраняем данные перед сбросом, если это необходимо
    if user_id in user_trip_data:
        # Сохранить данные, прежде чем сбросить, если это требуется
        save_trip_data(user_id, user_trip_data[user_id])

    # Теперь вызываем reset, если это действительно нужно
    reset_and_start_over(user_id)

    # Загружаем поездки (не сбрасывая их)
    user_trip_data[user_id] = load_trip_data(user_id)

    # Возвращаем пользователя в меню расчета топлива
    reset_and_start_over(message)

# (9.17) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (КОМАНДА "ПОСМОТРЕТЬ ПОЕЗДКИ") ---------------

@bot.message_handler(func=lambda message: message.text == "Посмотреть поездки")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть поездки')
@track_usage('Посмотреть поездки')  # Добавление отслеживания статистики
#@log_user_actions
def view_trips(message):
    user_id = message.chat.id
    trips = load_trip_data(user_id)  # Загрузим поездки из базы данных

    if trips:
        # Создаем кнопки для выбора поездок
        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        buttons = []

        for i, trip in enumerate(trips, start=1):
            start_address = trip['start_location']['address']
            end_address = trip['end_location']['address']
            date = trip['date'] if trip['date'] != "Без даты" else "Без даты"  # Обрабатываем специальное значение
            button_text = f"№{i}. {date}"
            buttons.append(types.KeyboardButton(button_text))

            # Разделяем кнопки на несколько рядов, чтобы было удобнее
            if len(buttons) == 3 or i == len(trips):
                markup.row(*buttons)
                buttons = []  # Очищаем список для следующего ряда

        # Добавляем дополнительные кнопки
        markup.add("Посмотреть в Excel")
        markup.add("Вернуться в меню расчета топлива")
        markup.add("В главное меню")

        # Отправляем сообщение с кнопками
        bot.send_message(user_id, "Выберите поездку для просмотра:", reply_markup=markup)
    else:
        bot.send_message(user_id, "У вас нет сохраненных поездок!")

@bot.message_handler(func=lambda message: message.text == "Посмотреть в Excel")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть в Excel')
@track_usage('Посмотреть в Excel')  # Добавление отслеживания статистики
#@log_user_actions
def send_excel_file(message):
    user_id = message.chat.id
    excel_file_path = f"data base/trip/excel/{user_id}_trips.xlsx"

    if os.path.exists(excel_file_path):
        with open(excel_file_path, 'rb') as excel_file:
            bot.send_document(user_id, excel_file)
    else:
        bot.send_message(user_id, "Файл Excel не найден. Убедитесь, что у вас есть сохраненные поездки.")

@bot.message_handler(func=lambda message: message.text and re.match(r"№\d+\.\s*\d{2}\.\d{2}\.\d{4}|№\d+\.\s*Без даты", message.text))
@restricted
@track_user_activity
@check_chat_state
#@log_user_actions
def show_trip_details(message):
    user_id = message.chat.id
    trips = load_trip_data(user_id)  # Загружаем поездки для пользователя

    try:
        # Извлекаем номер поездки из сообщения
        match = re.match(r"№(\d+)\.\s*(\d{2}\.\d{2}\.\d{4}|Без даты)", message.text)
        if match:
            trip_index = int(match.group(1)) - 1  # Получаем индекс поездки
            if 0 <= trip_index < len(trips):  # Проверка на корректность индекса
                trip = trips[trip_index]

                # Формируем сообщение с данными поездки
                start_address = trip['start_location']['address']
                end_address = trip['end_location']['address']
                date = trip['date'] if trip['date'] != "Без даты" else "Без даты"  # Обработка "Без даты"
                summary_message = f"*ИТОГОВЫЕ ДАННЫЕ ПОЕЗДКИ* *{trip_index + 1}* \n\n"
                summary_message += "-------------------------------------------------------------\n\n"
                summary_message += f"📍 *Начальное местоположение:*\n\n{start_address}\n\n"
                summary_message += f"🏁 *Конечное местоположение:*\n\n{end_address}\n\n"
                summary_message += f"🗓️ *Дата поездки:* {date}\n\n"
                summary_message += f"📏 *Расстояние:* {trip['distance']:.2f} км.\n\n"
                summary_message += f"⛽ *Тип топлива:* {trip['fuel_type']}\n\n"
                summary_message += f"💵 *Цена топлива за литр:* {trip['price_per_liter']:.2f} руб.\n\n"
                summary_message += f"⚙️ *Расход топлива на 100 км:* {trip['fuel_consumption']} л.\n\n"
                summary_message += f"👥 *Количество пассажиров:* {trip['passengers']}\n\n"
                summary_message += "-------------------------------------------------------------\n\n"
                summary_message += f"🛢️ *ПОТРАЧЕНО ЛИТРОВ ТОПЛИВА:* {trip['fuel_spent']:.2f} л.\n\n"
                summary_message += f"💰 *СТОИМОСТЬ ТОПЛИВА ДЛЯ ПОЕЗДКИ:* {trip['fuel_cost']:.2f} руб.\n\n"
                summary_message += f"👤 *СТОИМОСТЬ ТОПЛИВА НА ЧЕЛОВЕКА:* {trip['fuel_cost_per_person']:.2f} руб.\n\n"

                # Проверяем, есть ли 'route_link' в данных поездки
                if 'route_link' in trip:
                    summary_message += f"[ССЫЛКА НА МАРШРУТ]({trip['route_link']})\n\n"
                else:
                    summary_message += "Ссылка на маршрут недоступна.\n\n"

                # Отправляем подробную информацию о поездке
                bot.send_message(user_id, summary_message, parse_mode="Markdown")

                # Оставляем клавиатуру с кнопками после просмотра поездки
                markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
                markup.add("Посмотреть другие поездки")
                markup.add("Вернуться в меню расчета топлива")
                markup.add("В главное меню")

                bot.send_message(user_id, "Вы можете посмотреть другие поездки", reply_markup=markup)

            else:
                bot.send_message(user_id, "Поездка с таким номером не найдена! Попробуйте снова")
        else:
            bot.send_message(user_id, "Ошибка при выборе поездки! Попробуйте снова")

    except (IndexError, ValueError) as e:
        bot.send_message(user_id, "Ошибка при обработке данных! Попробуйте снова")

# Обработчик для кнопки "Посмотреть другие поездки"
@bot.message_handler(func=lambda message: message.text == "Посмотреть другие поездки")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть другие поездки')

#@log_user_actions
def view_other_trips(message):
    view_trips(message)  # Вызываем функцию для повторного отображения списка поездок

# # Обработчики для кнопок "Вернуться в меню расчета топлива" и "В главное меню"
# @bot.message_handler(func=lambda message: message.text == "Вернуться в меню расчета топлива")
# @restricted
# @track_user_activity
# @check_chat_state
# @check_function_state_decorator('Вернуться в меню расчета топлива')

# #@log_user_actions
# def return_to_fuel_calc_menu(message):
#     chat_id = message.chat.id
#     reset_and_start_over(chat_id)  # Ваша функция для сброса и возвращения в меню расчета топлива
#     bot.send_message(chat_id, "Вы вернулись в меню расчета топлива", reply_markup=types.ReplyKeyboardRemove())

# (9.18) --------------- КОД ДЛЯ "РАСХОД ТОПЛИВА" (КОМАНДА "УДАЛИТЬ ПОЕЗДКУ") ---------------

@bot.message_handler(func=lambda message: message.text == "Удалить поездку")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить поездку')
@track_usage('Удалить поездку')  # Добавление отслеживания статистики
#@log_user_actions
def ask_for_trip_to_delete(message):
    user_id = message.chat.id

    if user_id in user_trip_data:
        if user_trip_data[user_id]:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            buttons = []

            # Изменяем создание кнопок для отображения номера и даты поездки
            for i, trip in enumerate(user_trip_data[user_id], start=1):
                trip_date = trip.get('date', 'Дата не указана')  # Получаем дату поездки
                button_text = f"№{i}. {trip_date}"  # Создаем текст кнопки с номером и датой
                buttons.append(types.KeyboardButton(button_text))

                # Добавляем кнопку в ряд по 3
                if len(buttons) == 3 or i == len(user_trip_data[user_id]):
                    markup.row(*buttons)
                    buttons = []  # Очистить список для следующего ряда

            # Добавляем дополнительные кнопки
            markup.add(types.KeyboardButton("Удалить все поездки"))
            markup.add(types.KeyboardButton("Вернуться в меню расчета топлива"))
            markup.add(types.KeyboardButton("В главное меню"))

            bot.send_message(user_id, "Выберите номер поездки для удаления:", reply_markup=markup)
            bot.register_next_step_handler(message, confirm_trip_deletion)
        else:
            bot.send_message(user_id, "У вас нет поездок для удаления!")
    else:
        bot.send_message(user_id, "У вас нет сохраненных поездок!")

#@log_user_actions
def confirm_trip_deletion(message):
    user_id = message.chat.id

    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(user_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Удалить все поездки":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.add(types.KeyboardButton("Вернуться в меню расчета топлива"))
        markup.add(types.KeyboardButton("В главное меню"))
        
        bot.send_message(
            user_id,
            "*Вы уверены, что хотите удалить все поездки?*\n\nПожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены",
            parse_mode="Markdown",
            reply_markup=markup
        )
        bot.register_next_step_handler(message, confirm_delete_all)
        return

    # Проверяем, соответствует ли текст кнопки формату "№N. дата"
    if message.text.startswith("№") and "." in message.text:
        try:
            trip_number = int(message.text.split(".")[0][1:])  # Извлекаем номер поездки
            if 1 <= trip_number <= len(user_trip_data[user_id]):
                deleted_trip = user_trip_data[user_id].pop(trip_number - 1)
                bot.send_message(user_id, f"Поездка *№.{trip_number}* успешно удалена!", parse_mode="Markdown")
                
                # Обновляем Excel файл
                update_excel_file(user_id)

            else:
                bot.send_message(user_id, "Неверный номер поездки. Пожалуйста, укажите корректный номер")
        except ValueError:
            bot.send_message(user_id, "Произошла ошибка при обработке номера поездки")
    else:
        bot.send_message(user_id, "Пожалуйста, выберите номер поездки для удаления с помощью кнопок")

    reset_and_start_over(user_id)

# Функция для подтверждения удаления всех поездок
#@log_user_actions
def confirm_delete_all(message):
    user_id = message.chat.id
    
    if message.text is None:
        bot.send_message(user_id, "Пожалуйста, отправьте текстовое сообщение")
        bot.register_next_step_handler(message, confirm_delete_all)
        return

    # Проверяем нажатие на кнопки "Вернуться в меню расчета топлива" и "В главное меню"
    if message.text == "Вернуться в меню расчета топлива":
        reset_and_start_over(user_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    user_response = message.text.lower()

    if user_response == "да":
        if user_id in user_trip_data and user_trip_data[user_id]:
            user_trip_data[user_id].clear()
            bot.send_message(user_id, "Все поездки были успешно удалены")
            # Очистка Excel файла
            update_excel_file(user_id)
        else:
            bot.send_message(user_id, "У вас нет поездок для удаления!")
            reset_and_start_over(user_id)
        
        # Очищаем Excel файл, оставляя только заголовки
        excel_file = os.path.join('data base', f"{user_id}_trips.xlsx")
        if os.path.exists(excel_file):
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
            workbook.save(excel_file)

    elif user_response == "нет":
        bot.send_message(user_id, "Удаление всех поездок отменено!")
        reset_and_start_over(user_id)
    else:
        bot.send_message(user_id, "Пожалуйста, ответьте *ДА* для подтверждения или *НЕТ* для отмены", parse_mode="Markdown")
        bot.register_next_step_handler(message, confirm_delete_all)

# (10) --------------- КОД ДЛЯ "ТРАТ" ---------------

# (10.1) --------------- КОД ДЛЯ "ТРАТ" (ОБРАБОТЧИК "ТРАТЫ И РЕМОНТЫ") ---------------

@bot.message_handler(func=lambda message: message.text == "Траты и ремонты")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Траты и ремонты')
@track_usage('Траты и ремонты')  # Добавление отслеживания статистики
#@log_user_actions
def handle_expenses_and_repairs(message):

    user_id = message.from_user.id

    expense_data = load_expense_data(user_id).get(str(user_id), {})
    repair_data = load_repair_data(user_id).get(str(user_id), {})

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Записать трату")
    item2 = types.KeyboardButton("Посмотреть траты")
    item3 = types.KeyboardButton("Записать ремонт")
    item4 = types.KeyboardButton("Посмотреть ремонты")
    item5 = types.KeyboardButton("Удалить траты")
    item6 = types.KeyboardButton("Удалить ремонты")
    item7 = types.KeyboardButton("В главное меню")
    item8 = types.KeyboardButton("Ваш транспорт")

    markup.add(item8)
    markup.add(item1, item3)
    markup.add(item2, item4)
    markup.add(item5, item6)
    markup.add(item7)

    bot.clear_step_handler_by_chat_id(user_id)
    bot.send_message(user_id, "Меню для учета трат и ремонтов. Выберите действие:", reply_markup=markup)

# (10.3) --------------- КОД ДЛЯ "ТРАТ" (ПРОВЕРКА НА МУЛЬТИМЕДИЮ) ---------------

#@log_user_actions
def contains_media(message):
    return (message.photo or message.video or message.document or message.animation or
            message.sticker or message.location or message.audio or message.contact or
            message.voice or message.video_note)

# (10.4) --------------- КОД ДЛЯ "ТРАТ" (ВОЗВРАТ В МЕНЮ) ---------------

#@log_user_actions
def send_menu(user_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Записать трату")
    item2 = types.KeyboardButton("Посмотреть траты")
    item3 = types.KeyboardButton("Записать ремонт")
    item4 = types.KeyboardButton("Посмотреть ремонты")
    item5 = types.KeyboardButton("Удалить траты")
    item6 = types.KeyboardButton("Удалить ремонты")
    item7 = types.KeyboardButton("В главное меню")
    item8 = types.KeyboardButton("Ваш транспорт")
    
    markup.add(item8)
    markup.add(item1, item3)
    markup.add(item2, item4)
    markup.add(item5, item6)
    markup.add(item7)

    bot.send_message(user_id, "Вы вернулись в меню. Выберите действие:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == "Вернуться в меню трат и ремонтов")
@bot.message_handler(commands=['restart2'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Вернуться в меню трат и ремонтов')
#@log_user_actions
def return_to_menu_2(message):
    user_id = message.from_user.id
    send_menu(user_id)

# (10.5) --------------- КОД ДЛЯ "ТРАТ" (ОБРАБОТЧИК "ЗАПИСАТЬ ТРАТУ") ---------------

# Обработка данных о расходах
def save_expense_data(user_id, user_data, selected_transport=None):
    folder_path = os.path.join("data base", "expense")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    file_path = os.path.join(folder_path, f"{user_id}_expenses.json")

    # Загружаем текущие данные
    current_data = load_expense_data(user_id)

    # Сохраняем категории и расходы, не трогая другие данные
    user_data["user_categories"] = user_data.get("user_categories", current_data.get("user_categories", []))
    user_data["expenses"] = current_data.get("expenses", [])

    # Сохраняем только, если новый транспорт был передан
    if selected_transport is not None:
        user_data["selected_transport"] = selected_transport

    # Запись данных в файл
    with open(file_path, "w", encoding="utf-8") as file:
        json.dump(user_data, file, ensure_ascii=False, indent=4)

def load_expense_data(user_id):
    folder_path = os.path.join("data base", "expense")
    file_path = os.path.join(folder_path, f"{user_id}_expenses.json")
    
    if not os.path.exists(file_path):
        return {"user_categories": [], "selected_transport": "", "expenses": []}

    try:
        with open(file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
    except (FileNotFoundError, ValueError) as e:
        data = {"user_categories": [], "selected_transport": "", "expenses": []}
    
    return data

#@log_user_actions
def get_user_categories(user_id):
    data = load_expense_data(user_id)
    default_categories = ["Без категории", "АЗС", "Мойка", "Парковка", "Платная дорога", "Страховка", "Штрафы"]
    user_categories = data.get("user_categories", [])
    return default_categories + user_categories

#@log_user_actions
def add_user_category(user_id, new_category):
    data = load_expense_data(user_id)
    
    # Проверяем, есть ли категория, и добавляем её
    if "user_categories" not in data:
        data["user_categories"] = []
    if new_category not in data["user_categories"]:
        data["user_categories"].append(new_category)

    # Сохраняем обновленные данные
    save_expense_data(user_id, data)  # Теперь данные категорий будут сохранены

#@log_user_actions
def remove_user_category(user_id, category_to_remove, selected_transport=""):
    data = load_expense_data(user_id)
    if "user_categories" in data and category_to_remove in data["user_categories"]:
        data["user_categories"].remove(category_to_remove)
        save_expense_data(user_id, data, selected_transport)

#@log_user_actions
def get_user_transport_keyboard(user_id):
    transports = user_transport.get(str(user_id), [])  # Приведение к строке для совместимости
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    # Добавляем кнопки транспорта
    for i in range(0, len(transports), 2):
        transport_buttons = [
            types.KeyboardButton(f"{transport['brand']} {transport['model']} ({transport['license_plate']})")
            for transport in transports[i:i+2]
        ]
        markup.row(*transport_buttons)

    markup.add(types.KeyboardButton("Добавить транспорт"))
    return markup

# Основная функция для записи траты
@bot.message_handler(func=lambda message: message.text == "Записать трату")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Записать трату')
@track_usage('Записать трату')  # Добавление отслеживания статистики
#@log_user_actions
def record_expense(message):
    user_id = message.from_user.id

    # Обработка мультимедийных файлов
    if contains_media(message):
        sent = bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, record_expense)
        return

    markup = get_user_transport_keyboard(str(user_id))
    markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
    markup.add(types.KeyboardButton("В главное меню"))
    
    bot.send_message(user_id, "Выберите транспорт для записи траты:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_transport_selection_for_expense)

#@log_user_actions
def handle_transport_selection_for_expense(message):
    user_id = message.from_user.id

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    elif message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Добавить транспорт":
        add_transport(message)
        return

    selected_transport = message.text
    for transport in user_transport.get(str(user_id), []):
        formatted_transport = f"{transport['brand']} {transport['model']} ({transport['license_plate']})"
        if formatted_transport == selected_transport:
            brand, model, license_plate = transport['brand'], transport['model'], transport['license_plate']
            break
    else:
        # Отправляем сообщение об ошибке и добавляем кнопки возврата в меню
        bot.send_message(user_id, "Не удалось найти указанный транспорт. Пожалуйста, выберите снова.")

        # Формируем клавиатуру с кнопками выбора транспорта и кнопками возврата в меню
        markup = get_user_transport_keyboard(user_id)
        markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
        markup.add(types.KeyboardButton("В главное меню"))

        bot.send_message(user_id, "Выберите транспорт или добавьте новый:", reply_markup=markup)
        bot.register_next_step_handler(message, handle_transport_selection_for_expense)
        return

    # Продолжение процесса после выбора транспорта
    process_category_selection(user_id, brand, model, license_plate)

# Основной процесс выбора категории для записи траты
#@log_user_actions
def process_category_selection(user_id, brand, model, license_plate, prompt_message=None):
    categories = get_user_categories(user_id)

    # Смайлы для категорий
    system_emoji = "🔹"
    user_emoji = "🔸"

    # Добавление смайлов к категориям: первые 7 считаются системными
    category_list = "\n".join(
        f"{system_emoji if i < 7 else user_emoji} {i + 1}. {category}"
        for i, category in enumerate(categories)
    )
    category_list = f"*Выберите категорию:*\n\n{category_list}"

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(
        types.KeyboardButton("Добавить категорию"),
        types.KeyboardButton("Удалить категорию")
    )
    markup.add(
        types.KeyboardButton("Вернуться в меню трат и ремонтов")
    )
    markup.add(
        types.KeyboardButton("В главное меню")
    )
    if prompt_message:
        bot.send_message(user_id, category_list, reply_markup=markup, parse_mode="Markdown")
        bot.register_next_step_handler(prompt_message, get_expense_category, brand, model, license_plate)
    else:
        prompt_message = bot.send_message(user_id, category_list, reply_markup=markup, parse_mode="Markdown")
        bot.register_next_step_handler(prompt_message, get_expense_category, brand, model, license_plate)

# Обработка выбора категории для траты
#@log_user_actions
def get_expense_category(message, brand, model, license_plate):
    user_id = message.from_user.id
    
    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expense_category, brand, model, license_plate)
        return

    # Проверка на текстовое сообщение
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expense_category, brand, model, license_plate)
        return
    
    selected_index = message.text.strip()

    # Переход в меню
    if selected_index == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    elif selected_index == "В главное меню":
        return_to_menu(message)
        return

    if selected_index == 'Без категории':
        selected_category = "Без категории"
        proceed_to_expense_name(message, selected_category, brand, model, license_plate)
        return

    if selected_index == 'Добавить категорию':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(
            types.KeyboardButton("Вернуться в меню трат и ремонтов")
        )
        markup.add(
            types.KeyboardButton("В главное меню")
        )
        bot.send_message(user_id, "Введите название новой категории:", reply_markup=markup)
        bot.register_next_step_handler(message, add_new_category, brand, model, license_plate)
        return

    if selected_index == 'Удалить категорию':
        handle_category_removal(message, brand, model, license_plate)
        return

    if selected_index.isdigit():
        index = int(selected_index) - 1
        categories = get_user_categories(user_id)
        if 0 <= index < len(categories):
            selected_category = categories[index]
            proceed_to_expense_name(message, selected_category, brand, model, license_plate)
        else:
            bot.send_message(user_id, "Неверный номер категории. Попробуйте снова.")
            bot.register_next_step_handler(message, get_expense_category, brand, model, license_plate)
    else:
        bot.send_message(user_id, "Пожалуйста, введите номер категории.")
        bot.register_next_step_handler(message, get_expense_category, brand, model, license_plate)

# Добавление новой категории для трат
#@log_user_actions
def add_new_category(message, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, add_new_category, brand, model, license_plate)
        return

    # Проверка на текстовое сообщение
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, add_new_category, brand, model, license_plate)
        return

    # Основной код функции
    new_category = message.text.strip().lower()  # Приведение категории к нижнему регистру

    if new_category in ["вернуться в меню трат и ремонтов", "в главное меню"]:
        if new_category == "вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    if not new_category:
        bot.send_message(user_id, "Название категории не может быть пустым. Пожалуйста, введите корректное название")
        bot.register_next_step_handler(message, add_new_category, brand, model, license_plate)
        return

    user_categories = [cat.lower() for cat in get_user_categories(user_id)]  # Получение списка категорий в нижнем регистре
    if new_category in user_categories:
        bot.send_message(user_id, "Такая категория уже существует. Пожалуйста, введите уникальное название")
        bot.register_next_step_handler(message, add_new_category, brand, model, license_plate)
        return

    add_user_category(user_id, new_category)
    bot.send_message(user_id, f"Категория *{new_category}* успешно добавлена!", parse_mode="Markdown")
    process_category_selection(user_id, brand, model, license_plate)

#@log_user_actions
def proceed_to_expense_name(message, selected_category, brand, model, license_plate):
    user_id = message.from_user.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите название траты:", reply_markup=markup)
    bot.register_next_step_handler(message, get_expense_name, selected_category, brand, model, license_plate)

# Обработчик для ввода названия траты
#@log_user_actions
def get_expense_name(message, selected_category, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expense_name, selected_category, brand, model, license_plate)
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    expense_name = message.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_skip = types.KeyboardButton("Пропустить описание")  
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_skip)
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите описание траты или пропустите этот шаг:", reply_markup=markup)
    bot.register_next_step_handler(message, get_expense_description, selected_category, expense_name, brand, model, license_plate)

#@log_user_actions
def get_expense_description(message, selected_category, expense_name, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expense_description, selected_category, expense_name, brand, model, license_plate)
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    description = message.text if message.text != "Пропустить описание" else ""

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return) 
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите дату траты:", reply_markup=markup)
    bot.register_next_step_handler(message, get_expense_date, selected_category, expense_name, description, brand, model, license_plate)

#@log_user_actions
def get_expense_date(message, selected_category, expense_name, description, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expense_date, selected_category, expense_name, description, brand, model, license_plate)
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    expense_date = message.text

    # Проверка на формат даты
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", expense_date):
        try:
            day, month, year = map(int, expense_date.split('.'))
            # Проверка на корректность значений дня, месяца и года
            if 1 <= month <= 12 and 1 <= day <= 31 and 2000 <= year <= 3000:
                datetime.strptime(expense_date, "%d.%m.%Y")
            else:
                raise ValueError
        except ValueError:
            bot.send_message(user_id, "Неверный формат даты или значения. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ")
            bot.register_next_step_handler(message, get_expense_date, selected_category, expense_name, description, brand, model, license_plate)
            return
    else:
        bot.send_message(user_id, "Неверный формат даты. Попробуйте еще раз в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, get_expense_date, selected_category, expense_name, description, brand, model, license_plate)
        return

    # Если дата корректна, переходим к следующему шагу
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)
    
    bot.send_message(user_id, "Введите сумму траты:", reply_markup=markup)
    bot.register_next_step_handler(message, get_expense_amount, selected_category, expense_name, description, expense_date, brand, model, license_plate)

#@log_user_actions    
def get_expense_amount(message, selected_category, expense_name, description, expense_date, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expense_amount, selected_category, expense_name, description, expense_date, brand, model, license_plate)
        return

    # Получаем сумму траты
    expense_amount = message.text.replace(",", ".")
    if not is_numeric(expense_amount):
        bot.send_message(user_id, "Пожалуйста, введите сумму траты в числовом формате.")
        bot.register_next_step_handler(message, get_expense_amount, selected_category, expense_name, description, expense_date, brand, model, license_plate)
        return

    # Загружаем данные о расходах пользователя
    data = load_expense_data(user_id)
    if str(user_id) not in data:
        data[str(user_id)] = {"expenses": []}

    # Добавляем новый расход
    selected_transport = f"{brand} {model} {license_plate}"
    expenses = data[str(user_id)].get("expenses", [])
    new_expense = {
        "category": selected_category,
        "name": expense_name,
        "date": expense_date,
        "amount": expense_amount,
        "description": description,
        "transport": {"brand": brand, "model": model, "license_plate": license_plate}
    }
    expenses.append(new_expense)
    data[str(user_id)]["expenses"] = expenses
    save_expense_data(user_id, data, selected_transport)  # Передаем selected_transport здесь

    # Сохраняем расход в Excel
    save_expense_to_excel(user_id, new_expense)

    bot.send_message(user_id, "Трата успешно записана!")
    send_menu(user_id)

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

def save_expense_to_excel(user_id, expense_data):
    # Путь к Excel-файлу пользователя
    excel_path = os.path.join("data base", "expense", "excel", f"{user_id}_expenses.xlsx")

    # Проверяем, существует ли директория
    directory = os.path.dirname(excel_path)
    if not os.path.exists(directory):
        os.makedirs(directory)  # Создаем директорию, если она не существует

    # Загружаем или создаём рабочую книгу
    try:
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Удаляем стандартный лист
        
        # Лист для всех расходов (Summary)
        summary_sheet = workbook["Summary"] if "Summary" in workbook.sheetnames else workbook.create_sheet("Summary")
        
        # Лист для конкретного транспортного средства
        transport_sheet_name = f"{expense_data['transport']['brand']}_{expense_data['transport']['model']}_{expense_data['transport']['license_plate']}"
        if transport_sheet_name not in workbook.sheetnames:
            transport_sheet = workbook.create_sheet(transport_sheet_name)
        else:
            transport_sheet = workbook[transport_sheet_name]
        
        # Определяем заголовки
        headers = ["Транспорт", "Категория", "Название", "Дата", "Сумма", "Описание"]
        
        # Вспомогательная функция для настройки листов
        #@log_user_actions
        def setup_sheet(sheet):
            if sheet.max_row == 1:
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

        # Добавляем заголовки и данные
        for sheet in [summary_sheet, transport_sheet]:
            setup_sheet(sheet)
            row_data = [
                f"{expense_data['transport']['brand']} {expense_data['transport']['model']} {expense_data['transport']['license_plate']}",
                expense_data["category"],
                expense_data["name"],
                expense_data["date"],
                float(expense_data["amount"]),  # Сохраняем сумму как число
                expense_data["description"],
            ]
            sheet.append(row_data)
        
        # Автоподгонка столбцов
        for sheet in [summary_sheet, transport_sheet]:
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Сохраняем рабочую книгу
        workbook.save(excel_path)
    except Exception as e:
        pass

# Удаление категории
@bot.message_handler(func=lambda message: message.text == "Удалить категорию")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить категорию')
@track_usage('Удалить категорию')  # Добавление отслеживания статистики
#@log_user_actions
def handle_category_removal(message, brand=None, model=None, license_plate=None):
    user_id = message.from_user.id
    categories = get_user_categories(user_id)

    # Смайлы для категорий
    system_emoji = "🔹"
    user_emoji = "🔸"

    # Создаем текстовый список категорий с эмодзи
    category_list = "\n".join(
        f"{system_emoji if i < 7 else user_emoji} {i + 1}. {category}"
        for i, category in enumerate(categories)
    )
    category_list = f"*Выберите категорию для удаления или 0 для отмены:*\n\n{category_list}"

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
    markup.add(types.KeyboardButton("В главное меню"))

    bot.send_message(user_id, category_list, reply_markup=markup, parse_mode="Markdown")
    bot.register_next_step_handler(message, remove_selected_category, brand, model, license_plate)

#@log_user_actions
def remove_selected_category(message, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, remove_selected_category, brand, model, license_plate)
        return

    # Проверка на наличие текста в сообщении
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, remove_selected_category, brand, model, license_plate)
        return

    selected_index = message.text.strip()

    # Проверка на команды меню
    if selected_index == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    elif selected_index == "В главное меню":
        return_to_menu(message)
        return

    if selected_index == '0':
        bot.send_message(user_id, "Удаление категории отменено!")
        return process_category_selection(user_id, brand, model, license_plate)

    if selected_index.isdigit():
        index = int(selected_index) - 1
        categories = get_user_categories(user_id)
        default_categories = ["без категории", "азс", "мойка", "парковка", "платная дорога", "страховка", "штрафы"]

        if 0 <= index < len(categories):
            category_to_remove = categories[index].lower()  # Приведение к нижнему регистру
            if category_to_remove in default_categories:
                bot.send_message(user_id, f"Нельзя удалить системную категорию *{category_to_remove}*. Попробуйте еще раз", parse_mode="Markdown")
                return bot.register_next_step_handler(message, remove_selected_category, brand, model, license_plate)

            remove_user_category(user_id, category_to_remove)
            bot.send_message(user_id, f"Категория *{category_to_remove}* успешно удалена!", parse_mode="Markdown")
        else:
            bot.send_message(user_id, "Неверный номер категории. Попробуйте снова")
            return bot.register_next_step_handler(message, remove_selected_category, brand, model, license_plate)
    else:
        bot.send_message(user_id, "Пожалуйста, введите номер категории")
        return bot.register_next_step_handler(message, remove_selected_category, brand, model, license_plate)

    return process_category_selection(user_id, brand, model, license_plate)

#@log_user_actions
def is_numeric(s):
    if s is not None:
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False

# (10.9) --------------- КОД ДЛЯ "ТРАТ" (ОБРАБОТЧИК "ПОСМОТРЕТЬ ТРАТЫ") ---------------

#@log_user_actions
def create_transport_options_markup():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_add_transport = types.KeyboardButton("Добавить транспорт")
    item_cancel = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_add_transport)
    markup.add(item_cancel)
    markup.add(item_main)
    return markup

#@log_user_actions
def ask_add_transport(message):
    user_id = message.from_user.id

    if message.text == "Добавить транспорт":
        add_transport(message)
        return
    if message.text == "Вернуться в меню трат и ремонтов":
        send_menu(user_id)
        return
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    else:
        bot.send_message(user_id, "Пожалуйста, выберите вариант.", reply_markup=create_transport_options_markup())
        bot.register_next_step_handler(message, ask_add_transport)

# (10.10) --------------- КОД ДЛЯ "ТРАТ" (ФУНКЦИЯ МАКСИМАЛЬНОГО СООБЩЕНИЯ) ---------------

MAX_MESSAGE_LENGTH = 4096 

#@log_user_actions
def send_message_with_split(user_id, message_text):
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text)
    else:
        message_parts = [message_text[i:i + MAX_MESSAGE_LENGTH] for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)]
        for part in message_parts:
            bot.send_message(user_id, part)

# (10.11) --------------- КОД ДЛЯ "ТРАТ" (ОБРАБОТЧИК "ТРАТЫ ЗА МЕСЯЦ") ---------------

selected_transport_dict = {}

# Функция для фильтрации трат по транспорту
#@log_user_actions
def filter_expenses_by_transport(user_id, expenses):
    selected_transport = selected_transport_dict.get(user_id)
    if not selected_transport:
        return expenses

    # Фильтруем по транспорту
    filtered_expenses = [expense for expense in expenses if f"{expense['transport']['brand']} {expense['transport']['model']} ({expense['transport']['license_plate']})" == selected_transport]
    return filtered_expenses

# Функция для отправки сообщений, если сообщение длинное
#@log_user_actions
def send_message_with_split(user_id, message_text):
    bot.send_message(user_id, message_text, parse_mode="Markdown")

# Функция для отправки меню
#@log_user_actions
def send_menu1(user_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Траты (по категориям)")
    item2 = types.KeyboardButton("Траты (месяц)")
    item3 = types.KeyboardButton("Траты (год)")
    item4 = types.KeyboardButton("Траты (все время)")
    item_excel = types.KeyboardButton("Посмотреть траты в EXCEL")  # Новая кнопка
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item1, item2)
    markup.add(item3, item4)
    markup.add(item_excel)
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Выберите вариант просмотра трат:", reply_markup=markup)

# Обработчик для просмотра трат
@bot.message_handler(func=lambda message: message.text == "Посмотреть траты")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть траты')
@track_usage('Посмотреть траты')  # Добавление отслеживания статистики
#@log_user_actions
def view_expenses(message):
    user_id = message.from_user.id

    # Загружаем данные о транспорте
    transport_list = load_transport_data(user_id)

    if not transport_list:
        bot.send_message(user_id, "У вас нет сохраненного транспорта. Хотите добавить транспорт?", reply_markup=create_transport_options_markup())
        bot.register_next_step_handler(message, ask_add_transport)
        return

    # Если транспорт есть, продолжаем с выбором
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    
    # Разбиваем транспорт на группы по два элемента
    for i in range(0, len(transport_list), 2):
        transport_buttons = []
        for j in range(i, min(i + 2, len(transport_list))):
            transport = transport_list[j]
            transport_name = f"{transport['brand']} {transport['model']} ({transport['license_plate']})"
            transport_buttons.append(types.KeyboardButton(transport_name))
        
        # Добавляем пару кнопок в строку
        markup.add(*transport_buttons)
    
    # Добавляем кнопки возврата
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main)

    bot.send_message(user_id, "Выберите ваш транспорт:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_transport_selection)

# Обработчик выбора транспорта
#@log_user_actions
def handle_transport_selection(message):
    user_id = message.from_user.id
    selected_transport = message.text

    if selected_transport == "Вернуться в меню трат и ремонтов":
        send_menu(user_id)
        return

    if selected_transport == "В главное меню":
        return_to_menu(message)
        return

    # Сохраняем выбранный транспорт для пользователя
    selected_transport_dict[user_id] = selected_transport

    # Теперь можем показывать доступные фильтры для трат
    bot.send_message(user_id, f"Показываю траты для транспорта: *{selected_transport}*", parse_mode="Markdown")
    send_menu1(user_id)

@bot.message_handler(func=lambda message: message.text == "Посмотреть траты в EXCEL")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть траты в EXCEL')
@track_usage('Посмотреть траты в EXCEL')  # Добавление отслеживания статистики
#@log_user_actions
def send_expenses_excel(message):
    user_id = message.from_user.id

    # Путь к Excel файлу
    excel_path = os.path.join("data base", "expense", "excel", f"{user_id}_expenses.xlsx")

    # Проверяем наличие файла
    if not os.path.exists(excel_path):
        bot.send_message(user_id, "Файл с вашими тратами не найден")
        return

    # Отправка файла пользователю
    with open(excel_path, 'rb') as excel_file:
        bot.send_document(user_id, excel_file)

# Обработчик для просмотра трат по категориям
@bot.message_handler(func=lambda message: message.text == "Траты (по категориям)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Траты (по категориям)')
@track_usage('Траты (по категориям)')  # Добавление отслеживания статистики
#@log_user_actions
def view_expenses_by_category(message):
    user_id = message.from_user.id
    user_data = load_expense_data(user_id)
    expenses = user_data.get(str(user_id), {}).get("expenses", [])

    # Фильтруем траты по выбранному транспорту
    expenses = filter_expenses_by_transport(user_id, expenses)

    # Получаем уникальные категории трат для выбранного транспорта
    categories = set(expense['category'] for expense in expenses)
    if not categories:
        bot.send_message(user_id, "*Нет доступных категорий* для выбранного транспорта", parse_mode="Markdown")
        send_menu1(user_id)  # Возврат в меню трат и ремонтов
        return

    category_buttons = [types.KeyboardButton(category) for category in categories]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(*category_buttons)
    item_return_to_expenses = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_return_to_main = types.KeyboardButton("В главное меню")
    markup.add(item_return_to_expenses)
    markup.add(item_return_to_main)

    bot.send_message(user_id, "Выберите категорию для просмотра трат:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_category_selection)

# Обработчик выбора категории
#@log_user_actions
def handle_category_selection(message):
    user_id = message.from_user.id
    selected_category = message.text

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_category_selection)
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_category_selection)
        return

    user_data = load_expense_data(user_id)
    expenses = user_data.get(str(user_id), {}).get("expenses", [])

    # Фильтруем траты по выбранному транспорту
    expenses = filter_expenses_by_transport(user_id, expenses)

    # Проверяем, существует ли выбранная категория
    if selected_category not in {expense['category'] for expense in expenses}:
        bot.send_message(user_id, "Выбранная категория не найдена. Пожалуйста, выберите корректную категорию")
        view_expenses_by_category(message)  # Запрашиваем повторный ввод категории
        return

    # Фильтруем траты по выбранной категории
    category_expenses = [expense for expense in expenses if expense['category'] == selected_category]

    total_expenses = 0
    expense_details = []

    for index, expense in enumerate(category_expenses, start=1):
        expense_name = expense.get("name", "Без названия")
        expense_date = expense.get("date", "")
        amount = float(expense.get("amount", 0))
        total_expenses += amount

        expense_details.append(
            f"💸 *№ {index}*\n\n"
            f"📂 *Категория:* {selected_category}\n"
            f"📌 *Название:* {expense_name}\n"
            f"📅 *Дата:* {expense_date}\n"
            f"💰 *Сумма:* {amount} руб.\n"
            f"📝 *Описание:* {expense.get('description', 'Без описания')}\n"
        )

    if expense_details:
        message_text = f"Траты в категории *{selected_category.lower()}*:\n\n\n" + "\n\n".join(expense_details)
        send_message_with_split(user_id, message_text)
        bot.send_message(
            user_id, 
            f"Итоговая сумма трат в категории *{selected_category.lower()}*: *{total_expenses}* руб.", 
            parse_mode="Markdown"
        )
    else:
        bot.send_message(user_id, f"В категории *{selected_category.lower()}* трат не найдено", parse_mode="Markdown")

    # Возвращаем пользователя в главное меню после отображения информации
    send_menu1(user_id)

# Обработчик трат за месяц
@bot.message_handler(func=lambda message: message.text == "Траты (месяц)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Траты (месяц)')
@track_usage('Траты (месяц)')  # Добавление отслеживания статистики
#@log_user_actions
def view_expenses_by_month(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main)

    bot.send_message(user_id, "Введите месяц и год (ММ.ГГГГ) для просмотра трат за этот период:", reply_markup=markup)
    bot.register_next_step_handler(message, get_expenses_by_month)

#@log_user_actions
def get_expenses_by_month(message):
    user_id = message.from_user.id
    date = message.text.strip() if message.text else None

    # Проверяем, есть ли текст в сообщении
    if not date:
        bot.send_message(user_id, "Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expenses_by_month)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expenses_by_month)
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expenses_by_month)
        return

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверка формата и диапазонов месяца и года
    if "." in date:
        parts = date.split(".")
        if len(parts) == 2:
            month, year = parts
            if month.isdigit() and year.isdigit() and 1 <= int(month) <= 12 and 2000 <= int(year) <= 3000:
                month, year = map(int, parts)
                
                user_data = load_expense_data(user_id)
                expenses = user_data.get(str(user_id), {}).get("expenses", [])

                # Фильтруем траты по выбранному транспорту
                expenses = filter_expenses_by_transport(user_id, expenses)

                total_expenses = 0
                expense_details = []

                for index, expense in enumerate(expenses, start=1):
                    expense_date = expense.get("date", "")
                    expense_date_parts = expense_date.split(".")
                    if len(expense_date_parts) >= 2:
                        expense_month, expense_year = map(int, expense_date_parts[1:3])

                        if expense_month == month and expense_year == year:
                            amount = float(expense.get("amount", 0))
                            total_expenses += amount

                            expense_name = expense.get("name", "Без названия")
                            description = expense.get("description", "")
                            category = expense.get("category", "Без категории")

                            expense_details.append(
                                f"💸 *№ {index}*\n\n"
                                f"📂 *Категория:* {category}\n"
                                f"📌 *Название:* {expense_name}\n"
                                f"📅 *Дата:* {expense_date}\n"
                                f"💰 *Сумма:* {amount} руб.\n"
                                f"📝 *Описание:* {description}\n"
                            )

                if expense_details:
                    message_text = f"Траты за *{date}* месяц:\n\n\n" + "\n\n".join(expense_details)
                    send_message_with_split(user_id, message_text)
                    bot.send_message(user_id, f"Итоговая сумма трат за *{date}* месяц: *{total_expenses}* руб.", parse_mode="Markdown")
                else:
                    bot.send_message(user_id, f"За *{date}* месяц трат не найдено", parse_mode="Markdown")
                
                send_menu1(user_id)  # Возвращаемся в меню после отображения информации
            else:
                bot.send_message(user_id, "Пожалуйста, введите корректный месяц и год в формате ММ.ГГГГ")
                bot.register_next_step_handler(message, get_expenses_by_month)
                return
        else:
            bot.send_message(user_id, "Пожалуйста, введите дату в формате ММ.ГГГГ")
            bot.register_next_step_handler(message, get_expenses_by_month)
            return
    else:
        bot.send_message(user_id, "Пожалуйста, введите дату в формате ММ.ГГГГ")
        bot.register_next_step_handler(message, get_expenses_by_month)
        return

# Обработчик трат за год
@bot.message_handler(func=lambda message: message.text == "Траты (год)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Траты (год)')
@track_usage('Траты (год)')  # Добавление отслеживания статистики
#@log_user_actions
def view_expenses_by_license_plate(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main)

    bot.send_message(user_id, "Введите год в формате (ГГГГ) для просмотра трат за этот год:", reply_markup=markup)
    bot.register_next_step_handler(message, get_expenses_by_license_plate)

#@log_user_actions
def get_expenses_by_license_plate(message):
    user_id = message.from_user.id

    # Проверяем, что сообщение содержит текст
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expenses_by_license_plate)
        return

    year_input = message.text.strip()

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expenses_by_license_plate)
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_expenses_by_license_plate)
        return

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверка, что введено четырехзначное число
    if not year_input.isdigit() or len(year_input) != 4:
        bot.send_message(user_id, "Пожалуйста, введите год в формате ГГГГ")
        bot.register_next_step_handler(message, get_expenses_by_license_plate)
        return

    # Преобразуем год в число и проверяем диапазон
    year = int(year_input)
    if year < 2000 or year > 3000:
        bot.send_message(user_id, "Введите год в формате ГГГГ")
        bot.register_next_step_handler(message, get_expenses_by_license_plate)
        return

    user_data = load_expense_data(user_id)
    expenses = user_data.get(str(user_id), {}).get("expenses", [])

    # Фильтруем траты по выбранному транспорту
    expenses = filter_expenses_by_transport(user_id, expenses)

    total_expenses = 0
    expense_details = []

    for index, expense in enumerate(expenses, start=1):
        expense_date = expense.get("date", "")
        if "." in expense_date:
            date_parts = expense_date.split(".")
            if len(date_parts) >= 3:
                expense_year = int(date_parts[2])
                if expense_year == year:
                    amount = float(expense.get("amount", 0))
                    total_expenses += amount

                    expense_name = expense.get("name", "Без названия")
                    description = expense.get("description", "")
                    category = expense.get("category", "Без категории")

                    expense_details.append(
                        f"💸 *№ {index}*\n\n"
                        f"📂 *Категория:* {category}\n"
                        f"📌 *Название:* {expense_name}\n"
                        f"📅 *Дата:* {expense_date}\n"
                        f"💰 *Сумма:* {amount} руб.\n"
                        f"📝 *Описание:* {description}\n"
                    )
                    
    if expense_details:
        message_text = f"Траты за *{year}* год:\n\n\n" + "\n\n".join(expense_details)
        send_message_with_split(user_id, message_text)
        bot.send_message(user_id, f"Итоговая сумма трат за *{year}* год: *{total_expenses}* руб.", parse_mode="Markdown")
    else:
        bot.send_message(user_id, f"За *{year}* год трат не найдено", parse_mode="Markdown")

    # Возвращаемся в меню после отображения информации
    send_menu1(user_id)

# Обработчик всех трат
@bot.message_handler(func=lambda message: message.text == "Траты (все время)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Траты (все время)')
@track_usage('Траты (все время)')  # Добавление отслеживания статистики
#@log_user_actions
def view_all_expenses(message):
    user_id = message.from_user.id

    user_data = load_expense_data(user_id)
    expenses = user_data.get(str(user_id), {}).get("expenses", [])

    # Фильтруем траты по выбранному транспорту
    expenses = filter_expenses_by_transport(user_id, expenses)

    total_expenses = 0
    expense_details = []

    for index, expense in enumerate(expenses, start=1):
        amount = float(expense.get("amount", 0))
        total_expenses += amount

        expense_name = expense.get("name", "Без названия")
        expense_date = expense.get("date", "")
        description = expense.get("description", "")
        category = expense.get("category", "Без категории")

        expense_details.append(
            f"💸 *№ {index}*\n\n"
            f"📂 *Категория:* {category}\n"
            f"📌 *Название:* {expense_name}\n"
            f"📅 *Дата:* {expense_date}\n"
            f"💰 *Сумма:* {amount} руб.\n"
            f"📝 *Описание:* {description}\n"
        )

    if expense_details:
        message_text = "*Все* траты:\n\n\n" + "\n\n".join(expense_details)
        send_message_with_split(user_id, message_text)
        bot.send_message(user_id, f"Итоговая сумма всех трат: *{total_expenses}* руб.", parse_mode="Markdown")
    else:
        bot.send_message(user_id, "Трат не найдено", parse_mode="Markdown")

    # Возвращаемся в меню после отображения информации
    send_menu1(user_id)

# (10.18) --------------- КОД ДЛЯ "ТРАТ" (ОБРАБОТЧИК "УДАЛИТЬ ТРАТЫ") ---------------

# Глобальный словарь для хранения выбранного транспорта по user_id
selected_transports = {}

def save_selected_transport(user_id, selected_transport):
    user_data = load_expense_data(user_id)  # Загружаем данные пользователя
    # Изменяем только выбранный транспорт, не трогая остальные данные
    user_data["selected_transport"] = selected_transport
    save_expense_data(user_id, user_data)  # Сохраняем только изменения

@bot.message_handler(func=lambda message: message.text == "Удалить траты")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить траты')
@track_usage('Удалить траты')  # Добавление отслеживания статистики
#@log_user_actions
def delete_expenses_menu(message):
    user_id = message.from_user.id

    transport_data = load_transport_data(user_id)
    if not transport_data:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_add_transport = types.KeyboardButton("Добавить транспорт")
        item_cancel = types.KeyboardButton("Вернуться в меню трат и ремонтов")
        item_main = types.KeyboardButton("В главное меню")
        markup.add(item_add_transport)
        markup.add(item_cancel)
        markup.add(item_main)
        bot.send_message(user_id, "У вас нет сохраненного транспорта. Хотите добавить транспорт?", reply_markup=markup)
        bot.register_next_step_handler(message, ask_add_transport)
        return

    # Отображение транспорта для удаления в два столбца
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    transport_buttons = []  # Временный список для кнопок

    # Проходим по данным о транспорте и создаем кнопки
    for transport in transport_data:
        brand = transport.get("brand", "Без названия")
        model = transport.get("model", "Без названия")
        license_plate = transport.get("license_plate", "Неизвестно")
        # Добавляем номер транспорта в скобках
        button_label = f"{brand} {model} ({license_plate})"
        transport_buttons.append(types.KeyboardButton(button_label))

    # Формируем строки по две кнопки
    for i in range(0, len(transport_buttons), 2):
        markup.row(*transport_buttons[i:i + 2])

    # Добавляем кнопки возврата
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Выберите транспорт для удаления трат:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_transport_selection_for_deletion)

#@log_user_actions
def handle_transport_selection_for_deletion(message):
    user_id = message.from_user.id
    selected_transport = message.text.strip()
    
    # Сохраняем выбранный транспорт в глобальном словаре
    selected_transports[user_id] = selected_transport
    save_selected_transport(user_id, selected_transport)  # Сохраняем в БД

    if selected_transport == "Вернуться в меню трат и ремонтов":
        send_menu(user_id)
        return

    if selected_transport == "В главное меню":
        return_to_menu(message)
        return

    # Запрос на удаление трат
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_month = types.KeyboardButton("Del траты (месяц)")
    item_license_plate = types.KeyboardButton("Del траты (год)")
    item_all_time = types.KeyboardButton("Del траты (все время)")
    item_del_category_exp = types.KeyboardButton("Del траты (категория)")
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    
    markup.add(item_del_category_exp, item_month)
    markup.add(item_license_plate, item_all_time)
    markup.add(item_return)
    markup.add(item_main_menu)
    bot.send_message(user_id, "Выберите вариант удаления трат:", reply_markup=markup)

expenses_to_delete_dict = {}


MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_long_message(user_id, text):
    # Разбиваем текст на части, если он слишком длинный
    while len(text) > MAX_MESSAGE_LENGTH:
        # Отправляем первую часть текста
        bot.send_message(user_id, text[:MAX_MESSAGE_LENGTH], parse_mode="Markdown")
        # Оставшийся текст
        text = text[MAX_MESSAGE_LENGTH:]
    # Отправляем остаток
    bot.send_message(user_id, text, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Del траты (категория)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del траты (категория)')
@track_usage('Del траты (категория)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_expenses_by_category(message):
    user_id = message.from_user.id

    selected_transport = selected_transports.get(user_id)

    if not selected_transport:
        bot.send_message(user_id, "Не выбран транспорт")
        send_menu(user_id)
        return

    # Приводим транспорт в одинаковый формат и удаляем скобки
    selected_transport_info = selected_transport.strip().lower().replace('(', '').replace(')', '')

    # Загружаем данные пользователя
    user_data = load_expense_data(user_id).get(str(user_id), {})
    expenses = user_data.get("expenses", [])

    # Получаем категории для удаления по выбранному транспорту
    categories = list({
        expense.get("category")
        for expense in expenses
        if f"{expense.get('transport', {}).get('brand', '').strip().lower()} {expense.get('transport', {}).get('model', '').strip().lower()} {str(expense.get('transport', {}).get('license_plate', '')).strip().lower()}".replace('(', '').replace(')', '') == selected_transport_info
    })

    if not categories:
        bot.send_message(user_id, "У вас *нет категорий* для удаления трат по выбранному транспорту", parse_mode="Markdown")
        send_menu(user_id)
        return

    # Создаем клавиатуру
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    markup.add(*[types.KeyboardButton(category) for category in categories])
    markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
    markup.add(types.KeyboardButton("В главное меню"))

    bot.send_message(user_id, "Выберите категорию для удаления:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_category_selection_for_deletion)

# Глобальный словарь для хранения выбранной категории
selected_categories = {}

user_expenses_to_delete = {}

#@log_user_actions
def handle_category_selection_for_deletion(message):
    user_id = message.from_user.id

    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_category_selection_for_deletion)
        return

    selected_category = message.text.strip()

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_category_selection_for_deletion)
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_category_selection_for_deletion)
        return

    # Проверка возврата в меню
    if selected_category == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if selected_category == "В главное меню":
        return_to_menu(message)
        return

    selected_categories[user_id] = selected_category

    user_data = load_expense_data(user_id).get(str(user_id), {})
    expenses = user_data.get("expenses", [])

    selected_transport = selected_transports.get(user_id)
    selected_transport_info = selected_transport.strip().lower().replace('(', '').replace(')', '')

    # Фильтруем расходы
    expenses_to_delete = [
        expense for expense in expenses 
        if expense.get("category") == selected_category and 
           f"{expense.get('transport', {}).get('brand', '').strip().lower()} {expense.get('transport', {}).get('model', '').strip().lower()} {str(expense.get('transport', {}).get('license_plate', '')).strip().lower()}".replace('(', '').replace(')', '') == selected_transport_info
    ]

    if not expenses_to_delete:
        bot.send_message(user_id, f"Нет трат для удаления в категории *{selected_category.lower()}* по выбранному транспорту", parse_mode="Markdown")
        send_menu(user_id)
        return

    # Сохраняем список трат
    user_expenses_to_delete[user_id] = expenses_to_delete

    # Формируем текстовый список для отображения
    expense_list_text = f"Список трат для удаления по категории *{selected_category.lower()}*:\n\n\n"
    for index, expense in enumerate(expenses_to_delete, start=1):
        expense_name = expense.get("name", "Без названия")
        expense_date = expense.get("date", "Неизвестно")
        expense_list_text += f"📄 №{index}. {expense_name} - ({expense_date})\n\n"

    expense_list_text += "\nВведите номер траты для удаления:"

    # Создаем клавиатуру с кнопками
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add("Вернуться в меню трат и ремонтов")
    markup.add("В главное меню")

    # Отправляем сообщение с текстом и клавишами в одном вызове
    bot.send_message(user_id, expense_list_text, reply_markup=markup, parse_mode="Markdown")

    # Переход к следующему шагу (запрос номера траты для удаления)
    bot.register_next_step_handler(message, delete_expense_confirmation)

#@log_user_actions
def delete_expense_confirmation(message):
    user_id = message.from_user.id

    # Проверка, если текстовое сообщение отсутствует
    if not message.text:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expense_confirmation)
        return

    selected_option = message.text.strip()

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expense_confirmation)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', selected_option):
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expense_confirmation)
        return

    if selected_option == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if selected_option == "В главное меню":
        return_to_menu(message)
        return

    # Проверка, что введен номер
    if not selected_option.isdigit():
        bot.send_message(user_id, "Пожалуйста, введите номер траты из списка.")
        bot.register_next_step_handler(message, delete_expense_confirmation)
        return

    expense_index = int(selected_option) - 1
    expenses_to_delete = user_expenses_to_delete.get(user_id, [])

    # Проверка на правильность номера
    if 0 <= expense_index < len(expenses_to_delete):
        deleted_expense = expenses_to_delete.pop(expense_index)
        user_data = load_expense_data(user_id).get(str(user_id), {})
        user_expenses = user_data.get("expenses", [])

        # Удаление выбранной траты
        if deleted_expense in user_expenses:
            user_expenses.remove(deleted_expense)
            save_expense_data(user_id, {str(user_id): user_data})

        bot.send_message(
            user_id,
            f"Трата *{deleted_expense.get('name', 'Без названия').lower()}* успешно удалена!",
            parse_mode="Markdown"
        )

        # Возврат в меню после успешного удаления
        send_menu(user_id)
    else:
        bot.send_message(user_id, "Неверный номер траты. Попробуйте снова.")
        bot.register_next_step_handler(message, delete_expense_confirmation)
        return

# Удаление трат за месяц

MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_long_message(user_id, text):
    # Разбиваем текст на части, если он слишком длинный
    while len(text) > MAX_MESSAGE_LENGTH:
        # Отправляем первую часть текста
        bot.send_message(user_id, text[:MAX_MESSAGE_LENGTH], parse_mode="Markdown")
        # Оставшийся текст
        text = text[MAX_MESSAGE_LENGTH:]
    # Отправляем остаток
    bot.send_message(user_id, text, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Del траты (месяц)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del траты (месяц)')
@track_usage('Del траты (месяц)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_expense_by_month(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)
    bot.send_message(user_id, "Введите месяц и год (ММ.ГГГГ) для удаления трат за этот месяц:", reply_markup=markup)
    bot.register_next_step_handler(message, delete_expenses_by_month)

#@log_user_actions
def delete_expenses_by_month(message):
    user_id = message.from_user.id
    month_year = message.text.strip() if message.text else None

    if not month_year:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение")
        bot.register_next_step_handler(message, delete_expenses_by_month)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expenses_by_month)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expenses_by_month)
        return

    # Проверка формата месяца и года
    match = re.match(r"^(0[1-9]|1[0-2])\.(20[0-9]{2})$", month_year)
    if not match:
        bot.send_message(user_id, "Введен неверный месяц или год. Пожалуйста, введите корректные данные (ММ.ГГГГ)")
        bot.register_next_step_handler(message, delete_expenses_by_month)
        return

    selected_month, selected_year = match.groups()

    # Проверка выбранного транспорта
    selected_transport = selected_transports.get(user_id)
    if not selected_transport:
        bot.send_message(user_id, "Транспорт не выбран. Пожалуйста, выберите транспорт")
        send_menu(user_id)
        return

    transport_info = selected_transport.split(" ")
    selected_brand = transport_info[0].strip()
    selected_model = transport_info[1].strip()
    selected_license_plate = transport_info[2].strip().replace("(", "").replace(")", "")

    user_data = load_expense_data(user_id).get(str(user_id), {})
    expenses = user_data.get("expenses", [])

    if not expenses:
        bot.send_message(user_id, f"У вас нет сохраненных трат за *{month_year}* месяц", parse_mode="Markdown")
        send_menu(user_id)
        return

    expenses_to_delete = []
    for index, expense in enumerate(expenses, start=1):
        expense_date = expense.get("date", "")

        if not expense_date or len(expense_date.split(".")) != 3:
            continue  # Пропускаем траты с некорректной датой

        expense_day, expense_month, expense_year = expense_date.split(".")

        # Исправляем условие для сравнения месяца и года
        if expense_month == selected_month and expense_year == selected_year:
            expense_license_plate = expense.get("transport", {}).get("license_plate", "").strip()
            expense_brand = expense.get("transport", {}).get("brand", "").strip()
            expense_model = expense.get("transport", {}).get("model", "").strip()

            if (expense_license_plate == selected_license_plate and
                expense_brand == selected_brand and
                expense_model == selected_model):
                expenses_to_delete.append((index, expense))

    if expenses_to_delete:
        # Сохранение списка трат для удаления
        expenses_to_delete_dict[user_id] = expenses_to_delete

        # Формирование сообщения
        expense_list_text = f"Список трат для удаления за *{month_year}* месяц:\n\n\n"
        for index, expense in expenses_to_delete:
            expense_name = expense.get("name", "Без названия")
            expense_date = expense.get("date", "Неизвестно")
            expense_list_text += f"📄 №{index}. {expense_name} - ({expense_date})\n\n"

        expense_list_text += "\nВведите номер траты для удаления:"

        # Отправка сообщения
        send_long_message(user_id, expense_list_text)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("Вернуться в меню трат и ремонтов", "В главное меню")
        bot.register_next_step_handler(message, confirm_delete_expense_month)
    else:
        bot.send_message(user_id, f"Нет трат для удаления за *{month_year}* месяц", parse_mode="Markdown")
        send_menu(user_id)

#@log_user_actions
def confirm_delete_expense_month(message):
    user_id = message.from_user.id

    # Проверка, если текстовое сообщение отсутствует
    if not message.text:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение")
        bot.register_next_step_handler(message, confirm_delete_expense_month)
        return

    selected_option = message.text.strip()

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_expense_month)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', selected_option):
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_expense_month)
        return

    # Проверка на возврат в меню
    if selected_option == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if selected_option == "В главное меню":
        return_to_menu(message)
        return

    # Проверка ввода номера траты
    if selected_option.isdigit():
        expense_index = int(selected_option) - 1
        expenses_to_delete = expenses_to_delete_dict.get(user_id, [])

        if 0 <= expense_index < len(expenses_to_delete):
            # Удаление выбранной траты
            _, deleted_expense = expenses_to_delete.pop(expense_index)

            # Обновление данных пользователя
            user_data = load_expense_data(user_id).get(str(user_id), {})
            if "expenses" in user_data and deleted_expense in user_data["expenses"]:
                user_data["expenses"].remove(deleted_expense)
                save_expense_data(user_id, {str(user_id): user_data})

            # Уведомление об успешном удалении
            bot.send_message(
                user_id,
                f"Трата *{deleted_expense.get('name', 'Без названия').lower()}* успешно удалена!",
                parse_mode="Markdown"
            )

            # Обновление глобального списка
            expenses_to_delete_dict[user_id] = expenses_to_delete

            # Возврат в меню после успешного удаления
            send_menu(user_id)
        else:
            bot.send_message(user_id, "Неверный выбор. Пожалуйста, попробуйте снова")
            bot.register_next_step_handler(message, confirm_delete_expense_month)
            return
    else:
        bot.send_message(user_id, "Некорректный ввод. Пожалуйста, введите номер траты")
        bot.register_next_step_handler(message, confirm_delete_expense_month)
        return

# Удаление трат за год
MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_long_message(user_id, text):
    # Разбиваем текст на части, если он слишком длинный
    while len(text) > MAX_MESSAGE_LENGTH:
        # Отправляем первую часть текста
        bot.send_message(user_id, text[:MAX_MESSAGE_LENGTH], parse_mode="Markdown")
        # Оставшийся текст
        text = text[MAX_MESSAGE_LENGTH:]
    # Отправляем остаток
    bot.send_message(user_id, text, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Del траты (год)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del траты (год)')
@track_usage('Del траты (год)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_expense_by_license_plate(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)
    bot.send_message(user_id, "Введите год (ГГГГ) для удаления трат за этот год:", reply_markup=markup)
    bot.register_next_step_handler(message, delete_expenses_by_license_plate)

#@log_user_actions
def delete_expenses_by_license_plate(message):
    user_id = message.from_user.id
    license_plate = message.text.strip() if message.text else None  # Проверяем наличие текста

    # Проверка на отсутствие текста
    if not license_plate:
        bot.send_message(user_id, "Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expenses_by_license_plate)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expenses_by_license_plate)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', license_plate):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_expenses_by_license_plate)
        return

    if license_plate == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if license_plate == "В главное меню":
        return_to_menu(message)
        return

    # Проверка формата года (от 2000 до 3000)
    if not re.match(r"^(20[0-9]{2}|2[1-9][0-9]{2}|3000)$", license_plate):
        bot.send_message(user_id, "Введен неверный год. Пожалуйста, введите корректный год.")
        bot.register_next_step_handler(message, delete_expenses_by_license_plate)
        return

    user_data = load_expense_data(user_id).get(str(user_id), {})
    expenses = user_data.get("expenses", [])

    selected_transport = selected_transports.get(user_id, None)

    if selected_transport:
        transport_info = selected_transport.split(" ")
        selected_brand = transport_info[0].strip()
        selected_model = transport_info[1].strip()
        selected_license_plate = str(transport_info[2].strip())
    else:
        selected_brand = selected_model = selected_license_plate = None

    if not expenses:
        bot.send_message(user_id, f"У вас пока нет сохраненных трат за *{license_plate}* год для удаления", parse_mode="Markdown")
        send_menu(user_id)
        return

    expenses_to_delete = []
    for index, expense in enumerate(expenses, start=1):
        expense_date = expense.get("date", "")
        expense_license_plate = expense_date.split(".")[2]

        expense_brand = expense.get("transport", {}).get("brand", "").strip()
        expense_model = expense.get("transport", {}).get("model", "").strip()

        if (expense_license_plate == license_plate and
           expense_brand == selected_brand and
           expense_model == selected_model):
            expenses_to_delete.append((index, expense))

    if expenses_to_delete:
        expenses_to_delete_dict[user_id] = expenses_to_delete

        expense_list_text = f"Список трат для удаления за *{license_plate}* год:\n\n\n"
        for index, expense in expenses_to_delete:
            expense_name = expense.get("name", "Без названия")
            expense_date = expense.get("date", "Неизвестно")
            expense_list_text += f"📄 №{index}. {expense_name} - ({expense_date})\n\n"

        expense_list_text += "\nВведите номер траты для удаления:"
        send_long_message(user_id, expense_list_text)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("Вернуться в меню трат и ремонтов")
        markup.add("В главное меню")
        bot.register_next_step_handler(message, confirm_delete_expense_license_plate)
    else:
        bot.send_message(user_id, f"За *{license_plate}* год трат не найдено для удаления", parse_mode="Markdown")
        send_menu(user_id)

#@log_user_actions
def confirm_delete_expense_license_plate(message):
    user_id = message.from_user.id

    # Проверка, если текстовое сообщение отсутствует
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_expense_license_plate)
        return

    selected_option = message.text.strip()

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_expense_license_plate)
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', selected_option):
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_expense_license_plate)
        return

    if selected_option == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if selected_option == "В главное меню":
        return_to_menu(message)
        return

    # Получаем список трат для удаления из глобальной переменной
    expenses_to_delete = expenses_to_delete_dict.get(user_id, [])

    if selected_option.isdigit():
        expense_index = int(selected_option) - 1

        if 0 <= expense_index < len(expenses_to_delete):
            _, deleted_expense = expenses_to_delete.pop(expense_index)

            user_data = load_expense_data(user_id).get(str(user_id), {})
            user_data["expenses"].remove(deleted_expense)
            save_expense_data(user_id, {str(user_id): user_data})

            bot.send_message(user_id, f"Трата *{deleted_expense.get('name', 'Без названия').lower()}* успешно удалена!", parse_mode="Markdown")

            # Обновление списка трат для удаления
            expenses_to_delete_dict[user_id] = expenses_to_delete

            # Возврат в меню после успешного удаления
            send_menu(user_id)
        else:
            bot.send_message(user_id, "Неверный выбор. Пожалуйста, попробуйте снова")
            bot.register_next_step_handler(message, confirm_delete_expense_license_plate)
    else:
        bot.send_message(user_id, "Некорректный ввод. Пожалуйста, введите номер траты")
        bot.register_next_step_handler(message, confirm_delete_expense_license_plate)

# Удаление всех трат
@bot.message_handler(func=lambda message: message.text == "Del траты (все время)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del траты (все время)')
@track_usage('Del траты (все время)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_all_expenses_for_selected_transport(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    # Запрашиваем подтверждение
    bot.send_message(user_id, 
                    "Вы уверены, что хотите удалить все траты для выбранного транспорта?\n\n"
                    "Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены",
                    reply_markup=markup, parse_mode="Markdown")
    bot.register_next_step_handler(message, confirm_delete_all_expenses)

#@log_user_actions
def confirm_delete_all_expenses(message):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_all_expenses)
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_all_expenses)
        return

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверяем, что message.text не None
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_all_expenses)
        return

    # Приводим ответ пользователя к нижнему регистру для проверки
    response = message.text.strip().lower()

    # Проверка на "да" и "нет"
    if response == "да":
        # Загружаем данные пользователя
        user_data = load_expense_data(user_id).get(str(user_id), {})
        expenses = user_data.get("expenses", [])

        # Получаем выбранный транспорт
        selected_transport = selected_transports.get(user_id, None)
        if selected_transport:
            transport_info = selected_transport.split(" ")
            selected_brand = transport_info[0].strip()
            selected_model = transport_info[1].strip()
            selected_license_plate = transport_info[2].strip()
        else:
            selected_brand = selected_model = selected_license_plate = None

        if not expenses:
            bot.send_message(user_id, "У вас пока нет сохраненных трат для удаления")
            send_menu(user_id)
            return

        # Фильтруем траты для удаления
        expenses_to_keep = []
        for expense in expenses:
            expense_brand = expense.get("transport", {}).get("brand", "").strip()
            expense_model = expense.get("transport", {}).get("model", "").strip()

            # Сравниваем с выбранным транспортом
            if not (expense_brand == selected_brand and expense_model == selected_model):
                expenses_to_keep.append(expense)

        # Обновляем список расходов пользователя
        user_data["expenses"] = expenses_to_keep
        save_expense_data(user_id, {str(user_id): user_data}, selected_transport)

        update_excel_file(user_id)

        # Сообщение об успешном удалении
        bot.send_message(user_id, f"Все траты для транспорта *{selected_brand} {selected_model} {selected_license_plate}* успешно удалены", parse_mode="Markdown")
    elif response == "нет":
        bot.send_message(user_id, "Удаление трат отменено")
    else:
        # Ответ для неподходящих сообщений
        bot.send_message(user_id, "Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены", parse_mode="Markdown")
        bot.register_next_step_handler(message, confirm_delete_all_expenses)
        return

    send_menu(user_id)


import os
import pandas as pd
import openpyxl

#@log_user_actions
def delete_expense(user_id, deleted_expense):
    # Удаляем трату из базы данных
    user_data = load_expense_data(user_id).get(str(user_id), {})
    expenses = user_data.get("expenses", [])

    # Удаляем трату из списка расходов
    expenses = [expense for expense in expenses if not (
        expense["transport"]["brand"] == deleted_expense["transport"]["brand"] and
        expense["transport"]["model"] == deleted_expense["transport"]["model"] and
        expense["transport"]["license_plate"] == deleted_expense["transport"]["license_plate"] and
        expense["category"] == deleted_expense["category"] and
        expense["name"] == deleted_expense["name"] and
        expense["date"] == deleted_expense["date"] and
        expense["amount"] == deleted_expense["amount"] and
        expense["description"] == deleted_expense["description"]
    )]

    # Обновляем данные пользователя
    user_data["expenses"] = expenses
    save_expense_data(user_id, user_data)

    # Обновляем Excel файл
    update_excel_file(user_id)

#@log_user_actions
def update_excel_file(user_id):
    user_data = load_expense_data(user_id).get(str(user_id), {})
    expenses = user_data.get("expenses", [])

    # Путь к Excel файлу пользователя
    excel_file_path = f"data base/expense/excel/{user_id}_expenses.xlsx"

    # Проверяем, существует ли файл
    if not os.path.exists(excel_file_path):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        workbook.save(excel_file_path)

    # Открываем существующий файл
    workbook = load_workbook(excel_file_path)

    # Обновление общего листа (Summary)
    summary_sheet = workbook["Summary"] if "Summary" in workbook.sheetnames else workbook.create_sheet("Summary")
    headers = ["Транспорт", "Категория", "Название", "Дата", "Сумма", "Описание"]

    # Очистка всех данных на листе Summary (кроме заголовков)
    if summary_sheet.max_row > 1:
        summary_sheet.delete_rows(2, summary_sheet.max_row)

    # Добавляем заголовки, если они еще не добавлены
    if summary_sheet.max_row == 0:  # Если лист пустой
        summary_sheet.append(headers)
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Заполняем общий лист новыми данными
    for expense in expenses:
        transport = expense["transport"]
        row_data = [
            f"{transport['brand']} {transport['model']} {transport['license_plate']}",
            expense["category"],
            expense["name"],
            expense["date"],
            float(expense["amount"]),  # Сохраняем сумму как число
            expense["description"],
        ]
        summary_sheet.append(row_data)

    # Обновление индивидуальных листов для каждого транспорта
    unique_transports = set((exp["transport"]["brand"], exp["transport"]["model"], exp["transport"]["license_plate"]) for exp in expenses)

    # Удаляем старые листы для уникальных транспортов
    for sheet_name in workbook.sheetnames:
        if sheet_name != "Summary" and (sheet_name.split('_')[0], sheet_name.split('_')[1], sheet_name.split('_')[2]) not in unique_transports:
            del workbook[sheet_name]

    for brand, model, license_plate in unique_transports:
        sheet_name = f"{brand}_{model}_{license_plate}"
        if sheet_name not in workbook.sheetnames:
            transport_sheet = workbook.create_sheet(sheet_name)
            transport_sheet.append(headers)
            for cell in transport_sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        else:
            transport_sheet = workbook[sheet_name]
            # Очистка всех данных на листе транспорта, кроме заголовков
            if transport_sheet.max_row > 1:
                transport_sheet.delete_rows(2, transport_sheet.max_row)

        # Заполняем траты для этого транспорта
        for expense in expenses:
            if (expense["transport"]["brand"], expense["transport"]["model"], expense["transport"]["license_plate"]) == (brand, model, license_plate):
                row_data = [
                    f"{brand} {model} {license_plate}",
                    expense["category"],
                    expense["name"],
                    expense["date"],
                    float(expense["amount"]),  # Сохраняем сумму как число
                    expense["description"],
                ]
                transport_sheet.append(row_data)

    # Автоподгонка столбцов
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        for col in current_sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            current_sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Сохраняем изменения
    workbook.save(excel_file_path)
    workbook.close()

# (11) --------------- КОД ДЛЯ "РЕМОНТОВ" ---------------

# (11.1) --------------- КОД ДЛЯ "РЕМОНТОВ" (ОБРАБОТЧИК "ЗАПИСАТЬ РЕМОНТ") ---------------
# Системные категории ремонта

user_transport = {}  # Это должно быть вашим хранилищем данных о транспорте

# Обработка данных о ремонтах
def save_repair_data(user_id, user_data, selected_transport=None):
    # Ваш код сохранения
    if selected_transport:
        user_data["selected_transport"] = selected_transport
    # Задаем новый путь к папке и файлу
    folder_path = os.path.join("data base", "repairs")
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Измененное имя файла: "user_id_repairs.json"
    file_path = os.path.join(folder_path, f"{user_id}_repairs.json")

    # Загружаем текущие данные пользователя, чтобы сохранить user_categories
    current_data = load_repair_data(user_id)

    # Сохраняем только изменения, не перезаписывая user_categories
    if "user_categories" in current_data:
        user_data["user_categories"] = current_data["user_categories"]

    # Добавляем поле selected_transport для данного пользователя
    user_data["selected_transport"] = selected_transport

    with open(file_path, "w", encoding="utf-8") as file:
        json.dump(user_data, file, ensure_ascii=False, indent=4)

def load_repair_data(user_id):
    folder_path = os.path.join("data base", "repairs")
    file_path = os.path.join(folder_path, f"{user_id}_repairs.json")

    if not os.path.exists(file_path):
        return {"user_categories": [], "selected_transport": ""}

    try:
        with open(file_path, "r", encoding="utf-8") as file:
            data = json.load(file)
            data.setdefault("selected_transport", "")
            return data
    except:
        return {"user_categories": [], "selected_transport": ""}

#@log_user_actions
def get_user_repair_categories(user_id):
    data = load_repair_data(user_id)
    system_categories = ["Без категории", "ТО", "Ремонт", "Запчасть", "Диагностика", "Электрика", "Кузов"]
    user_categories = data.get("user_categories", [])
    return system_categories + user_categories

#@log_user_actions
def add_repair_category(user_id, new_category):
    data = load_repair_data(user_id)
    if "user_categories" not in data:
        data["user_categories"] = []
    if new_category not in data["user_categories"]:
        data["user_categories"].append(new_category)
    save_repair_data(user_id, data)

#@log_user_actions
def remove_repair_category(user_id, category_to_remove):
    data = load_repair_data(user_id)
    if "user_categories" in data and category_to_remove in data["user_categories"]:
        data["user_categories"].remove(category_to_remove)
        save_repair_data(user_id, data)

@bot.message_handler(func=lambda message: message.text == "Записать ремонт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Записать ремонт')
@track_usage('Записать ремонт')  # Добавление отслеживания статистики
#@log_user_actions
def record_repair(message):
    user_id = message.from_user.id
    if contains_media(message):
        sent = bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, record_repair)
        return

    markup = get_user_transport_keyboard(user_id)
    markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
    markup.add(types.KeyboardButton("В главное меню"))
    
    bot.send_message(user_id, "Выберите транспорт для записи ремонта:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_transport_selection_for_repair)

#@log_user_actions
def handle_transport_selection_for_repair(message):
    user_id = message.from_user.id

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    elif message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Добавить транспорт":
        add_transport(message)
        return

    selected_transport = message.text
    for transport in user_transport.get(str(user_id), []):
        formatted_transport = f"{transport['brand']} {transport['model']} ({transport['license_plate']})"
        if formatted_transport == selected_transport:
            brand, model, license_plate = transport['brand'], transport['model'], transport['license_plate']
            break
    else:
        # Отправляем сообщение об ошибке и добавляем кнопки возврата в меню
        bot.send_message(user_id, "Не удалось найти указанный транспорт. Пожалуйста, выберите снова")

        # Формируем клавиатуру с кнопками выбора транспорта и кнопками возврата в меню
        markup = get_user_transport_keyboard(user_id)
        markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
        markup.add(types.KeyboardButton("В главное меню"))

        bot.send_message(user_id, "Выберите транспорт или добавьте новый:", reply_markup=markup)
        bot.register_next_step_handler(message, handle_transport_selection_for_repair)
        return

    # Продолжение процесса после выбора транспорта
    process_category_selection_repair(user_id, brand, model, license_plate)

from functools import partial

#@log_user_actions
def process_category_selection_repair(user_id, brand, model, license_plate):
    categories = get_user_repair_categories(user_id)

    # Смайлы для категорий
    system_emoji = "🔹"
    user_emoji = "🔸"

    # Добавление смайлов к категориям: первые 7 считаются системными
    category_list = "\n".join(
        f"{system_emoji if i < 7 else user_emoji} {i + 1}. {category}"
        for i, category in enumerate(categories)
    )
    category_list = f"*Выберите категорию:*\n\n{category_list}"

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(
        types.KeyboardButton("Добавить категорию"),
        types.KeyboardButton("Удалить категорию")
    )
    markup.add(
        types.KeyboardButton("Вернуться в меню трат и ремонтов")
    )
    markup.add(
        types.KeyboardButton("В главное меню")
    )

    prompt_message = bot.send_message(user_id, category_list, reply_markup=markup, parse_mode="Markdown")
    
    # Используем partial для передачи дополнительных параметров
    bot.register_next_step_handler(prompt_message, partial(get_repair_category, brand=brand, model=model, license_plate=license_plate))

#@log_user_actions
def get_repair_category(message, brand, model, license_plate):
    user_id = message.from_user.id
    selected_index = message.text.strip() if message.text else ""

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, partial(get_repair_category, brand=brand, model=model, license_plate=license_plate))
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, partial(get_repair_category, brand=brand, model=model, license_plate=license_plate))
        return

    if selected_index == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    elif selected_index == "В главное меню":
        return_to_menu(message)
        return

    if selected_index == 'Добавить категорию':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
        markup.add(types.KeyboardButton("В главное меню"))
        bot.send_message(user_id, "Введите название новой категории:", reply_markup=markup)
        bot.register_next_step_handler(message, partial(add_new_repair_category, brand=brand, model=model, license_plate=license_plate))
        return

    if selected_index == 'Удалить категорию':
        handle_repair_category_removal(message, brand, model, license_plate)
        return

    if selected_index.isdigit():
        index = int(selected_index) - 1
        categories = get_user_repair_categories(user_id)
        if 0 <= index < len(categories):
            selected_category = categories[index]
            proceed_to_repair_name(message, selected_category, brand, model, license_plate)
        else:
            bot.send_message(user_id, "Неверный ввод категории. Попробуйте еще раз")
            bot.register_next_step_handler(message, partial(get_repair_category, brand=brand, model=model, license_plate=license_plate))
    else:
        bot.send_message(user_id, "Пожалуйста, введите номер категории")
        bot.register_next_step_handler(message, partial(get_repair_category, brand=brand, model=model, license_plate=license_plate))

#@log_user_actions
def add_new_repair_category(message, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на пустое или None значение
    if not message.text:
        bot.send_message(user_id, "Пожалуйста, введите название категории")
        bot.register_next_step_handler(message, partial(add_new_repair_category, brand=brand, model=model, license_plate=license_plate))
        return

    new_category = message.text.strip()

    system_categories = ["Без категории", "ТО", "Ремонт", "Запчасть", "Диагностика", "Электрика", "Кузов"]

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, partial(add_new_repair_category, brand=brand, model=model, license_plate=license_plate))
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, partial(add_new_repair_category, brand=brand, model=model, license_plate=license_plate))
        return


    if new_category in system_categories:
        bot.send_message(user_id, "Эта категория уже существует в системе")
        process_category_selection_repair(user_id, brand, model, license_plate)
        return

    # Сохранение новой категории
    data = load_repair_data(user_id)
    if new_category not in data["user_categories"]:
        data["user_categories"].append(new_category)
        # Сохраняем изменения в файл
        with open(os.path.join("data base", "repairs", f"{user_id}_repairs.json"), "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=4)

        bot.send_message(user_id, f"Категория *{new_category}* успешно добавлена!", parse_mode="Markdown")
    else:
        bot.send_message(user_id, "Эта категория уже существует")

    process_category_selection_repair(user_id, brand, model, license_plate)

#@log_user_actions
def handle_repair_category_removal(message, brand, model, license_plate):
    user_id = message.from_user.id
    categories = get_user_repair_categories(user_id)

    # Определение системных категорий
    system_categories = ["Без категории", "ТО", "Ремонт", "Запчасть", "Диагностика", "Электрика", "Кузов"]

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, partial(handle_repair_category_removal, brand=brand, model=model, license_plate=license_plate))
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, partial(handle_repair_category_removal, brand=brand, model=model, license_plate=license_plate))
        return

    if not categories:
        bot.send_message(user_id, "Нет доступных категорий для удаления")
        process_category_selection_repair(user_id, brand, model, license_plate)
        return

    # Смайлы для категорий
    system_emoji = "🔹"
    user_emoji = "🔸"

    # Создание текста с жирным шрифтом для сообщения с добавленными смайликами
    category_list = "\n".join(
        f"{system_emoji if category in system_categories else user_emoji} {i + 1}. {category}"
        for i, category in enumerate(categories)
    )
    bot.send_message(user_id, f"Выберите категорию для удаления или 0 для отмены:\n\n{category_list}")

    # Создание кнопок
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Вернуться в меню трат и ремонтов")
    markup.add("В главное меню")
    bot.send_message(user_id, "Выберите действие для категории:", reply_markup=markup)

    bot.register_next_step_handler(message, remove_repair_category, categories, system_categories, brand, model, license_plate)

#@log_user_actions
def remove_repair_category(message, categories, system_categories, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на нажатие кнопки '0' для отмены
    if message.text == "0":
        # Возвращаемся к выбору категории для записи
        process_category_selection_repair(user_id, brand, model, license_plate)
        return

    # Проверка на нажатие кнопок
    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    elif message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверяем, что message.text не является пустым или None
    if message.text:
        try:
            # Преобразуем текст в целое число
            index = int(message.text) - 1
            if 0 <= index < len(categories):
                removed_category = categories[index]

                # Проверка, является ли категория системной
                if removed_category in system_categories:
                    bot.send_message(user_id, "Это системная категория, удаление невозможно. Попробуйте еще раз")
                    # Ждем повторного ввода
                    bot.register_next_step_handler(message, remove_repair_category, categories, system_categories, brand, model, license_plate)
                    return

                # Удаление категории
                data = load_repair_data(user_id)
                data["user_categories"].remove(removed_category)

                # Сохраняем изменения в файл
                with open(os.path.join("data base", "repairs", f"{user_id}_repairs.json"), "w", encoding="utf-8") as file:
                    json.dump(data, file, ensure_ascii=False, indent=4)

                bot.send_message(user_id, f"Категория *{removed_category}* успешно удалена!", parse_mode="Markdown")
                # После успешного удаления, можно вернуть пользователя обратно к выбору категории
                process_category_selection_repair(user_id, brand, model, license_plate)

            else:
                bot.send_message(user_id, "Неверный номер категории. Попробуйте снова")
                # Ждем повторного ввода
                bot.register_next_step_handler(message, remove_repair_category, categories, system_categories, brand, model, license_plate)

        except ValueError:
            bot.send_message(user_id, "Пожалуйста, введите корректный номер категории")
            # Ждем повторного ввода
            bot.register_next_step_handler(message, remove_repair_category, categories, system_categories, brand, model, license_plate)
    else:
        bot.send_message(user_id, "Пожалуйста, введите номер категории")
        # Ждем повторного ввода
        bot.register_next_step_handler(message, remove_repair_category, categories, system_categories, brand, model, license_plate)

#@log_user_actions
def proceed_to_repair_name(message, selected_category, brand, model, license_plate):
    user_id = message.from_user.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите название ремонта:", reply_markup=markup)
    bot.register_next_step_handler(message, get_repair_name, selected_category, brand, model, license_plate)

#@log_user_actions
def get_repair_name(message, selected_category, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message,
            partial(get_repair_name, selected_category=selected_category, brand=brand, model=model, license_plate=license_plate)
        )
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message,
            partial(get_repair_name, selected_category=selected_category, brand=brand, model=model, license_plate=license_plate)
        )
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    repair_name = message.text
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_skip = types.KeyboardButton("Пропустить описание")  # Исправлено на "Пропустить описание"
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_skip)
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите описание ремонта или пропустите этот шаг:", reply_markup=markup)
    bot.register_next_step_handler(message, get_repair_description, selected_category, repair_name, brand, model, license_plate)

#@log_user_actions
def get_repair_description(message, selected_category, repair_name, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message,
            partial(get_repair_description, selected_category=selected_category, repair_name=repair_name, brand=brand, model=model, license_plate=license_plate)
        )
        return
    
    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message,
            partial(get_repair_description, selected_category=selected_category, repair_name=repair_name, brand=brand, model=model, license_plate=license_plate)
        )
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    repair_description = message.text if message.text != "Пропустить описание" else ""
    
    # Ввод даты
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите дату ремонта:", reply_markup=markup)
    bot.register_next_step_handler(message, get_repair_date, selected_category, repair_name, repair_description, brand, model, license_plate)

#@log_user_actions
def is_valid_date(date_str):
    # Проверяем формат даты: ДД.ММ.ГГГГ
    pattern = r'^(0[1-9]|[12][0-9]|3[01])\.(0[1-9]|1[0-2])\.(2000|20[01][0-9]|202[0-9]|203[0-9]|[2-9][0-9]{3})$'
    return bool(re.match(pattern, date_str))

#@log_user_actions
def get_repair_date(message, selected_category, repair_name, repair_description, brand, model, license_plate):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message, 
            partial(
                get_repair_date, 
                selected_category=selected_category, 
                repair_name=repair_name, 
                repair_description=repair_description, 
                brand=brand, 
                model=model, 
                license_plate=license_plate
            )
        )
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message, 
            partial(
                get_repair_date, 
                selected_category=selected_category, 
                repair_name=repair_name, 
                repair_description=repair_description, 
                brand=brand, 
                model=model, 
                license_plate=license_plate
            )
        )
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    repair_date = message.text

    # Проверка формата даты
    if not is_valid_date(repair_date):
        bot.send_message(user_id, "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(
            message, 
            get_repair_date, 
            selected_category, 
            repair_name, 
            repair_description, 
            brand, 
            model, 
            license_plate
        )
        return

    # Ввод суммы
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Введите сумму ремонта:", reply_markup=markup)
    bot.register_next_step_handler(
        message, 
        save_repair_data_final, 
        selected_category, 
        repair_name, 
        repair_description, 
        repair_date, 
        brand, 
        model, 
        license_plate, 
        f"{brand} {model} {license_plate}"  # Здесь передаем selected_transport
    )

def save_repair_data_final(message, selected_category, repair_name, repair_description, repair_date, brand, model, license_plate, selected_transport):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message,
            partial(
                save_repair_data_final,
                selected_category=selected_category,
                repair_name=repair_name,
                repair_description=repair_description,
                repair_date=repair_date,
                brand=brand,
                model=model,
                license_plate=license_plate,
                selected_transport=selected_transport
            )
        )
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(
            message,
            partial(
                save_repair_data_final,
                selected_category=selected_category,
                repair_name=repair_name,
                repair_description=repair_description,
                repair_date=repair_date,
                brand=brand,
                model=model,
                license_plate=license_plate,
                selected_transport=selected_transport
            )
        )
        return

    if message.text in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if message.text == "Вернуться в меню трат и ремонтов":
            return_to_menu_2(message)
        else:
            return_to_menu(message)
        return

    try:
        # Преобразуем введённую сумму в число с плавающей точкой
        repair_amount = float(message.text)

        # Формирование данных о ремонте
        repair_data = {
            "category": selected_category,
            "name": repair_name,
            "date": repair_date,
            "amount": repair_amount,
            "description": repair_description,
            "transport": {
                "brand": brand,
                "model": model,
                "license_plate": license_plate
            }
        }

        # Загрузка данных о ремонтах пользователя
        data = load_repair_data(user_id)
        if str(user_id) not in data:
            data[str(user_id)] = {"repairs": []}

        # Добавление нового ремонта
        repairs = data[str(user_id)].get("repairs", [])
        repairs.append(repair_data)
        data[str(user_id)]["repairs"] = repairs

        # Сохранение данных в базу и Excel
        save_repair_data(user_id, data, selected_transport)  # Сохранение в базу
        save_repair_to_excel(user_id, repair_data)          # Сохранение в Excel

        bot.send_message(user_id, "Ремонт успешно записан")
        send_menu(user_id)

    except ValueError:
        bot.send_message(user_id, "Пожалуйста, введите корректную сумму")
        bot.register_next_step_handler(
            message,
            partial(
                save_repair_data_final,
                selected_category=selected_category,
                repair_name=repair_name,
                repair_description=repair_description,
                repair_date=repair_date,
                brand=brand,
                model=model,
                license_plate=license_plate,
                selected_transport=selected_transport
            )
        )

def save_repair_to_excel(user_id, repair_data):
    # Путь к Excel-файлу пользователя для ремонта
    excel_path = os.path.join("data base", "repairs", "excel", f"{user_id}_repairs.xlsx")

    # Проверяем, существует ли директория
    directory = os.path.dirname(excel_path)
    if not os.path.exists(directory):
        os.makedirs(directory)  # Создаем директорию, если она не существует

    # Загружаем или создаём рабочую книгу
    try:
        if os.path.exists(excel_path):
            workbook = load_workbook(excel_path)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Удаляем стандартный лист
        
        # Лист для всех ремонтов (Summary)
        summary_sheet = workbook["Summary"] if "Summary" in workbook.sheetnames else workbook.create_sheet("Summary")
        
        # Лист для конкретного транспортного средства
        transport_sheet_name = f"{repair_data['transport']['brand']}_{repair_data['transport']['model']}_{repair_data['transport']['license_plate']}"
        if transport_sheet_name not in workbook.sheetnames:
            transport_sheet = workbook.create_sheet(transport_sheet_name)
        else:
            transport_sheet = workbook[transport_sheet_name]
        
        # Определяем заголовки
        headers = ["Транспорт", "Категория", "Название", "Дата", "Сумма", "Описание"]
        
        # Вспомогательная функция для настройки листов
        #@log_user_actions
        def setup_sheet(sheet):
            if sheet.max_row == 1:
                sheet.append(headers)
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

        # Добавляем заголовки и данные
        for sheet in [summary_sheet, transport_sheet]:
            setup_sheet(sheet)
            row_data = [
                f"{repair_data['transport']['brand']} {repair_data['transport']['model']} {repair_data['transport']['license_plate']}",
                repair_data["category"],
                repair_data["name"],
                repair_data["date"],
                float(repair_data["amount"]),  # Сохраняем сумму как число
                repair_data["description"],
            ]
            sheet.append(row_data)
        
        # Автоподгонка столбцов
        for sheet in [summary_sheet, transport_sheet]:
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Сохраняем рабочую книгу
        workbook.save(excel_path)
    except Exception as e:
        pass

        
# (11.5) --------------- КОД ДЛЯ "РЕМОНТОВ" (ОБРАБОТЧИК "ПОСМОТРЕТЬ РЕМОНТЫ") ---------------

selected_repair_transport_dict = {}

# Функция для фильтрации ремонтов по транспорту
#@log_user_actions
def filter_repairs_by_transport(user_id, repairs):
    selected_transport = selected_repair_transport_dict.get(user_id)
    
    # Если транспорт не выбран, возвращаем весь список ремонтов
    if not selected_transport:
        return repairs

    # Форматирование строки транспорта для фильтрации
    filtered_repairs = [
        repair for repair in repairs 
        if f"{repair['transport']['brand']} {repair['transport']['model']} ({repair['transport']['license_plate']})" == selected_transport
    ]
    return filtered_repairs

# Функция для отправки сообщений с разделением на части (общая)
#@log_user_actions
def send_message_with_split(user_id, message_text):
    bot.send_message(user_id, message_text, parse_mode="Markdown")

# Функция для отправки меню для ремонта
#@log_user_actions
def send_repair_menu(user_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Ремонты (месяц)")
    item2 = types.KeyboardButton("Ремонты (год)")
    item3 = types.KeyboardButton("Ремонты (все время)")
    item4 = types.KeyboardButton("Ремонты (по категориям)")
    item5 = types.KeyboardButton("Посмотреть ремонты в EXCEL")
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item4, item1)
    markup.add(item2, item3)
    markup.add(item5)
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Выберите вариант просмотра ремонтов:", reply_markup=markup)

# Обработчик для просмотра ремонтов
@bot.message_handler(func=lambda message: message.text == "Посмотреть ремонты")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть ремонты')
@track_usage('Посмотреть ремонты')  # Добавление отслеживания статистики
#@log_user_actions
def view_repairs(message):
    user_id = message.from_user.id

    # Загружаем данные о транспорте
    transport_list = load_transport_data(user_id)

    if not transport_list:
        bot.send_message(user_id, "У вас нет сохраненного транспорта. Хотите добавить транспорт?", reply_markup=create_transport_options_markup())
        bot.register_next_step_handler(message, ask_add_transport)
        return

    # Если транспорт есть, формируем клавиатуру
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    # Группируем кнопки по 2 на строку
    transport_buttons = [
        types.KeyboardButton(f"{transport['brand']} {transport['model']} ({transport['license_plate']})")
        for transport in transport_list
    ]
    for i in range(0, len(transport_buttons), 2):
        markup.add(*transport_buttons[i:i+2])

    # Добавляем кнопки возврата
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main)

    bot.send_message(user_id, "Выберите ваш транспорт для просмотра ремонтов:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_transport_selection_for_repairs)

# Обработчик выбора транспорта для ремонта
#@log_user_actions
def handle_transport_selection_for_repairs(message):
    user_id = message.from_user.id
    selected_transport = message.text

    # Обработка возврата в меню
    if selected_transport == "Вернуться в меню трат и ремонтов":
        send_menu(user_id)
        return

    if selected_transport == "В главное меню":
        return_to_menu(message)
        return

    # Сохраняем выбранный транспорт для пользователя
    selected_repair_transport_dict[user_id] = selected_transport

    # Уведомляем пользователя о выбранном транспорте
    bot.send_message(
        user_id,
        f"Показываю ремонты для транспорта: *{selected_transport}*",
        parse_mode="Markdown"
    )

    # Отображаем меню ремонтов
    send_repair_menu(user_id)


@bot.message_handler(func=lambda message: message.text == "Посмотреть ремонты в EXCEL")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть ремонты в EXCEL')
@track_usage('Посмотреть ремонты в EXCEL')  # Добавление отслеживания статистики
#@log_user_actions
def send_repairs_excel(message):
    user_id = message.from_user.id

    # Путь к Excel файлу
    excel_path = os.path.join("data base", "repairs", "excel", f"{user_id}_repairs.xlsx")

    # Проверяем наличие файла
    if not os.path.exists(excel_path):
        bot.send_message(user_id, "Файл с вашими ремонтами не найден")
        return

    # Отправка файла пользователю
    with open(excel_path, 'rb') as excel_file:
        bot.send_document(user_id, excel_file)

MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_message_with_split(user_id, message_text):
    """Отправляет сообщение пользователю, разбивая на части, если превышена максимальная длина."""
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text, parse_mode="Markdown")
    else:
        message_parts = [
            message_text[i:i + MAX_MESSAGE_LENGTH] 
            for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)
        ]
        for part in message_parts:
            bot.send_message(user_id, part, parse_mode="Markdown")

MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_message_with_split(user_id, message_text):
    """Отправляет сообщение пользователю, разбивая на части, если превышена максимальная длина."""
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text, parse_mode="Markdown")
    else:
        message_parts = [
            message_text[i:i + MAX_MESSAGE_LENGTH] 
            for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)
        ]
        for part in message_parts:
            bot.send_message(user_id, part, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Ремонты (по категориям)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Ремонты (по категориям)')
@track_usage('Ремонты (по категориям)')  # Добавление отслеживания статистики
#@log_user_actions
def view_repairs_by_category(message):
    user_id = message.from_user.id
    user_data = load_repair_data(user_id)
    repairs = user_data.get(str(user_id), {}).get("repairs", [])

    # Фильтруем ремонты по выбранному транспорту
    repairs = filter_repairs_by_transport(user_id, repairs)

    # Получаем уникальные категории ремонтов для выбранного транспорта
    categories = {repair['category'] for repair in repairs}  # Сохраняем категории как есть

    if not categories:
        bot.send_message(user_id, "*Нет доступных категорий* для выбранного транспорта", parse_mode="Markdown")
        send_repair_menu(user_id)  # Возврат в меню трат и ремонтов
        return

    category_buttons = [types.KeyboardButton(category) for category in categories]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(*category_buttons)
    item_return_to_repairs = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_return_to_main = types.KeyboardButton("В главное меню")
    markup.add(item_return_to_repairs)
    markup.add(item_return_to_main)

    bot.send_message(user_id, "Выберите категорию для просмотра ремонтов:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_repair_category_selection)

# Обработчик выбора категории ремонтов
#@log_user_actions
def handle_repair_category_selection(message):
    user_id = message.from_user.id
    selected_category = message.text.strip().lower()  # Преобразуем выбранную категорию в нижний регистр

    # Проверка на пустую строку или "Вернуться в меню"
    if not selected_category or selected_category == "вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return
    if selected_category == "в главное меню":
        return_to_menu(message)
        return

    # Загружаем данные
    user_data = load_repair_data(user_id)
    repairs = user_data.get(str(user_id), {}).get("repairs", [])

    # Фильтруем ремонты по выбранному транспорту
    repairs = filter_repairs_by_transport(user_id, repairs)

    # Преобразуем все категории в нижний регистр для корректного сравнения
    available_categories = {repair['category'].lower() for repair in repairs}

    # Проверяем, существует ли выбранная категория
    if selected_category not in available_categories:
        bot.send_message(user_id, "Выбранная категория не найдена. Пожалуйста, выберите корректную категорию")
        view_repairs_by_category(message)  # Запрашиваем повторный ввод категории
        return

    # Фильтруем ремонты по выбранной категории
    category_repairs = [repair for repair in repairs if repair['category'].lower() == selected_category]

    total_repairs_amount = 0
    repair_details = []

    for index, repair in enumerate(category_repairs, start=1):
        repair_name = repair.get("name", "Без названия")
        repair_date = repair.get("date", "")
        repair_amount = float(repair.get("amount", 0))
        total_repairs_amount += repair_amount

        repair_details.append(
            f"🔧 *№ {index}*\n\n"
            f"📂 *Категория:* {repair['category']}\n"  # Отображаем категорию как есть
            f"📌 *Название:* {repair_name}\n"
            f"📅 *Дата:* {repair_date}\n"
            f"💰 *Сумма:* {repair_amount:.2f} руб.\n"
            f"📝 *Описание:* {repair.get('description', 'Без описания')}\n"
        )

    if repair_details:
        message_text = f"*Ремонты* в категории *{selected_category}*:\n\n\n" + "\n\n".join(repair_details)
        send_message_with_split(user_id, message_text)
        bot.send_message(
            user_id,
            f"Итоговая сумма ремонтов в категории *{selected_category}*: *{total_repairs_amount:.2f}* руб.",
            parse_mode="Markdown"
        )
    else:
        bot.send_message(
            user_id,
            f"В категории *{selected_category}* ремонтов не найдено",
            parse_mode="Markdown"
        )

    # Возвращаем пользователя в меню выбора категорий
    send_repair_menu(user_id) # Запрашиваем выбор категории заново

# Обработчик ремонтов за месяц
MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_message_with_split(user_id, message_text):
    """Отправляет сообщение пользователю, разбивая на части, если превышена максимальная длина."""
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text, parse_mode="Markdown")
    else:
        message_parts = [
            message_text[i:i + MAX_MESSAGE_LENGTH] 
            for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)
        ]
        for part in message_parts:
            bot.send_message(user_id, part, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Ремонты (месяц)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Ремонты (месяц)')
@track_usage('Ремонты (месяц)')  # Добавление отслеживания статистики
#@log_user_actions
def view_repairs_by_month(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main)

    bot.send_message(user_id, "Введите месяц и год (ММ.ГГГГ) для просмотра ремонтов за этот период:", reply_markup=markup)
    bot.register_next_step_handler(message, get_repairs_by_month)

#@log_user_actions
def get_repairs_by_month(message):
    user_id = message.from_user.id
    date = message.text.strip() if message.text else None

    # Проверяем, есть ли текст в сообщении
    if not date:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_repairs_by_month)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_repairs_by_month)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):
        bot.send_message(user_id, "Извините, но отправка смайликов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_repairs_by_month)
        return

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверка формата и диапазонов месяца и года
    if "." in date:
        parts = date.split(".")
        if len(parts) == 2:
            month, year = parts
            if month.isdigit() and year.isdigit() and 1 <= int(month) <= 12 and 2000 <= int(year) <= 3000:
                month, year = map(int, parts)

                user_data = load_repair_data(user_id)
                repairs = user_data.get(str(user_id), {}).get("repairs", [])

                # Фильтруем ремонты по выбранному транспорту
                repairs = filter_repairs_by_transport(user_id, repairs)

                total_repairs = 0
                repair_details = []

                for index, repair in enumerate(repairs, start=1):
                    repair_date = repair.get("date", "")
                    repair_date_parts = repair_date.split(".")
                    if len(repair_date_parts) >= 2:
                        repair_month, repair_year = map(int, repair_date_parts[1:3])

                        if repair_month == month and repair_year == year:
                            amount = float(repair.get("amount", 0))
                            total_repairs += amount

                            repair_name = repair.get("name", "Без названия")
                            description = repair.get("description", "")
                            category = repair.get("category", "Без категории")

                            repair_details.append(
                                f"🔧 *№ {index}*\n\n"
                                f"📂 *Категория:* {category}\n"
                                f"📌 *Название:* {repair_name}\n"
                                f"📅 *Дата:* {repair_date}\n"
                                f"💰 *Сумма:* {amount} руб.\n"
                                f"📝 *Описание:* {description}\n"
                            )

                if repair_details:
                    message_text = f"Ремонты за *{date}* месяц:\n\n\n" + "\n\n".join(repair_details)
                    send_message_with_split(user_id, message_text)
                    bot.send_message(user_id, f"Итоговая сумма ремонтов за *{date}* месяц: *{total_repairs}* руб.", parse_mode="Markdown")
                else:
                    bot.send_message(user_id, f"За *{date}* месяц ремонтов не найдено", parse_mode="Markdown")

                send_repair_menu(user_id)  # Возвращаемся в меню после отображения информации
            else:
                bot.send_message(user_id, "Пожалуйста, введите корректный месяц и год в формате ММ.ГГГГ")
                bot.register_next_step_handler(message, get_repairs_by_month)
                return
        else:
            bot.send_message(user_id, "Пожалуйста, введите дату в формате ММ.ГГГГ")
            bot.register_next_step_handler(message, get_repairs_by_month)
            return
    else:
        bot.send_message(user_id, "Пожалуйста, введите дату в формате ММ.ГГГГ")
        bot.register_next_step_handler(message, get_repairs_by_month)
        return

# Обработчик ремонтов за год
MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_message_with_split(user_id, message_text):
    """Отправляет сообщение пользователю, разбивая на части, если превышена максимальная длина."""
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text, parse_mode="Markdown")
    else:
        message_parts = [
            message_text[i:i + MAX_MESSAGE_LENGTH] 
            for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)
        ]
        for part in message_parts:
            bot.send_message(user_id, part, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Ремонты (год)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Ремонты (год)')
@track_usage('Ремонты (год)')  # Добавление отслеживания статистики
#@log_user_actions
def view_repairs_by_year(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main)

    bot.send_message(user_id, "Введите год в формате (ГГГГ) для просмотра ремонтов за этот год:", reply_markup=markup)
    bot.register_next_step_handler(message, get_repairs_by_year)

#@log_user_actions
def get_repairs_by_year(message):
    user_id = message.from_user.id

    # Проверяем, что сообщение содержит текст
    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_repairs_by_year)
        return

    year_input = message.text.strip()

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_repairs_by_year)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, get_repairs_by_year)
        return

    if message.text == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Проверка, что введено четырехзначное число
    if not year_input.isdigit() or len(year_input) != 4:
        bot.send_message(user_id, "Пожалуйста, введите год в формате ГГГГ.")
        bot.register_next_step_handler(message, get_repairs_by_year)
        return

    # Преобразуем год в число и проверяем диапазон
    year = int(year_input)
    if year < 2000 or year > 3000:
        bot.send_message(user_id, "Введите год в формате ГГГГ.")
        bot.register_next_step_handler(message, get_repairs_by_year)
        return

    user_data = load_repair_data(user_id)
    repairs = user_data.get(str(user_id), {}).get("repairs", [])

    # Фильтруем ремонты по выбранному транспорту
    repairs = filter_repairs_by_transport(user_id, repairs)

    total_repairs = 0
    repair_details = []

    for index, repair in enumerate(repairs, start=1):
        repair_date = repair.get("date", "")
        if "." in repair_date:
            date_parts = repair_date.split(".")
            if len(date_parts) >= 3:
                repair_year = int(date_parts[2])
                if repair_year == year:
                    amount = float(repair.get("amount", 0))
                    total_repairs += amount

                    repair_name = repair.get("name", "Без названия")
                    description = repair.get("description", "")
                    category = repair.get("category", "Без категории")

                    repair_details.append(
                        f"🔧 *№ {index}*\n\n"
                        f"📂 *Категория:* {category}\n"
                        f"📌 *Название:* {repair_name}\n"
                        f"📅 *Дата:* {repair_date}\n"
                        f"💰 *Сумма:* {amount} руб.\n"
                        f"📝 *Описание:* {description}\n"
                    )

    if repair_details:
        message_text = f"Ремонты за *{year}* год:\n\n\n" + "\n\n".join(repair_details)
        send_message_with_split(user_id, message_text)
        bot.send_message(user_id, f"Итоговая сумма ремонтов за *{year}* год: *{total_repairs}* руб.", parse_mode="Markdown")
    else:
        bot.send_message(user_id, f"За *{year}* год ремонтов не найдено", parse_mode="Markdown")

    # Возвращаемся в меню после отображения информации
    send_repair_menu(user_id)
    
# Обработчик всех ремонтов
MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_message_with_split(user_id, message_text):
    """Отправляет сообщение пользователю, разбивая на части, если превышена максимальная длина."""
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text, parse_mode="Markdown")
    else:
        message_parts = [
            message_text[i:i + MAX_MESSAGE_LENGTH] 
            for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)
        ]
        for part in message_parts:
            bot.send_message(user_id, part, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Ремонты (все время)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Ремонты (все время)')
@track_usage('Ремонты (все время)')  # Добавление отслеживания статистики
#@log_user_actions
def view_all_repairs(message):
    user_id = message.from_user.id

    user_data = load_repair_data(user_id)
    repairs = user_data.get(str(user_id), {}).get("repairs", [])

    # Фильтруем ремонты по выбранному транспорту
    repairs = filter_repairs_by_transport(user_id, repairs)

    total_repairs = 0
    repair_details = []

    for index, repair in enumerate(repairs, start=1):
        amount = float(repair.get("amount", 0))
        total_repairs += amount

        repair_name = repair.get("name", "Без названия")
        repair_date = repair.get("date", "")
        description = repair.get("description", "")
        category = repair.get("category", "Без категории")

        repair_details.append(
            f"🔧 *№ {index}*\n\n"
            f"📂 *Категория:* {category}\n"
            f"📌 *Название:* {repair_name}\n"
            f"📅 *Дата:* {repair_date}\n"
            f"💰 *Сумма:* {amount} руб.\n"
            f"📝 *Описание:* {description}\n"
        )

    if repair_details:
        message_text = "*Все* ремонты:\n\n\n" + "\n\n".join(repair_details)
        send_message_with_split(user_id, message_text)
        bot.send_message(user_id, f"Итоговая сумма всех ремонтов: *{total_repairs}* руб.", parse_mode="Markdown")
    else:
        bot.send_message(user_id, "Ремонтов не найдено", parse_mode="Markdown")

    # Возвращаемся в меню после отображения информации
    send_repair_menu(user_id)

# (11.9) --------------- КОД ДЛЯ "РЕМОНТОВ" (ОБРАБОТЧИК "УДАЛИТЬ РЕМОНТЫ") ---------------

selected_repair_transports = {}

# Сохранение выбранного транспорта для ремонтов
def save_selected_repair_transport(user_id, selected_transport):
    user_data = load_repair_data(user_id).get(str(user_id), {})
    user_data["selected_transport"] = selected_transport
    save_repair_data(user_id, {str(user_id): user_data})

@bot.message_handler(func=lambda message: message.text == "Удалить ремонты")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить ремонты')
@track_usage('Удалить ремонты')  # Добавление отслеживания статистики
#@log_user_actions
def delete_repairs_menu(message):
    user_id = message.from_user.id

    # Загружаем данные о транспорте
    transport_data = load_transport_data(user_id)
    if not transport_data:
        # Если транспортов нет, предлагаем добавить
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_add_transport = types.KeyboardButton("Добавить транспорт")
        item_cancel = types.KeyboardButton("Вернуться в меню трат и ремонтов")
        markup.add(item_add_transport)
        markup.add(item_cancel)
        bot.send_message(user_id, "У вас нет сохраненного транспорта. Хотите добавить транспорт?", reply_markup=markup)
        bot.register_next_step_handler(message, ask_add_transport)
        return

    # Формируем клавиатуру с транспортами
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    transport_buttons = [
        types.KeyboardButton(f"{transport['brand']} {transport['model']} ({transport['license_plate']})")
        for transport in transport_data
    ]
    for i in range(0, len(transport_buttons), 2):
        markup.add(*transport_buttons[i:i+2])  # Группировка по 2 кнопки в строке

    # Добавляем кнопки возврата
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    bot.send_message(user_id, "Выберите транспорт для удаления ремонтов:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_repair_transport_selection_for_deletion)

#@log_user_actions
def handle_repair_transport_selection_for_deletion(message):
    user_id = message.from_user.id
    selected_transport = message.text.strip()

    # Проверка на возврат в меню
    if selected_transport == "Вернуться в меню трат и ремонтов":
        send_menu(user_id)
        return
    if selected_transport == "В главное меню":
        return_to_menu(message)
        return

    # Сохраняем выбранный транспорт
    selected_repair_transports[user_id] = selected_transport
    save_selected_repair_transport(user_id, selected_transport)

    # Меню выбора опции удаления
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_month = types.KeyboardButton("Del ремонты (месяц)")
    item_license_plate = types.KeyboardButton("Del ремонты (год)")
    item_all_time = types.KeyboardButton("Del ремонты (все время)")
    item_del_category_rep = types.KeyboardButton("Del ремонты (категория)")
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")

    markup.add(item_del_category_rep, item_month)
    markup.add(item_license_plate, item_all_time)
    markup.add(item_return)
    markup.add(item_main_menu)
    bot.send_message(user_id, "Выберите вариант удаления ремонтов:", reply_markup=markup)

repairs_to_delete_dict = {}
selected_repair_categories = {}
user_repairs_to_delete = {}

@bot.message_handler(func=lambda message: message.text == "Del ремонты (категория)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del ремонты (категория)')
@track_usage('Del ремонты (категория)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_repairs_by_category(message):
    user_id = message.from_user.id
    selected_transport = selected_repair_transports.get(user_id)

    if not selected_transport:
        selected_transport = load_repair_data(user_id).get("selected_transport")
        if not selected_transport:
            bot.send_message(user_id, "Не выбран транспорт")
            send_menu(user_id)
            return

    selected_transport_info = selected_transport.strip().lower().replace('(', '').replace(')', '')

    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    categories = list({
        repair.get("category")
        for repair in repairs
        if f"{repair.get('transport', {}).get('brand', '').strip().lower()} {repair.get('transport', {}).get('model', '').strip().lower()} {str(repair.get('transport', {}).get('license_plate', '')).strip().lower()}" == selected_transport_info
    })

    if not categories:
        bot.send_message(user_id, "Нет категорий ремонтов для выбранного транспорта")
        send_menu(user_id)
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    markup.add(*[types.KeyboardButton(category) for category in categories])
    markup.add(types.KeyboardButton("Вернуться в меню трат и ремонтов"))
    markup.add(types.KeyboardButton("В главное меню"))

    bot.send_message(user_id, "Выберите категорию для удаления ремонтов:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_repair_category_selection_for_deletion)

#@log_user_actions
def handle_repair_category_selection_for_deletion(message):
    user_id = message.from_user.id
    selected_category = message.text.strip()

    if selected_category in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if selected_category == "Вернуться в меню трат и ремонтов":
            send_menu(user_id)
        else:
            return_to_menu(message)
        return

    selected_repair_categories[user_id] = selected_category

    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    selected_transport = selected_repair_transports.get(user_id)
    if not selected_transport:
        selected_transport = load_repair_data(user_id).get("selected_transport")
        if not selected_transport:
            bot.send_message(user_id, "Не выбран транспорт")
            send_menu(user_id)
            return

    selected_transport_info = selected_transport.strip().lower().replace('(', '').replace(')', '')

    repairs_to_delete = [
        repair for repair in repairs
        if repair.get("category") == selected_category and
           f"{repair.get('transport', {}).get('brand', '').strip().lower()} {repair.get('transport', {}).get('model', '').strip().lower()} {str(repair.get('transport', {}).get('license_plate', '')).strip().lower()}" == selected_transport_info
    ]

    if not repairs_to_delete:
        bot.send_message(user_id, f"Нет ремонтов в категории *{selected_category.lower()}* для выбранного транспорта", parse_mode="Markdown")
        send_menu(user_id)
        return

    user_repairs_to_delete[user_id] = repairs_to_delete

    repair_list_text = f"Список ремонтов в категории *{selected_category.lower()}*:\n\n\n"
    for index, repair in enumerate(repairs_to_delete, start=1):
        repair_name = repair.get("name", "Без названия")
        repair_date = repair.get("date", "Неизвестно")
        repair_list_text += f"🔧 №{index}. {repair_name} - ({repair_date})\n\n"

    repair_list_text += "\nВведите номер ремонта для удаления:"

    send_long_message(user_id, repair_list_text)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Вернуться в меню трат и ремонтов")
    markup.add("В главное меню")

    bot.register_next_step_handler(message, delete_repair_confirmation)

#@log_user_actions
def delete_repair_confirmation(message):
    user_id = message.from_user.id

    selected_option = message.text.strip()

    if selected_option in ["Вернуться в меню трат и ремонтов", "В главное меню"]:
        if selected_option == "Вернуться в меню трат и ремонтов":
            send_menu(user_id)
        else:
            return_to_menu(message)
        return

    if not selected_option.isdigit():
        bot.send_message(user_id, "Введите корректный номер ремонта из списка")
        bot.register_next_step_handler(message, delete_repair_confirmation)
        return

    repair_index = int(selected_option) - 1
    repairs_to_delete = user_repairs_to_delete.get(user_id, [])

    if 0 <= repair_index < len(repairs_to_delete):
        user_data = load_repair_data(user_id).get(str(user_id), {})
        user_repairs = user_data.get("repairs", [])

        deleted_repair = repairs_to_delete.pop(repair_index)
        user_repairs.remove(deleted_repair)
        user_data["repairs"] = user_repairs
        save_repair_data(user_id, {str(user_id): user_data})

        bot.send_message(
            user_id,
            f"Ремонт *{deleted_repair.get('name', 'Без названия')}* успешно удален", parse_mode="Markdown"
        )

        send_menu(user_id)
    else:
        bot.send_message(user_id, "Неверный номер ремонта. Попробуйте снова")
        bot.register_next_step_handler(message, delete_repair_confirmation)

# Удаление ремонтов за месяц
@bot.message_handler(func=lambda message: message.text == "Del ремонты (месяц)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del ремонты (месяц)')
@track_usage('Del ремонты (месяц)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_repair_by_month(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)
    bot.send_message(user_id, "Введите месяц и год (ММ.ГГГГ) для удаления ремонтов за этот месяц:", reply_markup=markup)
    bot.register_next_step_handler(message, delete_repairs_by_month_handler)

#@log_user_actions
def delete_repairs_by_month_handler(message):
    user_id = message.from_user.id
    month_year = message.text.strip() if message.text else None

    if not month_year:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_repairs_by_month_handler)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_repairs_by_month_handler)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', month_year):
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_repairs_by_month_handler)
        return

    # Проверка формата месяца и года
    match = re.match(r"^(0[1-9]|1[0-2])\.(20[0-9]{2})$", month_year)
    if not match:
        bot.send_message(user_id, "Введен неверный месяц или год. Пожалуйста, введите корректные данные (ММ.ГГГГ)")
        bot.register_next_step_handler(message, delete_repairs_by_month_handler)
        return

    selected_month, selected_year = match.groups()

    # Проверка выбранного транспорта
    selected_transport = selected_repair_transports.get(user_id)
    if not selected_transport:
        bot.send_message(user_id, "Транспорт не выбран. Пожалуйста, выберите транспорт")
        send_menu(user_id)
        return

    transport_info = selected_transport.split(" ")
    selected_brand = transport_info[0].strip()
    selected_model = transport_info[1].strip()
    selected_license_plate = transport_info[2].strip().replace("(", "").replace(")", "")

    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    if not repairs:
        bot.send_message(user_id, f"У вас нет сохраненных ремонтов за *{month_year}* месяц", parse_mode="Markdown")
        send_menu(user_id)
        return

    repairs_to_delete = []
    for index, repair in enumerate(repairs, start=1):
        repair_date = repair.get("date", "")

        if not repair_date or len(repair_date.split(".")) != 3:
            continue

        repair_day, repair_month, repair_year = repair_date.split(".")

        if repair_month == selected_month and repair_year == selected_year:
            repair_license_plate = repair.get("transport", {}).get("license_plate", "").strip()
            repair_brand = repair.get("transport", {}).get("brand", "").strip()
            repair_model = repair.get("transport", {}).get("model", "").strip()

            if (repair_license_plate == selected_license_plate and
                repair_brand == selected_brand and
                repair_model == selected_model):
                repairs_to_delete.append((index, repair))

    if repairs_to_delete:
        repairs_to_delete_dict[user_id] = repairs_to_delete

        repair_list_text = f"Список ремонтов для удаления за *{month_year}* месяц:\n\n\n"
        for index, repair in repairs_to_delete:
            repair_name = repair.get("name", "Без названия")
            repair_date = repair.get("date", "Неизвестно")
            repair_list_text += f"🔧 №{index}. {repair_name} - ({repair_date})\n\n"

        repair_list_text += "\nВведите номер ремонта для удаления:"

        send_long_message(user_id, repair_list_text)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("Вернуться в меню трат и ремонтов")
        markup.add("В главное меню")
        bot.register_next_step_handler(message, confirm_delete_repair_month)
    else:
        bot.send_message(user_id, f"Нет ремонтов для удаления за *{month_year}* месяц", parse_mode="Markdown")
        send_menu(user_id)

#@log_user_actions
def confirm_delete_repair_month(message):
    user_id = message.from_user.id

    if not message.text:
        bot.send_message(user_id, "Извините, отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_repair_month)
        return

    selected_option = message.text.strip()

    if selected_option == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if selected_option == "В главное меню":
        return_to_menu(message)
        return

    if selected_option.isdigit():
        repair_index = int(selected_option) - 1
        repairs_to_delete = repairs_to_delete_dict.get(user_id, [])

        if 0 <= repair_index < len(repairs_to_delete):
            _, deleted_repair = repairs_to_delete.pop(repair_index)

            user_data = load_repair_data(user_id).get(str(user_id), {})
            if "repairs" in user_data and deleted_repair in user_data["repairs"]:
                user_data["repairs"].remove(deleted_repair)
                save_repair_data(user_id, {str(user_id): user_data})

            bot.send_message(
                user_id,
                f"Ремонт *{deleted_repair.get('name', 'Без названия').lower()}* успешно удален",
                parse_mode="Markdown"
            )

            repairs_to_delete_dict[user_id] = repairs_to_delete
            send_menu(user_id)
        else:
            bot.send_message(user_id, "Неверный выбор. Пожалуйста, попробуйте снова")
            bot.register_next_step_handler(message, confirm_delete_repair_month)
    else:
        bot.send_message(user_id, "Некорректный ввод. Пожалуйста, введите номер ремонта")
        bot.register_next_step_handler(message, confirm_delete_repair_month)

# Удаление ремонтов за год
MAX_MESSAGE_LENGTH = 4096

#@log_user_actions
def send_long_message(user_id, message_text):
    """Отправляет сообщение пользователю, разбивая на части, если оно слишком длинное."""
    if len(message_text) <= MAX_MESSAGE_LENGTH:
        bot.send_message(user_id, message_text, parse_mode="Markdown")
    else:
        message_parts = [
            message_text[i:i + MAX_MESSAGE_LENGTH] 
            for i in range(0, len(message_text), MAX_MESSAGE_LENGTH)
        ]
        for part in message_parts:
            bot.send_message(user_id, part, parse_mode="Markdown")

@bot.message_handler(func=lambda message: message.text == "Del ремонты (год)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del ремонты (год)')
@track_usage('Del ремонты (год)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_repairs_by_year(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)
    bot.send_message(user_id, "Введите год (ГГГГ) для удаления ремонтов за этот год:", reply_markup=markup)
    bot.register_next_step_handler(message, delete_repairs_by_year_handler)

#@log_user_actions
def delete_repairs_by_year_handler(message):
    user_id = message.from_user.id
    year = message.text.strip() if message.text else None

    if not year:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_repairs_by_year_handler)
        return

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_repairs_by_year_handler)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', year):  
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, delete_repairs_by_year_handler)
        return

    if year == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if year == "В главное меню":
        return_to_menu(message)
        return

    # Проверка формата года (от 2000 до 3000)
    if not re.match(r"^(20[0-9]{2}|2[1-9][0-9]{2}|3000)$", year):
        bot.send_message(user_id, "Введен неверный год. Пожалуйста, введите корректный год")
        bot.register_next_step_handler(message, delete_repairs_by_year_handler)
        return

    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    # Получаем данные выбранного транспорта
    selected_transport = selected_repair_transports.get(user_id, None)

    if selected_transport:
        transport_info = selected_transport.split(" ")
        selected_brand = transport_info[0].strip()
        selected_model = transport_info[1].strip()
        selected_license_plate = str(transport_info[2].strip())
    else:
        selected_brand = selected_model = selected_license_plate = None

    if not repairs:
        bot.send_message(user_id, f"У вас пока нет сохраненных ремонтов за *{year}* год для удаления", parse_mode="Markdown")
        send_menu(user_id)
        return

    repairs_to_delete = []
    for index, repair in enumerate(repairs, start=1):
        repair_date = repair.get("date", "")
        repair_year = repair_date.split(".")[2]

        repair_brand = repair.get("transport", {}).get("brand", "").strip()
        repair_model = repair.get("transport", {}).get("model", "").strip()

        if (repair_year == year and
           repair_brand == selected_brand and
           repair_model == selected_model):
            repairs_to_delete.append((index, repair))

    if repairs_to_delete:
        repairs_to_delete_dict[user_id] = repairs_to_delete

        repair_list_text = f"Список ремонтов для удаления за *{year}* год:\n\n\n"
        for index, repair in repairs_to_delete:
            repair_name = repair.get("name", "Без названия")
            repair_date = repair.get("date", "Неизвестно")
            repair_list_text += f"🔧 №{index}. {repair_name} - ({repair_date})\n\n"

        repair_list_text += "\nВведите номер ремонта для удаления:"
        send_long_message(user_id, repair_list_text)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add("Вернуться в меню трат и ремонтов")
        markup.add("В главное меню")
        bot.register_next_step_handler(message, confirm_delete_repair)
    else:
        bot.send_message(user_id, f"За *{year}* год ремонтов не найдено для удаления", parse_mode="Markdown")
        send_menu(user_id)

#@log_user_actions
def confirm_delete_repair(message):
    user_id = message.from_user.id

    if not message.text:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_repair)
        return

    selected_option = message.text.strip()

    if re.search(r'[^\w\s,.?!]', selected_option):
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_repair)
        return

    if selected_option == "Вернуться в меню трат и ремонтов":
        return_to_menu_2(message)
        return

    if selected_option == "В главное меню":
        return_to_menu(message)
        return

    repairs_to_delete = repairs_to_delete_dict.get(user_id, [])

    if selected_option.isdigit():
        repair_index = int(selected_option) - 1

        if 0 <= repair_index < len(repairs_to_delete):
            _, deleted_repair = repairs_to_delete.pop(repair_index)

            user_data = load_repair_data(user_id).get(str(user_id), {})
            user_data["repairs"].remove(deleted_repair)
            save_repair_data(user_id, {str(user_id): user_data})

            bot.send_message(user_id, f"Ремонт *{deleted_repair.get('name', 'Без названия').lower()}* успешно удален", parse_mode="Markdown")

            repairs_to_delete_dict[user_id] = repairs_to_delete
            send_menu(user_id)
        else:
            bot.send_message(user_id, "Неверный выбор. Пожалуйста, попробуйте снова")
            bot.register_next_step_handler(message, confirm_delete_repair)
    else:
        bot.send_message(user_id, "Некорректный ввод. Пожалуйста, введите номер ремонта")
        bot.register_next_step_handler(message, confirm_delete_repair)

# Удаление всех ремонтов
@bot.message_handler(func=lambda message: message.text == "Del ремонты (все время)")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del ремонты (все время)')
@track_usage('Del ремонты (все время)')  # Добавление отслеживания статистики
#@log_user_actions
def delete_all_repairs_for_selected_transport(message):
    user_id = message.from_user.id

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_return = types.KeyboardButton("Вернуться в меню трат и ремонтов")
    item_main_menu = types.KeyboardButton("В главное меню")
    markup.add(item_return)
    markup.add(item_main_menu)

    # Запрашиваем подтверждение
    bot.send_message(user_id,
                     "Вы уверены, что хотите удалить все ремонты для выбранного транспорта?\n\n"
                     "Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены",
                     reply_markup=markup, parse_mode="Markdown")
    bot.register_next_step_handler(message, confirm_delete_all_repairs)
    
#@log_user_actions
def confirm_delete_all_repairs(message):
    user_id = message.from_user.id

    # Проверка на мультимедиа
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_all_repairs)
        return

    # Проверка на смайлики
    if re.search(r'[^\w\s,.?!]', message.text):
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_all_repairs)
        return

    # Возврат в меню
    if message.text == "Вернуться в меню трат и ремонтов":
        send_menu(user_id)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    response = message.text.strip().lower()

    # Загружаем данные пользователя
    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    # Получаем выбранный транспорт
    selected_transport = user_data.get("selected_transport")
    if selected_transport:
        transport_info = selected_transport.split()
        selected_brand = transport_info[0].strip()
        selected_model = transport_info[1].strip()
        selected_license_plate = transport_info[2].strip()
    else:
        selected_brand = selected_model = selected_license_plate = None

    if not repairs:
        bot.send_message(user_id, "У вас пока нет сохраненных ремонтов для удаления")
        send_menu(user_id)
        return

    if response == "да":
        # Фильтруем ремонты для удаления
        repairs_to_keep = []
        for repair in repairs:
            repair_brand = repair.get("transport", {}).get("brand", "").strip()
            repair_model = repair.get("transport", {}).get("model", "").strip()

            # Сравниваем с выбранным транспортом
            if not (repair_brand == selected_brand and repair_model == selected_model):
                repairs_to_keep.append(repair)

        # Обновляем список ремонтов пользователя
        user_data["repairs"] = repairs_to_keep
        save_repair_data(user_id, {str(user_id): user_data})

        update_repairs_excel_file(user_id)

        # Сообщение об успешном удалении
        bot.send_message(user_id, f"Все ремонты для транспорта *{selected_brand} {selected_model} {selected_license_plate}* успешно удалены", parse_mode="Markdown")
    elif response == "нет":
        bot.send_message(user_id, "Удаление ремонтов отменено")
    else:
        # Ответ для неподходящих сообщений
        bot.send_message(user_id, "Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены.", parse_mode="Markdown")
        bot.register_next_step_handler(message, confirm_delete_all_repairs)
        return

    send_menu(user_id)

#@log_user_actions
def delete_repairs(user_id, deleted_repair):
    # Удаляем ремонт из базы данных
    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    # Удаляем ремонт из списка ремонтов
    repairs = [repair for repair in repairs if not (
        repair["transport"]["brand"] == deleted_repair["transport"]["brand"] and
        repair["transport"]["model"] == deleted_repair["transport"]["model"] and
        repair["transport"]["license_plate"] == deleted_repair["transport"]["license_plate"] and
        repair["category"] == deleted_repair["category"] and
        repair["name"] == deleted_repair["name"] and
        repair["date"] == deleted_repair["date"] and
        repair["amount"] == deleted_repair["amount"] and
        repair["description"] == deleted_repair["description"]
    )]

    # Обновляем данные пользователя
    user_data["repairs"] = repairs
    save_repair_data(user_id, user_data)

    # Обновляем Excel файл
    update_repairs_excel_file(user_id)

#@log_user_actions
def update_repairs_excel_file(user_id):
    user_data = load_repair_data(user_id).get(str(user_id), {})
    repairs = user_data.get("repairs", [])

    # Путь к Excel файлу пользователя
    excel_file_path = f"data base/repairs/excel/{user_id}_repairs.xlsx"

    # Проверяем существование файла, создаем новый, если нужно
    if not os.path.exists(excel_file_path):
        workbook = Workbook()
        workbook.save(excel_file_path)

    try:
        workbook = load_workbook(excel_file_path)
    except Exception as e:
        workbook = Workbook()
        workbook.save(excel_file_path)
        workbook = load_workbook(excel_file_path)

    # Обновление общего листа (Summary)
    if "Summary" not in workbook.sheetnames:
        summary_sheet = workbook.create_sheet("Summary")
    else:
        summary_sheet = workbook["Summary"]
        summary_sheet.delete_rows(2, summary_sheet.max_row)  # Очистка всех данных, кроме заголовков

    headers = ["Транспорт", "Категория", "Название", "Дата", "Сумма", "Описание"]

    if summary_sheet.max_row == 0:  # Добавляем заголовки, если их нет
        summary_sheet.append(headers)
        for cell in summary_sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for repair in repairs:
        transport = repair["transport"]
        row_data = [
            f"{transport['brand']} {transport['model']} {transport['license_plate']}",
            repair["category"],
            repair["name"],
            repair["date"],
            float(repair["amount"]),
            repair["description"],
        ]
        summary_sheet.append(row_data)

    # Обновление индивидуальных листов для каждого транспорта
    unique_transports = set((rep["transport"]["brand"], rep["transport"]["model"], rep["transport"]["license_plate"]) for rep in repairs)
    existing_sheets = set(workbook.sheetnames) - {"Summary"}

    # Удаление листов, которые больше не актуальны
    for sheet_name in existing_sheets:
        parts = sheet_name.split("_")
        if len(parts) < 3 or tuple(parts) not in unique_transports:
            del workbook[sheet_name]

    for brand, model, license_plate in unique_transports:
        sheet_name = f"{brand}_{model}_{license_plate}".replace(" ", "_")[:31]
        if sheet_name not in workbook.sheetnames:
            transport_sheet = workbook.create_sheet(sheet_name)
            transport_sheet.append(headers)
            for cell in transport_sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        else:
            transport_sheet = workbook[sheet_name]
            transport_sheet.delete_rows(2, transport_sheet.max_row)

        for repair in repairs:
            if (repair["transport"]["brand"], repair["transport"]["model"], repair["transport"]["license_plate"]) == (brand, model, license_plate):
                row_data = [
                    f"{brand} {model} {license_plate}",
                    repair["category"],
                    repair["name"],
                    repair["date"],
                    float(repair["amount"]),
                    repair["description"],
                ]
                transport_sheet.append(row_data)

    # Автоподгонка ширины колонок
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        for col in current_sheet.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            adjusted_width = max_length + 2
            current_sheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    workbook.save(excel_file_path)
    workbook.close()

# (12) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" ---------------

# (12.1) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ВВОДНЫЕ ФУНКЦИИ) ---------------

geolocator = Nominatim(user_agent="geo_bot")

user_locations = {}

# Функция для генерации ссылки на Yandex.Карты
#@log_user_actions
def get_yandex_maps_search_url(latitude, longitude, query):
    base_url = "https://yandex.ru/maps/?"
    search_params = {
        'll': f"{longitude},{latitude}",
        'z': 15,  # Уровень масштабирования карты
        'text': query,  # Тип объекта, например, 'АЗС'
        'mode': 'search'
    }
    query_string = "&".join([f"{key}={value}" for key, value in search_params.items()])
    return f"{base_url}{query_string}"

# Функция для сокращения URL
#@log_user_actions
def shorten_url(original_url):
    endpoint = 'https://clck.ru/--'
    response = requests.get(endpoint, params={'url': original_url})
    return response.text

# Обработчик для главного меню "Поиск мест"
@bot.message_handler(func=lambda message: message.text == "Поиск мест")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Поиск мест')
@track_usage('Поиск мест')  # Добавление отслеживания статистики
#@log_user_actions
def send_welcome(message):

    user_id = message.chat.id
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

    # Кнопки с категориями
    button_azs = types.KeyboardButton("АЗС")
    button_car_wash = types.KeyboardButton("Автомойки")
    button_auto_service = types.KeyboardButton("Автосервисы")
    button_parking = types.KeyboardButton("Парковки")
    button_evacuation = types.KeyboardButton("Эвакуация")
    button_gibdd_mreo = types.KeyboardButton("ГИБДД")
    button_accident_commissioner = types.KeyboardButton("Комиссары")
    button_impound = types.KeyboardButton("Штрафстоянка")  # Новая кнопка для штрафстоянки

    # Кнопка для возврата в главное меню
    item1 = types.KeyboardButton("В главное меню")

    # Добавление кнопок на клавиатуру
    markup.add(button_azs, button_car_wash, button_auto_service)
    markup.add(button_parking, button_evacuation, button_gibdd_mreo, button_accident_commissioner, button_impound)  # Добавляем кнопку штрафстоянки
    markup.add(item1)

    # Отправка пользователю
    bot.send_message(user_id, "Выберите категорию для ближайшего поиска:", reply_markup=markup)

# Обработчик для выбора категории заново
@bot.message_handler(func=lambda message: message.text == "Выбрать категорию заново")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Выбрать категорию заново')
#@log_user_actions
def handle_reset_category(message):
    global selected_category
    selected_category = None
    send_welcome(message)

selected_category = None 

# Обработчик для выбора категории
@bot.message_handler(func=lambda message: message.text in {"АЗС", "Автомойки", "Автосервисы", "Парковки", "Эвакуация", "ГИБДД", "Комиссары", "Штрафстоянка"})
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('АЗС')
@check_function_state_decorator('Автомойки')
@check_function_state_decorator('Автосервисы')
@check_function_state_decorator('Парковки')
@check_function_state_decorator('Эвакуация')
@check_function_state_decorator('ГИБДД')
@check_function_state_decorator('Комиссары')
@check_function_state_decorator('Штрафстоянка')
@track_usage('АЗС')
@track_usage('Автомойки')
@track_usage('Автосервисы')
@track_usage('Парковки')
@track_usage('Эвакуация')
@track_usage('ГИБДД')
@track_usage('Комиссары')
@track_usage('Штрафстоянка')
#@log_user_actions
def handle_menu_buttons(message):
    global selected_category 
    if message.text in {"АЗС", "Автомойки", "Автосервисы", "Парковки", "Эвакуация", "ГИБДД", "Комиссары", "Штрафстоянка"}:  # Включаем "Штрафстоянка"
        selected_category = message.text 
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button_send_location = types.KeyboardButton("Отправить геолокацию", request_location=True)
        button_reset_category = types.KeyboardButton("Выбрать категорию заново")
        item1 = types.KeyboardButton("В главное меню")
        keyboard.add(button_send_location)
        keyboard.add(button_reset_category)
        keyboard.add(item1)
        bot.send_message(message.chat.id, f"Отправьте свою геолокацию. Вам будет выдан список ближайших мест по категории - *{selected_category.lower()}:*", reply_markup=keyboard, parse_mode="Markdown")
    else:
        selected_category = None
        bot.send_message(message.chat.id, "Пожалуйста, выберите категорию из меню.")

# Обработчик для получения геолокации
@bot.message_handler(content_types=['location'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Функция для обработки локации')
@track_usage('Функция для обработки локации')  # Добавление отслеживания статистики
#@log_user_actions
def handle_location(message):
    global selected_category  # Указываем, что selected_category является глобальной переменной
    user_id = message.chat.id
    latitude = message.location.latitude
    longitude = message.location.longitude

    # Проверяем, включен ли режим анти-радара
    if user_tracking.get(user_id, {}).get('tracking', False):
        # Обновляем местоположение пользователя для анти-радара
        user_tracking[user_id]['location'] = message.location

        # Запускаем трекинг камер для анти-радара, если он не был запущен
        if not user_tracking[user_id].get('started', False):
            user_tracking[user_id]['started'] = True
            track_user_location(user_id, message.location)

        # Прерываем обработку, так как анти-радар активен
        return
    elif selected_category:
        # Логика обработки геолокации для поиска мест
        try:
            location = geolocator.reverse((latitude, longitude), language='ru', timeout=10)
            address = location.address

            # Создание ссылки на Yandex.Карты
            search_url = get_yandex_maps_search_url(latitude, longitude, selected_category)
            short_search_url = shorten_url(search_url)

            # Формирование сообщения
            message_text = f"🏙️ *Ближайшие {selected_category.lower()} по адресу:* \n\n{address}\n\n"
            message_text += f"🗺️ [Ссылка на карту]({short_search_url})"

            # Отправка сообщения
            bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

            # Сброс категории после использования
            selected_category = None

            # Клавиатура для выбора новой категории или возврата в главное меню
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            button_reset_category = types.KeyboardButton("Выбрать категорию заново")
            item1 = types.KeyboardButton("В главное меню")
            keyboard.add(button_reset_category)
            keyboard.add(item1)
            bot.send_message(message.chat.id, "Выберите категорию заново или вернитесь в главное меню", reply_markup=keyboard)

        except Exception as e:
            bot.send_message(message.chat.id, f"Произошла ошибка при обработке вашего запроса: {e}")
    else:
        # Если ни анти-радар, ни категория не выбраны, отправляем сообщение о выборе категории
        bot.send_message(message.chat.id, "Пожалуйста, выберите категорию из меню")


#------------------- (НАЧАЛО) КОД ДЛЯ ПОИСКА МЕСТ, РАБОЧИЙ, НЕДОДЕЛАННЫЙ В ПЛАНЕ КАТЕГОРИЙ И ПРОВЕРКИ ВЫДАЧИ, НУЖЕН API------------------------


# geolocator = Nominatim(user_agent="geo_bot")

# user_locations = {}

# #@log_user_actions
# def calculate_distance(origin, destination):
#     return geodesic(origin, destination).kilometers

# # (12.2) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (СОКРАЩЕНИЕ ССЫЛОК) ---------------

# #@log_user_actions
# def shorten_url(original_url):
#     endpoint = 'https://clck.ru/--'
#     response = requests.get(endpoint, params={'url': original_url})
#     return response.text

# # (12.3) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (АЗС) ---------------

# #@log_user_actions
# def get_nearby_fuel_stations(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "АЗС",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045", 
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     fuel_stations = []
#     for feature in data["features"]:
#         coordinates = feature["geometry"]["coordinates"]
#         name = feature["properties"]["name"]
#         address = feature["properties"]["description"]
#         fuel_stations.append({"name": name, "coordinates": coordinates, "address": address})

#     return fuel_stations

# # (12.4) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (АВТОМОЙКИ) ---------------

# #@log_user_actions
# def get_nearby_car_washes(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "Автомойка",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045", 
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     car_washes = []
#     for feature in data["features"]:
#         coordinates = feature["geometry"]["coordinates"]
#         name = feature["properties"]["name"]
#         address = feature["properties"]["description"]
#         car_washes.append({"name": name, "coordinates": coordinates, "address": address})

#     return car_washes

# # (12.5) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (АВТОСЕРВИС) ---------------

# #@log_user_actions
# def get_nearby_auto_services(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "Автосервис",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045",
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     auto_services = []
#     for feature in data.get("features", []):
#         coordinates = feature.get("geometry", {}).get("coordinates")
#         name = feature.get("properties", {}).get("name")
#         address = feature.get("properties", {}).get("description", "Адрес не указан")
#         if coordinates and name:
#             auto_services.append({"name": name, "coordinates": coordinates, "address": address})

#     return auto_services

# # (12.6) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ПАРКОВКИ) ---------------

# #@log_user_actions
# def get_nearby_parking(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "Парковки",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045",
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     parking_places = []
#     for feature in data["features"]:
#         coordinates = feature["geometry"]["coordinates"]
#         name = feature["properties"]["name"]
#         address = feature["properties"]["description"]
#         parking_places.append({"name": name, "coordinates": coordinates, "address": address})

#     return parking_places

# # (12.7) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ЭВАКУАТОР) ---------------

# #@log_user_actions
# def get_nearby_evacuation_services(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "Эвакуация",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045",
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     evacuation_services = []
#     for feature in data["features"]:
#         coordinates = feature["geometry"]["coordinates"]
#         name = feature["properties"]["name"]
#         address = feature["properties"]["description"]
#         evacuation_services.append({"name": name, "coordinates": coordinates, "address": address})

#     return evacuation_services

# # (12.8) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ГИБДД) ---------------

# #@log_user_actions
# def get_nearby_gibdd_mreo(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "ГИБДД",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045",
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     gibdd_mreo_offices = []
#     for feature in data["features"]:
#         coordinates = feature["geometry"]["coordinates"]
#         name = feature["properties"]["name"]
#         address = feature["properties"]["description"]
#         gibdd_mreo_offices.append({"name": name, "coordinates": coordinates, "address": address})

#     return gibdd_mreo_offices

# # (12.9) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (КОМИССАРЫ) ---------------

# #@log_user_actions
# def get_nearby_accident_commissioner(latitude, longitude, user_coordinates):
#     api_url = "https://search-maps.yandex.ru/v1/"
#     params = {
#         "apikey": "af145d41-4168-430b-b7d5-392df4d232cc",
#         "text": "Комиссары",
#         "lang": "ru_RU",
#         "ll": f"{longitude},{latitude}",
#         "spn": "0.045,0.045",
#         "type": "biz"
#     }
#     response = requests.get(api_url, params=params)
#     data = response.json()

#     accident_commissioners = []
#     for feature in data["features"]:
#         coordinates = feature["geometry"]["coordinates"]
#         name = feature["properties"]["name"]
#         address = feature["properties"]["description"]
#         accident_commissioners.append({"name": name, "coordinates": coordinates, "address": address})

#     return accident_commissioners

# # (12.10) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ОБРАБОТЧИК "ПОИСК МЕСТ") ---------------

# @bot.message_handler(func=lambda message: message.text == "Поиск мест")
# @restricted
# @track_user_activity
# @check_chat_state
# @track_usage('Поиск мест')  # Добавление отслеживания статистики
# #@log_user_actions
# def send_welcome(message):
#     user_id = message.chat.id
#     markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

#     button_azs = types.KeyboardButton("АЗС")
#     button_car_wash = types.KeyboardButton("Автомойки")
#     button_auto_service = types.KeyboardButton("Автосервисы")
#     button_parking = types.KeyboardButton("Парковки")
#     button_evacuation = types.KeyboardButton("Эвакуация")
#     button_gibdd_mreo = types.KeyboardButton("ГИБДД")
#     button_accident_commissioner = types.KeyboardButton("Комиссары")

#     item1 = types.KeyboardButton("В главное меню")

#     markup.add(button_azs, button_car_wash, button_auto_service)
#     markup.add(button_parking, button_evacuation, button_gibdd_mreo, button_accident_commissioner)
#     markup.add(item1)

#     bot.send_message(user_id, "Выберите категорию для ближайшего поиска:", reply_markup=markup)

# # (12.11) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ОБРАБОТЧИК "ВЫБОР КАТЕГОРИИ ЗАНОВО") ---------------

# @bot.message_handler(func=lambda message: message.text == "Выбрать категорию заново")
# @restricted
# @track_user_activity
# @check_chat_state
# #@log_user_actions
# def handle_reset_category(message):
#     global selected_category
#     selected_category = None
#     send_welcome(message)

# selected_category = None 

# # (12.12) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ОБРАБОТЧИК "ДЛЯ ВЫБОРА КАТЕГОРИИ") ---------------

# @bot.message_handler(func=lambda message: message.text in {"АЗС", "Автомойки", "Автосервисы", "Парковки", "Эвакуация", "ГИБДД", "Комиссары"})
# @restricted
# @track_user_activity
# @check_chat_state
# #@log_user_actions
# def handle_menu_buttons(message):
#     global selected_category 
#     if message.text in {"АЗС", "Автомойки", "Автосервисы", "Парковки", "Эвакуация", "ГИБДД", "Комиссары"}:
#         selected_category = message.text 
#         keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
#         button_send_location = types.KeyboardButton("Отправить геолокацию", request_location=True)
#         button_reset_category = types.KeyboardButton("Выбрать категорию заново")
#         item1 = types.KeyboardButton("В главное меню")
#         keyboard.add(button_send_location)
#         keyboard.add(button_reset_category)
#         keyboard.add(item1)
#         bot.send_message(message.chat.id, f"Отправьте свою геолокацию. Вам будет выдан список ближайших {selected_category.lower()}.", reply_markup=keyboard)
#     else:
#         selected_category = None
#         bot.send_message(message.chat.id, "Пожалуйста, выберите категорию из меню.")

# # (12.13) --------------- КОД ДЛЯ "ПОИСКА МЕСТ" (ОБРАБОТЧИК "ГЕОЛОКАЦИЯ") ---------------

# @bot.message_handler(content_types=['location'])
# @restricted
# @track_user_activity
# @check_chat_state
# @check_function_state_decorator('Функция для обработки локации')
# @track_usage('Функция для обработки локации')  # Добавление отслеживания статистики
# #@log_user_actions
# def handle_location(message):
#     global selected_category, selected_location, user_locations
#     latitude = message.location.latitude
#     longitude = message.location.longitude
#     user_id = message.from_user.id

#     try:
#         location = geolocator.reverse((latitude, longitude), language='ru', timeout=10)
#         address = location.address

#         if selected_category is None:
#             bot.send_message(message.chat.id, "Пожалуйста, выберите категорию из меню.")
#             return

#         user_locations[user_id] = {"address": address, "coordinates": (latitude, longitude)}

#         if selected_category == "АЗС":
#             locations = get_nearby_fuel_stations(latitude, longitude, user_locations[user_id]["coordinates"])
#         elif selected_category == "Автомойки":
#             locations = get_nearby_car_washes(latitude, longitude, user_locations[user_id]["coordinates"])
#         elif selected_category == "Автосервисы":
#             locations = get_nearby_auto_services(latitude, longitude, user_locations[user_id]["coordinates"])
#         elif selected_category == "Парковки":
#             locations = get_nearby_parking(latitude, longitude, user_locations[user_id]["coordinates"])
#         elif selected_category == "Эвакуация":
#             locations = get_nearby_evacuation_services(latitude, longitude, user_locations[user_id]["coordinates"])
#         elif selected_category == "ГИБДД":
#             locations = get_nearby_gibdd_mreo(latitude, longitude, user_locations[user_id]["coordinates"])
#         elif selected_category == "Комиссары":
#             locations = get_nearby_accident_commissioner(latitude, longitude, user_locations[user_id]["coordinates"])

#         selected_location = {"address": address, "coordinates": (latitude, longitude)}

#         # Формирование текста сообщения с использованием смайлов, гиперссылок и Markdown
#         message_text = f"📍 *Ближайшие объекты по адресу:*\n\n🏠 *{address}*\n\n"
#         message_text += f"🔍 *{selected_category}:*\n\n"

#         for location in locations:
#             name = location["name"]
#             coordinates = location["coordinates"]
#             address = location["address"]

#             # Создание ссылки на Яндекс.Карты
#             yandex_maps_url = f"https://yandex.ru/maps/?rtext={user_locations[user_id]['coordinates'][0]},{user_locations[user_id]['coordinates'][1]}~{coordinates[1]},{coordinates[0]}&l=map&rtt=auto&ruri=~ymapsbm1%3A%2F%2Forg%3Foid%3D1847883008%26name%3D{quote(name)}%26address%3D{quote(address)}\n\n"
#             short_yandex_maps_url = shorten_url(yandex_maps_url)

#             # Форматирование каждого пункта с гиперссылкой
#             message_text += f"🚗 *{name}* ({address}):\n[Посмотреть на карте]({short_yandex_maps_url})\n\n"

#         # Отправка отформатированного сообщения с Markdown
#         bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

#         selected_category = None

#         keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
#         button_reset_category = types.KeyboardButton("Выбрать категорию заново")
#         item1 = types.KeyboardButton("В главное меню")
#         keyboard.add(button_reset_category)
#         keyboard.add(item1)
#         bot.send_message(message.chat.id, "Выберите категорию заново или вернитесь в главное меню.", reply_markup=keyboard)

#     except Exception as e:
#         bot.send_message(message.chat.id, f"Произошла ошибка при обработке вашего запроса: {e}")


#-------------------(КОНЕЦ) КОД ДЛЯ ПОИСКА МЕСТ, РАБОЧИЙ, НЕДОДЕЛАННЫЙ В ПЛАНЕ КАТЕГОРИЙ И ПРОВЕРКИ ВЫДАЧИ, НУЖЕН API------------------------

#-----------------------------------------------------------------------------------------------------------------------------------


# (13) --------------- КОД ДЛЯ "ПОИСКА ТРАНСПОРТА" ---------------

# (13.1) --------------- КОД ДЛЯ "ПОИСКА ТРАНСПОРТА" (СОХРАНЕНИЕ И ЗАГРУЗКА ГЕОЛОКАЦИИ) ---------------

LATITUDE_KEY = 'latitude'
LONGITUDE_KEY = 'longitude'

def save_location_data(location_data):
    folder_path = "data base/findtransport"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Сохраняем JSON с отступами для удобства чтения
    with open(os.path.join(folder_path, "location_data.json"), "w") as json_file:
        json.dump(location_data, json_file, indent=4)


def load_location_data():
    try:
        with open(os.path.join("data base/findtransport", "location_data.json"), "r") as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

location_data = load_location_data()

# (13.2) --------------- КОД ДЛЯ "ПОИСКА ТРАНСПОРТА" (ОБРАБОТЧИК "НАЙТИ ТРАНСПОРТ") ---------------

# Обработчик для команды "Найти транспорт"
@bot.message_handler(func=lambda message: message.text == "Найти транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Найти транспорт')
@track_usage('Найти транспорт')  # Добавление отслеживания статистики
#@log_user_actions
def start_transport_search(message):

    global location_data
    user_id = str(message.from_user.id)

    # Если уже начат процесс (начальная геопозиция была сохранена)
    if user_id in location_data and location_data[user_id].get('start_location') is not None and location_data[user_id].get('end_location') is None:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Продолжить")
        item2 = types.KeyboardButton("Начать заново")
        item3 = types.KeyboardButton("В главное меню")
        markup.add(item1, item2)
        markup.add(item3)
        bot.send_message(message.chat.id, "Хотите продолжить с того места, где остановились?", reply_markup=markup)
        bot.register_next_step_handler(message, continue_or_restart)
    else:
        start_new_transport_search(message)

# Функция для начала нового поиска транспорта
#@log_user_actions
def start_new_transport_search(message):
    global location_data
    user_id = str(message.from_user.id)

    # Очищаем данные для нового поиска
    location_data[user_id] = {'start_location': None, 'end_location': None}
    save_location_data(location_data)

    request_transport_location(message)

# Функция для запроса геопозиции транспорта
#@log_user_actions
def request_transport_location(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Отправить геопозицию", request_location=True)
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    bot.send_message(message.chat.id, "Отправьте геопозицию транспорта:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_car_location)

# Обработка ответа на предложение продолжить или начать заново
#@log_user_actions
def continue_or_restart(message):
    if message.text == "Продолжить":
        # После продолжения вернем кнопки для отправки геопозиции
        request_user_location(message)
    elif message.text == "Начать заново":
        start_new_transport_search(message)
    else:
        return_to_menu(message)

# Функция для запроса геопозиции пользователя
#@log_user_actions
def request_user_location(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Отправить геопозицию", request_location=True)
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    bot.send_message(message.chat.id, "Теперь отправьте свою геопозицию:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_car_location)

# Обработка геопозиции транспорта и пользователя
@bot.message_handler(content_types=['location'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Функция для обработки локации')
@track_usage('Функция для обработки локации')  # Добавление отслеживания статистики
#@log_user_actions
def handle_car_location(message):
    global location_data
    user_id = str(message.from_user.id)

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if user_id not in location_data:
        location_data[user_id] = {'start_location': None, 'end_location': None}

    # Сохранение геопозиции транспорта
    if location_data[user_id]['start_location'] is None:
        if message.location is not None:
            location_data[user_id]['start_location'] = {
                LATITUDE_KEY: message.location.latitude,
                LONGITUDE_KEY: message.location.longitude
            }

            # Сохраняем, что процесс поиска начат
            location_data[user_id]['process'] = 'in_progress'
            save_location_data(location_data)

            # Запрашиваем геопозицию пользователя
            request_user_location(message)
        else:
            handle_location_error(message)

    # Сохранение геопозиции пользователя
    elif location_data[user_id]['end_location'] is None:
        if message.location is not None:
            location_data[user_id]['end_location'] = {
                LATITUDE_KEY: message.location.latitude,
                LONGITUDE_KEY: message.location.longitude
            }

            # Завершаем процесс, сохранив данные
            location_data[user_id]['process'] = 'completed'
            save_location_data(location_data)

            send_map_link(message.chat.id, location_data[user_id]['start_location'], location_data[user_id]['end_location'])
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton("Найти транспорт")
            item2 = types.KeyboardButton("В главное меню")
            markup.add(item1)
            markup.add(item2)
            bot.send_message(message.chat.id, "Вы можете найти транспорт еще раз", reply_markup=markup)
        else:
            handle_location_error(message)

# Функция обработки ошибки при получении геопозиции
#@log_user_actions
def handle_location_error(message):
    if message.text == "В главное меню":
        return_to_menu(message)
        return

    bot.send_message(message.chat.id, "Извините, не удалось получить геопозицию. Попробуйте еще раз")
    bot.register_next_step_handler(message, handle_car_location)

# Отправка карты с маршрутом
#@log_user_actions
def send_map_link(chat_id, start_location, end_location):
    map_url = f"https://yandex.ru/maps/?rtext={end_location['latitude']},{end_location['longitude']}~{start_location['latitude']},{start_location['longitude']}&rtt=pd"
    short_url = shorten_url(map_url)
    bot.send_message(chat_id, f"📍 *Маршрут для поиска:* [ссылка]({short_url})", parse_mode="Markdown")
    
# (14) --------------- КОД ДЛЯ "ПОИСК РЕГИОНА ПО ГОСНОМЕРУ" ---------------

ALLOWED_LETTERS = "АВЕКМНОРСТУХABEKMHOPCTYX"  # Допустимые буквы для российских госномеров

#@log_user_actions
def is_valid_car_number(car_number):
    """
    Проверяет, соответствует ли введенный госномер формату:
    1 буква, 3 цифры, 2 буквы, 2 или 3 цифры.
    """
    # Регулярное выражение для формата А123АА12 или А123АА123
    pattern = rf"^[{ALLOWED_LETTERS}]\d{{3}}[{ALLOWED_LETTERS}]{{2}}\d{{2,3}}$"
    return bool(re.match(pattern, car_number))

@bot.message_handler(func=lambda message: message.text == "Код региона")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Код региона')
@track_usage('Код региона')  # Добавление отслеживания статистики
#@log_user_actions
def handle_start4(message):

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    bot.send_message(
        message.chat.id,
        "Отправьте один или несколько государственных номеров (8 - 9 символов) или кодов региона (2 - 3 цифры), разделенных запятой:",
        reply_markup=markup
    )
    bot.register_next_step_handler(message, process_input)

#@log_user_actions
def process_input(message):
    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_input)
        return

    text = message.text.strip()
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Ввести еще")
    item2 = types.KeyboardButton("В главное меню")
    markup.add(item1)
    markup.add(item2)

    inputs = [i.strip() for i in text.split(',')]
    responses = []

    for input_item in inputs:
        if input_item.isdigit() and (2 <= len(input_item) <= 3):
            region_code = input_item
            if region_code in regions:
                region_name = regions[region_code]
                response = f"🔍 *Регион для кода {region_code}:*\n{region_name}\n\n\n"
            else:
                response = f"❌ Не удалось определить регион для кода: *{region_code}*\n\n\n"
        
        elif 8 <= len(input_item) <= 9 and is_valid_car_number(input_item.upper()):
            car_number = input_item.upper()
            region_code = car_number[-3:] if len(car_number) == 9 else car_number[-2:]

            if region_code in regions:
                region_name = regions[region_code]
                avtocod_url = f"https://avtocod.ru/proverkaavto/{car_number}?rd=GRZ"
                short_url = shorten_url(avtocod_url)
                
                response = (
                    f"🔍 Регион для номера `{car_number}`: {region_name}\n"
                    f"🔗 [Ссылка на AvtoCod с поиском]({short_url})\n\n\n"
                )
            else:
                response = f"❌ Не удалось определить регион для номера: `{car_number}`\n\n\n"
        
        else:
            response = f"❌ Неверный формат для `{input_item}`. Пожалуйста, введите правильный госномер или код региона\n\n\n"

        responses.append(response)

    final_response = "".join(responses)
    bot.send_message(message.chat.id, final_response, reply_markup=markup, parse_mode="Markdown")
    bot.send_message(message.chat.id, "Вы можете ввести еще или выйти в главное меню")
    bot.register_next_step_handler(message, handle_action_after_response)

#@log_user_actions
def handle_action_after_response(message):
    if message.text == "Ввести еще":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("В главное меню")
        markup.add(item1)
        bot.send_message(
            message.chat.id,
            "Отправьте один или несколько государственных номеров (8 - 9 символов) или кодов региона (2 - 3 цифры), разделённых запятой:",
            reply_markup=markup
        )
        bot.register_next_step_handler(message, process_input)
    elif message.text == "В главное меню":
        return_to_menu(message)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, выберите действие из меню:")
        bot.register_next_step_handler(message, handle_action_after_response)


# (15) --------------- КОД ДЛЯ "ПОГОДЫ" ---------------

# (15.1) --------------- КОД ДЛЯ "ПОГОДЫ" (ВВОДНЫЕ ФУНКЦИИ) ---------------

# API ключ от OpenWeatherMap
API_KEY = '2949ae1ef99c838462d16e7b0caf65b5'

WEATHER_URL = 'http://api.openweathermap.org/data/2.5/weather'

FORECAST_URL = 'http://api.openweathermap.org/data/2.5/forecast'

MAX_MESSAGE_LENGTH = 4096

user_data = {}

# (15.1) --------------- КОД ДЛЯ "ПОГОДЫ" (ОБРАБОТЧИК "ПОГОДА") ---------------

@bot.message_handler(func=lambda message: message.text == "Погода")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Погода')
@track_usage('Погода')  # Добавление отслеживания статистики
#@log_user_actions
def handle_start_5(message):
    try:
        markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(telebot.types.KeyboardButton("Отправить геопозицию", request_location=True))
        markup.row(telebot.types.KeyboardButton("В главное меню"))

        bot.send_message(message.chat.id, "Отправьте свою геопозицию для отображения погоды:", reply_markup=markup)
        bot.register_next_step_handler(message, handle_location_5)
    
    except Exception as e:
        bot.send_message(message.chat.id, "Произошла ошибка при обработке вашего запроса. Попробуйте позже")

@bot.message_handler(content_types=['location'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Функция для обработки локации')
@track_usage('Функция для обработки локации')  # Добавление отслеживания статистики
#@log_user_actions
def handle_location_5(message):
    try:
        if message.location:
            latitude = message.location.latitude
            longitude = message.location.longitude

            # Сохраняем координаты пользователя
            save_user_location(message.chat.id, latitude, longitude, None)  # city_code пока None

            markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
            markup.row('Сегодня', 'Завтра')
            markup.row('Неделя', 'Месяц')
            markup.row('Другое место')
            markup.row('В главное меню')

            bot.send_message(message.chat.id, "Выберите период:", reply_markup=markup)
        elif message.text == "В главное меню":
            return_to_menu(message)
        else:
            bot.send_message(message.chat.id, "Не удалось получить данные о местоположении. Пожалуйста, отправьте местоположение еще раз")
            bot.register_next_step_handler(message, handle_location_5)
    except Exception as e:
        bot.send_message(message.chat.id, "Произошла ошибка при обработке местоположения. Попробуйте позже")


import telebot
import threading
import schedule
import time
import json
import requests
import traceback
from datetime import datetime

# Путь к файлу notifications.json
notifications_file_path = 'data base/notifications/notifications.json'

def save_user_location(chat_id, latitude, longitude, city_code):
    # Загрузка существующих данных
    try:
        with open('data base/notifications/notifications.json', 'r') as f:
            notifications = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        notifications = {}

    # Проверка, существует ли запись для пользователя
    if str(chat_id) not in notifications:
        notifications[str(chat_id)] = {
            "latitude": None,
            "longitude": None,
            "city_code": None
        }

    # Обновление данных
    if latitude is not None:
        notifications[str(chat_id)]["latitude"] = latitude
    if longitude is not None:
        notifications[str(chat_id)]["longitude"] = longitude
    if city_code is not None:
        notifications[str(chat_id)]["city_code"] = city_code

    # Сохранение обратно в файл
    with open('data base/notifications/notifications.json', 'w') as f:
        json.dump(notifications, f, ensure_ascii=False, indent=4)

# Функции для загрузки координат и получения погоды
def load_user_locations():
    file_path = 'data base/notifications/notifications.json'
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        return {}

import traceback

#@log_user_actions
def get_city_name(latitude, longitude):
    try:
        geocode_url = "https://eu1.locationiq.com/v1/reverse.php"
        params = {
            'key': 'pk.fa5c52bb6b9e1b801d72b75d151aea63',  # Ваш API ключ LocationIQ
            'lat': latitude,
            'lon': longitude,
            'format': 'json',
            'accept-language': 'ru'
        }
        response = requests.get(geocode_url, params=params)
        data = response.json()

        if response.status_code == 200:
            city = data.get("address", {}).get("city", None)
            return city or f"неизвестное место ({latitude}, {longitude})"
        return None
    except:
        return None

#@log_user_actions
def get_current_weather(coords):
    try:
        city_name = get_city_name(coords['latitude'], coords['longitude'])
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric',
            'lang': 'ru'
        }
        response = requests.get(WEATHER_URL, params=params, timeout=30)
        data = response.json()

        if response.status_code == 200:
            temperature = round(data['main']['temp'])
            feels_like = round(data['main']['feels_like'])
            humidity = data['main']['humidity']
            pressure = data['main']['pressure']
            wind_speed = data['wind']['speed']
            description = translate_weather_description(data['weather'][0]['description'])

            current_time = datetime.now().strftime("%H:%M")
            current_date = datetime.now().strftime("%d.%m.%Y")

            return (
                f"*Вам пришло новое уведомление!*🔔\n\n\n"
                f"*Погода на {current_date} в {current_time}*:\n"
                f"*(г. {city_name}; {coords['latitude']}, {coords['longitude']})*\n\n"
                f"🌡️ *Температура:* {temperature}°C\n"
                f"🌬️ *Ощущается как:* {feels_like}°C\n"
                f"💧 *Влажность:* {humidity}%\n"
                f"〽️ *Давление:* {pressure} мм рт. ст.\n"
                f"💨 *Скорость ветра:* {wind_speed} м/с\n"
                f"☁️ *Описание:* {description}\n\n"
            )
        else:
            return None
    except:
        return None

#@log_user_actions
def get_average_fuel_prices(city_code):
    fuel_prices = {}
    file_path = f'data base/azs/{city_code}_table_azs_data.json'

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            prices_data = json.load(f)

            for entry in prices_data:
                fuel_type = entry[1]  # Тип топлива
                price = entry[2]  # Цена топлива

                try:
                    price = float(price)  # Преобразуем цену в число (если это возможно)
                except ValueError:
                    continue  # Пропускаем этот элемент, если цена не может быть преобразована

                if fuel_type not in fuel_prices:
                    fuel_prices[fuel_type] = []

                fuel_prices[fuel_type].append(price)

    except FileNotFoundError:
        pass
        return None
    except json.JSONDecodeError:
        pass
        return None

    # Вычисление средних цен
    average_prices = {fuel: sum(prices) / len(prices) for fuel, prices in fuel_prices.items()}

    return average_prices

def load_city_names(file_path):
    city_names = {}
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                # Разделяем строку на название города и его код
                city_data = line.strip().split(' - ')
                if len(city_data) == 2:
                    city_name, city_code = city_data
                    city_names[city_code] = city_name  # Добавляем в словарь
    except FileNotFoundError:
        pass
    except Exception as e:
        pass
    
    return city_names

#@log_user_actions
def send_weather_notifications():
    user_locations = load_user_locations()
    city_names = load_city_names('files/combined_cities.txt')  # Загружаем названия городов
    
    for chat_id, coords in user_locations.items():
        weather_message = get_current_weather(coords)
        
        if weather_message:
            city_code = coords.get('city_code')
            city_name = city_names.get(city_code, city_code)  # Получаем название города
            average_prices = get_average_fuel_prices(city_code)

             # Получаем текущую дату и время
            current_time = datetime.now().strftime("%d.%m.%Y в %H:%M") 

            fuel_prices_message = ""
            if average_prices:
                fuel_prices_message = "\n*Актуальные цены на топливо (г. {}) на дату {}:*\n\n".format(city_name, current_time)
                for fuel_type, price in average_prices.items():
                    fuel_prices_message += f"⛽ *{fuel_type}:* {price:.2f} руб./л.\n"

            try:
                bot.send_message(chat_id, weather_message + fuel_prices_message, parse_mode="Markdown")
            except Exception as e:
                pass

# Настройка расписания для отправки уведомлений
schedule.every().day.at("07:30").do(send_weather_notifications)
schedule.every().day.at("13:00").do(send_weather_notifications)
schedule.every().day.at("17:00").do(send_weather_notifications)
schedule.every().day.at("20:00").do(send_weather_notifications)

@bot.message_handler(func=lambda message: message.text in ['Сегодня', 'Завтра', 'Неделя', 'Месяц', 'Другое место'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Сегодня')
@check_function_state_decorator('Завтра')
@check_function_state_decorator('Неделя')
@check_function_state_decorator('Месяц')
@check_function_state_decorator('Другое место')
@track_usage('Сегодня')
@track_usage('Завтра')
@track_usage('Неделя')
@track_usage('Месяц')
@track_usage('Другое место')
#@log_user_actions
def handle_period_5(message):
    period = message.text.lower()
    chat_id = message.chat.id
    user_locations = load_user_locations()  # Загрузите пользовательские местоположения
    coords = user_locations.get(str(chat_id))  # Получите координаты для текущего пользователя

    if coords is None:
        bot.send_message(chat_id, "Не удалось получить координаты. Пожалуйста, отправьте местоположение еще раз")
        return

    if period == 'Другое место':
        user_data.pop(chat_id, None)
        handle_start_5(message)
        return

    if message.text == "Другое место":
        handle_start_5(message)
        return

    if coords:
        if period == 'сегодня':
            send_weather(chat_id, coords, WEATHER_URL)
        elif period == 'завтра':
            send_forecast_daily(chat_id, coords, FORECAST_URL, 1)
        elif period == 'неделя':
            send_forecast_weekly(chat_id, coords, FORECAST_URL, 8)
        elif period == 'месяц':
            send_forecast_monthly(chat_id, coords, FORECAST_URL, 31)

#@log_user_actions
def send_weather(chat_id, coords, url):
    try:
        # Запрос текущей погоды
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric'  # Изменено на 'metric' для получения температуры в Цельсиях
        }
        response = requests.get(url, params=params)
        data = response.json()

        if response.status_code == 200:
            temperature = round(data['main']['temp'])
            feels_like = round(data['main']['feels_like'])
            humidity = data['main']['humidity']
            pressure = data['main']['pressure']
            wind_speed = data['wind']['speed']
            description = translate_weather_description(data['weather'][0]['description'])

            current_time = datetime.now().strftime("%H:%M")
            current_date = datetime.now().strftime("%d.%m.%Y")

            message = (
                f"*Погода на {current_date} в {current_time}:*\n\n"
                f"🌡️ *Температура:* {temperature}°C\n"
                f"🌬️ *Ощущается как:* {feels_like}°C\n"
                f"💧 *Влажность:* {humidity}%\n"
                f"〽️ *Давление:* {pressure} мм рт. ст.\n"
                f"💨 *Скорость ветра:* {wind_speed} м/с\n"
                f"☁️ *Описание:* {description}\n"
            )
            bot.send_message(chat_id, message, parse_mode="Markdown")
            send_forecast_remaining_day(chat_id, coords, FORECAST_URL)
        else:
            bot.send_message(chat_id, "Не удалось получить текущую погоду. Проверьте, правильно ли указаны координаты")
    except Exception as e:
        bot.send_message(chat_id, "Произошла ошибка при запросе текущей погоды. Попробуйте позже")

#@log_user_actions
def send_forecast_remaining_day(chat_id, coords, url):
    try:
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric',
            'lang': 'ru'
        }
        response = requests.get(url, params=params, timeout=30)
        data = response.json()

        if response.status_code == 200:
            forecasts = data['list']
            now = datetime.now()
            message = "*Прогноз на оставшуюся часть дня:*\n\n"

            for forecast in forecasts:
                date_time = datetime.strptime(forecast['dt_txt'], "%Y-%m-%d %H:%M:%S")
                
                # Только прогноз на сегодня
                if now.date() == date_time.date() and date_time > now:
                    formatted_date = date_time.strftime("%d.%m.%Y")
                    formatted_time = date_time.strftime("%H:%M")
                    temperature = round(forecast['main']['temp'])
                    feels_like = round(forecast['main']['feels_like'])
                    humidity = forecast['main']['humidity']
                    pressure = forecast['main']['pressure']
                    wind_speed = forecast['wind']['speed']
                    description = translate_weather_description(forecast['weather'][0]['description'])

                    message += (
                        f"*Погода на {formatted_date} в {formatted_time}:*\n\n"
                        f"🌡️ *Температура:* {temperature}°C\n"
                        f"🌬 *Ощущается как:* {feels_like}°C\n"
                        f"💧 *Влажность:* {humidity}%\n"
                        f"〽️ *Давление:* {pressure} мм рт. ст.\n"
                        f"💨 *Скорость ветра:* {wind_speed} м/с\n"
                        f"☁️ *Описание:* {description}\n\n"
                    )

            if message == "*Прогноз на оставшуюся часть дня:*\n\n":
                message = "Нет доступного прогноза на оставшуюся часть дня"

            bot.send_message(chat_id, message, parse_mode="Markdown")
        else:
            bot.send_message(chat_id, "Не удалось получить прогноз на оставшуюся часть дня")
    except Exception as e:
        bot.send_message(chat_id, "Произошла ошибка при запросе прогноза на оставшуюся часть дня. Попробуйте позже")


# (15.6) --------------- КОД ДЛЯ "ПОГОДЫ" (ФУНКЦИЯ ПЕРЕВОДА) ---------------

#@log_user_actions
def translate_weather_description(english_description):
    translation_dict = {
        'clear sky': 'ясное небо',
        'few clouds': 'небольшая облачность',
        'scattered clouds': 'рассеянные облака',
        'broken clouds': 'облачно с прояснениями',
        'shower rain': 'небольшой дождь',
        'rain': 'дождь',
        'thunderstorm': 'гроза',
        'snow': 'снег',
        'mist': 'туман',
        'light snow': 'небольшой снег',
        'overcast clouds': 'пасмурно',
        'light snow': 'небольшой снег',
        'snow': 'снег',
        'heavy snow': 'сильный снегопад',
    }

    return translation_dict.get(english_description, english_description)

# (15.7) --------------- КОД ДЛЯ "ПОГОДЫ" (ФУНКЦИЯ ОТПРАВКИ ПРОГНОЗА) ---------------

#@log_user_actions
def send_forecast(chat_id, coords, url, days=1):
    try:
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
        }
        response = requests.get(url, params=params)
        data = response.json()

        if response.status_code == 200:
            forecasts = data['list'][:days * 8]
            message = "*Прогноз на завтра:*\n"

            for forecast in forecasts:
                date_time = datetime.strptime(forecast['dt_txt'], "%Y-%m-%d %H:%M:%S")
                formatted_date = date_time.strftime("%d-%m-%Y %H:%M:%S")
                temperature = round(forecast['main']['temp'] - 273.15)
                feels_like = round(forecast['main']['feels_like'] - 273.15)
                humidity = forecast['main']['humidity']
                pressure = forecast['main']['pressure']
                wind_speed = forecast['wind']['speed']
                description = translate_weather_description(forecast['weather'][0]['description'])

                message += (
                    f"{formatted_date}:\n"
                    f"🌡️ *Температура:* {temperature}°C\n"
                    f"🌬️ *Ощущается как:* {feels_like}°C\n"
                    f"💧 *Влажность:* {humidity}%\n"
                    f"〽️ *Давление:* {pressure} мм рт. ст\n"
                    f"💨 *Скорость ветра:* {wind_speed} м/с\n"
                    f"☁️ *Описание:* {description}\n\n"
                )

            message_chunks = [message[i:i + MAX_MESSAGE_LENGTH] for i in range(0, len(message), MAX_MESSAGE_LENGTH)]

            for chunk in message_chunks:
                bot.send_message(chat_id, chunk, parse_mode="Markdown")
        else:
            bot.send_message(chat_id, "Не удалось получить прогноз погоды на завтра")
    except Exception as e:
        bot.send_message(chat_id, "Произошла ошибка при запросе прогноза на завтра. Попробуйте позже")

# (15.8) --------------- КОД ДЛЯ "ПОГОДЫ" (ФУНКЦИЯ ПОГОДЫ НА ЗАВТРА) ---------------

# ---------- ПРОГНОЗ НА ЗАВТРА -----------
#@log_user_actions
def send_forecast_daily(chat_id, coords, url, days_ahead):
    try:
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric',
            'lang': 'ru'
        }
        response = requests.get(url, params=params, timeout=30)
        data = response.json()

        if response.status_code == 200:
            forecast = data['list'][days_ahead * 8]
            date_time = datetime.strptime(forecast['dt_txt'], "%Y-%m-%d %H:%M:%S")
            temperature = round(forecast['main']['temp'])
            feels_like = round(forecast['main']['feels_like'])
            humidity = forecast['main']['humidity']
            pressure = forecast['main']['pressure']
            wind_speed = forecast['wind']['speed']
            description = translate_weather_description(forecast['weather'][0]['description'])

            message = (
                f"*Прогноз на {date_time.strftime('%d.%m.%Y')}*\n\n"
                f"🌡️ *Температура:* {temperature}°C\n"
                f"🌬️ *Ощущается как:* {feels_like}°C\n"
                f"💧 *Влажность:* {humidity}%\n"
                f"〽️ *Давление:* {pressure} мм рт. ст.\n"
                f"💨 *Скорость ветра:* {wind_speed} м/с\n"
                f"☁️ *Описание:* {description}\n"
            )
            bot.send_message(chat_id, message, parse_mode="Markdown")

            send_hourly_forecast_tomorrow(chat_id, coords, url)
        else:
            bot.send_message(chat_id, "Не удалось получить прогноз на завтра")
    except Exception as e:
        bot.send_message(chat_id, "Произошла ошибка при запросе прогноза на завтра. Попробуйте позже")

# Функция для отправки почасового прогноза на завтра
#@log_user_actions
def send_hourly_forecast_tomorrow(chat_id, coords, url):
    try:
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric',
            'lang': 'ru'
        }
        response = requests.get(url, params=params, timeout=30)
        data = response.json()

        if response.status_code == 200:
            forecasts = data['list']
            now = datetime.now()
            tomorrow = now + timedelta(days=1)
            message = "*Почасовой прогноз на завтра:*\n\n\n"  # Добавлен жирный шрифт и пустая строка

            for forecast in forecasts:
                date_time = datetime.strptime(forecast['dt_txt'], "%Y-%m-%d %H:%M:%S")
                if tomorrow.date() == date_time.date():
                    formatted_time = date_time.strftime("%H:%M")
                    formatted_date = date_time.strftime("%d.%m.%Y")  # Форматирование даты
                    temperature = round(forecast['main']['temp'])
                    feels_like = round(forecast['main']['feels_like'])
                    humidity = forecast['main']['humidity']
                    pressure = forecast['main']['pressure']
                    wind_speed = forecast['wind']['speed']
                    description = translate_weather_description(forecast['weather'][0]['description'])

                    message += (
                        f"*Погода на {formatted_date} в {formatted_time}:*\n\n"
                        f"🌡️ *Температура:* {temperature}°C\n"
                        f"🌬️ *Ощущается как:* {feels_like}°C\n"
                        f"💧 *Влажность:* {humidity}%\n"
                        f"〽️ *Давление:* {pressure} мм рт. ст.\n"
                        f"💨 *Скорость ветра:* {wind_speed} м/с\n"
                        f"☁️ *Описание:* {description}\n\n"  # Пустая строка
                    )

            if message == "*Почасовой прогноз на завтра:*\n\n":
                message = "Нет доступного почасового прогноза на завтра"

            message_chunks = [message[i:i + MAX_MESSAGE_LENGTH] for i in range(0, len(message), MAX_MESSAGE_LENGTH)]
            for chunk in message_chunks:
                bot.send_message(chat_id, chunk, parse_mode="Markdown")  # Убедитесь, что используется Markdown
        else:
            bot.send_message(chat_id, "Не удалось получить почасовой прогноз на завтра")
    except Exception as e:
        bot.send_message(chat_id, "Произошла ошибка при запросе почасового прогноза на завтра! Попробуйте позже")
        
# (15.9) --------------- КОД ДЛЯ "ПОГОДЫ" (ФУНКЦИЯ ПОГОДЫ НА НЕДЕЛЮ) ---------------


# ---------- ПРОГНОЗ НА НЕДЕЛЮ -----------

from datetime import datetime, timedelta
from collections import defaultdict

#@log_user_actions
def send_forecast_weekly(chat_id, coords, url, retries=3):
    try:
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric',
            'lang': 'ru'
        }

        daily_forecasts = defaultdict(list)
        message = "*Прогноз на неделю:*\n\n\n"

        for attempt in range(retries):
            try:
                response = requests.get(url, params=params, timeout=10)

                if response.status_code == 200:
                    data = response.json()
                    forecasts = data['list']

                    # Собираем прогнозы за 7 дней
                    for forecast in forecasts:
                        date_time = datetime.strptime(forecast['dt_txt'], "%Y-%m-%d %H:%M:%S")
                        date_str = date_time.strftime('%d.%m.%Y')

                        if len(daily_forecasts) >= 7 and date_str not in daily_forecasts:
                            break

                        temperature = round(forecast['main']['temp'])
                        feels_like = round(forecast['main']['feels_like'])
                        humidity = forecast['main']['humidity']
                        pressure = forecast['main']['pressure']
                        wind_speed = forecast['wind']['speed']
                        description = translate_weather_description(forecast['weather'][0]['description'])

                        daily_forecasts[date_str].append({
                            'temperature': temperature,
                            'feels_like': feels_like,
                            'humidity': humidity,
                            'pressure': pressure,
                            'wind_speed': wind_speed,
                            'description': description
                        })

                    for date, forecasts in daily_forecasts.items():
                        temp_sum = sum(f['temperature'] for f in forecasts)
                        feels_like_sum = sum(f['feels_like'] for f in forecasts)
                        count = len(forecasts)
                        avg_temp = round(temp_sum / count)
                        avg_feels_like = round(feels_like_sum / count)

                        message += (
                            f"*Погода на {date}:*\n\n"
                            f"🌡️ *Температура:* {avg_temp}°C\n"
                            f"🌬️ *Ощущается как:* {avg_feels_like}°C\n"
                            f"💧 *Влажность:* {forecasts[0]['humidity']}%\n"
                            f"〽️ *Давление:* {forecasts[0]['pressure']} мм рт. ст.\n"
                            f"💨 *Скорость ветра:* {forecasts[0]['wind_speed']} м/с\n"
                            f"☁️ *Описание:* {forecasts[0]['description']}\n\n"
                        )

                    bot.send_message(chat_id, message, parse_mode="Markdown")
                    break
                else:
                    bot.send_message(chat_id, "Не удалось получить прогноз на неделю")
                    break
            except Exception as e:
                if attempt == retries - 1:
                    bot.send_message(chat_id, "Не удалось получить прогноз на неделю после нескольких попыток")
    except Exception as e:
        bot.send_message(chat_id, "Произошла ошибка при запросе прогноза на неделю. Попробуйте позже")

# ---------- ПРОГНОЗ НА МЕСЯЦ -----------
from collections import defaultdict
import requests
from datetime import datetime, timedelta
import traceback

#@log_user_actions
def send_forecast_monthly(chat_id, coords, url, days=31):
    try:
        params = {
            'lat': coords['latitude'],
            'lon': coords['longitude'],
            'appid': API_KEY,
            'units': 'metric',
            'lang': 'ru'
        }
        response = requests.get(url, params=params, timeout=30)
        data = response.json()

        if response.status_code == 200:
            forecasts = data['list']
            message = "*Прогноз на месяц:*\n\n\n"

            daily_forecasts = {}
            for forecast in forecasts:
                date_time = datetime.strptime(forecast['dt_txt'], "%Y-%m-%d %H:%M:%S")
                date_str = date_time.strftime('%d.%m.%Y')

                if date_str not in daily_forecasts:
                    daily_forecasts[date_str] = {
                        'temperature': round(forecast['main']['temp']),
                        'feels_like': round(forecast['main']['feels_like'])
                    }

            for date, values in daily_forecasts.items():
                message += (
                    f"*{date}:*\n\n"
                    f"🌡️ *Температура:* {values['temperature']}°C\n"
                    f"🌬️ *Ощущается как:* {values['feels_like']}°C\n\n"
                )

            unavailable_dates = [
                date for date in pd.date_range(start=datetime.now(), periods=days).strftime('%d.%m.%Y')
                if date not in daily_forecasts
            ]

            if unavailable_dates:
                start_date = unavailable_dates[0]
                end_date = unavailable_dates[-1]
                message += (
                    f"*С {start_date} по {end_date}:*\n\n_"
                    f"Данные недоступны из-за ограничений_\n\n"
                )

            if message == "*Прогноз на месяц:*\n\n":
                message = "Нет доступного прогноза на месяц"

            bot.send_message(chat_id, message, parse_mode="Markdown")
        else:
            bot.send_message(chat_id, "Не удалось получить прогноз на месяц")
    except:
        bot.send_message(chat_id, "Произошла ошибка при запросе прогноза на месяц. Попробуйте позже")

# ЦЕНЫ НА ТОПЛИВО

import threading

# Переменные для состояния пользователя и данных
user_state = {}
user_data = {}

# Создаём необходимые папки
os.makedirs(os.path.join('data base', 'azs'), exist_ok=True)
os.makedirs(os.path.join('data base', 'cityforprice'), exist_ok=True)

# Путь к файлу для сохранения данных
DATA_FILE_PATH = os.path.join('data base', 'cityforprice', 'city_for_the_price.json')

# Функция для загрузки данных пользователей
def load_citys_users_data():
    global user_data
    if os.path.exists(DATA_FILE_PATH):
        with open(DATA_FILE_PATH, 'r', encoding='utf-8') as f:
            user_data = json.load(f)
    else:
        user_data = {}

# Функция для сохранения данных пользователей
def save_citys_users_data():
    with open(DATA_FILE_PATH, 'w', encoding='utf-8') as f:
        json.dump(user_data, f, ensure_ascii=False, indent=4)

# Загружаем данные пользователей при запуске
load_citys_users_data()

# Функция для создания имени файла на основе города и даты
#@log_user_actions
def create_filename(city_code, date):
    date_str = date.strftime('%d_%m_%Y')
    return f"{city_code}_table_azs_data_{date_str}.json"

# Функция для сохранения данных в JSON
def save_fuel_data(city_code, fuel_prices):
    filename = f'{city_code}_table_azs_data.json'  # Уникальный файл для каждого города
    filepath = os.path.join('data base', 'azs', filename)  # Указываем путь к папке

    # Сохраняем данные в файл, заменяя содержимое
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(fuel_prices, f, ensure_ascii=False, indent=4)

# Функция для загрузки сохранённых данных из JSON
def load_saved_data(city_code):
    filename = f'{city_code}_table_azs_data.json'  # Уникальный файл для каждого города
    filepath = os.path.join('data base', 'azs', filename)  # Указываем путь к папке

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data
    except FileNotFoundError:
        return None

# Функция для чтения городов и создания словаря для их URL
def load_cities():
    cities = {}
    try:
        with open('files/combined_cities.txt', 'r', encoding='utf-8') as file:
            for line in file:
                if '-' in line:  # Проверяем, что в строке есть дефис
                    city, city_code = line.strip().split(' - ')
                    cities[city.lower()] = city_code  # Приводим название города к нижнему регистру
    except FileNotFoundError:
        pass  # Если файл не найден, просто продолжаем
    return cities

# Загружаем города при запуске бота
cities_dict = load_cities()

# Функция для получения кода города на английском по его названию на русском
#@log_user_actions
def get_city_code(city_name):
    # Здесь не нужно дополнительно приводить к нижнему регистру, так как он уже приведен в load_cities
    return cities_dict.get(city_name.lower())  # Приводим введённое название города к нижнему регистру

# Обработчик команды "Цены на топливо"
@bot.message_handler(func=lambda message: message.text == "Цены на топливо")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Цены на топливо')
@track_usage('Цены на топливо')  # Добавление отслеживания статистики
#@log_user_actions
def fuel_prices_command(message):
    chat_id = message.chat.id
    load_citys_users_data()  # Загружаем данные перед использованием
    user_state[chat_id] = "choosing_city"  # Устанавливаем состояние выбора города

    str_chat_id = str(chat_id)

    # Создаем клавиатуру с кнопкой "В главное меню" и последними городами
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

    if str_chat_id in user_data and 'recent_cities' in user_data[str_chat_id]:
        recent_cities = user_data[str_chat_id]['recent_cities']
        city_buttons = [types.KeyboardButton(city.capitalize()) for city in recent_cities]
        markup.row(*city_buttons)  # Все кнопки городов будут в одной строке
    else:
        pass  # Нет недавних городов

    markup.add(types.KeyboardButton("В главное меню"))

    # Отправляем сообщение с просьбой ввести город
    bot.send_message(chat_id, "Введите город или выберите из последних:", reply_markup=markup)
    # Регистрируем следующий шаг для обработки ввода города
    bot.register_next_step_handler(message, process_city_selection)

# Обработчик выбора города
#@log_user_actions
def process_city_selection(message):
    chat_id = message.chat.id
    str_chat_id = str(chat_id)

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if user_state.get(chat_id) != "choosing_city":
        bot.send_message(chat_id, "Пожалуйста, используйте доступные кнопки для навигации.")
        return

    city_name = message.text.strip().lower()
    city_code = get_city_code(city_name)

    if city_code:
        if str_chat_id not in user_data:
            user_data[str_chat_id] = {'recent_cities': [], 'city_code': None}

        update_recent_cities(str_chat_id, city_name)
        user_data[str_chat_id]['city_code'] = city_code  # Сохраняем код города

        # Получаем координаты пользователя
        notifications = load_user_locations()  # Функция для загрузки notifications.json
        user_info = notifications.get(str_chat_id)

        latitude = None
        longitude = None

        if user_info:
            latitude = user_info.get('latitude')
            longitude = user_info.get('longitude')

        # Сохраняем city_code и координаты в notifications.json, если координаты есть
        save_user_location(chat_id, latitude, longitude, city_code)

        # Сохранение данных города
        save_citys_users_data()

        # Показываем меню с ценами на топливо
        site_type = "default_site_type"  # Определите значение site_type
        show_fuel_price_menu(chat_id, city_code, site_type)
    else:
        bot.send_message(chat_id, "Город не найден. Пожалуйста, Попробуйте еще раз")

        # Создаем клавиатуру
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)  # row_width=3 — города в строку

        # Добавляем кнопки с последними городами, если они есть
        recent_cities = user_data.get(str_chat_id, {}).get('recent_cities', [])
        if recent_cities:
            markup.add(*[types.KeyboardButton(city.title()) for city in recent_cities])  # Добавляем города в одну строку

        # Добавляем кнопку "В главное меню" в конце
        markup.add(types.KeyboardButton("В главное меню"))

        bot.send_message(chat_id, "Введите город или выберите из последних:", reply_markup=markup)

        # Повторный вызов обработчика для ввода города
        bot.register_next_step_handler(message, process_city_selection)

# Функция для обновления списка последних городов
#@log_user_actions
def update_recent_cities(chat_id, city_name):
    # Убедитесь, что данные о пользователе существуют
    if chat_id not in user_data:
        user_data[chat_id] = {'recent_cities': [], 'city_code': None}

    # Извлекаем список последних городов
    recent_cities = user_data[chat_id].get('recent_cities', [])

    # Удаляем город, если он уже есть, и добавляем в конец
    if city_name in recent_cities:
        recent_cities.remove(city_name)
    elif len(recent_cities) >= 3:
        # Удаляем первый город, если уже 3 города
        recent_cities.pop(0)

    # Добавляем новый город в конец списка
    recent_cities.append(city_name)

    # Сохраняем обновлённые данные пользователей
    user_data[chat_id]['recent_cities'] = recent_cities
    save_citys_users_data()  # Сохраняем данные в файл

# Определяем доступные типы топлива
fuel_types = ["АИ-92", "АИ-95", "АИ-98", "АИ-100", "ДТ", "Газ"]

# Функция для создания клавиатуры с типами топлива
#@log_user_actions
def show_fuel_price_menu(chat_id, city_code, site_type):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

    row1 = [types.KeyboardButton(fuel_type) for fuel_type in fuel_types[:3]]
    row2 = [types.KeyboardButton(fuel_type) for fuel_type in fuel_types[3:]]
    row3 = [types.KeyboardButton("В главное меню")]

    markup.add(*row1, *row2, *row3)

    sent = bot.send_message(chat_id, "Выберите тип топлива для отображения актуальных цен:", reply_markup=markup)
    bot.register_next_step_handler(sent, lambda msg: process_fuel_price_selection(msg, city_code, site_type))

# Глобальная переменная для отслеживания прогресса
progress = 0
progress_lock = threading.Lock()

def update_progress(chat_id, message_id, bot, start_time):
    global progress
    while progress < 100:
        time.sleep(1)  # Обновляем прогресс каждую секунду
        elapsed_time = time.time() - start_time
        with progress_lock:
            current_progress = progress
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=f"Данные обрабатываются. Ожидайте! Никуда не выходите!\n\nВыполнено: {current_progress:.2f}%\nПрошло времени: {elapsed_time:.2f} секунд"
        )

def process_fuel_price_selection(message, city_code, site_type):
    global progress
    chat_id = message.chat.id

    if chat_id not in user_data:
        user_data[chat_id] = {'city_code': city_code, 'site_type': site_type}

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    selected_fuel_type = message.text.strip().lower()

    fuel_type_mapping = {
        "аи-92": ["АИ-92", "Премиум 92"],
        "аи-95": ["АИ-95", "Премиум 95"],
        "аи-98": ["АИ-98", "Премиум 98"],
        "аи-100": ["АИ-100", "Премиум 100"],
        "дт": ["ДТ", "Премиум ДТ"],
        "газ": ["Газ"],
    }

    if selected_fuel_type not in fuel_type_mapping:
        sent = bot.send_message(chat_id, "Пожалуйста, выберите тип топлива из предложенных вариантов.")
        bot.register_next_step_handler(sent, lambda msg: process_fuel_price_selection(msg, city_code, site_type))
        return

    actual_fuel_types = fuel_type_mapping[selected_fuel_type]

    # Отправляем сообщение о начале обработки данных
    progress_message = bot.send_message(chat_id, "Данные обрабатываются. Ожидайте! Никуда не выходите!")
    message_id = progress_message.message_id

    # Запускаем поток для обновления прогресса
    start_time = time.time()
    progress_thread = threading.Thread(target=update_progress, args=(chat_id, message_id, bot, start_time))
    progress_thread.start()

    try:
        # Проверяем наличие сохранённых данных или парсим сайт
        fuel_prices = []
        total_steps = len(actual_fuel_types) * 2  # Количество шагов для обновления прогресса
        current_step = 0

        saved_data = load_saved_data(city_code)
        today = datetime.now().date()

        if saved_data:
            file_modification_time = datetime.fromtimestamp(os.path.getmtime(os.path.join('data base', 'azs', f"{city_code}_table_azs_data.json"))).date()
            if file_modification_time >= today:
                print(f"Данные для города {city_code} уже обновлены сегодня. Пропускаем парсинг.")
                fuel_prices = [
                    item for item in saved_data
                    if item[1].lower() == selected_fuel_type.lower() or
                    (selected_fuel_type.lower() == "газ" and item[1].lower() == "газ спбт") or
                    any(premium_type in item[1].lower() for premium_type in ["премиум 92", "премиум 95", "премиум 98", "премиум 100", "премиум дт"])
                ]
                with progress_lock:
                    progress = 100
            else:
                for fuel_type in actual_fuel_types:
                    try:
                        print(f"Пытаемся получить данные с сайта AZCPRICE для города {city_code} и типа топлива {fuel_type}")
                        fuel_prices.extend(get_fuel_prices_from_site(fuel_type, city_code, "azcprice"))
                        current_step += 1
                    except ValueError:
                        try:
                            print(f"Сайт AZCPRICE недоступен. Пытаемся получить данные с сайта petrolplus для города {city_code} и типа топлива {fuel_type}")
                            fuel_prices.extend(get_fuel_prices_from_site(fuel_type, city_code, "petrolplus"))
                            current_step += 1
                        except ValueError:
                            print(f"Оба сайта недоступны для города {city_code} и типа топлива {fuel_type}")
                            current_step += 1

                    with progress_lock:
                        progress = (current_step / total_steps) * 100

                # Сохраняем данные в файл
                save_fuel_data(city_code, fuel_prices)
        else:
            for fuel_type in actual_fuel_types:
                try:
                    print(f"Пытаемся получить данные с сайта AZCPRICE для города {city_code} и типа топлива {fuel_type}")
                    fuel_prices.extend(get_fuel_prices_from_site(fuel_type, city_code, "azcprice"))
                    current_step += 1
                except ValueError:
                    try:
                        print(f"Сайт AZCPRICE недоступен. Пытаемся получить данные с сайта petrolplus для города {city_code} и типа топлива {fuel_type}")
                        fuel_prices.extend(get_fuel_prices_from_site(fuel_type, city_code, "petrolplus"))
                        current_step += 1
                    except ValueError:
                        print(f"Оба сайта недоступны для города {city_code} и типа топлива {fuel_type}")
                        current_step += 1

                with progress_lock:
                    progress = (current_step / total_steps) * 100

            # Сохраняем данные в файл
            save_fuel_data(city_code, fuel_prices)

        if not fuel_prices:
            raise ValueError("Нет данных по ценам.")

        brand_prices = {}
        for brand, fuel_type, price in fuel_prices:
            price = float(price)
            # Удаляем номер из названия бренда
            brand_name = brand.split(' №')[0]
            if brand_name not in brand_prices:
                brand_prices[brand_name] = []
            brand_prices[brand_name].append((fuel_type, price))

        # Разделяем данные на обычные и "премиум"
        normal_prices = {brand: [price for fuel_type, price in prices if 'Премиум' not in fuel_type] for brand, prices in brand_prices.items()}
        premium_prices = {brand: [price for fuel_type, price in prices if 'Премиум' in fuel_type] for brand, prices in brand_prices.items()}

        # Удаляем пустые списки
        normal_prices = {brand: prices for brand, prices in normal_prices.items() if prices}
        premium_prices = {brand: prices for brand, prices in premium_prices.items() if prices}

        # Вычисляем средние цены
        averaged_normal_prices = {brand: sum(prices) / len(prices) for brand, prices in normal_prices.items()}
        averaged_premium_prices = {brand: sum(prices) / len(prices) for brand, prices in premium_prices.items()}

        # Сортируем цены
        sorted_normal_prices = sorted(averaged_normal_prices.items(), key=lambda x: x[1])
        sorted_premium_prices = sorted(averaged_premium_prices.items(), key=lambda x: x[1])

        # Преобразуем actual_fuel_type обратно в читаемый формат
        readable_fuel_type = selected_fuel_type.upper()

        # Формируем сообщения
        normal_prices_message = "\n\n".join([f"⛽ {i + 1}. {brand} - {avg_price:.2f} руб./л." for i, (brand, avg_price) in enumerate(sorted_normal_prices)])
        premium_prices_message = "\n\n".join([f"⛽ {i + 1}. {brand} - {avg_price:.2f} руб./л." for i, (brand, avg_price) in enumerate(sorted_premium_prices)])

        # Разбиваем сообщения на части, если они слишком длинные
        max_length = 4000
        normal_parts = [normal_prices_message[i:i + max_length] for i in range(0, len(normal_prices_message), max_length)]
        premium_parts = [premium_prices_message[i:i + max_length] for i in range(0, len(premium_prices_message), max_length)]

        # Отправляем сообщения
        if normal_parts:
            for i, part in enumerate(normal_parts):
                if i == 0:
                    bot.send_message(chat_id, f"*Актуальные цены на {readable_fuel_type}:*\n\n\n{part}", parse_mode="Markdown")
                else:
                    bot.send_message(chat_id, part, parse_mode="Markdown")

        if premium_parts:
            for i, part in enumerate(premium_parts):
                if i == 0:
                    bot.send_message(chat_id, f"*Актуальные цены на {readable_fuel_type} Премиум:*\n\n\n{part}", parse_mode="Markdown")
                else:
                    bot.send_message(chat_id, part, parse_mode="Markdown")

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        another_fuel_button = types.KeyboardButton("Посмотреть цены на другое топливо")
        choose_another_city_button = types.KeyboardButton("Выбрать другой город")
        main_menu_button = types.KeyboardButton("В главное меню")
        markup.add(another_fuel_button)
        markup.add(choose_another_city_button)
        markup.add(main_menu_button)

        user_state[chat_id] = "next_action"
        sent = bot.send_message(chat_id, "Вы можете посмотреть цены на другое топливо или выбрать другой город", reply_markup=markup)
        bot.register_next_step_handler(sent, process_next_action)

        # Обновляем прогресс до 100%
        with progress_lock:
            progress = 100

    except Exception as e:
        print(f"Ошибка: {e}")
        bot.send_message(chat_id, f"Ошибка получения цен!\n\nНе найдена таблица с ценами.\nПопробуйте выбрать другой город или тип топлива")
        show_fuel_price_menu(chat_id, city_code, site_type)

# Обработчик следующих действий (посмотреть другое топливо или выбрать другой город)
#@log_user_actions
def process_next_action(message):
    chat_id = message.chat.id
    text = message.text.strip().lower()

    if text == "выбрать другой город":
        user_state[chat_id] = "choosing_city"  # Устанавливаем состояние выбора города

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

        # Загружаем последние города для пользователя
        recent_cities = user_data.get(str(chat_id), {}).get('recent_cities', [])

        # Если есть последние города, создаем кнопки
        city_buttons = [types.KeyboardButton(city.capitalize()) for city in recent_cities]
        if city_buttons:
            markup.row(*city_buttons)

        markup.add(types.KeyboardButton("В главное меню"))
        bot.send_message(chat_id, "Введите название города или выберите из последних:", reply_markup=markup)

        # Передаем управление process_city_selection для выбора города
        bot.register_next_step_handler(message, process_city_selection)
    elif text == "посмотреть цены на другое топливо":
        city_code = user_data[str(chat_id)]['city_code']
        site_type = "default_site_type"  # Определите значение site_type
        show_fuel_price_menu(chat_id, city_code, site_type)
    elif text == "в главное меню":
        return_to_menu(message)
    else:
        bot.send_message(chat_id, "Пожалуйста, выберите одно из предложенных действий.")
        bot.register_next_step_handler(message, process_next_action)

# Функция для обработки данных по всем видам топлива и их сохранения
#@log_user_actions
def process_city_fuel_data(city_code, selected_fuel_type, site_type, actual_fuel_types):
    today = datetime.now().date()
    saved_data = load_saved_data(city_code)

    # Проверяем дату файла
    filepath = os.path.join('data base', 'azs', f"{city_code}_table_azs_data.json")
    if saved_data:
        file_modification_time = datetime.fromtimestamp(os.path.getmtime(filepath)).date()
        if file_modification_time < today:  # Если файл устарел
            # Получаем новые данные
            all_fuel_prices = []
            for fuel_type in actual_fuel_types:
                try:
                    print(f"Пытаемся получить данные с сайта AZCPRICE для города {city_code} и типа топлива {fuel_type}")
                    fuel_prices = get_fuel_prices_from_site(fuel_type, city_code, "azcprice")
                except ValueError:
                    try:
                        print(f"Сайт AZCPRICE недоступен. Пытаемся получить данные с сайта petrolplus для города {city_code} и типа топлива {fuel_type}")
                        fuel_prices = get_fuel_prices_from_site(fuel_type, city_code, "petrolplus")
                    except ValueError:
                        print(f"Оба сайта недоступны для города {city_code} и типа топлива {fuel_type}")
                        # Если таблица не найдена, возвращаем последние сохраненные данные
                        if saved_data:
                            return [
                                item for item in saved_data
                                if item[1].lower() == selected_fuel_type.lower() or
                                (selected_fuel_type.lower() == "газ" and item[1].lower() == "газ спбт")
                            ]
                        # Если данных нет, вызываем ту же ошибку
                        raise ValueError("Ошибка получения цен!\n\nНе найдена таблица с ценами.\nПопробуйте выбрать другой город или тип топлива")

                fuel_prices = remove_duplicate_prices(fuel_prices)
                all_fuel_prices.extend(fuel_prices)

            # Сохраняем новые данные
            save_fuel_data(city_code, all_fuel_prices)
            saved_data = all_fuel_prices  # Обновляем saved_data

    # Если данных нет, пытаемся их получить
    if not saved_data:
        all_fuel_prices = []
        for fuel_type in actual_fuel_types:
            try:
                print(f"Пытаемся получить данные с сайта AZCPRICE для города {city_code} и типа топлива {fuel_type}")
                fuel_prices = get_fuel_prices_from_site(fuel_type, city_code, "azcprice")
            except ValueError:
                try:
                    print(f"Сайт AZCPRICE недоступен. Пытаемся получить данные с сайта petrolplus для города {city_code} и типа топлива {fuel_type}")
                    fuel_prices = get_fuel_prices_from_site(fuel_type, city_code, "petrolplus")
                except ValueError:
                    print(f"Оба сайта недоступны для города {city_code} и типа топлива {fuel_type}")
                    # Если таблица не найдена и файла тоже нет, вызываем ту же ошибку
                    raise ValueError("Ошибка получения цен!\n\nНе найдена таблица с ценами.\nПопробуйте выбрать другой город или тип топлива")

            fuel_prices = remove_duplicate_prices(fuel_prices)
            all_fuel_prices.extend(fuel_prices)

        save_fuel_data(city_code, all_fuel_prices)
        saved_data = all_fuel_prices  # Возвращаем полученные данные

    # Фильтруем и возвращаем данные по выбранному типу топлива
    filtered_prices = [
        item for item in saved_data
        if item[1].lower() == selected_fuel_type.lower() or
        (selected_fuel_type.lower() == "газ" and item[1].lower() == "газ спбт") or
        any(premium_type in item[1].lower() for premium_type in ["премиум 92", "премиум 95", "премиум 98", "премиум 100", "премиум дт"])
    ]
    print(f"Отфильтрованные данные для города {city_code} и типа топлива {selected_fuel_type}: {filtered_prices}")
    return remove_duplicate_prices(filtered_prices)

# Обновлённая функция для удаления дублирующихся цен для каждой АЗС и каждого типа топлива
#@log_user_actions
def remove_duplicate_prices(fuel_prices):
    unique_prices = {}
    for brand, fuel_type, price in fuel_prices:
        price = float(price)
        key = (brand, fuel_type)  # Используем комбинацию бренда и типа топлива как уникальный ключ
        if key not in unique_prices or price < unique_prices[key]:
            unique_prices[key] = price  # Сохраняем минимальную цену для каждого бренда и типа топлива
    return [(brand, fuel_type, price) for (brand, fuel_type), price in unique_prices.items()]

# Функция для парсинга данных с сайта
#@log_user_actions
def get_fuel_prices_from_site(selected_fuel_type, city_code, site_type):
    if site_type == "azcprice":
        print(f"Парсинг данных с сайта AZCPRICE для города {city_code}")
        url = f'https://azsprice.ru/benzin-{city_code}'
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')

        table = soup.find('table')
        if not table:
            raise ValueError("Не найдена таблица с ценами")

        fuel_prices = []

        rows = table.find_all('tr')
        for row in rows[1:]:
            columns = row.find_all('td')
            if len(columns) < 5:
                continue

            brand = columns[1].text.strip()
            fuel_type = columns[2].text.strip()
            today_price = clean_price(columns[3].text.strip())

            # Добавляем логику для обработки премиальных видов топлива
            if fuel_type.lower() == selected_fuel_type.lower() or (fuel_type.lower() == "газ спбт" and selected_fuel_type.lower() == "газ"):
                fuel_prices.append((brand, fuel_type, today_price))
            elif any(premium_type in fuel_type.lower() for premium_type in ["премиум 92", "премиум 95", "премиум 98", "премиум 100", "премиум дт"]):
                fuel_prices.append((brand, fuel_type, today_price))

        return fuel_prices
    elif site_type == "petrolplus":
        print(f"Парсинг данных с сайта petrolplus для города {city_code}")
        base_url = f'https://www.petrolplus.ru/fuelstations/{city_code}/?PAGEN_='
        page = 1
        all_fuel_prices = []

        while True:
            url = f'{base_url}{page}'
            response = requests.get(url)

            # Проверка успешности запроса
            if response.status_code != 200:
                print(f"Ошибка получения данных с сайта: {response.status_code}")
                break

            soup = BeautifulSoup(response.text, 'html.parser')

            # Находим таблицу с данными
            table = soup.find('table')
            if not table:
                print(f"Не найдена таблица с ценами для города {city_code} на странице {page}")
                break

            # Парсим строки таблицы
            for row in table.find_all('tr')[1:]:  # Пропускаем заголовок таблицы
                cols = row.find_all('td')
                if len(cols) >= 3:
                    address = cols[0].text.strip()
                    brand = cols[1].text.strip()
                    fuel_types = [ft.strip() for ft in cols[2].stripped_strings]
                    prices = [p.strip().replace(',', '.') for p in cols[3].stripped_strings]

                    for fuel_type, price in zip(fuel_types, prices):
                        all_fuel_prices.append((brand, fuel_type, clean_price(price)))

            page += 1

        print(f"Данные успешно получены для города {city_code}: {all_fuel_prices}")
        return all_fuel_prices
    else:
        raise ValueError("Неизвестный тип сайта")

# __________________________________________________ЕСЛИ САЙТ НЕ РАБОТАЕТ (ЗАКОМЕНТИРОВАНО)__________________________________________________

# # Функция для парсинга данных с сайта
# #@log_user_actions
# def get_fuel_prices_from_site(selected_fuel_type, city_code, site_type):
#     if site_type == "azcprice":
#         print(f"Парсинг данных с сайта AZCPRICE для города {city_code}")
#         # Имитация недоступности сайта AZCPRICE
#         raise ValueError("Сайт AZCPRICE временно недоступен")
#     elif site_type == "petrolplus":
#         print(f"Парсинг данных с сайта petrolplus для города {city_code}")
#         base_url = f'https://www.petrolplus.ru/fuelstations/{city_code}/?PAGEN_='
#         page = 1
#         all_fuel_prices = []

#         while True:
#             url = f'{base_url}{page}'
#             response = requests.get(url)

#             # Проверка успешности запроса
#             if response.status_code != 200:
#                 print(f"Ошибка получения данных с сайта: {response.status_code}")
#                 break

#             soup = BeautifulSoup(response.text, 'html.parser')

#             # Находим таблицу с данными
#             table = soup.find('table')
#             if not table:
#                 print(f"Не найдена таблица с ценами для города {city_code} на странице {page}")
#                 break

#             # Парсим строки таблицы
#             for row in table.find_all('tr')[1:]:  # Пропускаем заголовок таблицы
#                 cols = row.find_all('td')
#                 if len(cols) >= 3:
#                     address = cols[0].text.strip()
#                     brand = cols[1].text.strip()
#                     fuel_types = [ft.strip() for ft in cols[2].stripped_strings]
#                     prices = [p.strip().replace(',', '.') for p in cols[3].stripped_strings]

#                     for fuel_type, price in zip(fuel_types, prices):
#                         all_fuel_prices.append((brand, fuel_type, clean_price(price)))

#             page += 1

#         print(f"Данные успешно получены для города {city_code}: {all_fuel_prices}")
#         return all_fuel_prices
#     else:
#         raise ValueError("Неизвестный тип сайта")

# Функция для очистки цены
#@log_user_actions
def clean_price(price):
    cleaned_price = ''.join([ch for ch in price if ch.isdigit() or ch == '.'])
    return cleaned_price

# Функция для проверки обновления и парсинга данных
#@log_user_actions
def parse_fuel_prices():
    cities_to_parse = os.listdir(os.path.join('data base', 'azs'))  # Список городов для парсинга
    for city_code in cities_to_parse:
        city_code = city_code.replace('_table_azs_data.json', '')  # Убираем расширение для получения кода города
        saved_data = load_saved_data(city_code)

        today = datetime.now().date()
        if saved_data:
            file_modification_time = datetime.fromtimestamp(os.path.getmtime(os.path.join('data base', 'azs', f"{city_code}_table_azs_data.json"))).date()
            if file_modification_time >= today:
                print(f"Данные для города {city_code} уже обновлены сегодня. Пропускаем.")
                continue

        # Если файл устарел или не существует, парсим новые данные
        all_fuel_prices = []
        for fuel_type in fuel_types:
            try:
                fuel_prices = get_fuel_prices_from_site(fuel_type, city_code, "azcprice")
            except ValueError:
                try:
                    fuel_prices = get_fuel_prices_from_site(fuel_type, city_code, "petrolplus")
                except ValueError:
                    print(f"Оба сайта недоступны для города {city_code}")
                    continue

            fuel_prices = remove_duplicate_prices(fuel_prices)
            all_fuel_prices.extend(fuel_prices)

        print(f"Сохранение данных для города {city_code} с {len(all_fuel_prices)} записями.")
        save_fuel_data(city_code, all_fuel_prices)
        print(f"Данные для города {city_code} успешно обновлены.")


# _____________________________________________________ПАРСИНГ САЙТА RUSSIABASE ОТДЕЛЬНО (НЕ ВСТРОЕН)_____________________________________________________

# # Функция для загрузки данных пользователей
# def load_citys_users_data():
#     global user_data
#     print("Загрузка данных пользователей...")
#     if os.path.exists(DATA_FILE_PATH):
#         with open(DATA_FILE_PATH, 'r', encoding='utf-8') as f:
#             user_data = json.load(f)
#             print(f"Данные пользователей загружены: {user_data}")
#     else:
#         user_data = {}
#         print("Файл данных пользователей не найден, создаем новый.")

# # Функция для сохранения данных пользователей
# def save_citys_users_data():
#     print("Сохранение данных пользователей...")
#     with open(DATA_FILE_PATH, 'w', encoding='utf-8') as f:
#         json.dump(user_data, f, ensure_ascii=False, indent=4)
#         print("Данные пользователей сохранены.")

# # Загружаем данные пользователей при запуске
# load_citys_users_data()

# # Функция для создания имени файла на основе города и даты
# #@log_user_actions
# def create_filename(city_code, date):
#     date_str = date.strftime('%d_%m_%Y')
#     city_info = cities_dict.get(city_code.lower())
#     if city_info:
#         city_name_en = city_info[1]  # Получаем английское название региона
#         return f"{city_name_en}_table_azs_data_{date_str}.json"
#     return None


# # Функция для чтения городов и создания словаря для их URL
# def load_cities_from_file(file_path, site_type):
#     cities = {}
#     try:
#         with open(file_path, 'r', encoding='utf-8') as file:
#             for line in file:
#                 if '-' in line:
#                     parts = line.strip().split(' - ')
#                     if len(parts) != 3:
#                         print(f"Ошибка формата строки: {line.strip()}")
#                         continue
#                     city_code = parts[0]
#                     region_name_ru = parts[1]
#                     region_name_en = parts[2]
#                     cities[region_name_ru.lower()] = (city_code, region_name_en, site_type)
#                     print(f"Добавлен город: {region_name_ru} с кодом {city_code} и типом {site_type}")
#     except FileNotFoundError:
#         print(f"Файл {file_path} не найден.")
#     return cities

# # Функция для загрузки городов из всех файлов
# def load_cities():
#     cities_dict_2 = load_cities_from_file('files/combined_cities_2.txt', 'region')
#     cities_dict_3 = load_cities_from_file('files/combined_cities_3.txt', 'raion')
#     cities_dict_4 = load_cities_from_file('files/combined_cities_4.txt', 'city')

#     # Объединяем все словари в один
#     combined_cities_dict = {**cities_dict_2, **cities_dict_3, **cities_dict_4}
#     print(f"Загружены города: {combined_cities_dict}")
#     return combined_cities_dict

# # Загружаем города при запуске бота
# cities_dict = load_cities()
# print(f"Загружены города: {cities_dict}")

# # Функция для чтения городов из файла combined_cities_4.txt
# def load_cities_4():
#     cities = {}
#     try:
#         with open('files/combined_cities_4.txt', 'r', encoding='utf-8') as file:
#             for line in file:
#                 if '-' in line:
#                     parts = line.strip().split(' - ')
#                     if len(parts) != 3:
#                         print(f"Ошибка формата строки: {line.strip()}")
#                         continue
#                     city_code = parts[0]
#                     region_name_ru = parts[1]
#                     region_name_en = parts[2]
#                     cities[region_name_ru.lower()] = (city_code, region_name_en)
#                     print(f"Добавлен город: {region_name_ru} с кодом {city_code}")
#     except FileNotFoundError:
#         print("Файл с городами не найден.")
#     return cities

# # Функция для чтения городов из файла combined_cities_3.txt
# def load_cities_3():
#     cities = {}
#     try:
#         with open('files/combined_cities_3.txt', 'r', encoding='utf-8') as file:
#             for line in file:
#                 if '-' in line:
#                     parts = line.strip().split(' - ')
#                     if len(parts) != 3:
#                         print(f"Ошибка формата строки: {line.strip()}")
#                         continue
#                     city_code = parts[0]
#                     region_name_ru = parts[1]
#                     region_name_en = parts[2]
#                     cities[region_name_ru.lower()] = (city_code, region_name_en)
#                     print(f"Добавлен город: {region_name_ru} с кодом {city_code}")
#     except FileNotFoundError:
#         print("Файл с городами не найден.")
#     return cities

# # Загружаем города из новых файлов при запуске бота
# cities_dict_4 = load_cities_4()
# cities_dict_3 = load_cities_3()
# print(f"Загружены города из combined_cities_4.txt: {cities_dict_4}")
# print(f"Загружены города из combined_cities_3.txt: {cities_dict_3}")


# # Функция для получения кода города на английском по его названию на русском
# #@log_user_actions
# def get_city_code(region_name):
#     region_name_lower = region_name.lower()  # Приводим ввод пользователя к нижнему регистру
#     city_info = cities_dict.get(region_name_lower)
#     print(f"Получен код города для региона {region_name}: {city_info}")
#     if city_info:
#         return str(city_info[0]), city_info[1], city_info[2]  # Возвращаем city_code как строку, английское название и тип ссылки
#     return None, None, None


# # Функция для сохранения данных в JSON
# def save_fuel_data(city_name, fuel_prices):
#     filename = f'{city_name}_table_azs_data.json'  # Уникальный файл для каждого города
#     filepath = os.path.join('data base', 'azs', filename)  # Указываем путь к папке
#     print(f"Сохранение данных о ценах на топливо для города {city_name} в файл {filepath}")
#     with open(filepath, 'w', encoding='utf-8') as f:
#         json.dump(fuel_prices, f, ensure_ascii=False, indent=4)

# def load_saved_data(city_identifier):
#     filename = f'{city_identifier}_table_azs_data.json'  # Уникальный файл для каждого города
#     filepath = os.path.join('data base', 'azs', filename)  # Указываем путь к папке
#     print(f"Загрузка сохраненных данных для города {city_identifier} из файла {filepath}")
#     try:
#         with open(filepath, 'r', encoding='utf-8') as f:
#             return json.load(f)
#     except FileNotFoundError:
#         print(f"Файл данных для города {city_identifier} не найден.")
#         return None

# # Обработчик команды "Цены на топливо"
# @bot.message_handler(func=lambda message: message.text == "Цены на топливо")
# @restricted
# @track_user_activity
# @check_chat_state
# @check_function_state_decorator('Цены на топливо')
# @track_usage('Цены на топливо')  # Добавление отслеживания статистики
# #@log_user_actions
# def fuel_prices_command(message):
#     chat_id = message.chat.id
#     load_citys_users_data()  # Загружаем данные перед использованием
#     user_state[chat_id] = "choosing_city"  # Устанавливаем состояние выбора города

#     str_chat_id = str(chat_id)

#     # Создаем клавиатуру с кнопкой "В главное меню" и последними городами
#     markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

#     if str_chat_id in user_data and 'recent_cities' in user_data[str_chat_id]:
#         recent_cities = user_data[str_chat_id]['recent_cities']
#         city_buttons = [types.KeyboardButton(city.capitalize()) for city in recent_cities]
#         markup.row(*city_buttons)  # Все кнопки городов будут в одной строке
#     else:
#         pass  # Нет недавних городов

#     markup.add(types.KeyboardButton("В главное меню"))

#     # Отправляем сообщение с просьбой ввести город
#     bot.send_message(chat_id, "Введите регион или выберите из последних:", reply_markup=markup)
#     # Регистрируем следующий шаг для обработки ввода города
#     bot.register_next_step_handler(message, process_city_selection)

# # Обработчик выбора города
# #@log_user_actions
# def process_city_selection(message):
#     chat_id = message.chat.id
#     str_chat_id = str(chat_id)

#     if message.text == "В главное меню":
#         return_to_menu(message)
#         return

#     if user_state.get(chat_id) != "choosing_city":
#         bot.send_message(chat_id, "Пожалуйста, используйте доступные кнопки для навигации.")
#         return

#     region_name = message.text.strip()  # Приводим ввод пользователя к нижнему регистру
#     city_code, city_name_en, site_type = get_city_code(region_name)
#     print(f"Получен city_code для региона {region_name}: {city_code}")

#     if city_code:
#         if str_chat_id not in user_data:
#             user_data[str_chat_id] = {'recent_cities': [], 'city_code': None, 'site_type': None}

#         update_recent_cities(str_chat_id, region_name)
#         user_data[str_chat_id]['city_code'] = city_code  # Сохраняем city_code как строку
#         user_data[str_chat_id]['site_type'] = site_type  # Сохраняем site_type
#         print(f"Сохранен city_code для пользователя {str_chat_id}: {user_data[str_chat_id]['city_code']}")
#         print(f"Сохранен site_type для пользователя {str_chat_id}: {user_data[str_chat_id]['site_type']}")

#         # Получаем координаты пользователя
#         notifications = load_user_locations()  # Функция для загрузки notifications.json
#         user_info = notifications.get(str_chat_id)

#         latitude = None
#         longitude = None

#         if user_info:
#             latitude = user_info.get('latitude')
#             longitude = user_info.get('longitude')

#         # Сохраняем city_code и координаты в notifications.json, если координаты есть
#         save_user_location(chat_id, latitude, longitude, city_code)

#         # Сохранение данных города
#         save_citys_users_data()

#         # Показываем меню с ценами на топливо
#         show_fuel_price_menu(chat_id, city_code, site_type)
#     else:
#         bot.send_message(chat_id, "Регион не найден. Пожалуйста, Попробуйте еще раз")

#         # Создаем клавиатуру
#         markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)  # row_width=3 — регионы в строку

#         # Добавляем кнопки с последними регионами, если они есть
#         recent_cities = user_data.get(str_chat_id, {}).get('recent_cities', [])
#         if recent_cities:
#             markup.add(*[types.KeyboardButton(city.title()) for city in recent_cities])  # Добавляем регионы в одну строку

#         # Добавляем кнопку "В главное меню" в конце
#         markup.add(types.KeyboardButton("В главное меню"))

#         bot.send_message(chat_id, "Введите регион или выберите из последних:", reply_markup=markup)

#         # Повторный вызов обработчика для ввода региона
#         bot.register_next_step_handler(message, process_city_selection)


# # Функция для обновления списка последних городов
# #@log_user_actions
# def update_recent_cities(chat_id, region_name):
#     # Убедитесь, что данные о пользователе существуют
#     if chat_id not in user_data:
#         user_data[chat_id] = {'recent_cities': [], 'city_code': None}

#     # Извлекаем список последних регионов
#     recent_cities = user_data[chat_id].get('recent_cities', [])

#     # Удаляем регион, если он уже есть, и добавляем в конец
#     if region_name in recent_cities:
#         recent_cities.remove(region_name)
#     elif len(recent_cities) >= 3:
#         # Удаляем первый регион, если уже 3 региона
#         recent_cities.pop(0)

#     # Добавляем новый регион в конец списка
#     recent_cities.append(region_name)

#     # Сохраняем обновлённые данные пользователей
#     user_data[chat_id]['recent_cities'] = recent_cities
#     save_citys_users_data()  # Сохраняем данные в файл

# # Определяем доступные типы топлива
# fuel_types = ["АИ-92", "АИ-95", "АИ-98", "АИ-100", "ДТ", "Газ"]

# # Функция для создания клавиатуры с типами топлива
# #@log_user_actions
# def show_fuel_price_menu(chat_id, city_code, site_type):
#     markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)

#     row1 = [types.KeyboardButton(fuel_type) for fuel_type in fuel_types[:3]]
#     row2 = [types.KeyboardButton(fuel_type) for fuel_type in fuel_types[3:]]
#     row3 = [types.KeyboardButton("В главное меню")]

#     markup.add(*row1, *row2, *row3)

#     sent = bot.send_message(chat_id, "Выберите тип топлива для отображения актуальных цен:", reply_markup=markup)
#     bot.register_next_step_handler(sent, lambda msg: process_fuel_price_selection(msg, city_code, site_type))

# # Обработчик выбора топлива
# #@log_user_actions
# def process_fuel_price_selection(message, city_code, site_type):
#     chat_id = message.chat.id

#     if chat_id not in user_data:
#         user_data[chat_id] = {'city_code': city_code, 'site_type': site_type}

#     if message.text == "В главное меню":
#         return_to_menu(message)
#         return

#     selected_fuel_type = message.text.strip().lower()

#     fuel_type_mapping = {
#         "аи-92": ["ai92", "ai92plus"],
#         "аи-95": ["ai95", "ai95plus"],
#         "аи-98": ["ai98", "ai98plus"],
#         "аи-100": ["ai100", "ai100plus"],
#         "дт": ["dt", "dtplus"],
#         "газ": ["gas"],
#     }

#     if selected_fuel_type not in fuel_type_mapping:
#         sent = bot.send_message(chat_id, "Пожалуйста, выберите тип топлива из предложенных вариантов.")
#         bot.register_next_step_handler(sent, lambda msg: process_fuel_price_selection(msg, city_code, site_type))
#         return

#     actual_fuel_types = fuel_type_mapping[selected_fuel_type]

#     try:
#         # Проверяем наличие сохранённых данных или парсим сайт
#         fuel_prices = []
#         for fuel_type in actual_fuel_types:
#             fuel_prices.extend(process_city_fuel_data(city_code, fuel_type, site_type, actual_fuel_types))

#         if not fuel_prices:
#             raise ValueError("Нет данных по ценам.")

#         brand_prices = {}
#         for brand, fuel_type, price in fuel_prices:
#             price = float(price)
#             # Удаляем номер из названия бренда
#             brand_name = brand.split(' №')[0]
#             if brand_name not in brand_prices:
#                 brand_prices[brand_name] = []
#             brand_prices[brand_name].append((fuel_type, price))

#         # Разделяем данные на обычные и "плюс"
#         normal_prices = {brand: [price for fuel_type, price in prices if 'plus' not in fuel_type] for brand, prices in brand_prices.items()}
#         plus_prices = {brand: [price for fuel_type, price in prices if 'plus' in fuel_type] for brand, prices in brand_prices.items()}

#         # Удаляем пустые списки
#         normal_prices = {brand: prices for brand, prices in normal_prices.items() if prices}
#         plus_prices = {brand: prices for brand, prices in plus_prices.items() if prices}

#         # Вычисляем средние цены
#         averaged_normal_prices = {brand: sum(prices) / len(prices) for brand, prices in normal_prices.items()}
#         averaged_plus_prices = {brand: sum(prices) / len(prices) for brand, prices in plus_prices.items()}

#         # Сортируем цены
#         sorted_normal_prices = sorted(averaged_normal_prices.items(), key=lambda x: x[1])
#         sorted_plus_prices = sorted(averaged_plus_prices.items(), key=lambda x: x[1])

#         # Преобразуем actual_fuel_type обратно в читаемый формат
#         readable_fuel_type = selected_fuel_type.upper()

#         # Формируем сообщения
#         normal_prices_message = "\n\n".join([f"⛽ {i + 1}. {brand} - {avg_price:.2f} руб./л." for i, (brand, avg_price) in enumerate(sorted_normal_prices)])
#         plus_prices_message = "\n\n".join([f"⛽ {i + 1}. {brand} - {avg_price:.2f} руб./л." for i, (brand, avg_price) in enumerate(sorted_plus_prices)])

#         # Разбиваем сообщения на части, если они слишком длинные
#         max_length = 4000
#         normal_parts = [normal_prices_message[i:i + max_length] for i in range(0, len(normal_prices_message), max_length)]
#         plus_parts = [plus_prices_message[i:i + max_length] for i in range(0, len(plus_prices_message), max_length)]

#         # Отправляем сообщения
#         if normal_parts:
#             for i, part in enumerate(normal_parts):
#                 if i == 0:
#                     bot.send_message(chat_id, f"*Актуальные цены на {readable_fuel_type}:*\n\n\n{part}", parse_mode="Markdown")
#                 else:
#                     bot.send_message(chat_id, part, parse_mode="Markdown")

#         if plus_parts:
#             for i, part in enumerate(plus_parts):
#                 if i == 0:
#                     bot.send_message(chat_id, f"*Актуальные цены на {readable_fuel_type}+:*\n\n\n{part}", parse_mode="Markdown")
#                 else:
#                     bot.send_message(chat_id, part, parse_mode="Markdown")

#         markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
#         another_fuel_button = types.KeyboardButton("Посмотреть цены на другое топливо")
#         choose_another_city_button = types.KeyboardButton("Выбрать другой город")
#         main_menu_button = types.KeyboardButton("В главное меню")
#         markup.add(another_fuel_button)
#         markup.add(choose_another_city_button)
#         markup.add(main_menu_button)

#         user_state[chat_id] = "next_action"
#         sent = bot.send_message(chat_id, "Вы можете посмотреть цены на другое топливо или выбрать другой город.", reply_markup=markup)
#         bot.register_next_step_handler(sent, process_next_action)

#     except Exception as e:
#         print(f"Ошибка: {e}")
#         bot.send_message(chat_id, f"Ошибка получения цен!\n\nНе найдена таблица с ценами.\nПопробуйте выбрать другой город или тип топлива")
#         show_fuel_price_menu(chat_id, city_code, site_type)

# # Обработчик следующих действий (посмотреть другое топливо или выбрать другой город)
# #@log_user_actions
# def process_next_action(message):
#     chat_id = message.chat.id
#     text = message.text.strip().lower()

#     if text == "выбрать другой город":
#         user_state[chat_id] = "choosing_city"  # Устанавливаем состояние выбора города

#         markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

#         # Загружаем последние города для пользователя
#         recent_cities = user_data.get(str(chat_id), {}).get('recent_cities', [])

#         # Если есть последние города, создаем кнопки
#         city_buttons = [types.KeyboardButton(city.capitalize()) for city in recent_cities]
#         if city_buttons:
#             markup.row(*city_buttons)

#         markup.add(types.KeyboardButton("В главное меню"))
#         bot.send_message(chat_id, "Введите название города или выберите из последних:", reply_markup=markup)

#         # Передаем управление process_city_selection для выбора города
#         bot.register_next_step_handler(message, process_city_selection)
#     elif text == "в главное меню":
#         return_to_menu(message)
#     elif text == "посмотреть цены на другое топливо":
#         city_code = user_data[str(chat_id)]['city_code']
#         site_type = user_data[str(chat_id)]['site_type']
#         show_fuel_price_menu(chat_id, city_code, site_type)
#     else:
#         bot.send_message(chat_id, "Пожалуйста, выберите одно из предложенных действий.")
#         bot.register_next_step_handler(message, process_next_action)

# # Функция для обработки данных по всем видам топлива и их сохранения
# #@log_user_actions
# def process_city_fuel_data(city_code, selected_fuel_type, site_type, actual_fuel_types):
#     print(f"Обработка данных для city_code: {city_code}")
#     city_info = None
#     for region_name, info in cities_dict.items():
#         if info[0] == city_code:
#             city_info = info
#             break
#     print(f"city_info для city_code {city_code}: {city_info}")
#     if not city_info:
#         raise ValueError(f"Не найден город для кода {city_code}")

#     city_name = city_info[1]  # Используем название региона
#     print(f"Найден город: {city_name}")
#     today = datetime.now().date()
#     saved_data = load_saved_data(city_name)

#     # Проверяем дату файла
#     filepath = os.path.join('data base', 'azs', f"{city_name}_table_azs_data.json")
#     if saved_data:
#         file_modification_time = datetime.fromtimestamp(os.path.getmtime(filepath)).date()
#         if file_modification_time < today:  # Если файл устарел
#             # Получаем новые данные
#             all_fuel_prices = []
#             try:
#                 fuel_prices = get_fuel_prices_from_site(city_code, site_type)
#                 print(f"Получены новые данные для города {city_name}: {fuel_prices}")
#             except ValueError as e:
#                 print(f"Ошибка получения данных: {e}")
#                 # Если таблица не найдена, возвращаем последние сохраненные данные
#                 if saved_data:
#                     return [
#                         item for item in saved_data
#                         if any(ft.lower() == item[1].lower() for ft in actual_fuel_types)
#                     ]
#                 # Если данных нет, вызываем ту же ошибку
#                 raise ValueError("Ошибка получения цен!\n\nНе найдена таблица с ценами.\nПопробуйте выбрать другой город или тип топлива")

#             fuel_prices = remove_duplicate_prices(fuel_prices)
#             all_fuel_prices.extend(fuel_prices)

#             # Сохраняем новые данные
#             save_fuel_data(city_name, all_fuel_prices)
#             saved_data = all_fuel_prices  # Обновляем saved_data

#     # Если данных нет, пытаемся их получить
#     if not saved_data:
#         all_fuel_prices = []
#         try:
#             fuel_prices = get_fuel_prices_from_site(city_code, site_type)
#             print(f"Получены данные для города {city_name}: {fuel_prices}")
#         except ValueError as e:
#             print(f"Ошибка получения данных: {e}")
#             # Если таблица не найдена и файла тоже нет, вызываем ту же ошибку
#             raise ValueError("Ошибка получения цен!\n\nНе найдена таблица с ценами.\nПопробуйте выбрать другой город или тип топлива")

#         fuel_prices = remove_duplicate_prices(fuel_prices)
#         all_fuel_prices.extend(fuel_prices)

#         save_fuel_data(city_name, all_fuel_prices)
#         saved_data = all_fuel_prices  # Возвращаем полученные данные

#     # Фильтруем и возвращаем данные по выбранному типу топлива
#     filtered_prices = [
#         item for item in saved_data
#         if any(ft.lower() == item[1].lower() for ft in actual_fuel_types)
#     ]
#     print(f"Отфильтрованные данные для {selected_fuel_type}: {filtered_prices}")

#     # Проверяем, что данные не пустые
#     if not filtered_prices:
#         raise ValueError("Нет данных по ценам.")

#     return remove_duplicate_prices(filtered_prices)


# # Обновлённая функция для удаления дублирующихся цен для каждой АЗС и каждого типа топлива
# #@log_user_actions
# def remove_duplicate_prices(fuel_prices):
#     unique_prices = {}
#     for brand, fuel_type, price in fuel_prices:
#         price = float(price)
#         key = (brand, fuel_type)  # Используем комбинацию бренда и типа топлива как уникальный ключ
#         if key not in unique_prices or price < unique_prices[key]:
#             unique_prices[key] = price  # Сохраняем минимальную цену для каждого бренда и типа топлива
#     return [(brand, fuel_type, price) for (brand, fuel_type), price in unique_prices.items()]

# # Функция для парсинга данных с сайта
# #@log_user_actions
# def get_fuel_prices_from_site(city_code, site_type):
#     if site_type == 'region':
#         base_url = f'https://russiabase.ru/prices?region={city_code}&page='
#     elif site_type == 'city':
#         base_url = f'https://russiabase.ru/prices?city={city_code}&page='
#     elif site_type == 'raion':
#         base_url = f'https://russiabase.ru/prices?raion={city_code}&page='
#     else:
#         raise ValueError("Неверный тип сайта")

#     page = 1
#     all_fuel_prices = []

#     while True:
#         url = f'{base_url}{page}'
#         print(f"Отправляем запрос на URL: {url}")

#         response = requests.get(url)
#         print(f"Статус ответа: {response.status_code}")

#         if response.status_code != 200:
#             print(f"Ошибка при получении данных: {response.status_code}")
#             break

#         # Извлекаем JSON-данные из HTML
#         data_start = response.text.find('{"props":')
#         data_end = response.text.find('</script>', data_start)
#         if data_start == -1 or data_end == -1:
#             print("Не удалось найти JSON-данные в HTML")
#             break

#         json_data = response.text[data_start:data_end].strip()
#         print(f"Извлеченные JSON-данные: {json_data[:500]}...")  # Выводим первые 500 символов для отладки

#         try:
#             data = json.loads(json_data)
#         except json.JSONDecodeError as e:
#             print(f"Ошибка при парсинге JSON: {e}")
#             break

#         fuel_prices = []

#         print("Начинаем парсинг данных о ценах на топливо...")
#         for station in data['props']['pageProps']['listing']['listing']:
#             prices = station['prices']
#             for fuel_type, price_info in prices.items():
#                 price = price_info['value']
#                 if price:  # Проверяем, что цена не пустая
#                     fuel_prices.append((station['name'], fuel_type, price))
#                     print(f"Добавлена цена: {station['name']} - {fuel_type}: {price}")

#         if not fuel_prices:
#             print("Нет данных на этой странице")
#             break

#         all_fuel_prices.extend(fuel_prices)
#         page += 1

#     print(f"Всего найдено {len(all_fuel_prices)} цен на топливо")
#     return all_fuel_prices

# # # Пример вызова функции для отладки
# # city_code = 11  # Пример кода города
# # try:
# #     fuel_prices = get_fuel_prices_from_site(city_code)
# #     print(f"Итоговые данные о ценах на топливо: {fuel_prices}")
# # except Exception as e:
# #     print(f"Ошибка: {e}")


# # Функция для очистки цены
# #@log_user_actions
# def clean_price(price):
#     cleaned_price = ''.join([ch for ch in price if ch.isdigit() or ch == '.'])
#     return cleaned_price

# # Функция для проверки обновления и парсинга данных
# #@log_user_actions
# def parse_fuel_prices():
#     cities_to_parse = os.listdir(os.path.join('data base', 'azs'))  # Список городов для парсинга
#     for city_code in cities_to_parse:
#         city_code = city_code.replace('_table_azs_data.json', '')  # Убираем расширение для получения кода города
#         saved_data = load_saved_data(city_code)

#         today = datetime.now().date()
#         if saved_data:
#             file_modification_time = datetime.fromtimestamp(os.path.getmtime(os.path.join('data base', 'azs', f"{city_code}_table_azs_data.json"))).date()
#             if file_modification_time >= today:
#                 print(f"Данные для города {city_code} уже обновлены сегодня. Пропускаем.")
#                 continue

#         # Если файл устарел или не существует, парсим новые данные
#         all_fuel_prices = []
#         for fuel_type in fuel_types:
#             fuel_prices = get_fuel_prices_from_site(fuel_type, city_code)
#             all_fuel_prices.extend(fuel_prices)

#         print(f"Сохранение данных для города {city_code} с {len(all_fuel_prices)} записями.")
#         save_fuel_data(city_code, all_fuel_prices)
#         print(f"Данные для города {city_code} успешно обновлены.")

# !!!!!!!!!!!!!!!!!!!ЭТО ОБЩИЙ ПЛАНИРОВЩИК ЗАДАЧ!!!!!!!!!!!!!!!!!!!!!
# Общий планировщик задач с объединенной логикой для уведомлений и парсинга
def schedule_tasks():
    while True:
        now = datetime.now()
        
        # Запуск функции парсинга цен на топливо в 00:00
        if now.hour == 0 and now.minute == 0:
            parse_fuel_prices()
            time.sleep(60 * 5)  # Ожидание 5 минут перед следующим городом

        # Проверка запланированных уведомлений (запускает все задачи, добавленные в schedule)
        schedule.run_pending()
        
        # Ожидание перед следующей проверкой расписания
        time.sleep(300)

# Запуск объединенного потока
threading.Thread(target=schedule_tasks, daemon=True).start()



# Ваш транспорт


# Определение состояний
class States:
    ADDING_TRANSPORT = 1
    CONFIRMING_DELETE = 2

# Хранение данных о транспорте в памяти
user_transport = {}

# Функции для работы с файлами
def save_transport_data(user_id, user_data):
    folder_path = "data base/transport"  # Изменено на подкаталог transport
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    # Сохраняем данные о транспорте
    with open(os.path.join(folder_path, f"{user_id}_transport.json"), "w", encoding="utf-8") as file:
        json.dump(user_data, file, ensure_ascii=False, indent=4)

def load_transport_data(user_id):
    folder_path = "data base/transport"  # Изменено на подкаталог transport
    try:
        with open(os.path.join(folder_path, f"{user_id}_transport.json"), "r", encoding="utf-8") as file:
            data = json.load(file)
    except FileNotFoundError:
        data = []
    return data

# Загрузка данных о транспорте при старте бота
def load_all_transport():
    folder_path = "data base/transport"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    users = os.listdir(folder_path)
    for user_file in users:
        if user_file.endswith("_transport.json"):
            user_id = user_file.split("_")[0]
            user_transport[user_id] = load_transport_data(user_id)

# Пример вызова для загрузки данных
load_all_transport()

# Команда для управления транспортом
@bot.message_handler(func=lambda message: message.text == "Ваш транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Ваш транспорт')
@track_usage('Ваш транспорт')  # Добавление отслеживания статистики
#@log_user_actions
def manage_transport(message):
    user_id = str(message.chat.id)
    
    # Создаем клавиатуру с тремя кнопками в первом ряду и двумя в следующих строках
    keyboard = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True)
    
    # Добавляем три кнопки в первый ряд
    keyboard.add("Добавить транспорт", "Посмотреть транспорт", "Удалить транспорт")
    
    # Добавляем кнопки для возвращения в меню трат и в главное меню в следующих строках
    keyboard.add("Вернуться в меню трат и ремонтов")
    keyboard.add("В главное меню")
    
    bot.send_message(user_id, "Выберите действие для транспорта:", reply_markup=keyboard)

# Функция для проверки наличия мультимедийных файлов
#@log_user_actions
def check_media(message, user_id, next_step_handler=None, *args):
    if (message.photo or message.video or message.document or message.animation or 
        message.sticker or message.location or message.audio or 
        message.contact or message.voice or message.video_note):
        bot.send_message(user_id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        if next_step_handler:
            bot.register_next_step_handler(message, next_step_handler, *args)  # Вернуться к следующему шагу
        return True
    return False

# Функция для создания новой клавиатуры
#@log_user_actions
def create_transport_keyboard():
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    item_main_menu = types.KeyboardButton("В главное меню")
    item_return_transport = types.KeyboardButton("Вернуться в ваш транспорт")
    markup.add(item_return_transport)
    markup.add(item_main_menu)
    return markup

@bot.message_handler(func=lambda message: message.text == "Добавить транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Добавить транспорт')
@track_usage('Добавить транспорт')  # Добавление отслеживания статистики
#@log_user_actions
def add_transport(message):
    user_id = str(message.chat.id)
    if check_media(message, user_id): return
    bot.send_message(user_id, "Введите марку транспорта:", reply_markup=create_transport_keyboard())
    bot.register_next_step_handler(message, process_brand)

#@log_user_actions
def format_brand_model(text):
    """Приводит марку и модель к формату 'Первое слово с большой буквы, остальное с маленькой'."""
    return " ".join(word.capitalize() for word in text.split())

#@log_user_actions
def process_brand(message):
    user_id = str(message.chat.id)
    if check_media(message, user_id, process_brand): return

    # Обработка кнопок
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    if message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Форматируем бренд
    brand = format_brand_model(message.text)
    bot.send_message(user_id, "Введите модель транспорта:", reply_markup=create_transport_keyboard())
    bot.register_next_step_handler(message, process_model, brand)

#@log_user_actions
def process_model(message, brand):
    user_id = str(message.chat.id)
    if check_media(message, user_id, process_model, brand): return

    # Обработка кнопок
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    if message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Форматируем модель
    model = format_brand_model(message.text)
    bot.send_message(user_id, "Введите год транспорта:", reply_markup=create_transport_keyboard())
    bot.register_next_step_handler(message, process_year, brand, model)

#@log_user_actions
def process_year(message, brand, model):
    user_id = str(message.chat.id)
    if check_media(message, user_id, process_year, brand, model): return

    # Обработка кнопок
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    if message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Проверка на корректность года
    try:
        year = int(message.text)
        if year < 1960 or year > 3000:  # Проверка диапазона
            raise ValueError("Год должен быть от 1960 г. до 3000 г.")
    except ValueError:
        bot.send_message(user_id, "Ошибка! Пожалуйста, введите корректный год (от 1960 г. до 3000 г.). Попробуйте снова", reply_markup=create_transport_keyboard())
        bot.register_next_step_handler(message, process_year, brand, model)
        return

    bot.send_message(user_id, "Введите госномер:", reply_markup=create_transport_keyboard())
    bot.register_next_step_handler(message, process_license_plate, brand, model, year)

#@log_user_actions
def process_license_plate(message, brand, model, year):
    user_id = str(message.chat.id)
    if check_media(message, user_id, process_license_plate, brand, model, year): 
        return

    # Обработка кнопок
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    if message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Приведение госномера к верхнему регистру
    license_plate = message.text.upper()

    # Проверка длины и формата госномера
    import re
    pattern = r'^[АВЕКМНОРСТУХABEKMHOPCTYX]\d{3}[АВЕКМНОРСТУХABEKMHOPCTYX]{2}\d{2,3}$'
    if not re.match(pattern, license_plate):
        bot.send_message(user_id, "Ошибка! Госномер должен соответствовать формату госномеров РФ. Попробуйте снова", reply_markup=create_transport_keyboard())
        bot.register_next_step_handler(message, process_license_plate, brand, model, year)
        return

    # Проверка уникальности госномера
    if any(t["license_plate"] == license_plate for t in user_transport.get(user_id, [])):
        bot.send_message(user_id, "Ошибка! Такой госномер уже существует. Попробуйте снова", reply_markup=create_transport_keyboard())
        bot.register_next_step_handler(message, process_license_plate, brand, model, year)
        return

    # Сохраняем транспорт в память
    if user_id not in user_transport:
        user_transport[user_id] = []
    
    user_transport[user_id].append({"brand": brand, "model": model, "year": year, "license_plate": license_plate})
    save_transport_data(user_id, user_transport[user_id])  # Сохранение данных о транспорте

    bot.send_message(user_id, f"🚗 *Транспорт добавлен 🚗*\n\n*{brand} - {model} - {year} - {license_plate}*", parse_mode="Markdown", reply_markup=create_transport_keyboard())

    # Переход в меню "Ваш транспорт"
    manage_transport(message)

#@log_user_actions
def delete_expenses_related_to_transport(user_id, transport, selected_transport=""):
    expenses_data = load_expense_data(user_id)
    if user_id in expenses_data:
        updated_expenses = []
        for expense in expenses_data[user_id]['expenses']:
            transport_data = expense.get('transport', {})
            if (transport_data.get('brand') != transport.get('brand') or
                transport_data.get('model') != transport.get('model') or
                transport_data.get('year') != transport.get('year')):
                updated_expenses.append(expense)

        expenses_data[user_id]['expenses'] = updated_expenses
        save_expense_data(user_id, expenses_data, selected_transport)

#@log_user_actions
def delete_repairs_related_to_transport(user_id, transport):
    repair_data = load_repair_data(user_id)
    if user_id in repair_data:
        updated_repairs = []
        for repair in repair_data[user_id].get("repairs", []):
            transport_data = repair.get('transport', {})
            if (transport_data.get('brand') != transport.get('brand') or
                transport_data.get('model') != transport.get('model') or
                transport_data.get('year') != transport.get('year')):
                updated_repairs.append(repair)

        repair_data[user_id]["repairs"] = updated_repairs
        save_repair_data(user_id, repair_data, selected_transport="")


# @bot.message_handler(func=lambda message: message.text == "Изменить транспорт")
# @restricted
# @track_user_activity
# @check_chat_state
# @check_function_state_decorator('Изменить транспорт')
# @track_usage('Изменить транспорт')  # Добавление отслеживания статистики
# #@log_user_actions
# def edit_transport(message):
#     user_id = str(message.chat.id)
#     if user_id not in user_transport or not user_transport[user_id]:
#         bot.send_message(user_id, "У вас нет добавленного транспорта.", reply_markup=create_transport_keyboard())
#         return

#     # Отправляем список транспорта для выбора
#     transport_list = user_transport[user_id]
#     response = "\n".join([f"№{index + 1}. {item['brand']} - {item['model']} - {item['year']} - {item['license_plate']}" for index, item in enumerate(transport_list)])
#     bot.send_message(user_id, f"Выберите номер транспорта для изменения:\n\n{response}", reply_markup=create_transport_keyboard())
#     bot.register_next_step_handler(message, select_transport_for_editing)

# #@log_user_actions
# def select_transport_for_editing(message):
#     user_id = str(message.chat.id)
#     try:
#         index = int(message.text) - 1
#         if index < 0 or index >= len(user_transport[user_id]):
#             raise ValueError
#         selected_transport = user_transport[user_id][index]
#         bot.send_message(user_id, f"Вы выбрали:\n{selected_transport['brand']} - {selected_transport['model']} - {selected_transport['year']} - {selected_transport['license_plate']}\n\nЧто вы хотите изменить? (Марка/Модель/Год/Госномер)", reply_markup=create_transport_keyboard())
#         bot.register_next_step_handler(message, edit_transport_field, index)
#     except ValueError:
#         bot.send_message(user_id, "Ошибка! Введите корректный номер транспорта.", reply_markup=create_transport_keyboard())
#         bot.register_next_step_handler(message, select_transport_for_editing)

# #@log_user_actions
# def edit_transport_field(message, index):
#     user_id = str(message.chat.id)
#     field = message.text.lower()
#     valid_fields = {"марка": "brand", "модель": "model", "год": "year", "госномер": "license_plate"}
#     if field not in valid_fields:
#         bot.send_message(user_id, "Ошибка! Выберите одно из: Марка, Модель, Год, Госномер.", reply_markup=create_transport_keyboard())
#         bot.register_next_step_handler(message, edit_transport_field, index)
#         return

#     bot.send_message(user_id, f"Введите новое значение для поля '{field.capitalize()}':", reply_markup=create_transport_keyboard())
#     bot.register_next_step_handler(message, update_transport_field, index, valid_fields[field])

# #@log_user_actions
# def update_transport_field(message, index, field):
#     user_id = str(message.chat.id)
#     new_value = message.text.strip()

#     # Проверка нового значения
#     if field == "license_plate":
#         # Приведение госномера к верхнему регистру
#         new_value = new_value.upper()

#         # Проверка формата госномера
#         import re
#         pattern = r'^[АВЕКМНОРСТУХABEKMHOPCTYX]\d{3}[АВЕКМНОРСТУХABEKMHOPCTYX]{2}\d{2,3}$'
#         if not re.match(pattern, new_value):
#             bot.send_message(user_id, "Ошибка! Госномер должен соответствовать формату.", reply_markup=create_transport_keyboard())
#             bot.register_next_step_handler(message, update_transport_field, index, field)
#             return

#         # Проверка уникальности госномера
#         if any(t["license_plate"] == new_value for t in user_transport[user_id] if t != user_transport[user_id][index]):
#             bot.send_message(user_id, "Ошибка! Такой госномер уже существует.", reply_markup=create_transport_keyboard())
#             bot.register_next_step_handler(message, update_transport_field, index, field)
#             return

#     elif field == "year":
#         # Проверка корректности года
#         try:
#             new_value = int(new_value)
#             if new_value < 1960 or new_value > 3000:
#                 raise ValueError
#         except ValueError:
#             bot.send_message(user_id, "Ошибка! Введите корректный год (диапазон от 1960 до 3000).", reply_markup=create_transport_keyboard())
#             bot.register_next_step_handler(message, update_transport_field, index, field)
#             return

#     # Обновление данных
#     user_transport[user_id][index][field] = new_value
#     save_transport_data(user_id, user_transport[user_id])

#     bot.send_message(user_id, "Данные транспорта успешно обновлены.", reply_markup=create_transport_keyboard())

#     # Обновление данных о тратах и ремонтах
#     update_expense_repair_data(user_id, user_transport[user_id][index])

# #@log_user_actions
# def update_expense_repair_data(user_id, updated_transport):
#     # Обновляем данные в расходах
#     expense_data = load_expense_data(user_id)
#     for expense in expense_data.get("expenses", []):
#         if expense.get("transport", {}).get("license_plate") == updated_transport["license_plate"]:
#             expense["transport"] = updated_transport
#     save_expense_data(user_id, expense_data)

#     # Обновляем данные в ремонтах
#     repair_data = load_repair_data(user_id)
#     for repair in repair_data.get("repairs", []):
#         if repair.get("transport", {}).get("license_plate") == updated_transport["license_plate"]:
#             repair["transport"] = updated_transport
#     save_repair_data(user_id, repair_data)

@bot.message_handler(func=lambda message: message.text == "Удалить транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить транспорт')
@track_usage('Удалить транспорт')  # Добавление отслеживания статистики
#@log_user_actions
def delete_transport(message):
    user_id = str(message.chat.id)
    if user_id in user_transport and user_transport[user_id]:
        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        transport_list = user_transport[user_id]
        
        # Формируем список транспорта для удаления
        for index, item in enumerate(transport_list, start=1):
            keyboard.add(f"№{index}. {item['brand']} - {item['model']} - {item['year']} - {item['license_plate']}")
        
        # Кнопки "В главное меню" и "Вернуться в ваш транспорт"
        item_main_menu = types.KeyboardButton("В главное меню")
        item_return_transport = types.KeyboardButton("Вернуться в ваш транспорт")
        keyboard.add("Удалить весь транспорт")
        keyboard.add(item_return_transport)
        keyboard.add(item_main_menu)
        
        bot.send_message(user_id, "Выберите транспорт для удаления:", reply_markup=keyboard)
        bot.register_next_step_handler(message, process_transport_selection)
    else:
        bot.send_message(user_id, "У вас нет добавленного транспорта!")

#@log_user_actions
def process_transport_selection(message):
    user_id = str(message.chat.id)
    selected_transport = message.text.strip()

    # Проверяем, выбрал ли пользователь транспорт из списка или другую опцию
    if selected_transport == "В главное меню":
        return_to_menu(message)
        return

    if selected_transport == "Удалить весь транспорт":
        delete_all_transports(message)  # Переход к удалению всех транспортов
        return

    if message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    transport_list = user_transport.get(user_id, [])
    if transport_list:
        # Проверяем, начинается ли текст с номера транспорта
        try:
            index = int(selected_transport.split('.')[0].replace("№", "").strip()) - 1
            if 0 <= index < len(transport_list):
                transport_to_delete = transport_list[index]
                bot.send_message(
    user_id,
    f"*Вы точно хотите удалить данный транспорт?\n\n{transport_to_delete['brand']} - {transport_to_delete['model']} - {transport_to_delete['year']} - {transport_to_delete['license_plate']}*\n\n"
    "Удаление транспорта приведет к удалению всех трат и ремонтов!\n\n"
    "Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены",
    parse_mode="Markdown",
    reply_markup=get_return_menu_keyboard()
)
                bot.register_next_step_handler(message, lambda msg: process_confirmation(msg, transport_to_delete))
            else:
                raise ValueError("Индекс вне диапазона")
        except ValueError:
            bot.send_message(user_id, "Ошибка! Пожалуйста, выберите транспорт для удаления из списка")
            delete_transport(message)
    else:
        bot.send_message(user_id, "У вас нет добавленного транспорта!")

#@log_user_actions
def process_confirmation(message, transport_to_delete):
    user_id = str(message.chat.id)
    confirmation = message.text.strip().upper()

    # Проверка на "В главное меню" или "Вернуться в ваш транспорт"
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    elif message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Проверка подтверждения на "ДА" или "НЕТ"
    if confirmation == "ДА":
        if user_id in user_transport and transport_to_delete in user_transport[user_id]:
            user_transport[user_id].remove(transport_to_delete)
            delete_expenses_related_to_transport(user_id, transport_to_delete)
            delete_repairs_related_to_transport(user_id, transport_to_delete)
            save_transport_data(user_id, user_transport[user_id])
            update_excel_file(user_id)
            update_repairs_excel_file(user_id)
            bot.send_message(user_id, "Транспорт и связанные с ним траты и ремонты успешно удалены!")
            manage_transport(message)  # Возвращаем в меню транспорта
        else:
            bot.send_message(user_id, "Ошибка удаления: транспорт не найден")
    elif confirmation == "НЕТ":
        bot.send_message(user_id, "Удаление отменено!")
        manage_transport(message)  # Возвращаем в меню транспорта
    else:
        bot.send_message(user_id, "Ошибка! Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены", parse_mode="Markdown")
        bot.register_next_step_handler(message, lambda msg: process_confirmation(msg, transport_to_delete))

#@log_user_actions
def process_delete_all_confirmation(message):
    user_id = str(message.chat.id)
    confirmation = message.text.strip().upper()

    # Проверка на "В главное меню" или "Вернуться в ваш транспорт"
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    elif message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Проверка подтверждения на "ДА" или "НЕТ"
    if confirmation == "ДА":
        if user_id in user_transport:
            # Удаляем все транспорты
            transports = user_transport[user_id]
            user_transport[user_id] = []  # Очищаем список
            save_transport_data(user_id, user_transport[user_id])

            for transport in transports:
                delete_expenses_related_to_transport(user_id, transport)
                delete_repairs_related_to_transport(user_id, transport)

            bot.send_message(user_id, "Весь транспорт и связанные с ним траты и ремонты успешно удалены!")
            manage_transport(message)  # Возвращаем в меню транспорта
        else:
            bot.send_message(user_id, "У вас нет добавленного транспорта!")
    elif confirmation == "НЕТ":
        bot.send_message(user_id, "Удаление отменено!")
        manage_transport(message)  # Возвращаем в меню транспорта
    else:
        bot.send_message(user_id, "Ошибка! Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены", parse_mode="Markdown")
        bot.register_next_step_handler(message, process_delete_all_confirmation)

# Функция для создания клавиатуры с кнопками возврата
#@log_user_actions
def get_return_menu_keyboard():
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    item_main_menu = types.KeyboardButton("В главное меню")
    item_return_transport = types.KeyboardButton("Вернуться в ваш транспорт")
    markup.add(item_return_transport)
    markup.add(item_main_menu)
    return markup

@bot.message_handler(func=lambda message: message.text == "Удалить весь транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить весь транспорт')
@track_usage('Удалить весь транспорт')  # Добавление отслеживания статистики
#@log_user_actions
def delete_all_transports(message):
    user_id = str(message.chat.id)
    if user_id in user_transport and user_transport[user_id]:
        # Подтверждение удаления всех транспортов
        bot.send_message(
    user_id,
    "*Вы уверены, что хотите удалить весь транспорт?*\n\n"
    "Удаление транспорта приведет к удалению всех трат и ремонтов!\n\n"
    "Введите *ДА* для подтверждения или *НЕТ* для отмены",
    parse_mode="Markdown",
    reply_markup=get_return_menu_keyboard()
)
        bot.register_next_step_handler(message, process_delete_all_confirmation)
    else:
        bot.send_message(user_id, "У вас нет добавленного транспорта!")

#@log_user_actions
def process_delete_all_confirmation(message):
    user_id = str(message.chat.id)
    confirmation = message.text.strip().upper()

    # Проверка на "В главное меню" или "Вернуться в ваш транспорт"
    if message.text == "В главное меню":
        return_to_menu(message)
        return
    elif message.text == "Вернуться в ваш транспорт":
        manage_transport(message)
        return

    # Проверка подтверждения на "ДА" или "НЕТ"
    if confirmation == "ДА":
        if user_id in user_transport:
            # Удаляем все транспорты
            transports = user_transport[user_id]
            user_transport[user_id] = []  # Очищаем список
            save_transport_data(user_id, user_transport[user_id])

            for transport in transports:
                delete_expenses_related_to_transport(user_id, transport)
                delete_repairs_related_to_transport(user_id, transport)

            bot.send_message(user_id, "Весь транспорт и связанные с ним траты и ремонты успешно удалены!")
            manage_transport(message)  # Возвращаем в меню транспорта
        else:
            bot.send_message(user_id, "У вас нет добавленного транспорта!")
    elif confirmation == "НЕТ":
        bot.send_message(user_id, "Удаление отменено!")
        manage_transport(message)  # Возвращаем в меню транспорта
    else:
        bot.send_message(user_id, "Ошибка! Пожалуйста, введите *ДА* для подтверждения или *НЕТ* для отмены", parse_mode="Markdown")
        bot.register_next_step_handler(message, process_delete_all_confirmation)

@bot.message_handler(func=lambda message: message.text == "Посмотреть транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть транспорт')
@track_usage('Посмотреть транспорт')  # Добавление отслеживания статистики
#@log_user_actions
def view_transport(message):
    user_id = str(message.chat.id)
    if user_id in user_transport and user_transport[user_id]:
        transport_list = user_transport[user_id]
        response = "\n\n".join([
            f"№{index + 1}. {item['brand']} - {item['model']} - {item['year']} - `{item['license_plate']}`"
            for index, item in enumerate(transport_list)
        ])
        bot.send_message(
            user_id,
            f"🚙 *Ваш транспорт* 🚙\n\n{response}",
            parse_mode="Markdown"
        )
    else:
        bot.send_message(user_id, "У вас нет добавленного транспорта!")
    
    # Возвращаем пользователя в manage_transport
    manage_transport(message)

@bot.message_handler(func=lambda message: message.text == "Вернуться в ваш транспорт")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Вернуться в ваш транспорт')
#@log_user_actions
def return_to_transport_menu(message):
    manage_transport(message)  # Возвращаем пользователя в меню транспорта


# АНТИ-РАДАР 


# 📂 Загрузка данных о камерах из файла
camera_data = []
coordinates = []
try:
    with open('files/milestones.csv', mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file, delimiter=';')
        for row in reader:
            latitude = float(row['gps_y'])
            longitude = float(row['gps_x'])
            camera_data.append({
                'id': row['camera_id'],
                'latitude': latitude,
                'longitude': longitude,
                'description': row['camera_place'],
            })
            coordinates.append((latitude, longitude))
except Exception as e:
    pass

# 🌳 Создание KD-дерева для поиска ближайших камер
camera_tree = cKDTree(coordinates)
user_tracking = {}

# Обработчик команды для старта анти-радара
@bot.message_handler(func=lambda message: message.text == "Анти-радар")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Анти-радар')
@track_usage('Анти-радар')  # Добавление отслеживания статистики
#@log_user_actions
def start_antiradar(message):
    user_id = message.chat.id
    user_tracking[user_id] = {'tracking': True, 'notification_ids': [], 'last_notified_camera': {}, 'location': None, 'started': False}

    keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    button_geo = telebot.types.KeyboardButton(text="Отправить геопозицию", request_location=True)
    button_off_geo = telebot.types.KeyboardButton(text="Выключить анти-радар")
    keyboard.add(button_geo)
    keyboard.add(button_off_geo)
    bot.send_message(user_id, "Пожалуйста, разрешите доступ к геопозиции для запуска анти-радара. Нажмите кнопку, чтобы отправить геопозицию", reply_markup=keyboard)

    # Отправка сообщения с камерами
    message_text = "⚠️ Внимание! Камеры впереди:\n\n"  # Заголовок сообщения
    sent_message = bot.send_message(user_id, message_text, parse_mode="Markdown")
    
    # Закрепляем сообщение
    bot.pin_chat_message(user_id, sent_message.message_id)  # Закрепляем сообщение
    user_tracking[user_id]['last_camera_message'] = sent_message.message_id  # Сохраняем ID сообщения для дальнейшего использования

# Обработчик для получения геопозиции при активированном анти-радаре
@bot.message_handler(content_types=['location'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Функция для обработки локации')
@track_usage('Функция для обработки локации')  # Добавление отслеживания статистики
#@log_user_actions
def handle_antiradar_location(message):
    user_id = message.chat.id
    
    # Проверка на наличие геолокации
    if message.location:
        latitude = message.location.latitude
        longitude = message.location.longitude
    else:
        bot.send_message(user_id, "Геолокация недоступна. Попробуйте снова")
        return

    # Проверка: если анти-радар включен
    if user_tracking.get(user_id, {}).get('tracking', False):
        # Обновляем только координаты пользователя
        user_tracking[user_id]['location'] = message.location
        
        # Запускаем трекинг, если он еще не был запущен
        if not user_tracking[user_id].get('started', False):
            user_tracking[user_id]['started'] = True
            track_user_location(user_id, message.location)
    else:
        bot.send_message(user_id, "Пожалуйста, выберите категорию из меню")

# Обработчик для выключения анти-радара
@bot.message_handler(func=lambda message: message.text == "Выключить анти-радар")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Выключить анти-радар')
@track_usage('Выключить анти-радар')  # Добавление отслеживания статистики
def stop_antiradar(message):
    user_id = message.chat.id
    if user_id in user_tracking:
        user_tracking[user_id]['tracking'] = False
        bot.send_message(user_id, "Анти-радар остановлен!")

        # Удаление закрепленного сообщения с камерами
        if user_tracking[user_id].get('last_camera_message'):
            bot.unpin_chat_message(user_id, user_tracking[user_id]['last_camera_message'])  # Убираем закрепление
            bot.delete_message(user_id, user_tracking[user_id]['last_camera_message'])  # Удаляем сообщение

        # Возврат в главное меню
        return_to_menu(message)

    else:
        bot.send_message(user_id, "Анти-радар не был запущен")


def delete_messages(user_id, message_id):
    time.sleep(6)
    try:
        bot.delete_message(chat_id=user_id, message_id=message_id)
    except:
        pass


MAX_CAMERAS_IN_MESSAGE = 5  # Максимальное количество камер в одном сообщении
ALERT_DISTANCE = 150  # Расстояние для уведомления о приближении к камере
EXIT_DISTANCE = 50  # Расстояние для уведомления о выходе из зоны камеры
IN_ZONE_DISTANCE = 15  # Расстояние, при котором показывается сообщение "в зоне камеры"


def track_user_location(user_id, initial_location):
    def monitor():
        while user_tracking.get(user_id, {}).get('tracking', False):
            user_location = user_tracking[user_id]['location']
            user_position = (user_location.latitude, user_location.longitude)
            distances, indices = camera_tree.query(user_position, k=len(camera_data))
            nearest_cameras = []
            unique_addresses = set()

            for i, distance in enumerate(distances[:MAX_CAMERAS_IN_MESSAGE]):
                if distance <= 1000:
                    camera = camera_data[indices[i]]
                    actual_distance = int(geodesic(user_position, (camera['latitude'], camera['longitude'])).meters)
                    camera_address = camera['description']

                    if camera_address not in user_tracking[user_id]['last_notified_camera']:
                        nearest_cameras.append((actual_distance, camera))
                        unique_addresses.add(camera_address)
                        user_tracking[user_id]['last_notified_camera'][camera_address] = {
                            'entered': False,
                            'exited': False,
                            'in_zone': False
                        }

                    if actual_distance <= ALERT_DISTANCE and not user_tracking[user_id]['last_notified_camera'][camera_address]['entered']:
                        notification_message = (
                            f"⚠️ Вы приближаетесь к камере!\n\n"
                            f"📷 Расстояние - *{actual_distance} м.*\n"
                            f"🗺️ *{camera_address}*"
                        )
                        try:
                            sent_message = bot.send_message(user_id, notification_message, parse_mode="Markdown")
                            user_tracking[user_id]['notification_ids'].append(sent_message.message_id)
                            user_tracking[user_id]['last_notified_camera'][camera_address]['entered'] = True
                        except:
                            pass

                    if actual_distance <= IN_ZONE_DISTANCE and not user_tracking[user_id]['last_notified_camera'][camera_address]['in_zone']:
                        in_zone_message = (
                            f"📍 Вы находитесь в зоне действия камеры!\n\n"
                            f"🗺️ *{camera_address}*"
                        )
                        try:
                            in_zone_sent = bot.send_message(user_id, in_zone_message, parse_mode="Markdown")
                            user_tracking[user_id]['notification_ids'].append(in_zone_sent.message_id)
                            user_tracking[user_id]['last_notified_camera'][camera_address]['in_zone'] = True
                        except:
                            pass

                    if actual_distance > EXIT_DISTANCE and user_tracking[user_id]['last_notified_camera'][camera_address]['entered']:
                        exit_message = (
                            f"🔔 Вы вышли из зоны действия камеры!\n\n"
                            f"🗺️ *{camera_address}*"
                        )
                        try:
                            exit_sent_message = bot.send_message(user_id, exit_message, parse_mode="Markdown")
                            user_tracking[user_id]['notification_ids'].append(exit_sent_message.message_id)
                            user_tracking[user_id]['last_notified_camera'][camera_address]['exited'] = True
                        except:
                            pass

            message_text = "⚠️ Внимание! Камеры впереди:\n\n"
            for distance, camera in nearest_cameras:
                message_text += f"📷 Камера через *{distance} м.*\n🗺️ Адрес: *{camera['description']}*\n\n"

            if nearest_cameras:
                try:
                    if user_tracking[user_id].get('last_camera_message'):
                        bot.edit_message_text(chat_id=user_id, message_id=user_tracking[user_id]['last_camera_message'], text=message_text, parse_mode="Markdown")
                    else:
                        sent_message = bot.send_message(user_id, message_text, parse_mode="Markdown")
                        user_tracking[user_id]['last_camera_message'] = sent_message.message_id
                except:
                    pass

            time.sleep(3)

    threading.Thread(target=monitor, daemon=True).start()


# ----------------------------------КОД ДЛЯ НАПОМИНАНИЙ---------------------------------------

# Путь к файлу базы данных
DB_PATH = 'data base/reminders/reminders.json'

# Функция для загрузки данных
def load_data():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

    if not os.path.exists(DB_PATH):
        with open(DB_PATH, 'w', encoding='utf-8') as file:
            json.dump({"users": {}}, file, indent=4, ensure_ascii=False)

    with open(DB_PATH, 'r', encoding='utf-8') as file:
        data = json.load(file)

    if 'users' not in data:
        data['users'] = {}

    return data

# Функция для сохранения данных
def save_data(data):
    if 'users' not in data:
        data['users'] = {}

    with open(DB_PATH, 'w', encoding='utf-8') as file:
        json.dump(data, file, indent=4, ensure_ascii=False)

# Функция для возврата в меню напоминаний
@bot.message_handler(func=lambda message: message.text == "Вернуться в меню напоминаний")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Вернуться в меню напоминаний')
#@log_user_actions
def return_to_reminders_menu(message):
    reminders_menu(message)

# Функция для возврата в главное меню
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('В главное меню')
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == "В главное меню")
@bot.message_handler(commands=['mainmenu'])
def return_to_menu(message):
    start(message)

def send_reminders():
    data = load_data()
    current_time = datetime.now()
    logging.info(f"Current time: {current_time}")

    for user_id, user_data in data["users"].items():
        reminders = user_data.get("reminders", [])
        for reminder in reminders:
            reminder_type = reminder.get("type")
            reminder_datetime = datetime.strptime(reminder["date"] + " " + reminder["time"], "%d.%m.%Y %H:%M")
            logging.info(f"Checking reminder: {reminder}")

            if reminder["status"] == "active":
                if reminder_type == "один раз":
                    if reminder_datetime <= current_time:
                        bot.send_message(user_id, f"⏰ *У вас напоминание!* ⏰\n\n\n📝 Название: {reminder['title']} \n\n📅 Дата: {reminder['date']} \n🕒 Время: {reminder['time']} \n🔖 Тип: {reminder['type']}", parse_mode="Markdown")
                        reminder["status"] = "expired"
                elif reminder_type == "ежедневно":
                    if current_time.date() == reminder_datetime.date() and current_time.time() >= reminder_datetime.time():
                        bot.send_message(user_id, f"⏰ *У вас напоминание!* ⏰\n\n\n📝 Название: {reminder['title']} \n\n📅 Дата: {reminder['date']} \n🕒 Время: {reminder['time']} \n🔖 Тип: {reminder['type']}", parse_mode="Markdown")
                        reminder["date"] = (reminder_datetime + timedelta(days=1)).strftime("%d.%m.%Y")
                elif reminder_type == "еженедельно":
                    if current_time.date() == reminder_datetime.date() and current_time.time() >= reminder_datetime.time():
                        bot.send_message(user_id, f"⏰ *У вас напоминание!* ⏰\n\n\n📝 Название: {reminder['title']} \n\n📅 Дата: {reminder['date']} \n🕒 Время: {reminder['time']} \n🔖 Тип: {reminder['type']}", parse_mode="Markdown")
                        reminder["date"] = (reminder_datetime + timedelta(weeks=1)).strftime("%d.%m.%Y")
                elif reminder_type == "ежемесячно":
                    if current_time.date() == reminder_datetime.date() and current_time.time() >= reminder_datetime.time():
                        bot.send_message(user_id, f"⏰ *У вас напоминание!* ⏰\n\n\n📝 Название: {reminder['title']} \n\n📅 Дата: {reminder['date']} \n🕒 Время: {reminder['time']} \n🔖 Тип: {reminder['type']}", parse_mode="Markdown")
                        next_month = reminder_datetime.month % 12 + 1
                        next_year = reminder_datetime.year + (reminder_datetime.month // 12)
                        reminder["date"] = reminder_datetime.replace(day=reminder_datetime.day, month=next_month, year=next_year).strftime("%d.%m.%Y")

        save_data(data)

# Функция для фонового планировщика
def run_scheduler():
    while True:
        send_reminders()
        time.sleep(15)

# Запуск фона в отдельном потоке
threading.Thread(target=run_scheduler, daemon=True).start()

# Обработка кнопки "Напоминания"
@bot.message_handler(func=lambda message: message.text == "Напоминания")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Напоминания')
@track_usage('Напоминания')
#@log_user_actions
def reminders_menu(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Добавить напоминание', 'Посмотреть напоминания', 'Удалить напоминание')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите действие для напоминаний:", reply_markup=markup)

# Обработка кнопки "Добавить напоминание"
@bot.message_handler(func=lambda message: message.text == "Добавить напоминание")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Добавить напоминание')
@track_usage('Добавить напоминание')  # Добавление отслеживания статистики
#@log_user_actions
def add_reminder(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    msg = bot.send_message(message.chat.id, "Введите название напоминания:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_title_step)

#@log_user_actions
def process_title_step(message):
    user_id = str(message.from_user.id)
    data = load_data()

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_title_step)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    if user_id not in data["users"]:
        data["users"][user_id] = {"reminders": []}

    reminder = {"title": message.text}
    data["users"][user_id]["current_reminder"] = reminder
    save_data(data)

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Один раз', 'Ежедневно')
    markup.add('Еженедельно', 'Ежемесячно')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    msg = bot.send_message(message.chat.id, "Выберите тип напоминания:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_type_step)

#@log_user_actions
def process_type_step(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminder = data["users"][user_id]["current_reminder"]

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_type_step)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    reminder_type = message.text.lower()
    if reminder_type in ["ежедневно", "еженедельно", "ежемесячно", "один раз"]:
        reminder["type"] = reminder_type
    else:
        msg = bot.send_message(message.chat.id, "Неверный тип напоминания. Пожалуйста, выберите из предложенных вариантов.")
        bot.register_next_step_handler(msg, process_type_step)
        return

    save_data(data)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    msg = bot.send_message(message.chat.id, "Введите дату напоминания:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_date_step_for_repairs)

#@log_user_actions
def process_date_step_for_repairs(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminder = data["users"][user_id]["current_reminder"]

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_date_step_for_repairs)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    date_input = message.text

    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", date_input):
        try:
            day, month, year = map(int, date_input.split('.'))
            if 1 <= month <= 12 and 1 <= day <= 31 and 2000 <= year <= 3000:
                reminder_date = datetime.strptime(date_input, "%d.%m.%Y")
                current_date = datetime.now()
                if reminder_date.date() >= current_date.date():
                    reminder["date"] = date_input
                else:
                    raise ValueError("Дата не может быть меньше текущей даты.")
            else:
                raise ValueError
        except ValueError as e:
            msg = bot.send_message(message.chat.id, f"Ошибка: {e}. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ")
            bot.register_next_step_handler(msg, process_date_step_for_repairs)
            return
    else:
        msg = bot.send_message(message.chat.id, "Неверный формат даты. Попробуйте еще раз в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(msg, process_date_step_for_repairs)
        return

    save_data(data)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    msg = bot.send_message(message.chat.id, "Введите время напоминания:", reply_markup=markup)
    bot.register_next_step_handler(msg, process_time_step)

#@log_user_actions
def process_time_step(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminder = data["users"][user_id]["current_reminder"]

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_time_step)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    time_input = message.text

    if re.match(r"^\d{2}:\d{2}$", time_input):
        try:
            hour, minute = map(int, time_input.split(':'))
            if 0 <= hour <= 23 and 0 <= minute <= 59:
                reminder_time = datetime.strptime(time_input, "%H:%M").time()
                current_time = datetime.now().time()
                reminder_date = datetime.strptime(reminder["date"], "%d.%m.%Y").date()
                current_date = datetime.now().date()

                if reminder_date > current_date or (reminder_date == current_date and reminder_time >= current_time):
                    reminder["time"] = time_input
                    reminder["status"] = "active"
                else:
                    raise ValueError("Время не может быть меньше текущего времени.")
            else:
                raise ValueError
        except ValueError as e:
            msg = bot.send_message(message.chat.id, f"Ошибка: {e}. Пожалуйста, введите время в формате ЧЧ:ММ.")
            bot.register_next_step_handler(msg, process_time_step)
            return
    else:
        msg = bot.send_message(message.chat.id, "Неверный формат времени. Попробуйте еще раз в формате ЧЧ:ММ.")
        bot.register_next_step_handler(msg, process_time_step)
        return

    data["users"][user_id]["reminders"].append(reminder)
    del data["users"][user_id]["current_reminder"]
    save_data(data)

    bot.send_message(message.chat.id, "Напоминание добавлено!")
    reminders_menu(message)

# Обработка кнопки "Посмотреть напоминания"
@bot.message_handler(func=lambda message: message.text == "Посмотреть напоминания")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть напоминания')
@track_usage('Посмотреть напоминания')  # Добавление отслеживания статистики
#@log_user_actions
def view_reminders(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Активные', 'Истекшие')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите тип напоминаний:", reply_markup=markup)

# Обработка выбора "Активные" напоминания
@bot.message_handler(func=lambda message: message.text == "Активные")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Активные')
@track_usage('Активные')  # Добавление отслеживания статистики
#@log_user_actions
def view_active_reminders(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Один раз (активные)', 'Ежедневно (активные)')
    markup.add('Еженедельно (активные)', 'Ежемесячно (активные)')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите тип активных напоминаний:", reply_markup=markup)

# Обработка выбора "Истекшие" напоминания
@bot.message_handler(func=lambda message: message.text == "Истекшие")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Истекшие')
@track_usage('Истекшие')  # Добавление отслеживания статистики
#@log_user_actions
def view_expired_reminders(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Один раз (истекшие)', 'Ежедневно (истекшие)')
    markup.add('Еженедельно (истекшие)', 'Ежемесячно (истекшие)')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите тип истекших напоминаний:", reply_markup=markup)


# Обработка выбора типа активных напоминаний
@bot.message_handler(func=lambda message: message.text in ['Один раз (активные)', 'Ежедневно (активные)', 'Еженедельно (активные)', 'Ежемесячно (активные)'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Один раз (активные)')
@check_function_state_decorator('Ежедневно (активные)')
@check_function_state_decorator('Еженедельно (активные)')
@check_function_state_decorator('Ежемесячно (активные)')
@track_usage('Один раз (активные)')
@track_usage('Ежедневно (активные)')
@track_usage('Еженедельно (активные)')
@track_usage('Ежемесячно (активные)')
#@log_user_actions
def view_active_reminders_by_type(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminders = data["users"].get(user_id, {}).get("reminders", [])
    current_date = datetime.now()

    # Извлекаем тип напоминания из сообщения
    reminder_type = message.text.split(' ')[0].lower() + ' ' + message.text.split(' ')[1].lower()
    reminder_type = reminder_type.replace("(активные)", "").strip()

    active = []
    for reminder in reminders:
        reminder_datetime = datetime.strptime(reminder["date"] + ' ' + reminder["time"], "%d.%m.%Y %H:%M")

        if reminder["status"] == "active" and reminder["type"] == reminder_type:
            if reminder_type == "один раз" and reminder_datetime >= current_date:
                active.append(reminder)
            elif reminder_type == "ежедневно" and reminder_datetime.date() >= current_date.date():
                active.append(reminder)
            elif reminder_type == "еженедельно" and reminder_datetime.weekday() == current_date.weekday():
                active.append(reminder)
            elif reminder_type == "ежемесячно" and reminder_datetime.day == current_date.day:
                active.append(reminder)

    if active:
        response = f"*Активные напоминания ({reminder_type}):*\n\n"
        for i, reminder in enumerate(active, 1):
            response += (
                f"\n⭐ №{i} ⭐\n\n\n"
                f"📝 Название: {reminder['title']}\n\n"
                f"📅 Дата: {reminder['date']}\n"
                f"🕒 Время: {reminder['time']}\n"
                f"✅ Статус: Активное\n"
                f"🔖 Тип: {reminder['type']}\n\n"
            )
    else:
        response = f"*Активные напоминания ({reminder_type}):*\n\nНет активных напоминаний"

    bot.send_message(message.chat.id, response, parse_mode="Markdown")

# Обработка выбора типа истекших напоминаний
@bot.message_handler(func=lambda message: message.text in ['Один раз (истекшие)', 'Ежедневно (истекшие)', 'Еженедельно (истекшие)', 'Ежемесячно (истекшие)'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Один раз (истекшие)')
@check_function_state_decorator('Ежедневно (истекшие)')
@check_function_state_decorator('Еженедельно (истекшие)')
@check_function_state_decorator('Ежемесячно (истекшие)')
@track_usage('Один раз (истекшие)')
@track_usage('Ежедневно (истекшие)')
@track_usage('Еженедельно (истекшие)')
@track_usage('Ежемесячно (истекшие)')
#@log_user_actions
def view_expired_reminders_by_type(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminders = data["users"].get(user_id, {}).get("reminders", [])

    # Извлекаем тип напоминания из сообщения
    reminder_type = message.text.split(' ')[0].lower() + ' ' + message.text.split(' ')[1].lower()
    reminder_type = reminder_type.replace("(истекшие)", "").strip()

    expired = []
    for reminder in reminders:
        if reminder["status"] == "expired" and reminder["type"] == reminder_type:
            expired.append(reminder)

    if expired:
        response = f"*Истекшие напоминания ({reminder_type}):*\n\n"
        for i, reminder in enumerate(expired, 1):
            response += (
                f"\n❌ №{i} ❌\n\n\n"
                f"📝 Название: {reminder['title']}\n\n"
                f"📅 Дата: {reminder['date']}\n"
                f"🕒 Время: {reminder['time']}\n"
                f"✅ Статус: Истекло\n"
                f"🔖 Тип: {reminder['type']}\n\n"
            )
    else:
        response = f"*Истекшие напоминания ({reminder_type}):*\n\nНет истекших напоминаний"

    bot.send_message(message.chat.id, response, parse_mode="Markdown")

# Обработка кнопки "Удалить"
@bot.message_handler(func=lambda message: message.text == "Удалить напоминание")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить напоминание')
@track_usage('Удалить напоминание')  # Добавление отслеживания статистики
#@log_user_actions
def delete_reminder(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Del Активные', 'Del Истекшие')
    markup.add('Удалить все напоминания')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите тип напоминаний для удаления:", reply_markup=markup)

# Обработка выбора "Del Активные" напоминания
@bot.message_handler(func=lambda message: message.text == "Del Активные")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del Активные')
@track_usage('Del Активные')  # Добавление отслеживания статистики
#@log_user_actions
def delete_active_reminders(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Del Один раз (активные)', 'Del Ежедневно (активные)')
    markup.add('Del Еженедельно (активные)', 'Del Ежемесячно (активные)')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите тип активных напоминаний для удаления:", reply_markup=markup)

# Обработка выбора "Del Истекшие" напоминания
@bot.message_handler(func=lambda message: message.text == "Del Истекшие")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del Истекшие')
@track_usage('Del Истекшие')  # Добавление отслеживания статистики
#@log_user_actions
def delete_expired_reminders(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Del Один раз (истекшие)', 'Del Ежедневно (истекшие)')
    markup.add('Del Еженедельно (истекшие)', 'Del Ежемесячно (истекшие)')
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите тип истекших напоминаний для удаления:", reply_markup=markup)

# Обработка выбора типа активных напоминаний для удаления
@bot.message_handler(func=lambda message: message.text in ['Del Один раз (активные)', 'Del Ежедневно (активные)', 'Del Еженедельно (активные)', 'Del Ежемесячно (активные)'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del Один раз (активные)')
@check_function_state_decorator('Del Ежедневно (активные)')
@check_function_state_decorator('Del Еженедельно (активные)')
@check_function_state_decorator('Del Ежемесячно (активные)')
@track_usage('Del Один раз (активные)')
@track_usage('Del Ежедневно (активные)')
@track_usage('Del Еженедельно (активные)')
@track_usage('Del Ежемесячно (активные)')
#@log_user_actions
def delete_active_reminders_by_type(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminders = data["users"].get(user_id, {}).get("reminders", [])

    reminder_type = message.text.split(' ')[1].lower() + ' ' + message.text.split(' ')[2].lower()
    reminder_type = reminder_type.replace("(активные)", "").strip()

    active_reminders = [reminder for reminder in reminders if reminder["status"] == "active" and reminder["type"] == reminder_type]

    if not active_reminders:
        bot.send_message(
            message.chat.id,
            f"*Активные напоминания ({reminder_type}) для удаления:*\n\nНет активных напоминаний",
            parse_mode="Markdown",
        )
        delete_active_reminders(message)  # Автоматический возврат в меню выбора активных напоминаний
        return

    response = f"*Активные напоминания ({reminder_type}) для удаления:*\n\n"
    for i, reminder in enumerate(active_reminders, 1):
        response += (
            f"\n❌ №{i} ❌\n\n\n"
            f"📝 Название: {reminder['title']}\n\n"
            f"📅 Дата: {reminder['date']}\n"
            f"🕒 Время: {reminder['time']}\n"
            f"✅ Статус: Активное\n"
            f"🔖 Тип: {reminder['type']}\n\n"
        )

    bot.send_message(message.chat.id, response, parse_mode="Markdown")

    # Сохраняем контекст
    data["users"][user_id]["current_reminder_type"] = reminder_type
    data["users"][user_id]["current_reminders"] = active_reminders
    save_data(data)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Введите номер для удаления напоминания:", reply_markup=markup)
    bot.register_next_step_handler(message, confirm_delete_active_step)

# Обработка выбора типа истекших напоминаний для удаления
@bot.message_handler(func=lambda message: message.text in ['Del Один раз (истекшие)', 'Del Ежедневно (истекшие)', 'Del Еженедельно (истекшие)', 'Del Ежемесячно (истекшие)'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Del Один раз (истекшие)')
@check_function_state_decorator('Del Ежедневно (истекшие)')
@check_function_state_decorator('Del Еженедельно (истекшие)')
@check_function_state_decorator('Del Ежемесячно (истекшие)')
@track_usage('Del Один раз (истекшие)')
@track_usage('Del Ежедневно (истекшие)')
@track_usage('Del Еженедельно (истекшие)')
@track_usage('Del Ежемесячно (истекшие)')
#@log_user_actions
def delete_expired_reminders_by_type(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminders = data["users"].get(user_id, {}).get("reminders", [])

    reminder_type = message.text.split(' ')[1].lower() + ' ' + message.text.split(' ')[2].lower()
    reminder_type = reminder_type.replace("(истекшие)", "").strip()

    expired_reminders = [reminder for reminder in reminders if reminder["status"] == "expired" and reminder["type"] == reminder_type]

    if not expired_reminders:
        bot.send_message(
            message.chat.id,
            f"*Истекшие напоминания ({reminder_type}) для удаления:*\n\nНет истекших напоминаний",
            parse_mode="Markdown",
        )
        delete_expired_reminders(message)  # Автоматический возврат в меню выбора истекших напоминаний
        return

    response = f"*Истекшие напоминания ({reminder_type}) для удаления:*\n\n"
    for i, reminder in enumerate(expired_reminders, 1):
        response += (
            f"\n❌ №{i} ❌\n\n\n"
            f"📝 Название: {reminder['title']}\n\n"
            f"📅 Дата: {reminder['date']}\n"
            f"🕒 Время: {reminder['time']}\n"
            f"✅ Статус: Истекло\n"
            f"🔖 Тип: {reminder['type']}\n\n"
        )

    bot.send_message(message.chat.id, response, parse_mode="Markdown")

    # Сохраняем контекст
    data["users"][user_id]["current_reminder_type"] = reminder_type
    data["users"][user_id]["current_reminders"] = expired_reminders
    save_data(data)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Введите номер для удаления напоминания:", reply_markup=markup)
    bot.register_next_step_handler(message, confirm_delete_expired_step)

#@log_user_actions
def confirm_delete_active_step(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminders = data["users"][user_id].get("current_reminders", [])
    reminder_type = data["users"][user_id].get("current_reminder_type")

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_active_step)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    try:
        reminder_index = int(message.text) - 1
        if 0 <= reminder_index < len(reminders) and reminders[reminder_index]["type"] == reminder_type:
            user_reminders = data["users"][user_id]["reminders"]
            user_reminders.remove(reminders[reminder_index])
            save_data(data)
            bot.send_message(message.chat.id, f"Напоминание *№{reminder_index + 1}* удалено!", parse_mode="Markdown")
            reminders_menu(message)
        else:
            raise IndexError
    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный номер напоминания. Пожалуйста, введите правильный номер")
        bot.register_next_step_handler(message, confirm_delete_active_step)

#@log_user_actions
def confirm_delete_expired_step(message):
    user_id = str(message.from_user.id)
    data = load_data()
    reminders = data["users"][user_id].get("current_reminders", [])
    reminder_type = data["users"][user_id].get("current_reminder_type")

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_expired_step)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    try:
        reminder_index = int(message.text) - 1
        if 0 <= reminder_index < len(reminders) and reminders[reminder_index]["type"] == reminder_type:
            user_reminders = data["users"][user_id]["reminders"]
            user_reminders.remove(reminders[reminder_index])
            save_data(data)
            bot.send_message(message.chat.id, f"Напоминание *№{reminder_index + 1}* удалено!", parse_mode="Markdown")
            reminders_menu(message)
        else:
            raise IndexError
    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный номер напоминания. Пожалуйста, введите правильный номер")
        bot.register_next_step_handler(message, confirm_delete_expired_step)

# Обработка кнопки "Удалить все напоминания"
@bot.message_handler(func=lambda message: message.text == "Удалить все напоминания")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Удалить все напоминания')
@track_usage('Удалить все напоминания')  # Добавление отслеживания статистики
#@log_user_actions
def delete_all_reminders(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Вернуться в меню напоминаний')
    markup.add('В главное меню')

    bot.send_message(message.chat.id, "Вы уверены, что хотите удалить все напоминания? Напишите *ДА* или *НЕТ*", reply_markup=markup, parse_mode="Markdown")

    bot.register_next_step_handler(message, confirm_delete_all_step)

#@log_user_actions
def confirm_delete_all_step(message):
    user_id = str(message.from_user.id)
    data = load_data()

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, confirm_delete_all_step)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Вернуться в меню напоминаний":
        reminders_menu(message)
        return

    user_reminders = data["users"].get(user_id, {}).get("reminders", [])

    if message.text.strip().upper() == "ДА":
        if not user_reminders:
            bot.send_message(message.chat.id, "У вас нет напоминаний!", parse_mode="Markdown")
        else:
            # Удаляем все напоминания и очищаем контекст
            data["users"][user_id]["reminders"] = []
            data["users"][user_id]["current_reminder_type"] = None
            data["users"][user_id]["current_reminders"] = []
            save_data(data)
            bot.send_message(message.chat.id, "Все напоминания удалены!", parse_mode="Markdown")
        
        reminders_menu(message)
    elif message.text.strip().upper() == "НЕТ":
        bot.send_message(message.chat.id, "Удаление всех напоминаний отменено!", parse_mode="Markdown")
        reminders_menu(message)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, напишите *ДА* или *НЕТ*", parse_mode="Markdown")
        bot.register_next_step_handler(message, confirm_delete_all_step)


#----------------------------------------- КОДЫ OBD2-------------------------------

# Загрузка кодов ошибок OBD-II из файла
def load_error_codes():
    error_codes = {}
    with open("files/codes_obd2.txt", "r", encoding="utf-8") as file:
        for line in file:
            parts = line.strip().split(" ", 1)
            if len(parts) == 2:
                code, description = parts
                error_codes[code] = description
    return error_codes

# Загружаем коды ошибок
error_codes = load_error_codes()

# Обработчик нажатия кнопки "OBD2"
@bot.message_handler(func=lambda message: message.text == "Коды OBD2")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Коды OBD2')
@track_usage('Коды OBD2')  # Добавление отслеживания статистики
#@log_user_actions
def obd2_request(message):
    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        # Сообщение об ошибке
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        # Повторный запрос на ввод, но без отправки повторного текста
        msg = bot.send_message(message.chat.id, "Введите код ошибки (или несколько через запятую):")
        bot.register_next_step_handler(msg, process_error_codes)
        return
    
    # Устанавливаем клавиатуру с кнопкой "В главное меню"
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(telebot.types.KeyboardButton("В главное меню"))
    
    # Запрашиваем ввод кода ошибки
    msg = bot.send_message(message.chat.id, "Введите код ошибки (можно несколько через запятую):", reply_markup=markup)
    bot.register_next_step_handler(msg, process_error_codes)

# Обработка введенных кодов ошибок
#@log_user_actions
def process_error_codes(message):
    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        # Сообщение об ошибке
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        # Не отправляем повторное сообщение, а только регистрируем следующий шаг
        bot.register_next_step_handler(message, process_error_codes)
        return
    
    # Проверка, если пользователь нажал "В главное меню"
    if message.text == "В главное меню":
        return_to_menu(message)
        return

    # Разделение и поиск описания для введенных кодов
    codes = [code.strip().upper() for code in message.text.split(",")]
    response = ""
    
    for code in codes:
        if code in error_codes:
            response += f"🔧 *Код ошибки*: {code}\n📋 *Описание*: {error_codes[code]}\n\n\n"
        else:
            response += f"🔧 *Код ошибки*: {code}\n❌ *Описание*: Не найдено\n\n\n"
    # Отправляем ответ пользователю
    bot.send_message(message.chat.id, response, parse_mode="Markdown")

    # Кнопки для повторного ввода кода ошибки или возврата в главное меню
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(telebot.types.KeyboardButton("Посмотреть другие ошибки"))
    markup.add(telebot.types.KeyboardButton("В главное меню"))
    bot.send_message(message.chat.id, "Вы можете посмотреть другие ошибки или вернуться в меню", reply_markup=markup)

# Обработчик кнопки "Посмотреть другие ошибки"
@bot.message_handler(func=lambda message: message.text == "Посмотреть другие ошибки")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Посмотреть другие ошибки')
#@log_user_actions
def another_error_request(message):
    # Проверка на мультимедийные файлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        # Сообщение об ошибке
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        # Не отправляем повторное сообщение, а только регистрируем следующий шаг
        bot.register_next_step_handler(message, process_error_codes)
        return

    obd2_request(message)




# (ADMIN) ----------------------------------------------- КОД ДЛЯ "АНМИН-ПАНЕЛИ" ------------------------------------------------------

# (ADMIN 1) -------------------------------------------------"ВХОД ДЛЯ АДМИНА" --------------------------------------------------------

# (ADMIN 2) ---------------------------------------------- "МЕНЮ ДЛЯ АДМИН-ПАНЕЛИ" ------------------------------------------------

# (ADMIN 3) --------------------------------- " ФУНКЦИЯ "АДМИН" " ---------------------------------------------

ADMIN_USERNAME = "EvgenyAlex21"
ADMIN_PASSWORD = "HH1515az!"

admin_sessions = set()

ADMIN_SESSIONS_PATH = 'data base/admin/admin_sessions.json'

# Глобальная переменная для отслеживания изменения данных входа
credentials_changed = False

admin_sessions = set()

# Функции для работы с единой базой данных
# В функции load_admin_data добавляем загрузку удалённых админов
def load_admin_data():
    if os.path.exists(ADMIN_SESSIONS_PATH):
        with open(ADMIN_SESSIONS_PATH, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return {
                "admin_sessions": set(data.get("admin_sessions", [])),  # Загружаем как множество
                "admins_data": data.get("admins_data", {}),
                "removed_admins": {k: v for k, v in data.get("removed_admins", {}).items()},
                "login_password_hash": data.get("login_password_hash", "")
            }
    return {
        "admin_sessions": set(),
        "admins_data": {},
        "removed_admins": {},
        "login_password_hash": ""
    }

def save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins=None):
    with open(ADMIN_SESSIONS_PATH, 'w', encoding='utf-8') as file:
        json.dump({
            "admin_sessions": list(admin_sessions),  # Сохраняем как список
            "admins_data": admins_data,
            "removed_admins": removed_admins or {},
            "login_password_hash": login_password_hash
        }, file, ensure_ascii=False, indent=4)


admin_sessions = set()

# Загрузка данных из единой базы
data = load_admin_data()
admin_sessions = data["admin_sessions"]
admins_data = data["admins_data"]
removed_admins = data["removed_admins"]
login_password_hash = data["login_password_hash"]


login_password_hash = hashlib.sha256(f"{ADMIN_USERNAME}:{ADMIN_PASSWORD}".encode()).hexdigest()

# Функция для получения хеша
#@log_user_actions
def get_login_password_hash():
    return hashlib.sha256(f"{ADMIN_USERNAME}:{ADMIN_PASSWORD}".encode()).hexdigest()

#@log_user_actions
def update_login_password(chat_id, new_username=None, new_password=None):
    global admins_data, login_password_hash

    # Используем chat_id для получения admin_id
    admin_id = str(chat_id)

    # Проверяем, существует ли администратор с таким ID
    if admin_id not in admins_data:
        return f"Администратор с ID {admin_id} не найден."

    # Получаем текущие данные администратора
    admin_data = admins_data[admin_id]
    current_username = admin_data["admins_username"]
    current_password_hash = admin_data["login_password_hash_for_user_id"]

    # Если новый логин задан, обновляем его
    if new_username:
        # Сохраняем старый логин для сравнения
        old_username = current_username

        # Обновляем логин
        admin_data["admins_username"] = new_username

        # Если логин действительно изменился, пересчитываем хэш
        if old_username != new_username:
            admin_data["login_password_hash_for_user_id"] = hashlib.sha256(f"{new_username}:{new_password or current_password_hash}".encode()).hexdigest()

    # Если задан новый пароль, пересчитываем хэш
    if new_password:
        new_hash = hashlib.sha256(f"{new_username or current_username}:{new_password}".encode()).hexdigest()
        admin_data["login_password_hash_for_user_id"] = new_hash
    else:
        # Если пароль не изменился, оставляем текущий хэш
        admin_data["login_password_hash_for_user_id"] = current_password_hash

    # Обновляем глобальный хэш
    login_password_hash = hashlib.sha256(f"{new_username or current_username}:{new_password or current_password_hash}".encode()).hexdigest()

    # Сохраняем обновлённые данные
    admins_data[admin_id] = admin_data

    # Обновляем данные в кэше
    save_admin_data(admin_sessions, admins_data, login_password_hash)

    # Возвращаем результат
    if new_username and new_password:
        return f"Логин обновлён на {new_username}, пароль обновлён."
    elif new_username:
        return f"Логин обновлён на {new_username}. Пароль остался прежним."
    elif new_password:
        return "Пароль обновлён."
    else:
        return "Изменений не было."

def verify_login_password_hash():
    global login_password_hash
    # Получаем актуальный хеш
    current_hash = hashlib.sha256(f"{ADMIN_USERNAME}:{ADMIN_PASSWORD}".encode()).hexdigest()

    # Сравниваем с сохранённым хешем
    if current_hash != login_password_hash:
        login_password_hash = current_hash
        save_admin_data(admin_sessions, admins_data, login_password_hash)

verify_login_password_hash()

# Функция изменения логина и пароля
#@log_user_actions
def change_admin_credentials(new_username=None, new_password=None):
    global ADMIN_USERNAME, ADMIN_PASSWORD, login_password_hash
    
    # Получаем текущие значения для обновления
    current_username = ADMIN_USERNAME
    current_password = ADMIN_PASSWORD
    
    if new_username:
        ADMIN_USERNAME = new_username
        current_username = new_username  # обновляем текущий логин
    if new_password:
        ADMIN_PASSWORD = new_password
        current_password = new_password  # обновляем текущий пароль
    
    # Генерация хеша с актуальными данными логина и пароля
    login_password_hash = hashlib.sha256(f"{current_username}:{current_password}".encode()).hexdigest()
    
    # Обновляем данные в базе
    save_admin_data(admin_sessions, admins_data, login_password_hash)

# Функции управления администраторами. Добавление администратора
#@log_user_actions
def add_admin(admin_id, username, permissions=None, initiator_chat_id=None):
    admin_id = str(admin_id)
    if permissions is None:
        permissions = ["Админ", "Смена данных входа", "Сменить пароль", "Сменить логин и пароль", "Статистика", "Пользователи", "Использование функций", "Список ошибок", "Версия и аптайм"]  # Права по умолчанию
    user_data = {
        "user_id": admin_id,
        "first_name": " ",
        "last_name": " ",
        "username": username,
        "phone": " ",
        "permissions": permissions,
        "is_new": True  # Флаг, указывающий, что администратор новый
    }
    admins_data[admin_id] = user_data
    admin_sessions.append(admin_id)  # Добавляем в текущие сессии
    save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)
    bot.send_message(admin_id, "✅ Вы стали администратором! Быстрый вход по команде /admin доступен...")
    if initiator_chat_id:
        bot.send_message(initiator_chat_id, f"Администратор {escape_markdown(username)} - `{admin_id}` успешно добавлен!", parse_mode="Markdown")

#@log_user_actions
def remove_admin(admin_id, initiator_chat_id):
    admin_id = str(admin_id)
    if admin_id in admins_data:
        admin_username = admins_data[admin_id]["username"]
        # Добавляем в список удалённых администраторов
        removed_admins[admin_id] = {"username": admin_username}

        # Удаляем данные администратора
        del admins_data[admin_id]
        if admin_id in admin_sessions:
            admin_sessions.remove(admin_id)  # Удаляем из текущих сессий

        # Сохраняем изменения
        save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

        # Отправляем сообщение удалённому администратору
        bot.send_message(admin_id, "🚫 Вас удалили из администраторов!")

        # Отправляем сообщение инициатору
        bot.send_message(initiator_chat_id, f"Администратор {escape_markdown(admin_username)} - {admin_id} успешно удалён!")
    else:
        bot.send_message(initiator_chat_id, f"Администратор с ID {admin_id} не найден.")

#@log_user_actions
def check_permission(admin_id, permission):
    return permission in admins_data.get(str(admin_id), {}).get("permissions", [])

# Функция для получения данных о пользователе Telegram
#@log_user_actions
def get_user_data(message):
    user = message.from_user
    return {
        "user_id": user.id,
        "first_name": user.first_name if user.first_name else " ",
        "last_name": user.last_name if user.last_name else " ",
        "username": f"@{user.username}" if user.username else " ",
        "phone": user.phone_number if hasattr(user, 'phone_number') else " "
    }

# Функция обновления данных администратора в базе
#@log_user_actions
def update_admin_data(user_data):
    admin_id = str(user_data["user_id"])
    
    # Проверяем, не находится ли администратор в списке удалённых
    if admin_id in removed_admins:
        print(f"Пользователь с ID {admin_id} был ранее удалён и не может быть добавлен обратно!")
        return
    
    if admin_id in admins_data:
        existing_data = admins_data[admin_id]
        # Обновляем только если данные изменились
        if (existing_data.get("first_name") != user_data["first_name"] or
            existing_data.get("last_name") != user_data["last_name"] or
            existing_data.get("username") != user_data["username"] or
            existing_data.get("phone") != user_data["phone"]):
            admins_data[admin_id].update(user_data)
            save_admin_data(admin_sessions, admins_data, login_password_hash)
    else:
        # Если пользователя нет в базе и он не удалён, добавляем его
        admins_data[admin_id] = user_data
        save_admin_data(admin_sessions, admins_data, login_password_hash)

# Обработчик входа в админ-панель
# Функция для обработки входа в админ-панель
@bot.message_handler(commands=['admin'])
@check_chat_state
#@log_user_actions
def handle_admin_login(message):
    global credentials_changed
    user_data = get_user_data(message)
    admin_id = str(user_data["user_id"])

    # Проверяем, есть ли пользователь в сессиях
    if admin_id in admin_sessions:  # Если сессия активна
        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        if not credentials_changed:
            markup.add('Быстрый вход')
        if credentials_changed:
            markup.add('Ввести логин и пароль заново')
        markup.add('В главное меню')
        bot.send_message(
            user_data["user_id"],
            "Выберите способ входа:",
            reply_markup=markup
        )
        bot.register_next_step_handler(message, process_login_choice)
    else:
        # Проверяем, является ли администратор новым
        if is_new_admin(admin_id):
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add('Смена данных входа')
            bot.send_message(message.chat.id, "Вы новый администратор! Пожалуйста, смените данные входа:", reply_markup=markup)
            bot.register_next_step_handler(message, handle_change_credentials)
        else:
            # Предлагаем авторизоваться заново
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item_main = types.KeyboardButton("В главное меню")
            markup.add(item_main)
            bot.send_message(message.chat.id, "Введите логин:", reply_markup=markup)
            bot.register_next_step_handler(message, verify_username)

#@log_user_actions
def is_new_admin(admin_id):
    return admins_data.get(admin_id, {}).get("is_new", False)

#@log_user_actions
def process_login_choice(message):
    global credentials_changed
    user_id = str(message.chat.id)

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    if message.text == "Быстрый вход":
        if user_id in admin_sessions and not credentials_changed:
            # Проверяем, является ли администратор новым
            if is_new_admin(user_id):
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                markup.add('Смена данных входа')
                bot.send_message(message.chat.id, "Вы новый администратор! Пожалуйста, смените данные входа:", reply_markup=markup)
                bot.register_next_step_handler(message, handle_change_credentials)
            else:
                session_data = admins_data.get(user_id, {})
                bot.send_message(message.chat.id, f"Добро пожаловать в админ-панель, {session_data.get('username', 'Админ')}!")
                show_admin_panel(message)
        else:
            bot.send_message(message.chat.id, "Сессия недействительна. Пожалуйста, авторизуйтесь заново.")
            handle_admin_login(message)
    elif message.text == "Ввести логин и пароль заново":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item_main = types.KeyboardButton("В главное меню")
        markup.add(item_main)
        bot.send_message(message.chat.id, "Введите логин:", reply_markup=markup)
        bot.register_next_step_handler(message, verify_username)
    else:
        bot.send_message(message.chat.id, "Неверный выбор. Попробуйте снова.")
        handle_admin_login(message)

#@log_user_actions
def verify_username(message):
    if message.text == "В главное меню":
        return_to_menu(message)
        return

    username = message.text
    is_valid, error_message = is_valid_username(username)
    if not is_valid:
        bot.send_message(message.chat.id, f"Ошибка: {error_message}. Попробуйте снова.")
        bot.register_next_step_handler(message, verify_username)
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_main = types.KeyboardButton("В главное меню")
    markup.add(item_main)
    bot.send_message(message.chat.id, "Введите пароль:", reply_markup=markup)
    bot.register_next_step_handler(message, verify_password, username)

#@log_user_actions
def verify_password(message, username):
    global credentials_changed

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    password = message.text
    admin_id = str(message.chat.id)

    # Проверяем, удалён ли администратор
    if admin_id in removed_admins:
        bot.send_message(message.chat.id, "Ваш доступ был отключён. Обратитесь к корневому администратору.")
        return

    # Проверяем общий логин/пароль
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        # Добавляем админа в активные сессии
        admin_sessions.append(admin_id)
        save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

        # Обновляем данные администратора в базе
        user_data = get_user_data(message)
        update_admin_data(user_data)

        session_data = admins_data.get(admin_id, {})
        bot.send_message(message.chat.id, f"Добро пожаловать в админ-панель, {session_data.get('username', 'Админ')}!")
        show_admin_panel(message)

        # Сбрасываем флаг изменения данных входа
        credentials_changed = False
        return

    # Проверяем индивидуальный хеш
    admin_hash = admins_data.get(admin_id, {}).get("login_password_hash_for_user_id", "")
    if generate_login_password_hash(username, password) == admin_hash:
        # Добавляем админа в активные сессии
        admin_sessions.append(admin_id)
        save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

        # Обновляем данные администратора в базе
        user_data = get_user_data(message)
        update_admin_data(user_data)

        session_data = admins_data.get(admin_id, {})
        bot.send_message(message.chat.id, f"Добро пожаловать в админ-панель, {session_data.get('username', 'Админ')}!")
        show_admin_panel(message)

        # Сбрасываем флаг изменения данных входа
        credentials_changed = False

    # Проверяем индивидуальный хеш
    admin_hash = admins_data.get(admin_id, {}).get("login_password_hash_for_user_id", "")
    if generate_login_password_hash(username, password) == admin_hash:
        # Добавляем админа в активные сессии
        admin_sessions.append(admin_id)
        save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

        # Обновляем данные администратора в базе
        user_data = get_user_data(message)
        update_admin_data(user_data)

        # Сбрасываем флаг изменения данных входа
        credentials_changed = False
    else:
        bot.send_message(message.chat.id, "Неверные логин или пароль. Попробуйте снова.")
        handle_admin_login(message)

# Функция выхода из админ-панели
@bot.message_handler(func=lambda message: message.text == 'Выход' and str(message.chat.id) in admin_sessions)
#@log_user_actions
def admin_logout(message):
    # Админ-сессия сохраняется, просто информируем о выходе
    bot.send_message(message.chat.id, "Вы вышли из админ-панели!\nБыстрый вход сохранен")
    return_to_menu(message)

# (ADMIN 2) ---------------------------------------------- "МЕНЮ ДЛЯ АДМИН-ПАНЕЛИ" ------------------------------------------------

# Показ админ-панели
#@log_user_actions
def show_admin_panel(message):
    admin_id = str(message.chat.id)
    if is_new_admin(admin_id):
        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Смена данных входа')
        bot.send_message(message.chat.id, "Вы новый администратор! Пожалуйста, смените данные входа:", reply_markup=markup)
        bot.register_next_step_handler(message, handle_change_credentials)
    else:
        markup = types.ReplyKeyboardMarkup(row_width=3)
        markup.add('Админ', 'Бан', 'Функции')
        markup.add('Общение', 'Реклама', 'Статистика')
        markup.add('Файлы', 'Резервная копия', 'Редакция')
        markup.add('Экстренная остановка')
        markup.add('Выход')
        bot.send_message(message.chat.id, "Выберите действие из админ-панели:", reply_markup=markup)

# Обработчик для кнопки "В меню админ-панели"
@bot.message_handler(func=lambda message: message.text == 'В меню админ-панели')
#@log_user_actions
def handle_return_to_admin_panel(message):
    show_admin_panel(message)


# (ADMIN 3) --------------------------------- " ФУНКЦИЯ "АДМИН" " --------------------------------------------- 

# Функция для получения идентификатора корневого администратора
#@log_user_actions
def get_root_admin_id():
    if admins_data:
        # Возвращаем первый ключ из словаря admins_data
        return next(iter(admins_data))
    return None

# Функция для проверки, является ли пользователь корневым администратором
#@log_user_actions
def is_root_admin(admin_id):
    return admin_id == get_root_admin_id()

#@log_user_actions
def check_permission(admin_id, permission):
    if is_root_admin(admin_id):
        return True  # Корневой администратор имеет доступ ко всем функциям
    return permission in admins_data.get(str(admin_id), {}).get("permissions", [])   


# Функция для проверки логина на соответствие требованиям
#@log_user_actions
def is_valid_username(username):
    if len(username) < 3:
        return False, "Логин должен содержать не менее 3 символов."
    if not re.search(r'[A-Z]', username):
        return False, "Логин должен содержать хотя бы одну заглавную букву."
    if not re.search(r'[a-z]', username):
        return False, "Логин должен содержать хотя бы одну строчную букву."
    if not re.search(r'[0-9]', username):
        return False, "Логин должен содержать хотя бы одну цифру."
    return True, ""

# Функция для проверки пароля на соответствие требованиям
#@log_user_actions
def is_valid_password(password):
    if len(password) < 8:
        return False, "Пароль должен содержать не менее 8 символов."
    if not re.search(r'[A-Z]', password):
        return False, "Пароль должен содержать хотя бы одну заглавную букву."
    if not re.search(r'[a-z]', password):
        return False, "Пароль должен содержать хотя бы одну строчную букву."
    if not re.search(r'[0-9]', password):
        return False, "Пароль должен содержать хотя бы одну цифру."
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False, "Пароль должен содержать хотя бы один специальный символ."
    return True, ""

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Админ' and check_admin_access(message))
def show_settings_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Админ'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True)
    markup.add('Смена данных входа')
    markup.add('Добавить админа', 'Удалить админа', 'Права доступа')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите настройку:", reply_markup=markup)

# Функция для генерации хеша логина и пароля
#@log_user_actions
def generate_login_password_hash(username, password):
    return hashlib.sha256(f"{username}:{password}".encode()).hexdigest()

# Функция для изменения данных входа для конкретного admin_id
#@log_user_actions
def update_admin_login_credentials(message, admin_id, new_username=None, new_password=None):
    global credentials_changed
    admin_id = str(admin_id)
    if admin_id not in admins_data:
        bot.send_message(admin_id, "Администратор не найден.")
        return

    current_username = admins_data[admin_id].get("admins_username", "")
    current_password_hash = admins_data[admin_id].get("login_password_hash_for_user_id", "")

    # Если задаётся новый логин, обновляем его
    if new_username:
        current_username = new_username

    # Если задаётся новый пароль, обновляем хеш
    if new_password:
        current_password_hash = generate_login_password_hash(current_username, new_password)

    # Обновляем данные
    admins_data[admin_id]["admins_username"] = current_username
    admins_data[admin_id]["login_password_hash_for_user_id"] = current_password_hash
    admins_data[admin_id]["is_new"] = False  # Сбрасываем флаг нового администратора
    save_admin_data(admin_sessions, admins_data, login_password_hash)
    bot.send_message(admin_id, "Данные входа успешно обновлены!")

    # Устанавливаем флаг изменения данных входа
    credentials_changed = True

    # Перенаправляем админа в главное меню и просим авторизоваться заново
    bot.send_message(admin_id, "Данные входа изменены!\n\nПожалуйста, авторизуйтесь заново, используя команду /admin")
    return_to_menu(message)


# Обработчики смены логина и пароля
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Смена данных входа' and check_admin_access(message))
def handle_change_credentials(message):
    global credentials_changed
    admin_id = str(message.chat.id)
    admin_data = admins_data.get(admin_id, {})

    if not check_permission(admin_id, 'Смена данных входа'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    # Проверка, есть ли у админа логин и пароль
    has_credentials = admin_data.get("admins_username") and admin_data.get("login_password_hash_for_user_id")

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    if has_credentials:
        markup.add('Сменить пароль')
    markup.add('Сменить логин и пароль')
    bot.send_message(message.chat.id, "Выберите, что хотите изменить:", reply_markup=markup)


#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Сменить пароль' and check_admin_access(message))
def handle_change_password(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Сменить пароль'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню админ-панели')
    password_requirements = (
        "🔒 Введите новый пароль\n\n"
        "Требования к паролю:\n"
        "- 🔢 Не менее 8 символов\n"
        "- 🔡 Хотя бы одна заглавная буква\n"
        "- 🔠 Хотя бы одна строчная буква\n"
        "- 🔢 Хотя бы одна цифра\n"
        "- 🔣 Хотя бы один специальный символ (например, !@#$%^&*(),.?\":{}|<>)"
    )
    bot.send_message(message.chat.id, password_requirements, reply_markup=markup)
    bot.register_next_step_handler(message, process_new_password)

#@log_user_actions
def is_password_unique(new_password):
    # Проверяем, используется ли новый пароль в данный момент
    for admin_data in admins_data.values():
        current_password_hash = admin_data.get("login_password_hash_for_user_id", "")
        current_username = admin_data.get("admins_username", "")
        if generate_login_password_hash(current_username, new_password) == current_password_hash:
            return False
    return True

#@log_user_actions
def process_new_password(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_new_password)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    new_password = message.text
    is_valid, error_message = is_valid_password(new_password)
    if not is_valid:
        bot.send_message(message.chat.id, f"Ошибка: {error_message}. Попробуйте снова.")
        bot.register_next_step_handler(message, process_new_password)
        return

    if not is_password_unique(new_password):
        bot.send_message(message.chat.id, "Ошибка: Этот пароль уже используется. Попробуйте другой.")
        bot.register_next_step_handler(message, process_new_password)
        return

    update_admin_login_credentials(message, message.chat.id, new_password=new_password)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Сменить логин и пароль' and check_admin_access(message))
def handle_change_login_and_password(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Сменить логин и пароль'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню админ-панели')
    login_requirements = (
        "🔒 Введите новый логин\n\n"
        "Требования к логину:\n"
        "- 🔢 Не менее 3 символов\n"
        "- 🔡 Хотя бы одна заглавная буква\n"
        "- 🔠 Хотя бы одна строчная буква\n"
        "- 🔢 Хотя бы одна цифра"
    )
    bot.send_message(message.chat.id, login_requirements, reply_markup=markup)
    bot.register_next_step_handler(message, process_new_login_and_password_step1)

#@log_user_actions
def process_new_login_and_password_step1(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_new_login_and_password_step1)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return
    new_login = message.text
    is_valid, error_message = is_valid_username(new_login)
    if not is_valid:
        bot.send_message(message.chat.id, f"Ошибка: {error_message}. Попробуйте снова")
        bot.register_next_step_handler(message, process_new_login_and_password_step1)
        return
    if any(admin.get("admins_username") == new_login for admin in admins_data.values()):
        bot.send_message(message.chat.id, "Логин уже существует. Попробуйте другой")
        bot.register_next_step_handler(message, process_new_login_and_password_step1)
        return
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню админ-панели')
    password_requirements = (
        "🔒 Введите новый пароль\n\n"
        "Требования к паролю:\n"
        "- 🔢 Не менее 8 символов\n"
        "- 🔡 Хотя бы одна заглавная буква\n"
        "- 🔠 Хотя бы одна строчная буква\n"
        "- 🔢 Хотя бы одна цифра\n"
        "- 🔣 Хотя бы один специальный символ (например, !@#$%^&*(),.?\":{}|<>)"
    )
    bot.send_message(message.chat.id, password_requirements, reply_markup=markup)
    bot.register_next_step_handler(message, process_new_login_and_password_step2, new_login)

#@log_user_actions
def process_new_login_and_password_step2(message, new_login):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_new_login_and_password_step2, new_login)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    new_password = message.text
    is_valid, error_message = is_valid_password(new_password)
    if not is_valid:
        bot.send_message(message.chat.id, f"Ошибка: {error_message}. Попробуйте снова.")
        bot.register_next_step_handler(message, process_new_login_and_password_step2, new_login)
        return

    if not is_password_unique(new_password):
        bot.send_message(message.chat.id, "Ошибка: Этот пароль уже используется. Попробуйте другой")
        bot.register_next_step_handler(message, process_new_login_and_password_step2, new_login)
        return

    update_admin_login_credentials(message, message.chat.id, new_username=new_login, new_password=new_password)

#@log_user_actions    
def escape_markdown(text):
    # Экранируем специальные символы Markdown
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

# Функция для отображения списка администраторов для удаления
#@log_user_actions
def list_admins_for_removal(message):
    admin_list = []
    for admin_id, data in admins_data.items():
        username = escape_markdown(data['username'])
        admin_list.append(f"№{len(admin_list) + 1}. {username} - `{admin_id}`")

    if admin_list:
        response_message = "📋 Список *всех* администраторов:\n\n\n" + "\n\n".join(admin_list)
        if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
            bot.send_message(message.chat.id, "📜 Список администраторов слишком большой для отправки в одном сообщении!")
        else:
            bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номера, ID или username администраторов для удаления:", reply_markup=markup)

#@log_user_actions
def list_removed_admins(message):
    removed_admin_list = []
    for admin_id, data in removed_admins.items():
        username = escape_markdown(data['username'])
        removed_admin_list.append(f"№{len(removed_admin_list) + 1}. {username} - `{admin_id}`")

    if removed_admin_list:
        response_message = "📋 Список *удалённых* администраторов:\n\n\n" + "\n\n".join(removed_admin_list)
        if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
            bot.send_message(message.chat.id, "📜 Список удалённых администраторов слишком большой для отправки в одном сообщении!")
        else:
            bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номера, ID или username администраторов для добавления:", reply_markup=markup)

#@log_user_actions
def add_admin(admin_id, username, permissions=None, initiator_chat_id=None):
    admin_id = str(admin_id)
    if permissions is None:
        permissions = ["Админ", "Смена данных входа", "Сменить пароль", "Сменить логин и пароль", "Статистика", "Пользователи", "Использование функций", "Список ошибок", "Версия и аптайм"]  # Права по умолчанию
    user_data = {
        "user_id": admin_id,
        "first_name": " ",
        "last_name": " ",
        "username": username,
        "phone": " ",
        "permissions": permissions,
        "is_new": True  # Флаг, указывающий, что администратор новый
    }
    admins_data[admin_id] = user_data
    admin_sessions.append(admin_id)  # Добавляем в текущие сессии
    save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)
    bot.send_message(admin_id, "✅ Вы стали администратором! Быстрый вход по команде /admin доступен...")
    if initiator_chat_id:
        bot.send_message(initiator_chat_id, f"Администратор {escape_markdown(username)} - `{admin_id}` успешно добавлен!", parse_mode="Markdown")

#@log_user_actions
def remove_admin(admin_id, initiator_chat_id):
    admin_id = str(admin_id)
    if admin_id in admins_data:
        admin_username = admins_data[admin_id]["username"]
        # Добавляем в список удалённых администраторов
        removed_admins[admin_id] = {"username": admin_username}

        # Удаляем данные администратора
        del admins_data[admin_id]
        if admin_id in admin_sessions:
            admin_sessions.remove(admin_id)  # Удаляем из текущих сессий

        # Сохраняем изменения
        save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

        # Отправляем сообщение удалённому администратору
        bot.send_message(admin_id, "🚫 Вас удалили из администраторов!")

        # Отправляем сообщение инициатору
        bot.send_message(initiator_chat_id, f"Администратор {escape_markdown(admin_username)} - `{admin_id}` успешно удалён!", parse_mode="Markdown")
    else:
        bot.send_message(initiator_chat_id, f"Администратор с ID `{admin_id}` не найден!", parse_mode="Markdown")

#@log_user_actions
def check_permission(admin_id, permission):
    if is_root_admin(admin_id):
        return True  # Корневой администратор имеет доступ ко всем функциям
    return permission in admins_data.get(str(admin_id), {}).get("permissions", [])

#@log_user_actions
def is_root_admin(admin_id):
    return admin_id == get_root_admin_id()

#@log_user_actions
def get_root_admin_id():
    if admins_data:
        # Возвращаем первый ключ из словаря admins_data
        return next(iter(admins_data))
    return None

# Путь к файлу users.json
users_db_path = os.path.join('data base', 'admin', 'users.json')

# Функция для загрузки данных из users.json
def load_users_data():
    with open(users_db_path, 'r', encoding='utf-8') as file:
        return json.load(file)

# Загрузка данных из users.json
users_data = load_users_data()

@bot.message_handler(func=lambda message: message.text == 'Удалить админа' and check_admin_access(message))
#@log_user_actions
def handle_remove_admin(message):
    # Проверяем, является ли текущий пользователь корневым администратором
    root_admin_id = get_root_admin_id()  # Первый ID в списке `admins_data`
    if str(message.chat.id) != root_admin_id:
        bot.send_message(message.chat.id, "⛔️ Недостаточно прав! Только корневой администратор может удалять администраторов!")
        return

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить админа'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    list_admins_for_removal(message)
    bot.register_next_step_handler(message, process_remove_admin, root_admin_id, message.chat.id)

# Путь к файлу admin_sessions.json
admin_sessions_db_path = os.path.join('data base', 'admin', 'admin_sessions.json')

# Функция для загрузки данных из admin_sessions.json
def load_admin_sessions_data():
    with open(admin_sessions_db_path, 'r', encoding='utf-8') as file:
        return json.load(file)

# Загрузка данных из admin_sessions.json
admin_sessions_data = load_admin_sessions_data()

#@log_user_actions
def process_remove_admin(message, root_admin_id, initiator_chat_id):
    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    input_data = message.text.strip()
    admin_ids = []

    # Разбиваем ввод на части
    parts = input_data.split(',')
    for part in parts:
        part = part.strip()
        if part.isdigit() and len(part) < 5:  # Если ввод - число меньше 5 символов, это номер из списка
            index = int(part) - 1  # Номера в списке начинаются с 1
            if 0 <= index < len(admins_data):
                admin_id = list(admins_data.keys())[index]
                admin_ids.append(admin_id)
            else:
                bot.send_message(message.chat.id, f"Такого администратора с *номером* `{part}` не существует! Попробуйте еще раз", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_remove_admin, root_admin_id, initiator_chat_id)
                return
        elif part.isdigit():  # Если ввод - это ID
            admin_id = part
            if admin_id not in admins_data:
                bot.send_message(message.chat.id, f"Такого администратора с *ID* `{admin_id}` не существует! Попробуйте еще раз", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_remove_admin, root_admin_id, initiator_chat_id)
                return
            admin_ids.append(admin_id)
        else:  # Если ввод - это username
            username = part
            admin_id = next(
                (user_id for user_id, data in admins_data.items() if data.get("username") == username),
                None
            )
            if not admin_id:
                bot.send_message(message.chat.id, f"Такого администратора с *username* {escape_markdown(username)} не существует! Попробуйте еще раз", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_remove_admin, root_admin_id, initiator_chat_id)
                return
            admin_ids.append(admin_id)

    for admin_id in admin_ids:
        if str(message.chat.id) == admin_id:
            bot.send_message(message.chat.id, "Невозможно удалить самого себя!")
            continue
        if admin_id == root_admin_id:
            bot.send_message(message.chat.id, "Невозможно удалить корневого администратора!")
            continue
        remove_admin(admin_id, initiator_chat_id)

    show_settings_menu(message)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Добавить админа' and check_admin_access(message))
def handle_add_admin(message):
    # Проверяем, является ли текущий пользователь корневым администратором
    root_admin_id = get_root_admin_id()  # Первый ID в списке `admins_data`
    if str(message.chat.id) != root_admin_id:
        bot.send_message(message.chat.id, "⛔️ Недостаточно прав! Только корневой администратор может добавлять администраторов!")
        return

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Добавить админа'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    # Выводим список пользователей
    users_data = load_users_data()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data.get('username', 'Неизвестный пользователь'))
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`")

    if user_list:
        response_message = "📋 Список *пользователей*:\n\n\n" + "\n\n".join(user_list)
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    # Выводим список удалённых администраторов
    if removed_admins:
        removed_admin_list = []
        for admin_id, data in removed_admins.items():
            username = escape_markdown(data['username'])
            removed_admin_list.append(f"№{len(removed_admin_list) + 1}. {username} - `{admin_id}`")

        response_message = "📋 Список *удалённых администраторов*:\n\n" + "\n".join(removed_admin_list)
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    # Сообщение для добавления
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню админ-панели')
    bot.send_message(
        message.chat.id,
        "Введите номера, ID или username пользователей или удалённых администраторов для добавления:",
        reply_markup=markup
    )
    bot.register_next_step_handler(message, process_add_admin, root_admin_id, message.chat.id)

#@log_user_actions
def process_add_admin(message, root_admin_id, initiator_chat_id):
    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    input_data = message.text.strip()
    admin_ids = []

    # Разбиваем ввод на части
    parts = input_data.split(',')
    for part in parts:
        part = part.strip()
        if part.isdigit() and len(part) < 5:  # Если ввод - число меньше 5 символов, это номер из списка
            index = int(part) - 1  # Номера в списке начинаются с 1
            if 0 <= index < len(users_data):
                user_id = list(users_data.keys())[index]
                admin_ids.append(user_id)
            else:
                bot.send_message(message.chat.id, f"Такого пользователя с *номером* `{part}` не существует! Попробуйте еще раз", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_add_admin, root_admin_id, initiator_chat_id)
                return
        elif part.isdigit():  # Если ввод - это ID
            user_id = part
            if user_id not in users_data and user_id not in removed_admins:
                bot.send_message(message.chat.id, f"Такого пользователя с *ID* `{user_id}` не существует! Попробуйте еще раз", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_add_admin, root_admin_id, initiator_chat_id)
                return
            admin_ids.append(user_id)
        else:  # Если ввод - это username
            username = part
            user_id = next(
                (user_id for user_id, data in users_data.items() if data.get("username") == username),
                None
            )
            if not user_id:
                bot.send_message(message.chat.id, f"Такого пользователя с *username* {escape_markdown(username)} не существует! Попробуйте еще раз", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_add_admin, root_admin_id, initiator_chat_id)
                return
            admin_ids.append(user_id)

    for admin_id in admin_ids:
        if admin_id in admins_data:
            username = users_data.get(admin_id, {}).get("username", "Неизвестный пользователь")
            bot.send_message(message.chat.id, f"Администратор с {escape_markdown(username)} - `{admin_id}` уже существует!", parse_mode="Markdown")
            continue
        username = users_data[admin_id]["username"]
        add_admin(admin_id, username, permissions=["Админ", "Смена данных входа", "Сменить пароль", "Сменить логин и пароль", "Статистика", "Пользователи", "Использование функций", "Список ошибок", "Версия и аптайм"], initiator_chat_id=initiator_chat_id)
        del removed_admins[admin_id]

    save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)
    show_settings_menu(message)


# Функция для проверки прав доступа
#@log_user_actions
def check_permission(admin_id, permission):
    if is_root_admin(admin_id):
        return True  # Корневой администратор имеет доступ ко всем функциям

    # Получаем текущие права администратора
    current_permissions = admins_data.get(str(admin_id), {}).get("permissions", [])

    # Проверяем, есть ли право в списке, игнорируя часть до двоеточия
    for perm in current_permissions:
        if perm.split(':')[-1].strip() == permission.split(':')[-1].strip():
            return True
    return False

# Функция для экранирования специальных символов Markdown
#@log_user_actions
def escape_markdown(text):
    # Экранируем специальные символы Markdown
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

# Функция для отображения списка администраторов
#@log_user_actions
def list_admins(message):
    admin_list = []
    for admin_id, data in admins_data.items():
        username = data['username']
        # Экранируем имя пользователя
        escaped_username = escape_markdown(username)
        admin_list.append(f"№{len(admin_list) + 1}. {escaped_username} - `{admin_id}`")

    if admin_list:
        response_message = "📋 Список администраторов:\n\n\n" + "\n\n".join(admin_list)
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('Вернуться в админ')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер, ID или username администратора для просмотра его прав:", reply_markup=markup)
    bot.register_next_step_handler(message, process_admin_selection)

# Функция для обработки выбора администратора
#@log_user_actions
def process_admin_selection(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_admin_selection)
        return

    if message.text == "Вернуться в админ":
        show_settings_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    input_data = message.text.strip()

    # Проверяем ввод: номер администратора, ID или имя пользователя
    try:
        if input_data.isdigit() and len(input_data) < 5:  # Если длина ID меньше 5 символов, это номер из списка
            admin_number = int(input_data)
            if admin_number < 1 or admin_number > len(admins_data):
                raise ValueError
            admin_id = list(admins_data.keys())[admin_number - 1]
        elif input_data.isdigit():  # Если ввод длиннее 5 символов, это ID
            admin_id = input_data
            if admin_id not in admins_data:
                bot.send_message(message.chat.id, f"Такого администратора с ID `{admin_id}` не существует. Попробуйте снова.", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_admin_selection)
                return
        else:  # Если ввод - это username
            admin_id = next((key for key, data in admins_data.items() if data.get('username') == input_data), None)
            if not admin_id:
                bot.send_message(message.chat.id, f"Такого администратора с именем пользователя `{input_data}` не существует. Попробуйте снова.", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_admin_selection)
                return

        admin_data = admins_data[admin_id]
        permissions = admin_data.get("permissions", [])

        if is_root_admin(admin_id):
            bot.send_message(message.chat.id, "*Корневой администратор* обладает *всеми правами*!", parse_mode="Markdown")
            show_settings_menu(message)
            return

        # Экранируем имя пользователя
        escaped_username = escape_markdown(admin_data['username'])
        permissions_list = format_permissions_with_headers(permissions)  # Форматирование списка прав с нумерацией
        bot.send_message(message.chat.id, f"Текущие права администратора {escaped_username} - `{admin_id}`:\n\n{permissions_list}", parse_mode="Markdown")

        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Добавить права', 'Удалить права')
        markup.add('Вернуться в админ')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Выберите действие для прав:", reply_markup=markup)
        bot.register_next_step_handler(message, process_permission_action, admin_id)

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный ввод. Пожалуйста, введите номер, ID или имя администратора.")
        bot.register_next_step_handler(message, process_admin_selection)

#@log_user_actions
def get_available_permissions(admin_id):
    all_permissions = [
        "Админ", "Бан", "Функции", "Общение", "Реклама", "Статистика", "Файлы", "Резервная копия", "Редакция", "Экстренная остановка",
        "Админ: Смена данных входа", "Админ: Сменить пароль", "Админ: Сменить логин и пароль", "Админ: Добавить админа", "Админ: Удалить админа", "Админ: Права доступа", "Админ: Добавить права", "Админ: Удалить права",
        "Бан: Заблокировать", "Бан: Разблокировать", "Бан: Удалить данные", "Бан: Удалить пользователя",
		"Функции: Включение", "Функции: Выключение",
        "Общение: Чат", "Общение: Запросы", "Общение: Оповещения", "Общение: Диалоги", 
		"Общение: По времени", "Общение: Отправить по времени", "Общение: Просмотр (по времени)", "Общение: Удалить (по времени)", 
		"Общение: Всем", "Общение: Отправить сообщение", "Общение: Отправленные", "Общение: Удалить отправленные", 
		"Общение: Отдельно", "Общение: Отправить отдельно", "Общение: Посмотреть отдельно", "Общение: Удалить отдельно", 
		"Общение: Просмотр диалогов", "Общение: Удалить диалоги", "Общение: Удалить диалог", "Общение: Удалить все диалоги",
		"Реклама: Запросы на рекламу", "Реклама: Удалить рекламу",
		"Статистика: Пользователи", "Статистика: Использование функций", "Статистика: Список ошибок", "Статистика: Версия и аптайм",
		"Файлы: Просмотр файлов", "Файлы: Поиск файлов по ID", "Файлы: Добавить файлы", "Файлы: Замена файлов", "Файлы: Удалить файлы",
		"Резервная копия: Создать копию", "Резервная копия: Восстановить данные",
		"Редакция: Опубликовать новость", "Редакция: Отредактировать новость", "Редакция: Посмотреть новость", "Редакция: Удалить новость"
		"Экстренная остановка: Подтвердить остановку", "Экстренная остановка: Подтвердить остановку", "Экстренная остановка: Отмена остановки"
    ]

    # Получаем текущие права администратора
    current_permissions = admins_data.get(admin_id, {}).get("permissions", [])

    # Создаем множество для хранения уникальных прав
    unique_permissions = set()

    # Добавляем текущие права в множество, игнорируя часть до двоеточия
    for perm in current_permissions:
        unique_permissions.add(perm.split(':')[-1].strip())

    # Возвращаем доступные права, исключая те, которые уже есть у администратора
    available_permissions = [perm for perm in all_permissions if perm.split(':')[-1].strip() not in unique_permissions]
    return available_permissions

#@log_user_actions
def format_permissions_with_headers(permissions):
    main_functions = ["Админ", "Бан", "Функции", "Общение", "Реклама", "Статистика", "Файлы", "Резервная копия", "Редакция"]
    formatted_permissions = []
    counter = 1

    # Добавляем основные функции в начало списка
    formatted_permissions.append("*Основные права:*")
    for main_func in main_functions:
        if any(perm.split(':')[-1].strip() == main_func for perm in permissions):
            formatted_permissions.append(f"№{counter}. {main_func}")
            counter += 1
    formatted_permissions.append("")  # Добавляем пустую строку для абзаца

    for main_func in main_functions:
        sub_permissions = [perm.split(':')[-1].strip() for perm in permissions if perm.startswith(main_func) and perm.split(':')[-1].strip() != main_func]
        if sub_permissions:
            formatted_permissions.append(f"*Права в \"{main_func}\":*")
            for perm in sub_permissions:
                formatted_permissions.append(f"№{counter}. {perm}")
                counter += 1
            formatted_permissions.append("")  # Добавляем пустую строку для абзаца

    # Добавляем права, которые не начинаются с основных функций
    other_permissions = [perm.split(':')[-1].strip() for perm in permissions if not any(perm.startswith(main_func) for main_func in main_functions)]
    if other_permissions:
        formatted_permissions.append("*Другие права:*")
        for perm in other_permissions:
            formatted_permissions.append(f"№{counter}. {perm}")
            counter += 1

    return "\n".join(formatted_permissions)

#@log_user_actions
def format_permissions_as_list(permissions):
    formatted_permissions = []
    counter = 1

    for perm in permissions:
        formatted_permissions.append(f"№{counter}. {perm}")
        counter += 1

    return "\n".join(formatted_permissions)

# Функция для обработки выбора действия с правами
#@log_user_actions
def process_permission_action(message, admin_id):
    if message.text == "Вернуться в админ":
        show_settings_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.text == 'Добавить права':
        if not check_permission(str(message.chat.id), 'Добавить права'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        available_permissions = get_available_permissions(admin_id)
        permissions_list = format_permissions_with_headers(available_permissions)
        bot.send_message(message.chat.id, f"*Доступные* права для *добавления*:\n\n{permissions_list}", parse_mode="Markdown")

        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Вернуться в админ')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите номера прав через запятую для *добавления*:", parse_mode="Markdown", reply_markup=markup)
        bot.register_next_step_handler(message, process_add_permissions, admin_id, available_permissions)

    elif message.text == 'Удалить права':
        if not check_permission(str(message.chat.id), 'Удалить права'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        current_permissions = admins_data[admin_id].get("permissions", [])
        permissions_list = format_permissions_as_list(current_permissions)
        bot.send_message(message.chat.id, f"*Текущие права* администратора:\n\n{permissions_list}", parse_mode="Markdown")

        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Вернуться в админ')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите номера прав через запятую для *удаления*:", parse_mode="Markdown", reply_markup=markup)
        bot.register_next_step_handler(message, process_remove_permissions, admin_id, current_permissions)


# Функция для форматирования списка прав с основными функциями
#@log_user_actions
def format_permissions_with_main_functions(permissions):
    main_functions = ["Админ", "Бан", "Функции", "Общение", "Реклама", "Статистика", "Файлы", "Резервная копия", "Редакция"]
    formatted_permissions = []
    counter = 1

    # Добавляем основные функции в начало списка
    formatted_permissions.append("*Основные функции:*")
    for main_func in main_functions:
        formatted_permissions.append(f"№{counter}. {main_func}")
        counter += 1
    formatted_permissions.append("")  # Добавляем пустую строку для абзаца

    for main_func in main_functions:
        sub_permissions = [perm for perm in permissions if perm.startswith(main_func) and perm != main_func]
        if sub_permissions:
            formatted_permissions.append(f"*{main_func}:*")
            for perm in sub_permissions:
                formatted_permissions.append(f"№{counter}. {perm.split(': ')[1]}")
                counter += 1
            formatted_permissions.append("")  # Добавляем пустую строку для абзаца

    return "\n".join(formatted_permissions)

# Функция для обработки добавления прав
#@log_user_actions
def process_add_permissions(message, admin_id, available_permissions):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_add_permissions, admin_id, available_permissions)
        return

    if message.text == "Вернуться в админ":
        show_settings_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        permission_numbers = [int(num.strip()) - 1 for num in message.text.split(',')]
        permissions_to_add = []

        for num in permission_numbers:
            if 0 <= num < len(available_permissions):
                permission = available_permissions[num].split(':')[-1].strip()
                if permission in admins_data[admin_id].get("permissions", []):
                    bot.send_message(message.chat.id, f"Право *{escape_markdown(permission)}* уже добавлено! Попробуйте снова", parse_mode="Markdown")
                    bot.register_next_step_handler(message, process_add_permissions, admin_id, available_permissions)
                    return
                permissions_to_add.append(permission)
            else:
                bot.send_message(message.chat.id, f"Неверный номер права: *{num + 1}*! Попробуйте снова", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_add_permissions, admin_id, available_permissions)
                return

        if permissions_to_add:
            admins_data[admin_id].setdefault("permissions", []).extend(permissions_to_add)
            save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

            admin_data = admins_data[admin_id]
            # Экранируем имя пользователя
            escaped_username = escape_markdown(admin_data['username'])
            # Преобразуем новые права в нижний регистр и экранируем их
            escaped_permissions_to_add = [escape_markdown(permission.lower()) for permission in permissions_to_add]

            bot.send_message(message.chat.id, f"Права для админа {escaped_username} - `{admin_id}` обновлены!", parse_mode="Markdown")
            bot.send_message(admin_id, f"⚠️ Ваши права были *изменены*!\n\n➕ *Добавлены* новые права: _{', '.join(escaped_permissions_to_add)}_", parse_mode="Markdown")
            show_settings_menu(message)
        else:
            bot.send_message(message.chat.id, "Неверные номера прав! Попробуйте снова")
            bot.register_next_step_handler(message, process_add_permissions, admin_id, available_permissions)

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверные номера прав! Попробуйте снова")
        bot.register_next_step_handler(message, process_add_permissions, admin_id, available_permissions)

#@log_user_actions
def process_remove_permissions(message, admin_id, current_permissions):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_remove_permissions, admin_id, current_permissions)
        return

    if message.text == "Вернуться в админ":
        show_settings_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        # Обработка введенных номеров
        permission_numbers = [int(num.strip()) - 1 for num in message.text.split(',')]

        # Проверка номеров
        if any(num < 0 or num >= len(current_permissions) for num in permission_numbers):
            bot.send_message(message.chat.id, "Ошибка: Один или несколько номеров прав некорректны! Попробуйте снова", parse_mode="Markdown")
            bot.register_next_step_handler(message, process_remove_permissions, admin_id, current_permissions)
            return

        # Определяем права для удаления
        permissions_to_remove = [current_permissions[num] for num in permission_numbers]

        # Удаляем права из текущих
        updated_permissions = [perm for perm in current_permissions if perm not in permissions_to_remove]

        # Сохраняем изменения
        admins_data[admin_id]["permissions"] = updated_permissions
        save_admin_data(admin_sessions, admins_data, login_password_hash, removed_admins)

        # Уведомляем пользователя
        escaped_username = escape_markdown(admins_data[admin_id]['username'])
        escaped_removed = ', '.join(escape_markdown(perm.lower()) for perm in permissions_to_remove)

        bot.send_message(message.chat.id, f"✅ Права для администратора {escaped_username} - `{admin_id}` обновлены!", parse_mode="Markdown")
        bot.send_message(admin_id, f"⚠️ Ваши права были *изменены*!\n\n❌ *Удалены* права: _{escaped_removed}_", parse_mode="Markdown")

        show_settings_menu(message)

    except ValueError:
        bot.send_message(message.chat.id, "Ошибка: Введите номера прав через запятую", parse_mode="Markdown")
        bot.register_next_step_handler(message, process_remove_permissions, admin_id, current_permissions)
    except Exception as e:
        bot.send_message(message.chat.id, f"Произошла ошибка: {str(e)}")


# Функция для форматирования списка прав
#@log_user_actions
def format_permissions(permissions):
    formatted_permissions = []
    counter = 1

    for perm in permissions:
        formatted_permissions.append(f"№{counter}. {perm}")
        counter += 1

    return "\n".join(formatted_permissions)

# Обработчик для команды "Права доступа"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Права доступа' and str(message.chat.id) in admin_sessions)
def handle_permissions(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Права доступа'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    list_admins(message)

# (ADMIN 3) ------------------------------------------ "БАН ПОЛЬЗОВАТЕЛЙ ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

# Максимальная длина сообщения в Telegram
TELEGRAM_MESSAGE_LIMIT = 4096

# Определение корневой директории на основе исполняемого файла
EXECUTABLE_FILE = '(93 update ИСПРАВЛЕНИЕ25  ( (  )) CAR MANAGER TG BOT (official) v0924.py'
BASE_DIR = os.path.dirname(os.path.abspath(EXECUTABLE_FILE))

# Пути к директориям и файлам
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')
FILES_PATH = os.path.join(BASE_DIR, 'data base')
ADDITIONAL_FILES_PATH = os.path.join(BASE_DIR, 'files')
ADMIN_SESSIONS_FILE = os.path.join(BASE_DIR, 'data base', 'admin', 'admin_sessions.json')
USER_DATA_PATH = os.path.join(BASE_DIR, 'data base', 'admin', 'users.json')

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data['admin_sessions']

# Функция для проверки прав доступа
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "⚠️ У вас нет *прав доступа* для выполнения этой операции!", parse_mode="Markdown")
        return False

# Декоратор для ограничения доступа
def restricted(func):
    def wrapper(message, *args, **kwargs):
        user_id = message.from_user.id
        if is_user_blocked(user_id):
            bot.send_message(message.chat.id, "🚫 Вы *заблокированы* и не можете выполнять это действие!", parse_mode="Markdown")
            return
        if not check_admin_access(message):
            return
        return func(message, *args, **kwargs)
    return wrapper

def save_user_data(user_data):
    with open(USER_DATA_PATH, 'w', encoding='utf-8') as file:
        json.dump(user_data, file, ensure_ascii=False, indent=4)

def load_user_data():
    if os.path.exists(USER_DATA_PATH):
        with open(USER_DATA_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    return {}

#@log_user_actions
def is_user_blocked(user_id):
    users_data = load_user_data()
    return users_data.get(str(user_id), {}).get('blocked', False)

#@log_user_actions
def block_user(user_id):
    users_data = load_user_data()
    if str(user_id) in users_data:
        users_data[str(user_id)]['blocked'] = True
        save_user_data(users_data)
        return users_data[str(user_id)]['username'], user_id
    return None, None

#@log_user_actions
def unblock_user(user_id):
    users_data = load_user_data()
    if str(user_id) in users_data:
        users_data[str(user_id)]['blocked'] = False
        save_user_data(users_data)
        return users_data[str(user_id)]['username'], user_id
    return None, None

#@log_user_actions
def get_user_id_by_username(username):
    users_data = load_user_data()
    for user_id, data in users_data.items():
        if data['username'] == f"@{username}":
            return user_id
    return None

#@log_user_actions
def escape_markdown(text):
    # Экранируем специальные символы Markdown
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

#@log_user_actions
def get_root_admin():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    root_admin_id = data['admin_sessions'][0]
    return root_admin_id

#@log_user_actions
def list_users_for_ban(message):
    users_data = load_user_data()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *заблокирован* 🚫" if data.get('blocked', False) else " - *разблокирован* ✅"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Заблокировать", "Разблокировать")
    markup.add("Удалить данные", "Удалить пользователя")
    markup.add("В меню админ-панели")
    bot.send_message(message.chat.id, "Выберите действие для пользователя:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_ban_action)

#@log_user_actions
def choose_ban_action(message):

    if message.text == "Заблокировать":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Заблокировать'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        choose_block_method(message)

    elif message.text == "Разблокировать":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Разблокировать'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        choose_unblock_method(message)

    elif message.text == "Удалить данные":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Удалить данные'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        delete_user_data(message)

    elif message.text == "Удалить пользователя":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Удалить пользователя'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        delete_user(message)

    elif message.text == "Вернуться в бан":
        ban_user_prompt(message)       

    elif message.text == "В меню админ-панели":
        show_admin_panel(message)


#@log_user_actions
def choose_block_method(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Заблокировать'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("По ID", "По USERNAME")
    markup.add("Вернуться в бан")
    markup.add("В меню админ-панели")
    bot.send_message(message.chat.id, "Выберите способ блокировки:", reply_markup=markup)
    bot.register_next_step_handler(message, process_block_method)

#@log_user_actions
def process_block_method(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Заблокировать'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "По ID":
        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в бан")
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите *id* пользователей для блокировки:", reply_markup=markup, parse_mode="Markdown")
        bot.register_next_step_handler(message, block_user_by_id)
    elif message.text == "По USERNAME":
        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в бан")
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите *username* пользователей для блокировки:", reply_markup=markup, parse_mode="Markdown")
        bot.register_next_step_handler(message, block_user_by_username)
    elif message.text == "Вернуться в бан":
        ban_user_prompt(message)       
    elif message.text == "В меню админ-панели":
        show_admin_panel(message)

#@log_user_actions
def choose_unblock_method(message):
    
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Разблокировать'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("По ID", "По USERNAME")
    markup.add("Вернуться в бан")
    markup.add("В меню админ-панели")
    bot.send_message(message.chat.id, "Выберите способ разблокировки:", reply_markup=markup)
    bot.register_next_step_handler(message, process_unblock_method)

#@log_user_actions
def process_unblock_method(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Разблокировать'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "По ID":
        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в бан")
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите *id* пользователей для разблокировки:", reply_markup=markup, parse_mode="Markdown")
        bot.register_next_step_handler(message, unblock_user_by_id)
    elif message.text == "По USERNAME":
        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в бан")
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите *username* пользователей для разблокировки:", reply_markup=markup, parse_mode="Markdown")
        bot.register_next_step_handler(message, unblock_user_by_username)
    elif message.text == "Вернуться в бан":
        ban_user_prompt(message)       
    elif message.text == "В меню админ-панели":
        show_admin_panel(message)

#@log_user_actions
def block_user_by_username(message):
    # Проверка на мультимедиа и смайлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, block_user_by_username)
        return

    if message.text == "Вернуться в бан":
        ban_user_prompt(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    usernames = [username.strip().lstrip('@') for username in message.text.split(',')]
    success_users = []
    failed_users = []
    already_blocked_users = []

    root_admin_id = get_root_admin()
    admin_id = str(message.chat.id)

    for username in usernames:
        user_id = get_user_id_by_username(username)
        if user_id:
            if str(user_id) == root_admin_id:
                failed_users.append(username)
                continue
            if str(user_id) == admin_id:
                failed_users.append(username)
                continue
            if is_user_blocked(user_id):
                already_blocked_users.append(username)
            else:
                username, user_id = block_user(user_id)
                if username and user_id:
                    success_users.append((username, user_id))
                else:
                    failed_users.append(username)
        else:
            failed_users.append(username)

    if success_users:
        response_message = "🚫 Пользователи заблокированы!\n\n"
        for username, user_id in success_users:
            escaped_username = escape_markdown(username)
            response_message += f"*USERNAME* - {escaped_username}\n*ID* - `{user_id}`\n\n"
            bot.send_message(user_id, "🚫 Ваш аккаунт был *заблокирован* администратором!", parse_mode="Markdown")
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    if failed_users:
        failed_users_message = f"Ошибка при блокировке пользователей: {', '.join([f'@{escape_markdown(username)}' for username in failed_users])}! Попробуйте снова"
        bot.send_message(message.chat.id, failed_users_message, parse_mode="Markdown")

    if already_blocked_users:
        already_blocked_users_message = f"Пользователи уже заблокированы: {', '.join([f'@{escape_markdown(username)}' for username in already_blocked_users])}! Попробуйте снова"
        bot.send_message(message.chat.id, already_blocked_users_message, parse_mode="Markdown")

    if failed_users or already_blocked_users:
        bot.register_next_step_handler(message, block_user_by_username)
        return

    ban_user_prompt(message)

#@log_user_actions
def block_user_by_id(message):
    # Проверка на мультимедиа и смайлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, block_user_by_id)
        return

    if message.text == "Вернуться в бан":
        ban_user_prompt(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    user_ids = [user_id.strip() for user_id in message.text.split(',')]
    success_users = []
    failed_users = []
    already_blocked_users = []

    root_admin_id = get_root_admin()
    admin_id = str(message.chat.id)

    for user_id in user_ids:
        if user_id.isdigit():
            user_id = int(user_id)
            if str(user_id) == root_admin_id:
                failed_users.append(user_id)
                continue
            if str(user_id) == admin_id:
                failed_users.append(user_id)
                continue
            if is_user_blocked(user_id):
                already_blocked_users.append(user_id)
            else:
                username, user_id = block_user(user_id)
                if username and user_id:
                    success_users.append((username, user_id))
                else:
                    failed_users.append(user_id)
        else:
            failed_users.append(user_id)

    if success_users:
        response_message = "🚫 Пользователи заблокированы!\n\n"
        for username, user_id in success_users:
            escaped_username = escape_markdown(username)
            response_message += f"*ID* - `{user_id}`\n*USERNAME* - {escaped_username}\n\n"
            bot.send_message(user_id, "🚫 Ваш аккаунт был *заблокирован* администратором!", parse_mode="Markdown")
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    if failed_users:
        failed_users_message = f"Ошибка при блокировке пользователей: {', '.join([f'`{user_id}`' for user_id in failed_users])}! Попробуйте снова"
        bot.send_message(message.chat.id, failed_users_message, parse_mode="Markdown")

    if already_blocked_users:
        already_blocked_users_message = f"Пользователи уже заблокированы: {', '.join([f'`{user_id}`' for user_id in already_blocked_users])}! Попробуйте снова"
        bot.send_message(message.chat.id, already_blocked_users_message, parse_mode="Markdown")

    if failed_users or already_blocked_users:
        bot.register_next_step_handler(message, block_user_by_id)
        return

    ban_user_prompt(message)

#@log_user_actions
def unblock_user_by_username(message):
    # Проверка на мультимедиа и смайлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, unblock_user_by_username)
        return

    if message.text == "Вернуться в бан":
        ban_user_prompt(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    usernames = [username.strip().lstrip('@') for username in message.text.split(',')]
    success_users = []
    failed_users = []
    already_unblocked_users = []

    root_admin_id = get_root_admin()
    admin_id = str(message.chat.id)

    for username in usernames:
        user_id = get_user_id_by_username(username)
        if user_id:
            if str(user_id) == root_admin_id:
                failed_users.append(username)
                continue
            if str(user_id) == admin_id:
                failed_users.append(username)
                continue
            if not is_user_blocked(user_id):
                already_unblocked_users.append(username)
            else:
                username, user_id = unblock_user(user_id)
                if username and user_id:
                    success_users.append((username, user_id))
                else:
                    failed_users.append(username)
        else:
            failed_users.append(username)

    if success_users:
        response_message = "✅ Пользователи разблокированы!\n\n"
        for username, user_id in success_users:
            escaped_username = escape_markdown(username)
            response_message += f"*USERNAME* - {escaped_username}\n*ID* - `{user_id}`\n\n"
            bot.send_message(user_id, "✅ Ваш аккаунт был *разблокирован* администратором!", parse_mode="Markdown")
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    if failed_users:
        failed_users_message = f"Ошибка при разблокировке пользователей: {', '.join([f'@{escape_markdown(username)}' for username in failed_users])}. Попробуйте снова"
        bot.send_message(message.chat.id, failed_users_message, parse_mode="Markdown")

    if already_unblocked_users:
        already_unblocked_users_message = f"Пользователи уже разблокированы: {', '.join([f'@{escape_markdown(username)}' for username in already_unblocked_users])}. Попробуйте снова"
        bot.send_message(message.chat.id, already_unblocked_users_message, parse_mode="Markdown")

    if failed_users or already_unblocked_users:
        bot.register_next_step_handler(message, unblock_user_by_username)
        return

    ban_user_prompt(message)

#@log_user_actions
def unblock_user_by_id(message):
    # Проверка на мультимедиа и смайлы
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, unblock_user_by_id)
        return

    if message.text == "Вернуться в бан":
        ban_user_prompt(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    user_ids = [user_id.strip() for user_id in message.text.split(',')]
    success_users = []
    failed_users = []
    already_unblocked_users = []

    root_admin_id = get_root_admin()
    admin_id = str(message.chat.id)

    for user_id in user_ids:
        if user_id.isdigit():
            user_id = int(user_id)
            if str(user_id) == root_admin_id:
                failed_users.append(user_id)
                continue
            if str(user_id) == admin_id:
                failed_users.append(user_id)
                continue
            if not is_user_blocked(user_id):
                already_unblocked_users.append(user_id)
            else:
                username, user_id = unblock_user(user_id)
                if username and user_id:
                    success_users.append((username, user_id))
                else:
                    failed_users.append(user_id)
        else:
            failed_users.append(user_id)

    if success_users:
        response_message = "✅ Пользователи разблокированы!\n\n"
        for username, user_id in success_users:
            escaped_username = escape_markdown(username)
            response_message += f"*ID* - `{user_id}`\n*USERNAME* - {escaped_username}\n\n"
            bot.send_message(user_id, "✅ Ваш аккаунт был *разблокирован* администратором!", parse_mode="Markdown")
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    if failed_users:
        failed_users_message = f"Ошибка при разблокировке пользователей: {', '.join([f'`{user_id}`' for user_id in failed_users])}! Попробуйте снова"
        bot.send_message(message.chat.id, failed_users_message, parse_mode="Markdown")

    if already_unblocked_users:
        already_unblocked_users_message = f"Пользователи уже разблокированы: {', '.join([f'`{user_id}`' for user_id in already_unblocked_users])}! Попробуйте снова"
        bot.send_message(message.chat.id, already_unblocked_users_message, parse_mode="Markdown")

    if failed_users or already_unblocked_users:
        bot.register_next_step_handler(message, unblock_user_by_id)
        return

    ban_user_prompt(message)


# Обработка команд админа
@bot.message_handler(func=lambda message: message.text == 'Бан' and check_admin_access(message))
#@log_user_actions
def ban_user_prompt(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Бан'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    list_users_for_ban(message)

#@log_user_actions
def delete_user_data_by_id(message, user_id):
    if not os.path.exists(USER_DATA_PATH):
        bot.send_message(message.chat.id, "База данных пользователей не существует!")
        return

    users_data = load_user_data()
    if str(user_id) not in users_data:
        bot.send_message(message.chat.id, f"Пользователь с *ID* `{user_id}` не найден!", parse_mode="Markdown")
        return

    root_admin_id = get_root_admin()
    admin_id = str(message.chat.id)

    if str(user_id) == root_admin_id:
        bot.send_message(message.chat.id, f"Ошибка при удалении данных пользователя с *ID* `{user_id}`: Нельзя удалить данные корневого администратора!", parse_mode="Markdown")
        return

    if str(user_id) == admin_id:
        bot.send_message(message.chat.id, f"Ошибка при удалении данных пользователя с *ID* `{user_id}`: Нельзя удалить данные самого себя!", parse_mode="Markdown")
        return

    for root, dirs, files in os.walk(BASE_DIR):
        for file in files:
            if file.endswith('.json'):
                file_path = os.path.join(root, file)

                # Проверяем, содержит ли название файла ID пользователя
                if str(user_id) in file:
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                    except json.JSONDecodeError:
                        continue
                    except UnicodeDecodeError:
                        try:
                            with open(file_path, 'r', encoding='windows-1251') as f:
                                data = json.load(f)
                        except json.JSONDecodeError:
                            continue
                        except UnicodeDecodeError:
                            continue

                    if isinstance(data, list):
                        updated_data = [item for item in data if item != str(user_id)]
                        if len(updated_data) < len(data):
                            data = updated_data
                    elif isinstance(data, dict):
                        if 'admin_sessions' in data and str(user_id) in data['admin_sessions']:
                            data['admin_sessions'].remove(str(user_id))
                        if 'admins_data' in data and str(user_id) in data['admins_data']:
                            data['admins_data'].pop(str(user_id))
                    else:
                        continue

                else:
                    # Проверяем содержимое файла
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                    except json.JSONDecodeError:
                        continue
                    except UnicodeDecodeError:
                        try:
                            with open(file_path, 'r', encoding='windows-1251') as f:
                                data = json.load(f)
                        except json.JSONDecodeError:
                            continue
                        except UnicodeDecodeError:
                            continue

                    if isinstance(data, dict):
                        if str(user_id) in data:
                            data.pop(str(user_id), None)
                        else:
                            continue
                    elif isinstance(data, list):
                        updated_data = []
                        for item in data:
                            if isinstance(item, dict) and item.get('user_id') == user_id:
                                continue
                            else:
                                updated_data.append(item)
                        data = updated_data
                    else:
                        continue

                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=4)

    users_data.pop(str(user_id), None)
    save_user_data(users_data)

#@log_user_actions
def delete_user_from_users_db(message, user_id=None, username=None):
    if not os.path.exists(USER_DATA_PATH):
        bot.send_message(message.chat.id, "База данных пользователей не существует!")
        return

    try:
        with open(USER_DATA_PATH, 'r', encoding='utf-8') as file:
            users_data = json.load(file)
    except json.JSONDecodeError:
        bot.send_message(message.chat.id, f"Файл {USER_DATA_PATH} содержит некорректные данные!")
        return
    except UnicodeDecodeError:
        try:
            with open(USER_DATA_PATH, 'r', encoding='windows-1251') as file:
                users_data = json.load(file)
        except json.JSONDecodeError:
            bot.send_message(message.chat.id, f"Файл {USER_DATA_PATH} содержит некорректные данные!")
            return
        except UnicodeDecodeError:
            bot.send_message(message.chat.id, f"Файл {USER_DATA_PATH} содержит некорректную кодировку!")
            return

    if user_id:
        if str(user_id) not in users_data:
            bot.send_message(message.chat.id, f"Пользователь с ID `{user_id}` не найден!", parse_mode="Markdown")
            return
        users_data.pop(str(user_id), None)
    elif username:
        username = username.lstrip('@')  # Убираем @ перед сравнением
        found = False
        for user_id, data in users_data.items():
            if data.get('username') == f"@{username}":  # Сравниваем с @ в базе данных
                users_data.pop(user_id, None)
                found = True
                break
        if not found:
            bot.send_message(message.chat.id, f"Пользователь с *username* @{username} не найден!", parse_mode="Markdown")
            return

    with open(USER_DATA_PATH, 'w', encoding='utf-8') as file:
        json.dump(users_data, file, ensure_ascii=False, indent=4)

#@log_user_actions
def delete_user_data(message):
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в бан")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите *username* или *id*  пользователя для удаления данных:", reply_markup=markup, parse_mode="Markdown")
    bot.register_next_step_handler(message, process_delete_user_data)

#@log_user_actions
def process_delete_user_data(message):
    if message.text == "Вернуться в бан":
        ban_user_prompt(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    user_inputs = message.text.strip().split(',')

    # Load user data if database exists
    if not os.path.exists(USER_DATA_PATH):
        bot.send_message(message.chat.id, "База данных пользователей не существует!")
        return

    users_data = load_user_data()

    for user_input in user_inputs:
        user_input = user_input.strip()

        # Determine if the input is an ID or username
        user_id = None
        if user_input.isdigit():
            user_id = int(user_input)
        else:
            for user_data in users_data.values():
                if 'username' in user_data and user_data['username'] == user_input:
                    user_id = user_data['user_id']
                    break

        if user_id is None:
            bot.send_message(message.chat.id, f"Пользователь с таким *username* или *id* не найден: {user_input}!", parse_mode="Markdown")
            continue

        # Check access rights
        admin_id = str(message.chat.id)
        root_admin_id = get_root_admin()

        if str(user_id) == root_admin_id:
            bot.send_message(message.chat.id, f"Ошибка: нельзя удалить данные корневого администратора {escape_markdown(user_input)} - `{user_id}`!", parse_mode="Markdown")
            continue

        if str(user_id) == admin_id:
            bot.send_message(message.chat.id, f"Ошибка: нельзя удалить данные самого себя {escape_markdown(user_input)} - `{user_id}`!", parse_mode="Markdown")
            continue

        # Delete user data
        for root, dirs, files in os.walk(BASE_DIR):
            for file in files:
                if file.endswith('.json'):
                    file_path = os.path.join(root, file)

                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                    except (json.JSONDecodeError, UnicodeDecodeError):
                        try:
                            with open(file_path, 'r', encoding='windows-1251') as f:
                                data = json.load(f)
                        except Exception as e:
                            continue

                    if isinstance(data, dict) and str(user_id) in data:
                        data.pop(str(user_id), None)
                    elif isinstance(data, list):
                        data = [item for item in data if item != str(user_id)]

                    with open(file_path, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=4)

        # Update user data in users.json
        username = users_data[str(user_id)]['username']
        users_data.pop(str(user_id), None)
        save_user_data(users_data)

        bot.send_message(message.chat.id, f"Данные пользователя {escape_markdown(username)} - `{user_id}` успешно удалены!", parse_mode="Markdown")

    ban_user_prompt(message)

#@log_user_actions
def delete_user(message):
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в бан")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите *id* или *username* пользователя для удаления:", reply_markup=markup, parse_mode="Markdown")
    bot.register_next_step_handler(message, process_delete_user)

#@log_user_actions
def process_delete_user(message):
    if message.text == "Вернуться в бан":
        ban_user_prompt(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    user_inputs = message.text.split(",")  # Разделяем введенные значения по запятой
    user_inputs = [input.strip() for input in user_inputs]  # Убираем пробелы

    root_admin_id = get_root_admin()
    admin_id = str(message.chat.id)

    for user_input in user_inputs:
        if user_input.isdigit():  # Если введен ID
            user_id = int(user_input)
            if str(user_id) == root_admin_id:
                bot.send_message(
                    message.chat.id,
                    f"Ошибка: пользователь с *ID* `{user_id}` является корневым администратором и не может быть удален!",
                    parse_mode="Markdown"
                )
                continue
            if str(user_id) == admin_id:
                bot.send_message(
                    message.chat.id,
                    f"Ошибка: вы не можете удалить самого себя по ID: `{user_id}`!",
                    parse_mode="Markdown"
                )
                continue

            # Проверяем, существует ли пользователь с таким ID
            users_data = load_user_data()
            if str(user_id) not in users_data:
                bot.send_message(
                    message.chat.id,
                    f"Пользователь с *ID* `{user_id}` не найден! Попробуйте снова ввести верный *id* или *username*",
                    parse_mode="Markdown"
                )
                # Ожидаем новый ввод
                bot.register_next_step_handler(message, process_delete_user)
                return

            delete_user_from_users_db(message, user_id=user_id)
            bot.send_message(
                message.chat.id,
                f"Пользователь с *ID* `{user_id}` успешно удален!",
                parse_mode="Markdown"
            )
        else:  # Если введено имя пользователя
            username = user_input.lstrip('@')
            escaped_username = escape_markdown(username)  # Экранируем username
            user_id = get_user_id_by_username(username)
            if str(user_id) == root_admin_id:
                bot.send_message(
                    message.chat.id,
                    f"Ошибка: пользователь с *username* @{escaped_username} является корневым администратором и не может быть удален!",
                    parse_mode="Markdown"
                )
                continue
            if str(user_id) == admin_id:
                bot.send_message(
                    message.chat.id,
                    f"Ошибка: вы не можете удалить самого себя по @{escaped_username}!",
                    parse_mode="Markdown"
                )
                continue
            if user_id is None:
                bot.send_message(
                    message.chat.id,
                    f"Пользователь с *username* @{escaped_username} не найден! Попробуйте снова ввести верный *id* или *username*",
                    parse_mode="Markdown"
                )
                # Ожидаем новый ввод
                bot.register_next_step_handler(message, process_delete_user)
                return

            delete_user_from_users_db(message, username=username)
            bot.send_message(
                message.chat.id,
                f"Пользователь с *username* @{escaped_username} успешно удален!",
                parse_mode="Markdown"
            )

    # После успешного удаления пользователя, возвращаем в ban_user_prompt(message)
    ban_user_prompt(message)

# (ADMIN 4) ------------------------------------------ "СТАТИСТИКА ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

# Глобальные переменные для отслеживания пользователей
active_users = {}  # Словарь для отслеживания времени последней активности пользователей
total_users = set()  # Множество для хранения уникальных пользователей

# Пути к JSON-файлам
ADMIN_SESSIONS_FILE = 'data base/admin/admin_sessions.json'
USER_DATA_FILE = 'data base/admin/users.json'
STATS_FILE = 'data base/admin/stats.json'
ERRORS_LOG_FILE = 'data base/log/errors_log.json'

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data.get('admin_sessions', [])

# Загрузка данных пользователей
def load_user_data():
    try:
        with open(USER_DATA_FILE, 'r', encoding='utf-8') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}

# Сохранение данных пользователей
def save_user_data(data):
    with open(USER_DATA_FILE, 'w', encoding='utf-8') as file:
        json.dump(data, file, indent=4, ensure_ascii=False)

# Загрузка статистики из файла
def load_statistics():
    try:
        with open(STATS_FILE, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return {date: {'users': set(data[date]['users']), 'functions': data[date].get('functions', {})} for date in data}
    except FileNotFoundError:
        return {}

# Сохранение статистики в файл
def save_statistics(data):
    with open(STATS_FILE, 'w', encoding='utf-8') as file:
        json.dump({date: {'users': list(data[date]['users']), 'functions': data[date]['functions']} for date in data}, file, indent=4, ensure_ascii=False)

# Проверка прав администратора
#@log_user_actions
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "У вас нет прав доступа для выполнения этой операции.")
        return False

# Обновление данных о пользователе с экранированием username
#@log_user_actions
def update_user_activity(user_id, username=None, first_name="", last_name="", phone="", function_name=None):
    active_users[user_id] = datetime.now()
    total_users.add(user_id)

    user_data = load_user_data()
    current_time = datetime.now().strftime('%d.%m.%Y в %H:%M:%S')  # Формат времени

    if user_id in user_data:
        user_data[user_id]['username'] = f"@{username}" if username else ""
        user_data[user_id]['first_name'] = first_name
        user_data[user_id]['last_name'] = last_name
        user_data[user_id]['phone'] = phone
        user_data[user_id]['last_active'] = current_time
    else:
        user_data[user_id] = {
            'user_id': user_id,
            'first_name': first_name,
            'last_name': last_name,
            'username': f"@{username}" if username else "",
            'phone': phone,
            'last_active': current_time,
            'blocked': False,
            'actions': 0,  # Добавлено для отслеживания действий
            'session_time': 0,  # Добавлено для отслеживания времени сессии
            'returning': False  # Добавлено для отслеживания возвращающихся пользователей
        }
    save_user_data(user_data)

    # Обновление статистики
    statistics = load_statistics()
    today = datetime.now().strftime('%d.%m.%Y')
    if today not in statistics:
        statistics[today] = {'users': set(), 'functions': {}}
    statistics[today]['users'].add(user_id)
    if function_name:
        if function_name not in statistics[today]['functions']:
            statistics[today]['functions'][function_name] = 0
        statistics[today]['functions'][function_name] += 1
    save_statistics(statistics)

    
# Проверить, активен ли пользователь
#@log_user_actions
def is_user_active(last_active):
    try:
        last_active_time = datetime.strptime(last_active, '%d.%m.%Y в %H:%M:%S')
    except ValueError:
        return False
    return (datetime.now() - last_active_time).total_seconds() < 1 * 60  # 5 минут

# Получение статистики за определённый период
#@log_user_actions
def get_aggregated_statistics(period='all'):
    statistics = load_statistics()
    today = datetime.now()
    user_result = defaultdict(int)
    function_result = defaultdict(int)

    for date_str, usage in statistics.items():
        record_date = datetime.strptime(date_str, '%d.%m.%Y')

        if period == 'day' and record_date.date() == today.date():
            user_result['users'] += len(usage['users'])
            if 'functions' in usage:
                for func_name, count in usage['functions'].items():
                    function_result[func_name] += count
        elif period == 'week' and today - timedelta(days=today.weekday()) <= record_date <= today:
            user_result['users'] += len(usage['users'])
            if 'functions' in usage:
                for func_name, count in usage['functions'].items():
                    function_result[func_name] += count
        elif period == 'month' and record_date.year == today.year and record_date.month == today.month:
            user_result['users'] += len(usage['users'])
            if 'functions' in usage:
                for func_name, count in usage['functions'].items():
                    function_result[func_name] += count
        elif period == 'year' and record_date.year == today.year:
            user_result['users'] += len(usage['users'])
            if 'functions' in usage:
                for func_name, count in usage['functions'].items():
                    function_result[func_name] += count
        elif period == 'all':
            user_result['users'] += len(usage['users'])
            if 'functions' in usage:
                for func_name, count in usage['functions'].items():
                    function_result[func_name] += count

    return dict(user_result), dict(function_result)


# Получить статистику пользователей
#@log_user_actions
def get_statistics():
    users_data = load_user_data()
    online_users = len([user for user in users_data.values() if is_user_active(user["last_active"]) and not user['blocked']])
    total_users = len(users_data)

    statistics = load_statistics()
    today = datetime.now().strftime('%d.%m.%Y')
    week_start = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime('%d.%m.%Y')
    month_start = datetime.now().replace(day=1).strftime('%d.%m.%Y')
    year_start = datetime.now().replace(month=1, day=1).strftime('%d.%m.%Y')

    users_today = len(statistics.get(today, {}).get('users', set()))
    users_week = len(set().union(*[data.get('users', set()) for date, data in statistics.items() if week_start <= date <= today]))
    users_month = len(set().union(*[data.get('users', set()) for date, data in statistics.items() if month_start <= date <= today]))
    users_year = len(set().union(*[data.get('users', set()) for date, data in statistics.items() if year_start <= date <= today]))

    return online_users, total_users, users_today, users_week, users_month, users_year


# Функция для экранирования специальных символов Markdown
#@log_user_actions
def escape_markdown(text):
    # Экранируем специальные символы Markdown
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

# Получить список активных пользователей
#@log_user_actions
def list_active_users():
    users_data = load_user_data()
    active_users = [
        f"{index + 1}) `{user_id}`: {escape_markdown(user.get('username', 'Неизвестный'))}"
        for index, (user_id, user) in enumerate(users_data.items())
        if is_user_active(user["last_active"]) and not user['blocked']
    ]
    return "\n".join(active_users) if active_users else None

# Получить ТОП пользователей
#@log_user_actions
def get_top_users(top_n=10):
    users_data = load_user_data()
    user_activity = {user_id: user['last_active'] for user_id, user in users_data.items() if not user['blocked']}
    sorted_users = sorted(user_activity.items(), key=lambda x: x[1], reverse=True)
    top_users = sorted_users[:top_n]
    return [f"{index + 1}) {user_id}: {escape_markdown(users_data[user_id].get('username', 'Неизвестный'))}" for index, (user_id, _) in enumerate(top_users)]

# Получить последние действия пользователей
#@log_user_actions
def get_recent_actions(limit=10):
    users_data = load_user_data()
    recent_actions = sorted(users_data.items(), key=lambda x: x[1]['last_active'], reverse=True)
    return [f"{user_id}: {escape_markdown(user['username'])} - {user['last_active']}" for user_id, user in recent_actions[:limit]]

# Получить наиболее активное время использования бота
#@log_user_actions
def get_peak_usage_time():
    statistics = load_statistics()
    usage_times = defaultdict(int)

    for date_str, usage in statistics.items():
        record_date = datetime.strptime(date_str, '%d.%m.%Y')
        for func_name, count in usage.items():
            usage_times[record_date.hour] += count

    peak_hour = max(usage_times, key=usage_times.get)
    return peak_hour, usage_times[peak_hour]

# Получить версию бота
#@log_user_actions
def get_bot_version():
    return "1.0"  # Пример версии

# Получить аптайм бота
#@log_user_actions
def get_uptime():
    start_time = datetime(2025, 1, 1)  # Пример времени запуска
    uptime = datetime.now() - start_time
    days, seconds = uptime.days, uptime.seconds
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return f"{days} дней, {hours}:{minutes} часов"

# Загрузка ошибок из файла
def load_errors():
    with open(ERRORS_LOG_FILE, 'r', encoding='utf-8') as file:
        return json.load(file)

# Формирование нумерованного списка ошибок
#@log_user_actions
def get_error_list():
    errors = load_errors()
    error_list = []
    for index, error in enumerate(errors, start=1):
        error_details = error.get('error_details', '')
        error_details_more = "\n".join(error.get('error_details_more', []))
        error_list.append(f"🛑 *ОШИБКА №{index}* 🛑\n\n{error_details}\n\n{error_details_more}")
    return error_list

# Получить список пользователей и время их последней активности
#@log_user_actions
def get_user_last_active():
    users_data = load_user_data()
    user_last_active = [
        f"📌 {index + 1}) `{user_id}`: {escape_markdown(user.get('username', 'Неизвестный'))} - {user['last_active'][:-3]}"
        for index, (user_id, user) in enumerate(users_data.items())
        if not user['blocked']
    ]
    return "\n".join(user_last_active) if user_last_active else None

# Создание кнопок для подменю
#@log_user_actions
def create_submenu_buttons():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    buttons = [
        types.KeyboardButton("Пользователи"),
        types.KeyboardButton("Использование функций"),
        types.KeyboardButton("Список ошибок"),
        types.KeyboardButton("Версия и аптайм"),
        types.KeyboardButton("В меню админ-панели")
    ]
    markup.row(buttons[0], buttons[1])
    markup.row(buttons[2], buttons[3])
    markup.row(buttons[4])
    return markup

# Обработчик команды "Статистика"
@bot.message_handler(func=lambda message: message.text == 'Статистика' and check_admin_access(message))
#@log_user_actions
def show_statistics(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Статистика'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if not check_admin_access(message):
        return

    bot.send_message(message.chat.id, "Выберите категорию статистики:", reply_markup=create_submenu_buttons())

# Функция для экранирования специальных символов Markdown
#@log_user_actions
def escape_markdown(text):
    # Экранируем специальные символы Markdown
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text in ["Пользователи", "Версия и аптайм", "Использование функций", "Список ошибок", "В меню админ-панели"])
def handle_submenu_buttons(message):
    if not check_admin_access(message):
        return

    if message.text == "Пользователи":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Пользователи'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        online_count, total_count, users_today, users_week, users_month, users_year = get_statistics()
        active_user_list = list_active_users()
        user_last_active_list = get_user_last_active()
        response_message = (
            f"*🌐 Пользователи онлайн:* {online_count}\n"
            f"*👥 Всего пользователей:* {total_count}\n\n"
            f"*📅 Пользователи за день:* {users_today}\n"
            f"*📅 Пользователи за неделю:* {users_week}\n"
            f"*📅 Пользователи за месяц:* {users_month}\n"
            f"*📅 Пользователи за год:* {users_year}\n\n"
        )
        if active_user_list:
            response_message += "*🌐 Пользователи онлайн:*\n\n"
            for user in active_user_list.split('\n'):
                response_message += f"👤 {user}\n"
        else:
            response_message += "*🌐 Нет активных пользователей*\n\n"

        if user_last_active_list:
            response_message += "*\n🕒 Последняя активность пользователей:*\n\n"
            response_message += user_last_active_list
        else:
            response_message += "*\n🕒 Нет данных о последней активности пользователей*"

        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")
    elif message.text == "Версия и аптайм":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Версия и аптайм'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        bot_version = get_bot_version()
        uptime = get_uptime()
        bot.send_message(message.chat.id, f"*🤖 Версия бота:* {bot_version}\n\n*⏳ Аптайм бота:* {uptime}", parse_mode="Markdown")
    elif message.text == "Использование функций":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Использование функций'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        stats_day_users, stats_day_functions = get_aggregated_statistics('day')
        stats_week_users, stats_week_functions = get_aggregated_statistics('week')
        stats_month_users, stats_month_functions = get_aggregated_statistics('month')
        stats_year_users, stats_year_functions = get_aggregated_statistics('year')
        stats_all_users, stats_all_functions = get_aggregated_statistics('all')

        response_message = (
            "*📊 Использование функций:*\n\n\n"
            "☀️ *[За день]* ☀️\n\n" +
            "\n".join([f"{escape_markdown(key)}: {value}" for key, value in stats_day_functions.items()]) +
            "\n\n7️⃣ *[За неделю]* 7️⃣\n\n" +
            "\n".join([f"{escape_markdown(key)}: {value}" for key, value in stats_week_functions.items()]) +
            "\n\n🗓️ *[За месяц]* 🗓️\n\n" +
            "\n".join([f"{escape_markdown(key)}: {value}" for key, value in stats_month_functions.items()]) +
            "\n\n⌛ *[За год]* ⌛\n\n" +
            "\n".join([f"{escape_markdown(key)}: {value}" for key, value in stats_year_functions.items()]) +
            "\n\n♾️ *[За всё время]* ♾️\n\n" +
            "\n".join([f"{escape_markdown(key)}: {value}" for key, value in stats_all_functions.items()])
        )
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")
    elif message.text == "Список ошибок":

        admin_id = str(message.chat.id)
        if not check_permission(admin_id, 'Список ошибок'):
            bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
            return

        error_list = get_error_list()
        if not error_list:
            bot.send_message(message.chat.id, "Данные не найдены! Активные ошибки не обнаружены!", parse_mode="Markdown")
        else:
            full_message = "\n\n".join(error_list)
            if len(full_message) > 4096:
                parts = [full_message[i:i + 4096] for i in range(0, len(full_message), 4096)]
                for part in parts:
                    bot.send_message(message.chat.id, part, parse_mode="Markdown")
            else:
                bot.send_message(message.chat.id, full_message, parse_mode="Markdown")
    elif message.text == "В меню админ-панели":
        bot.send_message(message.chat.id, "Выберите категорию статистики:", reply_markup=create_submenu_buttons())


# (ADMIN 5) ------------------------------------------ "РЕЗЕРВНАЯ КОПИЯ ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

import zipfile

# Путь к директории для бэкапов и текущего исполняемого файла
BACKUP_DIR = 'backups'
SOURCE_DIR = '.'
EXECUTABLE_FILE = '(104 update ИСПРАВЛЕНИЕ34  ( ( )) CAR MANAGER TG BOT (official) v0924.py'

# Путь к JSON файлу с админскими сессиями
ADMIN_SESSIONS_FILE = 'data base/admin/admin_sessions.json'

#@log_user_actions
def normalize_name(name):
    return re.sub(r'[<>:"/\\|?*]', '_', name)

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data['admin_sessions']

# Функция для проверки прав доступа
#@log_user_actions
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "У вас нет прав доступа для выполнения этой операции.")
        return False

# Обработчик для кнопки "Резервная копия"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Резервная копия' and check_admin_access(message))
def show_backup_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Резервная копия'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Создать копию', 'Восстановить данные')
    markup.add('В меню админ-панели')

    bot.send_message(message.chat.id, "Выберите действие с резервной копией:", reply_markup=markup)

# Обработчик для кнопки "Создать копию"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Создать копию' and check_admin_access(message))
def handle_create_backup(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Создать копию'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    backup_path = create_backup()
    if backup_path:
        backup_message = f"Резервная копия создана!\n\nПуть к резервной копии:\n{backup_path}"
    else:
        backup_message = "Ошибка при создании резервной копии."
    bot.send_message(message.chat.id, backup_message)
    show_admin_panel(message)

# Обработчик для кнопки "Восстановить данные"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Восстановить данные' and check_admin_access(message))
def handle_restore_backup(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Восстановить данные'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    success = restore_latest_backup()
    if success:
        bot.send_message(message.chat.id, "Данные успешно восстановлены из последней резервной копии!")
    else:
        bot.send_message(message.chat.id, "Ошибка: последний бэкап не найден!")
    show_admin_panel(message)

# Функция для создания резервной копии
#@log_user_actions
def create_backup():
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(BACKUP_DIR, f'backup_{timestamp}.zip')

    os.makedirs(BACKUP_DIR, exist_ok=True)

    try:
        with zipfile.ZipFile(backup_file, 'w', zipfile.ZIP_STORED) as zipf:
            for root, dirs, files in os.walk(SOURCE_DIR):
                if 'backups' in dirs:
                    dirs.remove('backups')  # Исключаем папку backups

                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, SOURCE_DIR)

                    if EXECUTABLE_FILE in file:
                        continue

                    if len(file_path) > 260:
                        continue

                    try:
                        zipf.write(file_path, arcname)
                    except Exception:
                        pass

        # Проверяем целостность архива
        with zipfile.ZipFile(backup_file, 'r') as zipf:
            if zipf.testzip() is not None:
                return None

        return backup_file
    except Exception:
        return None

# Функция для восстановления из последнего бэкапа
#@log_user_actions
def restore_latest_backup():
    backups = sorted(os.listdir(BACKUP_DIR), reverse=True)
    if not backups:
        return False

    latest_backup = os.path.join(BACKUP_DIR, backups[0])

    if not os.path.exists(latest_backup):
        return False

    with zipfile.ZipFile(latest_backup, 'r') as zipf:
        zipf.extractall(SOURCE_DIR)

    return True

# Функция для отправки уведомления администратору
#@log_user_actions
def notify_admin(backup_path):
    admin_sessions = load_admin_sessions()
    current_time = datetime.now().strftime('%d.%m.%Y в %H:%M')
    for admin_id in admin_sessions:
        bot.send_message(admin_id, f"Резервная копия создана!\n\nВремя создания: {current_time}\n\nПуть к резервной копии:\n{backup_path}", parse_mode="Markdown")

# Функция для автоматического создания резервной копии и отправки уведомления
#@log_user_actions
def scheduled_backup():
    backup_path = create_backup()
    notify_admin(backup_path)

# Планирование задачи на каждый день в 00:00
schedule.every().day.at("00:00").do(scheduled_backup)


# (ADMIN n) ------------------------------------------ "ВКЛ/ВЫКЛ ФУУНКЦИЙ ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

# Путь к JSON файлу с админскими сессиями
ADMIN_SESSIONS_FILE = 'data base/admin/admin_sessions.json'

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data['admin_sessions']

# Функция для проверки прав доступа
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "У вас нет прав доступа для выполнения этой операции!")
        return False

# Путь к файлу с состоянием функций
FUNCTIONS_STATE_PATH = 'data base/admin/functions_state.json'

# Функция загрузки состояния функций из файла
def load_function_states():
    if os.path.exists(FUNCTIONS_STATE_PATH):
        with open(FUNCTIONS_STATE_PATH, 'r', encoding='utf-8') as file:
            data = json.load(file)
            if data:
                return data
    # Если файл отсутствует или пуст, создаем его с начальным состоянием
    save_function_states({})
    return {}

# Сохранение состояния функций в файл
def save_function_states(states):
    with open(FUNCTIONS_STATE_PATH, 'w', encoding='utf-8') as file:
        json.dump(states, file, ensure_ascii=False, indent=4)

# Загрузка состояний функций при старте
function_states = load_function_states()

# Добавление новых функций, если они отсутствуют
new_functions = {
    "Общее меню": [
        "Расход топлива", "Траты и ремонты", "Ваш транспорт", "Найти транспорт", "Поиск мест",
        "Погода", "Цены на топливо", "Код" "Анти-радар", "Напоминания", "Коды OBD2", "Прочее", "Сайт"
    ],
    "Кнопки": [
        "В главное меню", "Вернуться в меню расчета топлива", "Вернуться в меню трат и ремонтов",
        "Выбрать категорию заново", "Вернуться в меню напоминаний", "Посмотреть другие ошибки",
        "Вернуться в ваш транспорт", "еще новости"
    ],
    "Расход топлива": [
        "Рассчитать расход топлива", "Сохранить поездку", "Посмотреть поездки",
        "Посмотреть в Excel", "Посмотреть другие поездки", "Удалить поездку"
    ],
    "Меню трат": [
        "Записать трату", "Посмотреть траты", "Удалить траты", "Удалить категорию"
    ],
    "Меню просмотра трат": [
        "Траты (по категориям)", "Траты (месяц)", "Траты (год)", "Траты (все время)",
        "Посмотреть траты в EXCEL"
    ],
    "Меню удаления трат": [
        "Del траты (категория)", "Del траты (месяц)", "Del траты (год)", "Del траты (все время)"
    ],
    "Меню записи ремонтов": [
        "Записать ремонт", "Посмотреть ремонты", "Удалить ремонты"
    ],
    "Меню просмотра ремонтов": [
        "Ремонты (по категориям)", "Ремонты (месяц)", "Ремонты (год)", "Ремонты (все время)",
        "Посмотреть ремонты в EXCEL"
    ],
    "Меню удаления ремонтов": [
        "Del ремонты (категория)", "Del ремонты (месяц)", "Del ремонты (год)", "Del ремонты (все время)"
    ],
    "Меню ваш транспорт": [
        "Добавить транспорт", "Посмотреть транспорт", "Изменить транспорт"
    ],
    "Меню удаления ваш транспорт": [
        "Удалить транспорт", "Удалить весь транспорт"
    ],
    "Меню поиск мест": [
        "АЗС", "Автомойки", "Автосервисы", "Парковки", "Эвакуация", "ГИБДД", "Комиссары", "Штрафстоянка"
    ],
    "Меню погоды": [
        "Сегодня", "Завтра", "Неделя", "Месяц", "Другое место"
    ],
    "Меню напоминаний": [
        "Добавить напоминание", "Посмотреть напоминания", "Активные", "Истекшие"
    ],
    "Меню удаления напоминаний": [
        "Удалить напоминание", "Удалить напоминание", "Удалить все напоминания"
    ],
    "Другие функции": [
        "Выключить анти-радар", "Функция для обработки локации"
    ],
    "Прочее": [
        "Чат с админом", "Для рекламы", "Новости"
    ],
    "Меню новости": ['3 новости', '5 новостей', '7 новостей', '10 новостей', '15 новостей'],
    "Для рекламы": [
        "Заявка на рекламу", "Ваши заявки"
    ],
        "Код региона": [
        "Код региона"
    ]
}

# Функция для обновления состояния функций
def update_function_states():
    global function_states
    for category, functions in new_functions.items():
        for function_name in functions:
            if function_name not in function_states:
                function_states[function_name] = {"state": True, "deactivation_time": None}
    save_function_states(function_states)

# Обновляем состояние функций при загрузке
update_function_states()

for category, functions in new_functions.items():
    for function_name in functions:
        if function_name not in function_states:
            function_states[function_name] = {"state": True, "deactivation_time": None}

save_function_states(function_states)  # Передаем аргумент function_states

#@log_user_actions
def set_function_state(function_name, state, deactivation_time=None):
    if function_name in function_states:
        function_states[function_name]['state'] = state
        if deactivation_time:
            function_states[function_name]['deactivation_time'] = deactivation_time
        else:
            function_states[function_name]['deactivation_time'] = None
        save_function_states(function_states)  # Передаем аргумент function_states
        return f"Функция *{function_name}* успешно {'активирована' if state else 'деактивирована'}!"
    else:
        return "Ошибка: функция не найдена!"

#@log_user_actions
def activate_function(function_name):
    return set_function_state(function_name, True)

#@log_user_actions
def deactivate_function(function_name, deactivation_time):
    return set_function_state(function_name, False, deactivation_time)

# Функция для активации функции через некоторое время
def activate_function_later(function_name, delay):
    threading.Timer(delay.total_seconds(), lambda: notify_admin_and_activate(function_name)).start()

#@log_user_actions
def notify_admin_and_activate(function_name):
    deactivation_time = function_states[function_name]['deactivation_time']
    date_part, time_part = deactivation_time.split('; ')
    admin_sessions = load_admin_sessions()
    for admin_id in admin_sessions:
        if len([fn for fn, data in function_states.items() if not data['state']]) == 1:
            bot.send_message(admin_id, f"Функция *{function_name.lower()}* была *включена* по истечению времени (до {date_part} в {time_part})!", parse_mode="Markdown")
        else:
            bot.send_message(admin_id, f"Функции *{function_name.lower()}* была *включена* по истечению времени (до {date_part} в {time_part})!", parse_mode="Markdown")
    activate_function(function_name)

# Проверка корректности даты и времени
#@log_user_actions
def is_valid_date_time(date_str, time_str):
    try:
        date = datetime.strptime(date_str, "%d.%m.%Y")
        time = datetime.strptime(time_str, "%H:%M")
        if 2000 <= date.year <= 3000 and 1 <= date.month <= 12 and 1 <= date.day <= 31:
            if 0 <= time.hour <= 23 and 0 <= time.minute <= 59:
                return True
        return False
    except ValueError:
        return False

# Обработчик временной деактивации
#@log_user_actions
def handle_time_deactivation(time_spec, function_names, message):
    try:
        end_time = datetime.strptime(time_spec, "%d.%m.%Y; %H:%M")
        now = datetime.now()

        # Проверяем, что время в будущем
        if end_time > now:
            for function_name in function_names:
                deactivate_function(function_name, time_spec)  # Деактивируем сразу
                delay = end_time - now  # Вычисляем задержку
                activate_function_later(function_name, delay)  # Запускаем активацию позже
            date_part, time_part = time_spec.split('; ')
            if len(function_names) == 1:
                bot.send_message(message.chat.id, f"Функция *{', '.join(function_names).lower()}* *отключена* до {date_part} в {time_part}!", parse_mode="Markdown")
            else:
                bot.send_message(message.chat.id, f"Функции *{', '.join(function_names).lower()}* *отключены* до {date_part} в {time_part}!", parse_mode="Markdown")
            toggle_functions(message)  
        else:
            bot.send_message(message.chat.id, "Указанное время уже прошло! Пожалуйста, введите дату и время снова")
            bot.register_next_step_handler(message, process_disable_function_date_step, function_names)
    except ValueError:
        bot.send_message(message.chat.id, "Неверный формат! Попробуйте снова")
        bot.register_next_step_handler(message, process_disable_function_date_step, function_names)

# Обработчик для команды "Функции"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Функции' and check_admin_access(message))
def toggle_functions(message): 

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Функции'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if check_admin_access(message):
        if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
            bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
            bot.register_next_step_handler(message, toggle_functions)
            return
        
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(types.KeyboardButton('Включение'), types.KeyboardButton('Выключение'))
    markup.add(types.KeyboardButton('В меню админ-панели'))
    bot.send_message(message.chat.id, "Выберите действие для функций:", reply_markup=markup)

# Обработчик для кнопки "Включение"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Включение' and check_admin_access(message))
def enable_function(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Включение'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if check_admin_access(message):
        if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
            bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
            bot.register_next_step_handler(message, enable_function)
            return
        disabled_functions = [(name, data['deactivation_time']) for name, data in function_states.items() if not data['state']]
        if disabled_functions:
            response = "*Выключенные* функции:\n\n\n"
            index = 1
            for category, functions in new_functions.items():
                response += f"*{category}*:\n"
                for function in functions:
                    if function in [name for name, _ in disabled_functions]:
                        deactivation_time = next((data for name, data in disabled_functions if name == function), None)
                        if deactivation_time:
                            date_part, time_part = deactivation_time.split('; ')
                            response += f"❌ {index}. *{function}* (до {date_part} в {time_part})\n"
                        else:
                            response += f"❌ {index}. *{function}*\n"
                        index += 1
                response += "\n"
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('Вернуться в функции'))
            markup.add(types.KeyboardButton('В меню админ-панели'))
            bot.send_message(message.chat.id, response, parse_mode="Markdown", reply_markup=markup)
            bot.send_message(message.chat.id, "Введите номера функций для включения:")
            bot.register_next_step_handler(message, process_enable_function_step)
        else:
            bot.send_message(message.chat.id, "*Все* функции уже *включены*!", parse_mode="Markdown")
            toggle_functions(message) 

#@log_user_actions
def process_enable_function_step(message):

    if message.text == "Вернуться в функции":
        toggle_functions(message) 
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_enable_function_step)
        return

    try:
        function_indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        disabled_functions = [name for name, data in function_states.items() if not data['state']]
        valid_indices = [i for i in function_indices if 0 <= i < len(disabled_functions)]
        if valid_indices:
            function_names = [disabled_functions[i] for i in valid_indices]
            for function_name in function_names:
                activate_function(function_name)
            bot.send_message(message.chat.id, f"Функции *{', '.join(function_names).lower()}* успешно *активированы*!", parse_mode="Markdown")
            toggle_functions(message) 
        else:
            bot.send_message(message.chat.id, "Неверные номера функций!")
    except ValueError:
        bot.send_message(message.chat.id, "Неверный формат! Введите номера функций через запятую")

# Обработчик для кнопки "Выключение"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Выключение' and check_admin_access(message))
def disable_function(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Выключение'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if check_admin_access(message):
        if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
            bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
            bot.register_next_step_handler(message, disable_function)
            return
        enabled_functions = [name for name, data in function_states.items() if data['state']]
        if enabled_functions:
            response = "*Включенные* функции:\n\n\n"
            index = 1
            for category, functions in new_functions.items():
                response += f"*{category}*:\n"
                for function in functions:
                    if function in enabled_functions:
                        response += f"✅ {index}. *{function}*\n"
                        index += 1
                response += "\n"
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('Вернуться в функции'))
            markup.add(types.KeyboardButton('В меню админ-панели'))
            bot.send_message(message.chat.id, response, parse_mode="Markdown", reply_markup=markup)
            bot.send_message(message.chat.id, "Введите номера функций для выключения:")
            bot.register_next_step_handler(message, process_disable_function_step)
        else:
            bot.send_message(message.chat.id, "*Все* функции уже *выключены*!", parse_mode="Markdown")
            toggle_functions(message)

#@log_user_actions
def process_disable_function_step(message):

    if message.text == "Вернуться в функции":
        toggle_functions(message) 
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_disable_function_step)
        return

    try:
        function_indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        enabled_functions = [name for name, data in function_states.items() if data['state']]
        valid_indices = [i for i in function_indices if 0 <= i < len(enabled_functions)]
        if valid_indices:
            function_names = [enabled_functions[i] for i in valid_indices]
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('Вернуться в функции'))
            markup.add(types.KeyboardButton('В меню админ-панели'))
            bot.send_message(message.chat.id, "Введите дату для выключения:", reply_markup=markup)
            bot.register_next_step_handler(message, process_disable_function_date_step, function_names)
        else:
            bot.send_message(message.chat.id, "Неверные номера функций!")
            bot.register_next_step_handler(message, process_disable_function_step)
    except ValueError:
        bot.send_message(message.chat.id, "Неверный формат! Введите номера функций через запятую")
        bot.register_next_step_handler(message, process_disable_function_step)

#@log_user_actions
def process_disable_function_date_step(message, function_names):

    if message.text == "Вернуться в функции":
        toggle_functions(message) 
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_disable_function_date_step, function_names)
        return

    date_str = message.text
    if is_valid_date_time(date_str, "00:00"):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('Вернуться в функции'))
        markup.add(types.KeyboardButton('В меню админ-панели'))
        bot.send_message(message.chat.id, "Введите время для выключения:", reply_markup=markup)
        bot.register_next_step_handler(message, process_disable_function_time_step, function_names, date_str, message)
    else:
        bot.send_message(message.chat.id, "Неверный формат даты! Попробуйте снова")
        bot.register_next_step_handler(message, process_disable_function_date_step, function_names)

#@log_user_actions
def process_disable_function_time_step(message, function_names, date_str, original_message):

    if message.text == "Вернуться в функции":
        toggle_functions(message) 
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_disable_function_time_step, function_names, date_str, original_message)
        return

    time_str = message.text
    if is_valid_date_time(date_str, time_str):
        time_spec = f"{date_str}; {time_str}"
        handle_time_deactivation(time_spec, function_names, original_message)
    else:
        bot.send_message(message.chat.id, "Неверный формат времени! Попробуйте снова")
        bot.register_next_step_handler(message, process_disable_function_time_step, function_names, date_str, original_message)

# (ADMIN n) ------------------------------------------ "ОПОВЕЩЕНИЯ ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

# Путь к файлу
DATABASE_PATH = 'data base/admin/chats/alerts.json'
ADMIN_SESSIONS_FILE = 'data base/admin/admin_sessions.json'
USER_DATA_PATH = 'data base/admin/users.json'  # Путь к файлу с данными пользователей

# Глобальные переменные
alerts = {"sent_messages": {}, "notifications": {}}  # Инициализация структуры alerts
admin_sessions = []

# Загрузка пользователей
def load_users():
    if os.path.exists(USER_DATA_PATH):
        with open(USER_DATA_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    return {}

# Сохранение базы данных
def save_database():
    # Преобразование datetime объектов в строки
    for key, value in alerts['notifications'].items():
        if 'time' in value and isinstance(value['time'], datetime):
            value['time'] = value['time'].strftime("%d.%m.%Y в %H:%M")
    for key, value in alerts['sent_messages'].items():
        if 'time' in value and isinstance(value['time'], datetime):
            value['time'] = value['time'].strftime("%d.%m.%Y в %H:%M")

    # Убедиться, что директория для базы данных существует
    os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)

    # Сохранение данных в файл
    try:
        with open(DATABASE_PATH, 'w', encoding='utf-8') as file:
            json.dump(alerts, file, ensure_ascii=False, indent=4)
    except Exception as e:
        pass  # Игнорировать ошибки

    # Обратное преобразование строк в datetime
    for key, value in alerts['notifications'].items():
        if 'time' in value and isinstance(value['time'], str):
            value['time'] = datetime.strptime(value['time'], "%d.%m.%Y в %H:%M")
    for key, value in alerts['sent_messages'].items():
        if 'time' in value and isinstance(value['time'], str):
            value['time'] = datetime.strptime(value['time'], "%d.%m.%Y в %H:%M")

# Загрузка базы данных
def load_database():
    if os.path.exists(DATABASE_PATH):
        try:
            with open(DATABASE_PATH, 'r', encoding='utf-8') as file:
                data = json.load(file)

            # Преобразование строк во временные метки
            for key, value in data['notifications'].items():
                if 'time' in value and value['time']:
                    value['time'] = datetime.strptime(value['time'], "%d.%m.%Y в %H:%M")

            # Преобразование списка sent_messages в словарь (если нужно)
            if isinstance(data['sent_messages'], list):
                sent_messages_dict = {}
                for i, msg in enumerate(data['sent_messages']):
                    msg_id = str(i + 1)
                    sent_messages_dict[msg_id] = msg
                    if 'time' in msg and msg['time']:
                        msg['time'] = datetime.strptime(msg['time'], "%d.%m.%Y в %H:%M")
                data['sent_messages'] = sent_messages_dict
            else:
                for key, value in data['sent_messages'].items():
                    if 'time' in value and value['time']:
                        value['time'] = datetime.strptime(value['time'], "%d.%m.%Y в %H:%M")

            return data
        except Exception:
            pass  # Игнорировать ошибки

    return {"sent_messages": {}, "notifications": {}}

# Инициализация базы данных при запуске
alerts = load_database()

def check_notifications():
    while True:
        now = datetime.now()
        for key, n in alerts['notifications'].items():
            if n['status'] == 'active' and 'time' in n and n['time'] <= now:
                user_id = n.get('user_id')
                if user_id:
                    user_ids = [user_id]
                else:
                    user_ids = load_users().keys()

                for user_id in user_ids:
                    if n.get('text'):
                        bot.send_message(user_id, n['text'])
                    else:
                        for file in n.get('files', []):
                            if file['type'] == 'photo':
                                bot.send_photo(user_id, file['file_id'], caption=file.get('caption'))
                            elif file['type'] == 'video':
                                bot.send_video(user_id, file['file_id'], caption=file.get('caption'))
                            elif file['type'] == 'document':
                                bot.send_document(user_id, file['file_id'], caption=file.get('caption'))
                            elif file['type'] == 'animation':
                                bot.send_animation(user_id, file['file_id'], caption=file.get('caption'))
                            elif file['type'] == 'sticker':
                                bot.send_sticker(user_id, file['file_id'])
                            elif file['type'] == 'audio':
                                bot.send_audio(user_id, file['file_id'], caption=file.get('caption'))
                            elif file['type'] == 'voice':
                                bot.send_voice(user_id, file['file_id'], caption=file.get('caption'))
                            elif file['type'] == 'video_note':
                                bot.send_video_note(user_id, file['file_id'])
                n['status'] = 'sent'
        save_database()
        time.sleep(60)

# Запускаем проверку уведомлений в отдельном потоке
threading.Thread(target=check_notifications, daemon=True).start()

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    if os.path.exists(ADMIN_SESSIONS_FILE):
        with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
            data = json.load(file)
        return data.get('admin_sessions', [])
    return []

admin_sessions = load_admin_sessions()

#@log_user_actions
def check_admin_access(message):
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return False

# Показ меню уведомлений
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Оповещения' and check_admin_access(message))
def show_notifications_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Оповещения'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True)
    markup.add('Всем', 'По времени', 'Отдельно')
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите тип оповещения:", reply_markup=markup)

# Обработчик для "По времени"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'По времени' and check_admin_access(message))
def handle_time_notifications(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'По времени'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Отправить по времени')
    markup.add('Просмотр (по времени)', 'Удалить (по времени)')
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Управление оповещениями по времени:", reply_markup=markup)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Отправить по времени' and check_admin_access(message))
def schedule_notification(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Отправить по времени'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Отправить всем', 'Отправить отдельно')
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите действие для отправки по времени:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_send_action)

@check_user_blocked
#@log_user_actions
def choose_send_action(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_send_action)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return

    if message.text == 'Отправить всем':
        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в общение")
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите тему оповещения:", reply_markup=markup)
        bot.register_next_step_handler(message, set_theme_for_notification)
    elif message.text == 'Отправить отдельно':
        list_users_for_time_notification(message)

@check_user_blocked
#@log_user_actions
def set_theme_for_notification(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_theme_for_notification)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    notification_theme = message.text
    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите текст оповещения или отправьте мультимедийный файл:", reply_markup=markup)
    bot.register_next_step_handler(message, set_time_for_notification, notification_theme)

#@log_user_actions
@check_user_blocked
def set_time_for_notification(message, notification_theme):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    notification_text = message.text or message.caption
    content_type = message.content_type
    file_id = None
    caption = message.caption

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите дату оповещения:", reply_markup=markup)
    bot.register_next_step_handler(message, process_notification_date, notification_theme, notification_text, content_type, file_id, caption)

#@log_user_actions
@check_user_blocked
def process_notification_date(message, notification_theme, notification_text, content_type, file_id, caption):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_notification_date, notification_theme, notification_text, content_type, file_id, caption)
        return

    date_str = message.text

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if not validate_date_format(date_str):
        bot.send_message(message.chat.id, "Неверный формат даты! Введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, process_notification_date, notification_theme, notification_text, content_type, file_id, caption)
        return

    try:
        notification_date = datetime.strptime(date_str, "%d.%m.%Y")
        if notification_date.date() < datetime.now().date():
            bot.send_message(message.chat.id, "Введенная дата уже прошла! Пожалуйста, введите корректную дату")
            bot.register_next_step_handler(message, process_notification_date, notification_theme, notification_text, content_type, file_id, caption)
            return
    except ValueError:
        bot.send_message(message.chat.id, "Неверный формат даты! Введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, process_notification_date, notification_theme, notification_text, content_type, file_id, caption)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите время оповещения:", reply_markup=markup)
    bot.register_next_step_handler(message, process_notification_time, notification_theme, notification_text, date_str, content_type, file_id, caption)

@check_user_blocked
#@log_user_actions
def validate_date_format(date_str):
    try:
        datetime.strptime(date_str, "%d.%m.%Y")
        return True
    except ValueError:
        return False

@check_user_blocked
#@log_user_actions
def process_notification_time(message, notification_theme, notification_text, date_str, content_type, file_id, caption):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_notification_time, notification_theme, notification_text, date_str, content_type, file_id, caption)
        return

    time_str = message.text

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if not validate_time_format(time_str):
        bot.send_message(message.chat.id, "Неверный формат времени! Введите время в формате ЧЧ:ММ")
        bot.register_next_step_handler(message, process_notification_time, notification_theme, notification_text, date_str, content_type, file_id, caption)
        return

    try:
        notification_time = datetime.strptime(f"{date_str}, {time_str}", "%d.%m.%Y, %H:%M")
        if notification_time < datetime.now():
            bot.send_message(message.chat.id, "Введенное время уже прошло! Пожалуйста, введите корректное время")
            bot.register_next_step_handler(message, process_notification_time, notification_theme, notification_text, date_str, content_type, file_id, caption)
            return
    except ValueError:
        bot.send_message(message.chat.id, "Неверный формат времени! Введите время в формате ЧЧ:ММ")
        bot.register_next_step_handler(message, process_notification_time, notification_theme, notification_text, date_str, content_type, file_id, caption)
        return

    notification_id = str(len(alerts['notifications']) + 1)
    alerts['notifications'][notification_id] = {
        'theme': notification_theme,
        'text': notification_text if content_type == 'text' else None,
        'time': notification_time,
        'status': 'active',
        'category': 'time',
        'user_id': None,
        'files': [
            {
                'type': content_type,
                'file_id': file_id,
                'caption': caption if content_type != 'text' else None
            }
        ],
        'content_type': content_type
    }
    save_database()
    bot.send_message(message.chat.id, f"Оповещение *{notification_theme.lower()}* запланировано на {notification_time.strftime('%d.%m.%Y в %H:%M')}!", parse_mode="Markdown")
    show_communication_menu(message)

@check_user_blocked  
#@log_user_actions
def validate_time_format(time_str):
    try:
        datetime.strptime(time_str, "%H:%M")
        return True
    except ValueError:
        return False

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Просмотр (по времени)' and check_admin_access(message))
def show_view_notifications(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Просмотр (по времени)'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Активные (по времени)', 'Остановленные (по времени)')
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите тип просмотра оповещений:", reply_markup=markup)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Активные (по времени)' and check_admin_access(message))
def show_active_notifications(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Активные (по времени)'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if alerts['notifications']:
        active_notifications = [
            f"⭐ №{i + 1} ⭐\n\n📝 *Тема*: {n['theme'].lower() if n['theme'] else 'без темы'}\n📅 *Дата*: {n['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {n['time'].strftime('%H:%M')}\n🔄 *Статус*: {'отправлено' if n['status'] == 'sent' else 'отложено'}\n"
            for i, n in enumerate([n for n in alerts['notifications'].values() if n['status'] == 'active' and n['category'] == 'time'])
        ]
        if active_notifications:
            bot.send_message(message.chat.id, "Список *активных оповещений (по времени)*:\n\n\n" + "\n\n".join(active_notifications), parse_mode="Markdown")

            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add("Вернуться в общение")
            markup.add('В меню админ-панели')

            bot.send_message(message.chat.id, "Введите номера оповещений для просмотра через запятую:", reply_markup=markup)
            bot.register_next_step_handler(message, show_notification_details, 'active')
        else:
            bot.send_message(message.chat.id, "Нет активных оповещений!")
    else:
        bot.send_message(message.chat.id, "Нет уведомлений!")

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Остановленные (по времени)' and check_admin_access(message))
def show_stopped_notifications(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Остановленные (по времени)'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    stopped_notifications = [
        f"⭐ №{i + 1} ⭐\n\n📝 *Тема*: {n['theme'].lower() if n['theme'] else 'без темы'}\n📅 *Дата*: {n['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {n['time'].strftime('%H:%M')}\n🔄 *Статус*: {'отправлено' if n['status'] == 'sent' else 'отложено'}\n"
        for i, n in enumerate([n for n in alerts['notifications'].values() if n['status'] == 'sent' and n['category'] == 'time'])
    ]
    if stopped_notifications:
        bot.send_message(message.chat.id, "Список *остановленных оповещений (по времени)*:\n\n\n" + "\n\n".join(stopped_notifications), parse_mode="Markdown")

        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в общение")
        markup.add('В меню админ-панели')

        bot.send_message(message.chat.id, "Введите номера оповещений для просмотра через запятую:", reply_markup=markup)
        bot.register_next_step_handler(message, show_notification_details, 'sent')
    else:
        bot.send_message(message.chat.id, "Нет остановленных оповещений!")

#@log_user_actions
def show_notification_details(message, status):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, show_notification_details, status)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        notifications = [n for n in alerts['notifications'].values() if n['status'] == status and n['category'] == 'time']
        valid_indices = [index for index in indices if 0 <= index < len(notifications)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверный номер оповещения! Попробуйте снова")
            bot.register_next_step_handler(message, show_notification_details, status)
            return

        for index in valid_indices:
            notification = notifications[index]
            theme = notification['theme'].lower() if notification['theme'] else 'без темы'
            content_type = notification.get('content_type', 'текст')
            status_text = 'отправлено' if notification['status'] == 'sent' else 'активно'

            notification_details = (
                f"*Основная информация*:\n\n"
                f"📝 *Тема*: {theme}\n"
                f"📁 *Тип контента*: {content_type}\n"
                f"📅 *Дата*: {notification['time'].strftime('%d.%m.%Y')}\n"
                f"🕒 *Время*: {notification['time'].strftime('%H:%M')}\n"
                f"🔄 *Статус*: {status_text}\n"
            )

            if content_type == 'text':
                notification_details += f"\n*Текст сообщения*:\n\n{notification['text']}\n"

            bot.send_message(message.chat.id, notification_details, parse_mode="Markdown")

            if content_type != 'text':
                for file in notification.get('files', []):
                    if file['type'] == 'photo':
                        bot.send_photo(message.chat.id, file['file_id'], caption=file.get('caption'))
                    elif file['type'] == 'video':
                        bot.send_video(message.chat.id, file['file_id'], caption=file.get('caption'))
                    elif file['type'] == 'document':
                        bot.send_document(message.chat.id, file['file_id'], caption=file.get('caption'))
                    elif file['type'] == 'animation':
                        bot.send_animation(message.chat.id, file['file_id'], caption=file.get('caption'))
                    elif file['type'] == 'sticker':
                        bot.send_sticker(message.chat.id, file['file_id'])
                    elif file['type'] == 'audio':
                        bot.send_audio(message.chat.id, file['file_id'], caption=file.get('caption'))
                    elif file['type'] == 'voice':
                        bot.send_voice(message.chat.id, file['file_id'], caption=file.get('caption'))
                    elif file['type'] == 'video_note':
                        bot.send_video_note(message.chat.id, file['file_id'])

        show_communication_menu(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер оповещения!")
        bot.register_next_step_handler(message, show_notification_details, status)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить (по времени)' and check_admin_access(message))
def delete_notification(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить (по времени)'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    notifications_list = [
        f"⭐ *№{i + 1}* ⭐\n\n📝 *Тема*: {n['theme'].lower() if n['theme'] else 'без темы'}\n📅 *Дата*: {n['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {n['time'].strftime('%H:%M')}\n🔄 *Статус*: {'отправлено' if n['status'] == 'sent' else 'активно'}"
        for i, n in enumerate([n for n in alerts['notifications'].values() if n['category'] == 'time'])
    ]
    if notifications_list:
        bot.send_message(message.chat.id, "Список для *удаления (по времени)*:\n\n\n" + "\n\n".join(notifications_list), parse_mode="Markdown")

        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в общение")
        markup.add('В меню админ-панели')

        bot.send_message(message.chat.id, "Введите номера оповещений для удаления через запятую:", reply_markup=markup)
        bot.register_next_step_handler(message, process_delete_notification)
    else:
        bot.send_message(message.chat.id, "Нет оповещений для удаления!")

#@log_user_actions
def process_delete_notification(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_delete_notification)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        notifications = list(alerts['notifications'].values())
        valid_indices = [index for index in indices if 0 <= index < len(notifications)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверный номер оповещения! Попробуйте снова")
            bot.register_next_step_handler(message, process_delete_notification)
            return

        deleted_notifications = []
        for index in sorted(valid_indices, reverse=True):
            notification_id = list(alerts['notifications'].keys())[index]
            deleted_notification = alerts['notifications'].pop(notification_id)
            deleted_notifications.append(deleted_notification)

        # Перенумерация оставшихся оповещений
        new_notifications = {}
        for i, (notification_id, notification) in enumerate(alerts['notifications'].items(), start=1):
            new_notifications[str(i)] = notification

        alerts['notifications'] = new_notifications
        save_database()  # Сохраняем изменения после удаления и перенумерации

        deleted_themes = ", ".join([f"*{msg['theme'].lower()}*" for msg in deleted_notifications])
        bot.send_message(message.chat.id, f"Оповещения по темам {deleted_themes} были удалены!", parse_mode="Markdown")

        show_communication_menu(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер уведомления!")
        bot.register_next_step_handler(message, process_delete_notification)

#@log_user_actions
def list_users_for_time_notification(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, list_users_for_time_notification)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    users_data = load_users()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *заблокирован* 🚫" if data.get('blocked', False) else " - *разблокирован* ✅"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер пользователя для отправки оповещения:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_user_for_time_notification)

#@log_user_actions
def choose_user_for_time_notification(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_user_for_time_notification)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return

    try:
        index = int(message.text) - 1
        users_data = load_users()
        user_list = list(users_data.keys())
        if 0 <= index < len(user_list):
            user_id = user_list[index]
            bot.send_message(message.chat.id, "Введите тему оповещения:")
            bot.register_next_step_handler(message, set_theme_for_time_notification, user_id)
        else:
            bot.send_message(message.chat.id, "Неверный номер пользователя! Попробуйте снова")
            bot.register_next_step_handler(message, choose_user_for_time_notification)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер пользователя!")
        bot.register_next_step_handler(message, choose_user_for_time_notification)

#@log_user_actions
def set_theme_for_time_notification(message, user_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_theme_for_time_notification, user_id)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    individual_theme = message.text
    bot.send_message(message.chat.id, "Введите текст оповещения или отправьте мультимедийный файл:")
    bot.register_next_step_handler(message, set_time_for_time_notification, user_id, individual_theme)

#@log_user_actions
def set_time_for_time_notification(message, user_id, individual_theme):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    notification_text = message.text or message.caption
    content_type = message.content_type
    file_id = None
    caption = message.caption

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите дату оповещения:", reply_markup=markup)
    bot.register_next_step_handler(message, process_time_notification_date, user_id, individual_theme, notification_text, content_type, file_id, caption)

#@log_user_actions
def process_time_notification_date(message, user_id, individual_theme, notification_text, content_type, file_id, caption):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_time_notification_date, user_id, individual_theme, notification_text, content_type, file_id, caption)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    date_str = message.text
    if not validate_date_format(date_str):
        bot.send_message(message.chat.id, "Неверный формат даты! Введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, process_time_notification_date, user_id, individual_theme, notification_text, content_type, file_id, caption)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите время оповещения:", reply_markup=markup)
    bot.register_next_step_handler(message, process_time_notification_time, user_id, individual_theme, notification_text, date_str, content_type, file_id, caption)

#@log_user_actions
def process_time_notification_time(message, user_id, individual_theme, notification_text, date_str, content_type, file_id, caption):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена на этом этапе. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_time_notification_time, user_id, individual_theme, notification_text, date_str, content_type, file_id, caption)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    time_str = message.text
    if not validate_time_format(time_str):
        bot.send_message(message.chat.id, "Неверный формат времени! Введите время в формате ЧЧ:ММ")
        bot.register_next_step_handler(message, process_time_notification_time, user_id, individual_theme, notification_text, date_str, content_type, file_id, caption)
        return

    try:
        notification_time = datetime.strptime(f"{date_str}, {time_str}", "%d.%m.%Y, %H:%M")
        if notification_time < datetime.now():
            bot.send_message(message.chat.id, "Введенное время уже прошло! Пожалуйста, введите корректное время")
            bot.register_next_step_handler(message, process_time_notification_time, user_id, individual_theme, notification_text, date_str, content_type, file_id, caption)
            return
    except ValueError:
        bot.send_message(message.chat.id, "Неверный формат времени! Введите время в формате ЧЧ:ММ")
        bot.register_next_step_handler(message, process_time_notification_time, user_id, individual_theme, notification_text, date_str, content_type, file_id, caption)
        return

    notification_id = str(len(alerts['notifications']) + 1)
    alerts['notifications'][notification_id] = {
        'theme': individual_theme,
        'text': notification_text if content_type == 'text' else None,
        'time': notification_time,
        'status': 'active',
        'category': 'time',
        'user_id': user_id,
        'files': [
            {
                'type': content_type,
                'file_id': file_id,
                'caption': caption if content_type != 'text' else None
            }
        ],
        'content_type': content_type
    }
    save_database()

    # Загрузка данных пользователя
    users_data = load_users()
    username = escape_markdown(users_data.get(user_id, {}).get('username', 'Неизвестный пользователь'))

    # Формирование сообщения с темой в нижнем регистре и выделение её жирным шрифтом
    theme = individual_theme.lower()
    formatted_time = notification_time.strftime("%d.%m.%Y в %H:%M")
    bot.send_message(message.chat.id, f"Оповещение *{theme.lower()}* запланировано на {formatted_time} для пользователя {username} - `{user_id}`!", parse_mode="Markdown")
    show_communication_menu(message)

# Обработчик для "Всем"
#@log_user_actions
@check_user_blocked
@bot.message_handler(func=lambda message: message.text == 'Всем' and check_admin_access(message))
def handle_broadcast_notifications(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Всем'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Отправить сообщение')
    markup.add('Отправленные', 'Удалить отправленные')
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Управление оповещения для всех:", reply_markup=markup)

@check_user_blocked
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Отправить сообщение' and check_admin_access(message))
def send_message_to_all(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, send_message_to_all)
        return

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Отправить сообщение'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите тему для оповещения:", reply_markup=markup)
    bot.register_next_step_handler(message, set_theme_for_broadcast)

@check_user_blocked
#@log_user_actions
def set_theme_for_broadcast(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_theme_for_broadcast)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    broadcast_theme = message.text
    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите текст оповещения или отправьте мультимедийный файл:", reply_markup=markup)
    bot.register_next_step_handler(message, process_broadcast_message, broadcast_theme)

@check_user_blocked
#@log_user_actions
def process_broadcast_message(message, broadcast_theme):

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    broadcast_text = message.text or message.caption
    content_type = message.content_type
    file_id = None
    caption = message.caption

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    users = load_users()
    user_ids = []
    for user_id in users.keys():
        if content_type == 'text':
            bot.send_message(user_id, broadcast_text)
        elif content_type == 'photo':
            bot.send_photo(user_id, file_id, caption=caption)
        elif content_type == 'video':
            bot.send_video(user_id, file_id, caption=caption)
        elif content_type == 'document':
            bot.send_document(user_id, file_id, caption=caption)
        elif content_type == 'animation':
            bot.send_animation(user_id, file_id, caption=caption)
        elif content_type == 'sticker':
            bot.send_sticker(user_id, file_id)
        elif content_type == 'audio':
            bot.send_audio(user_id, file_id, caption=caption)
        elif content_type == 'voice':
            bot.send_voice(user_id, file_id, caption=caption)
        elif content_type == 'video_note':
            bot.send_video_note(user_id, file_id)
        user_ids.append(user_id)

    # Сохраняем информацию о отправленных сообщениях
    notification_id = str(len(alerts['sent_messages']) + 1)
    alerts['sent_messages'][notification_id] = {
        'theme': broadcast_theme,
        'text': broadcast_text if content_type == 'text' else None,
        'time': datetime.now().strftime("%d.%m.%Y в %H:%M"),
        'status': 'sent',
        'category': 'all',
        'user_ids': user_ids,
        'files': [
            {
                'type': content_type,
                'file_id': file_id,
                'caption': caption if content_type != 'text' else None
            }
        ]
    }
    save_database()
    bot.send_message(message.chat.id, "Оповещение отправлено всем пользователям!")
    show_communication_menu(message)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Отправленные' and check_admin_access(message))
def show_sent_messages(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Отправленные'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if alerts['sent_messages']:  # Используем загруженный список отправленных сообщений
        sent_messages_list = [
            f"⭐ *№{i + 1}* ⭐\n\n📝 *Тема*: {msg['theme'].lower()}\n👤 *Пользователи*: {', '.join(msg.get('user_ids', []))}\n📅 *Дата*: {msg['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {msg['time'].strftime('%H:%M')}\n"
            for i, msg in enumerate(alerts['sent_messages'].values()) if msg['category'] == 'all'
        ]

        # Форматирование заголовка
        header = "Список *отправленных* оповещений:\n\n"
        max_length = 4096  # Максимальная длина сообщения в Telegram
        message_text = header

        for sent_message in sent_messages_list:
            if len(message_text) + len(sent_message) > max_length:
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                message_text = sent_message
            else:
                message_text += sent_message + "\n\n"

        if message_text:
            bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

        # Создаем клавиатуру с кнопкой "В меню админ-панели"
        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в общение")
        markup.add('В меню админ-панели')

        bot.send_message(message.chat.id, "Введите номер оповещения для просмотра:", reply_markup=markup)
        bot.register_next_step_handler(message, show_sent_message_details)
    else:
        bot.send_message(message.chat.id, "Нет отправленных оповещений!")

#@log_user_actions
def show_sent_message_details(message):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        sent_messages = list(alerts['sent_messages'].values())
        valid_indices = [index for index in indices if 0 <= index < len(sent_messages)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверные номера оповещений! Попробуйте снова")
            bot.register_next_step_handler(message, show_sent_message_details)
            return

        for index in valid_indices:
            sent_message = sent_messages[index]
            theme = sent_message['theme'].lower() if sent_message['theme'] else 'без темы'
            content_type = sent_message.get('content_type', 'текст')
            status_text = 'отправлено'

            # Форматирование даты и времени
            formatted_time = sent_message['time'].strftime("%d.%m.%Y в %H:%M")

            # Получение текста сообщения, если оно текстовое
            message_text = sent_message.get('text', '')

            sent_message_details = (
                f"*Основная информация*:\n\n"
                f"📝 *Тема*: {theme}\n"
                f"📁 *Тип контента*: {content_type}\n"
                f"📅 *Дата*: {formatted_time}\n"
                f"🔄 *Статус*: {status_text}\n"
            )

            if message_text:
                sent_message_details += f"\n*Текст сообщения*:\n\n{message_text}\n"

            bot.send_message(message.chat.id, sent_message_details, parse_mode="Markdown")

            # Отправка мультимедийного содержимого
            for file in sent_message.get('files', []):
                if file['type'] == 'photo':
                    bot.send_photo(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'video':
                    bot.send_video(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'document':
                    bot.send_document(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'animation':
                    bot.send_animation(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'sticker':
                    bot.send_sticker(message.chat.id, file['file_id'])
                elif file['type'] == 'audio':
                    bot.send_audio(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'voice':
                    bot.send_voice(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'video_note':
                    bot.send_video_note(message.chat.id, file['file_id'])

        show_communication_menu(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректные номера оповещений через запятую!")
        bot.register_next_step_handler(message, show_sent_message_details)
    except IndexError:
        bot.send_message(message.chat.id, "Ошибка: неверные номера оповещений! Попробуйте снова")
        bot.register_next_step_handler(message, show_sent_message_details)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить отправленные' and check_admin_access(message))
def delete_sent_messages(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить отправленные'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if alerts['sent_messages']:  # Используем загруженный список отправленных сообщений
        sent_messages_list = [
            f"⭐ *№{i + 1}* ⭐\n\n📝 *Тема*: {msg['theme'].lower()}\n👤 *Пользователи*: {', '.join(msg.get('user_ids', []))}\n📅 *Дата*: {msg['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {msg['time'].strftime('%H:%M')}\n"
            for i, msg in enumerate(alerts['sent_messages'].values()) if msg['category'] == 'all'
        ]

        # Форматирование заголовка
        header = "Список *отправленных* оповещений:\n\n"
        max_length = 4096  # Максимальная длина сообщения в Telegram
        message_text = header

        for sent_message in sent_messages_list:
            if len(message_text) + len(sent_message) > max_length:
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                message_text = sent_message
            else:
                message_text += sent_message + "\n\n"

        if message_text:
            bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

        # Создаем клавиатуру с кнопкой "В меню админ-панели"
        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в общение")
        markup.add('В меню админ-панели')

        bot.send_message(message.chat.id, "Введите номер оповещения для удаления:", reply_markup=markup)
        bot.register_next_step_handler(message, process_delete_sent_message)
    else:
        bot.send_message(message.chat.id, "Нет отправленных оповещений!")

#@log_user_actions
def process_delete_sent_message(message):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        sent_messages = list(alerts['sent_messages'].values())
        valid_indices = [index for index in indices if 0 <= index < len(sent_messages)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверные номера оповещений! Попробуйте снова")
            bot.register_next_step_handler(message, process_delete_sent_message)
            return

        deleted_messages = []
        for index in sorted(valid_indices, reverse=True):
            notification_id = list(alerts['sent_messages'].keys())[index]
            deleted_message = alerts['sent_messages'].pop(notification_id)
            deleted_messages.append(deleted_message)

        # Перенумерация оповещений после удаления
        new_sent_messages = {}
        for i, (key, value) in enumerate(alerts['sent_messages'].items(), start=1):
            new_sent_messages[str(i)] = value
        alerts['sent_messages'] = new_sent_messages

        save_database()  # Сохраняем изменения после удаления

        # Собираем все темы удаленных сообщений в одну строку
        deleted_themes = ", ".join([f"*{msg['theme'].lower()}*" for msg in deleted_messages])
        bot.send_message(message.chat.id, f"Оповещения (всем) по темам {deleted_themes} были удалены!", parse_mode="Markdown")

        show_communication_menu(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректные номера оповещений через запятую!")
        bot.register_next_step_handler(message, process_delete_sent_message)
    except IndexError:
        bot.send_message(message.chat.id, "Ошибка: неверные номера оповещений! Попробуйте снова")
        bot.register_next_step_handler(message, process_delete_sent_message)


# Обработчик для "Отдельно"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Отдельно' and check_admin_access(message))
def handle_individual_notifications(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Отдельно'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Отправить отдельно')
    markup.add('Посмотреть отдельно', 'Удалить отдельно')
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Управление оповещениями для отдельных пользователей:", reply_markup=markup)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Отправить отдельно' and check_admin_access(message))
def send_message_to_individual(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Отправить отдельно'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    list_users(message)

#@log_user_actions
def list_users(message):
    users_data = load_users()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *разблокирован* ✅" if not data.get('blocked', False) else " - *заблокирован* 🚫"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер пользователя для отправки сообщения:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_user_for_send)

#@log_user_actions
def choose_user_for_send(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_user_for_send)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        index = int(message.text) - 1
        users_data = load_users()
        user_list = list(users_data.keys())
        if 0 <= index < len(user_list):
            user_id = user_list[index]
            send_individual_message(message, user_id)
        else:
            bot.send_message(message.chat.id, "Неверный номер пользователя! Попробуйте снова")
            bot.register_next_step_handler(message, choose_user_for_send)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер пользователя!")
        bot.register_next_step_handler(message, choose_user_for_send)

#@log_user_actions
def send_individual_message(message, user_id):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    bot.send_message(message.chat.id, "Введите тему для оповещения:")
    bot.register_next_step_handler(message, set_theme_for_individual_broadcast, user_id)

#@log_user_actions
def set_theme_for_individual_broadcast(message, user_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_theme_for_individual_broadcast, user_id)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    broadcast_theme = message.text
    bot.send_message(message.chat.id, "Введите текст оповещения или отправьте мультимедийный файл:")
    bot.register_next_step_handler(message, process_individual_broadcast_message, user_id, broadcast_theme)

#@log_user_actions
def process_individual_broadcast_message(message, user_id, broadcast_theme):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    broadcast_text = message.text or message.caption
    content_type = message.content_type
    file_id = None
    caption = message.caption

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    if content_type == 'text':
        bot.send_message(user_id, broadcast_text)
    elif content_type == 'photo':
        bot.send_photo(user_id, file_id, caption=caption)
    elif content_type == 'video':
        bot.send_video(user_id, file_id, caption=caption)
    elif content_type == 'document':
        bot.send_document(user_id, file_id, caption=caption)
    elif content_type == 'animation':
        bot.send_animation(user_id, file_id, caption=caption)
    elif content_type == 'sticker':
        bot.send_sticker(user_id, file_id)
    elif content_type == 'audio':
        bot.send_audio(user_id, file_id, caption=caption)
    elif content_type == 'voice':
        bot.send_voice(user_id, file_id, caption=caption)
    elif content_type == 'video_note':
        bot.send_video_note(user_id, file_id)

    # Сохраняем информацию о отправленных сообщениях
    notification_id = str(len(alerts['sent_messages']) + 1)
    alerts['sent_messages'][notification_id] = {
        'theme': broadcast_theme,
        'text': broadcast_text if content_type == 'text' else None,
        'time': datetime.now().strftime("%d.%m.%Y в %H:%M"),
        'status': 'sent',
        'category': 'individual',
        'user_id': [user_id],
        'files': [
            {
                'type': content_type,
                'file_id': file_id,
                'caption': caption if content_type != 'text' else None
            }
        ]
    }
    save_database()

    # Загрузка данных пользователя
    users_data = load_users()
    username = escape_markdown(users_data.get(user_id, {}).get('username', 'Неизвестный пользователь'))

    bot.send_message(message.chat.id, f"Оповещение отправлено пользователю {username} - `{user_id}`!", parse_mode="Markdown")
    show_communication_menu(message)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Посмотреть отдельно' and check_admin_access(message))
def show_individual_messages(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Посмотреть отдельно'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    users_data = load_users()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *разблокирован* ✅" if not data.get('blocked', False) else " - *заблокирован* 🚫"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер пользователя для просмотра:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_user_for_view)

#@log_user_actions
def choose_user_for_view(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_user_for_view)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        users_data = load_users()
        user_list = list(users_data.keys())
        valid_indices = [index for index in indices if 0 <= index < len(user_list)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверный номер пользователя! Попробуйте снова")
            bot.register_next_step_handler(message, choose_user_for_view)
            return

        for index in valid_indices:
            user_id = user_list[index]
            view_individual_messages_for_user(message, user_id)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер пользователя!")
        bot.register_next_step_handler(message, choose_user_for_view)

#@log_user_actions
def view_individual_messages_for_user(message, user_id):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    sent_messages = [msg for msg in alerts['sent_messages'].values() if msg['category'] == 'individual' and user_id in msg.get('user_id', [])]
    if sent_messages:
        sent_messages_list = [
            f"⭐ №{i + 1} ⭐\n\n📝 *Тема*: {msg['theme'].lower()}\n👤 *Пользователи*: {', '.join(msg.get('user_id', []))}\n📅 *Дата*: {msg['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {msg['time'].strftime('%H:%M')}\n"
            for i, msg in enumerate(sent_messages)
        ]

        # Форматирование заголовка
        header = "Список *отправленных* оповещений:\n\n"
        max_length = 4096  # Максимальная длина сообщения в Telegram
        message_text = header

        for sent_message in sent_messages_list:
            if len(message_text) + len(sent_message) > max_length:
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                message_text = sent_message
            else:
                message_text += sent_message + "\n\n"

        if message_text:
            bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add("Вернуться в общение")
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите номер оповещения для просмотра:", reply_markup=markup)
        bot.register_next_step_handler(message, show_individual_message_details, user_id)
    else:
        bot.send_message(message.chat.id, "Нет отправленных оповещений для этого пользователя!")
        show_communication_menu(message)

#@log_user_actions
def show_individual_message_details(message, user_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, show_individual_message_details, user_id)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        sent_messages = [msg for msg in alerts['sent_messages'].values() if msg['category'] == 'individual' and user_id in msg.get('user_id', [])]
        valid_indices = [index for index in indices if 0 <= index < len(sent_messages)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверный номер оповещения! Попробуйте снова")
            bot.register_next_step_handler(message, show_individual_message_details, user_id)
            return

        for index in valid_indices:
            sent_message = sent_messages[index]
            theme = sent_message['theme'].lower() if sent_message['theme'] else 'без темы'
            content_type = sent_message.get('content_type', 'текст')
            status_text = 'отправлено' if sent_message.get('status') == 'sent' else 'активно'

            # Форматирование даты и времени
            formatted_time = sent_message['time'].strftime("%d.%m.%Y в %H:%M")

            # Получение текста сообщения, если оно текстовое
            message_text = sent_message.get('text', '')

            sent_message_details = (
                f"*Основная информация*:\n\n"
                f"📝 *Тема*: {theme}\n"
                f"📁 *Тип контента*: {content_type}\n"
                f"📅 *Дата*: {formatted_time}\n"
                f"🔄 *Статус*: {status_text}\n"
            )

            if message_text:
                sent_message_details += f"\n*Текст сообщения*:\n\n{message_text}\n"

            bot.send_message(message.chat.id, sent_message_details, parse_mode="Markdown")

            # Отправка мультимедийного содержимого
            for file in sent_message.get('files', []):
                if file['type'] == 'photo':
                    bot.send_photo(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'video':
                    bot.send_video(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'document':
                    bot.send_document(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'animation':
                    bot.send_animation(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'sticker':
                    bot.send_sticker(message.chat.id, file['file_id'])
                elif file['type'] == 'audio':
                    bot.send_audio(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'voice':
                    bot.send_voice(message.chat.id, file['file_id'], caption=file.get('caption'))
                elif file['type'] == 'video_note':
                    bot.send_video_note(message.chat.id, file['file_id'])

        show_communication_menu(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер оповещения!")
        bot.register_next_step_handler(message, show_individual_message_details, user_id)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить отдельно' and check_admin_access(message))
def delete_individual_messages(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить отдельно'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    users_data = load_users()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *разблокирован* ✅" if not data.get('blocked', False) else " - *заблокирован* 🚫"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add("Вернуться в общение")
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер пользователя для удаления:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_user_for_delete)

#@log_user_actions
def choose_user_for_delete(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_user_for_delete)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        index = int(message.text) - 1
        users_data = load_users()
        user_list = list(users_data.keys())
        if 0 <= index < len(user_list):
            user_id = user_list[index]
            delete_individual_messages_for_user(message, user_id)
        else:
            bot.send_message(message.chat.id, "Неверный номер пользователя! Попробуйте снова")
            bot.register_next_step_handler(message, choose_user_for_delete)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер пользователя!")
        bot.register_next_step_handler(message, choose_user_for_delete)

#@log_user_actions
def delete_individual_messages_for_user(message, user_id):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    sent_messages = [msg for msg in alerts['sent_messages'].values() if msg['category'] == 'individual' and user_id in msg.get('user_id', [])]
    if sent_messages:
        sent_messages_list = [
            f"⭐ №{i + 1} ⭐\n\n📝 *Тема*: {msg['theme'].lower()}\n👤 *Пользователи*: {', '.join(msg.get('user_id', []))}\n📅 *Дата*: {msg['time'].strftime('%d.%m.%Y')}\n🕒 *Время*: {msg['time'].strftime('%H:%M')}\n"
            for i, msg in enumerate(sent_messages)
        ]

        # Форматирование заголовка
        header = "Список *отправленных* оповещений:\n\n"
        max_length = 4096  # Максимальная длина сообщения в Telegram
        message_text = header

        for sent_message in sent_messages_list:
            if len(message_text) + len(sent_message) > max_length:
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                message_text = sent_message
            else:
                message_text += sent_message + "\n\n"

        if message_text:
            bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

        bot.send_message(message.chat.id, "Введите номер оповещения для удаления:")
        bot.register_next_step_handler(message, process_delete_individual_message, user_id)
    else:
        bot.send_message(message.chat.id, "Нет отправленных оповещений для этого пользователя!")
        show_communication_menu(message)

#@log_user_actions
def process_delete_individual_message(message, user_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, process_delete_individual_message, user_id)
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(index.strip()) - 1 for index in message.text.split(',')]
        sent_messages = [msg for msg in alerts['sent_messages'].values() if msg['category'] == 'individual' and user_id in msg.get('user_id', [])]
        valid_indices = [index for index in indices if 0 <= index < len(sent_messages)]

        if len(valid_indices) != len(indices):
            bot.send_message(message.chat.id, "Неверный номер оповещения! Попробуйте снова")
            bot.register_next_step_handler(message, process_delete_individual_message, user_id)
            return

        deleted_messages = []
        for index in sorted(valid_indices, reverse=True):
            notification_id = list(alerts['sent_messages'].keys())[index]
            deleted_message = alerts['sent_messages'].pop(notification_id)
            deleted_messages.append(deleted_message)

        # Перенумерация оповещений после удаления
        new_sent_messages = {}
        for i, (key, value) in enumerate(alerts['sent_messages'].items(), start=1):
            new_sent_messages[str(i)] = value
        alerts['sent_messages'] = new_sent_messages

        save_database()  # Сохраняем изменения после удаления

        # Собираем все темы удаленных сообщений в одну строку
        deleted_themes = ", ".join([f"*{msg['theme'].lower()}*" for msg in deleted_messages])
        bot.send_message(message.chat.id, f"Оповещения (всем) по темам {deleted_themes} были удалены!", parse_mode="Markdown")

        show_communication_menu(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер оповещения!")
        bot.register_next_step_handler(message, process_delete_individual_message, user_id)





# -----------------ДЛЯ РЕКЛАМЫ------------


# Путь к файлу
ADVERTISEMENT_PATH = 'data base/admin/chats/advertisement.json'

# Глобальные переменные
advertisements = {}
temp_advertisement = {
    'text': None,
    'caption': None,
    'files': [],
    'chat_id': None
}

# Сохранение данных рекламы
def save_advertisements():
    with open(ADVERTISEMENT_PATH, 'w', encoding='utf-8') as file:
        json.dump(advertisements, file, ensure_ascii=False, indent=4)

def load_advertisements():
    if os.path.exists(ADVERTISEMENT_PATH):
        with open(ADVERTISEMENT_PATH, 'r', encoding='utf-8') as file:
            data = json.load(file)
            for adv in data['advertisements'].values():
                if 'expected_date' in adv and 'expected_time' in adv:
                    adv['expected_date'] = datetime.strptime(adv['expected_date'], "%d.%m.%Y").strftime("%d.%m.%Y")
                    adv['expected_time'] = datetime.strptime(adv['expected_time'], "%H:%M").strftime("%H:%M")
                if 'end_date' in adv and 'end_time' in adv:
                    adv['end_date'] = datetime.strptime(adv['end_date'], "%d.%m.%Y").strftime("%d.%m.%Y")
                    adv['end_time'] = datetime.strptime(adv['end_time'], "%H:%M").strftime("%H:%M")
            return data
    return {"advertisements": {}}

# Инициализация данных рекламы при запуске
advertisements = load_advertisements()

# Обработчик для пользователя "Заявка на рекламу"
@bot.message_handler(func=lambda message: message.text == 'Заявка на рекламу')
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Заявка на рекламу')
@track_usage('Заявка на рекламу')  # Добавление отслеживания статистики
#@log_user_actions
def handle_advertisement_request(message):
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('Вернуться в меню для рекламы')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Введите тему рекламы и кратко о чем она:", reply_markup=markup)
    bot.register_next_step_handler(message, set_advertisement_theme)

#@log_user_actions
def set_advertisement_theme(message):

    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return

    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.content_type != 'text':
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_advertisement_theme)
        return

    advertisement_theme = message.text
    bot.send_message(message.chat.id, "Введите дату, на которую вы хотите разместить рекламу:")
    bot.register_next_step_handler(message, set_advertisement_date, advertisement_theme)

#@log_user_actions
def set_advertisement_date(message, advertisement_theme):

    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return

    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.content_type != 'text':
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_advertisement_date, advertisement_theme)
        return

    expected_date = message.text
    if expected_date is None or not validate_date_format(expected_date):
        bot.send_message(message.chat.id, "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, set_advertisement_date, advertisement_theme)
        return

    if not validate_future_date(expected_date):
        bot.send_message(message.chat.id, "Дата не может быть раньше текущей даты. Пожалуйста, введите корректную дату:")
        bot.register_next_step_handler(message, set_advertisement_date, advertisement_theme)
        return

    bot.send_message(message.chat.id, "Введите время, на которое вы хотите разместить рекламу:")
    bot.register_next_step_handler(message, set_advertisement_time, advertisement_theme, expected_date)

#@log_user_actions
def set_advertisement_time(message, advertisement_theme, expected_date):

    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return

    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.content_type != 'text':
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_advertisement_time, advertisement_theme, expected_date)
        return

    expected_time = message.text
    if not validate_time_format(expected_time):
        bot.send_message(message.chat.id, "Неверный формат времени. Пожалуйста, введите время в формате ЧЧ:ММ")
        bot.register_next_step_handler(message, set_advertisement_time, advertisement_theme, expected_date)
        return

    if not validate_future_time(expected_date, expected_time):
        bot.send_message(message.chat.id, "Время не может быть раньше текущего времени. Пожалуйста, введите корректное время:")
        bot.register_next_step_handler(message, set_advertisement_time, advertisement_theme, expected_date)
        return

    bot.send_message(message.chat.id, "Введите дату окончания действия рекламы:")
    bot.register_next_step_handler(message, set_advertisement_end_date, advertisement_theme, expected_date, expected_time)

#@log_user_actions
def set_advertisement_end_date(message, advertisement_theme, expected_date, expected_time):
    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.text == 'Вернуться в меню для рекламы':
        view_add_menu(message)
        return

    if message.content_type != 'text':
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_advertisement_end_date, advertisement_theme, expected_date, expected_time)
        return

    end_date = message.text
    if not validate_date_format(end_date):
        bot.send_message(message.chat.id, "Неверный формат даты. Пожалуйста, введите дату в формате ДД.ММ.ГГГГ")
        bot.register_next_step_handler(message, set_advertisement_end_date, advertisement_theme, expected_date, expected_time)
        return

    if not validate_future_date(end_date):
        bot.send_message(message.chat.id, "Дата окончания не может быть раньше текущей даты. Пожалуйста, введите корректную дату:")
        bot.register_next_step_handler(message, set_advertisement_end_date, advertisement_theme, expected_date, expected_time)
        return

    bot.send_message(message.chat.id, "Введите время окончания действия рекламы:")
    bot.register_next_step_handler(message, set_advertisement_end_time, advertisement_theme, expected_date, expected_time, end_date)

#@log_user_actions
def set_advertisement_end_time(message, advertisement_theme, expected_date, expected_time, end_date):
    
    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return    
    
    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.content_type != 'text':
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_advertisement_end_time, advertisement_theme, expected_date, expected_time, end_date)
        return

    end_time = message.text
    if not validate_time_format(end_time):
        bot.send_message(message.chat.id, "Неверный формат времени. Пожалуйста, введите время в формате ЧЧ:ММ:")
        bot.register_next_step_handler(message, set_advertisement_end_time, advertisement_theme, expected_date, expected_time, end_date)
        return

    bot.send_message(message.chat.id, "Отправьте текст рекламы:")
    bot.register_next_step_handler(message, collect_advertisement_text, advertisement_theme, expected_date, expected_time, end_date, end_time)

#@log_user_actions
def collect_advertisement_text(message, advertisement_theme, expected_date, expected_time, end_date, end_time):

    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return

    if message.text == 'В главное меню':
        return_to_menu(message)
        return

    if message.content_type != 'text':
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, collect_advertisement_text, advertisement_theme, expected_date, expected_time, end_date, end_time)
        return

    temp_advertisement['text'] = message.text
    temp_advertisement['chat_id'] = message.chat.id

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('Пропустить медиафайлы')
    markup.add('Вернуться в меню для рекламы')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Отправьте мультимедийные файлы (если есть):", reply_markup=markup)
    bot.register_next_step_handler(message, collect_advertisement_media, advertisement_theme, expected_date, expected_time, end_date, end_time)

#@log_user_actions
def collect_advertisement_media(message, advertisement_theme, expected_date, expected_time, end_date, end_time):
 
    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return

    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.text == "Пропустить медиафайлы":
        temp_advertisement['files'] = []  # Инициализируем 'files' пустым списком
        save_advertisement_request(message, advertisement_theme, expected_date, expected_time, end_date, end_time)
        return

    content_type = message.content_type
    file_id = None

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    if file_id:
        if 'files' not in temp_advertisement:
            temp_advertisement['files'] = []
        temp_advertisement['files'].append({
            'type': content_type,
            'file_id': file_id,
            'caption': temp_advertisement.get('text', '')  # Используем get для безопасного доступа
        })

        if len(temp_advertisement['files']) >= 10:
            save_advertisement_request(message, advertisement_theme, expected_date, expected_time, end_date, end_time)
            return

        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Добавить еще', 'Завершить отправку')
        markup.add('Вернуться в меню для рекламы')
        markup.add('В главное меню')
        bot.send_message(message.chat.id, "Медиафайл добавлен! Хотите добавить еще?", reply_markup=markup)
        bot.register_next_step_handler(message, handle_advertisement_media_options, advertisement_theme, expected_date, expected_time, end_date, end_time)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, отправьте мультимедийный файл!")
        bot.register_next_step_handler(message, collect_advertisement_media, advertisement_theme, expected_date, expected_time, end_date, end_time)

#@log_user_actions
def handle_advertisement_media_options(message, advertisement_theme, expected_date, expected_time, end_date, end_time):
    
    if message.text == 'Вернуться в меню для рекламы':
        temp_advertisement.clear()  # Очистка временных данных
        view_add_menu(message)
        return   
    
    if message.text == 'В главное меню':
        temp_advertisement.clear()  # Очистка временных данных
        return_to_menu(message)
        return

    if message.text == "Добавить еще":
        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Пропустить медиафайлы')
        markup.add('Вернуться в меню для рекламы')
        markup.add('В главное меню')
        bot.send_message(message.chat.id, "Отправьте следующий мультимедийный файл:", reply_markup=markup)
        bot.register_next_step_handler(message, collect_advertisement_media, advertisement_theme, expected_date, expected_time, end_date, end_time)
    elif message.text == "Завершить отправку":
        save_advertisement_request(message, advertisement_theme, expected_date, expected_time, end_date, end_time)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, выберите действие!")
        bot.register_next_step_handler(message, handle_advertisement_media_options, advertisement_theme, expected_date, expected_time, end_date, end_time)

#@log_user_actions
def save_advertisement_request(message, advertisement_theme, expected_date, expected_time, end_date, end_time):
    user_id = message.chat.id
    username = message.from_user.username
    advertisement_id = str(len(advertisements['advertisements']) + 1)
    advertisements['advertisements'][advertisement_id] = {
        'user_id': user_id,
        'username': username,
        'theme': advertisement_theme,
        'expected_date': expected_date,
        'expected_time': expected_time,
        'end_date': end_date,
        'end_time': end_time,
        'text': temp_advertisement['text'],
        'files': temp_advertisement['files'],
        'status': 'pending',
        'user_ids': [],
        'message_ids': []
    }

    save_advertisements()
    bot.send_message(message.chat.id, "Ваша заявка на рекламу была успешно сформирована и отправлена администратору!")

    # Отправка уведомления админам
    with open('data base/admin/admin_sessions.json', 'r', encoding='utf-8') as file:
        admin_data = json.load(file)
        admin_ids = admin_data['admin_sessions']

    for admin_id in admin_ids:
        bot.send_message(admin_id, f"У вас новая заявка на рекламу от пользователя `{user_id}` по теме *{advertisement_theme.lower()}* на {expected_date} в {expected_time} до {end_date} в {end_time}!", parse_mode="Markdown")

    temp_advertisement.clear()  # Очистка временных данных после сохранения
    return_to_menu(message)

#@log_user_actions
def schedule_advertisement_deletion(advertisement_id, end_date, end_time):
    end_datetime = datetime.strptime(f"{end_date} {end_time}", "%d.%m.%Y %H:%M")
    delay = (end_datetime - datetime.now()).total_seconds()
    threading.Timer(delay, delete_advertisement_messages, [advertisement_id]).start()

#@log_user_actions
def delete_advertisement_messages(advertisement_id):
    advertisement = advertisements['advertisements'][advertisement_id]
    user_ids = advertisement['user_ids']
    message_ids = advertisement['message_ids']

    for user_id, message_id in zip(user_ids, message_ids):
        try:
            bot.delete_message(user_id, message_id)
        except ApiTelegramException as e:
            if e.result.status_code == 400 and 'message to delete not found' in str(e.result):
                pass  # Пропустить, если сообщение не найдено
            else:
                pass  # Игнорировать остальные ошибки

    del advertisements['advertisements'][advertisement_id]
    save_advertisements()

# Функция для проверки формата даты
#@log_user_actions
def validate_date_format(date_str):
    if date_str is None:
        return False
    try:
        datetime.strptime(date_str, "%d.%m.%Y")
        return True
    except ValueError:
        return False

# Функция для проверки будущей даты
#@log_user_actions
def validate_future_date(date_str):
    today = datetime.now().date()
    input_date = datetime.strptime(date_str, "%d.%m.%Y").date()
    return input_date >= today

# Функция для проверки формата времени
#@log_user_actions
def validate_time_format(time_str):
    try:
        datetime.strptime(time_str, "%H:%M")
        return True
    except ValueError:
        return False

# Функция для проверки будущего времени
#@log_user_actions
def validate_future_time(date_str, time_str):
    now = datetime.now()
    input_datetime = datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M")
    return input_datetime >= now

# Функция для проверки срока
#@log_user_actions
def validate_duration(duration_str):
    try:
        duration = int(duration_str)
        return 1 <= duration <= 7
    except ValueError:
        return False

# Функция для обработки запросов на рекламу от администратора
#@log_user_actions
def handle_admin_advertisement_requests(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Запросы на рекламу', 'Удалить рекламу')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите действие для рекламы:", reply_markup=markup)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Запросы на рекламу' and check_admin_access(message))
def show_advertisement_requests(message):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Запросы на рекламу'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    pending_advertisements = [adv for adv in advertisements['advertisements'].values() if adv['status'] == 'pending']
    current_time = datetime.now()

    for adv in pending_advertisements:
        end_datetime = datetime.strptime(f"{adv['end_date']} {adv['end_time']}", "%d.%m.%Y %H:%M")
        if current_time >= end_datetime:
            advertisement_id = next(key for key, value in advertisements['advertisements'].items() if value == adv)
            user_id = adv['user_id']
            theme = adv['theme']
            del advertisements['advertisements'][advertisement_id]
            save_advertisements()
            bot.send_message(user_id, f"Ваша заявка по теме *{theme.lower()}* была отклонена, так как срок истек!", parse_mode="Markdown")

    pending_advertisements = [adv for adv in advertisements['advertisements'].values() if adv['status'] == 'pending']
    if pending_advertisements:
        advertisement_list = [
            f"⭐ *№{i + 1}*\n\n"
            f"👤 *Пользователь*: `{adv['user_id']}`\n"
            f"📝 *Тема*: {adv['theme'].lower()}\n"
            f"📅 *Начало*: {adv['expected_date']} в {adv['expected_time']}\n"
            f"⌛ *Конец*: {adv.get('end_date', 'N/A')} в {adv.get('end_time', 'N/A')}\n\n"
            for i, adv in enumerate(pending_advertisements)
        ]
        full_message = "*Список запросов* на рекламу:\n\n" + "\n\n".join(advertisement_list)

        # Разбиваем сообщение на части, если оно слишком длинное
        max_length = 4096
        if len(full_message) > max_length:
            parts = [full_message[i:i + max_length] for i in range(0, len(full_message), max_length)]
            for part in parts:
                bot.send_message(message.chat.id, part, parse_mode="Markdown")
        else:
            bot.send_message(message.chat.id, full_message, parse_mode="Markdown")

        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Вернуться в рекламу')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Введите номер запроса для просмотра:", reply_markup=markup)
        bot.register_next_step_handler(message, show_advertisement_request_details)
    else:
        bot.send_message(message.chat.id, "*Активных* запросов на рекламу нет!", parse_mode="Markdown")

#@log_user_actions
def show_advertisement_request_details(message):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        index = int(message.text) - 1
        advertisement_list = list(advertisements['advertisements'].values())
        if 0 <= index < len(advertisement_list):
            advertisement = advertisement_list[index]
            text = advertisement['text']

            # Основная информация о рекламе
            info_message = (
                f"⭐ *Основная информация о рекламе*:\n\n"
                f"📝 *Тема*: {advertisement['theme'].lower()}\n"
                f"📅 *Начало*: {advertisement['expected_date']} в {advertisement['expected_time']}\n"
                f"⌛ *Конец*: {advertisement.get('end_date', 'N/A')} в {advertisement.get('end_time', 'N/A')}\n\n"
            )

            bot.send_message(message.chat.id, info_message, parse_mode="Markdown")

            # Формируем текст сообщения в зависимости от наличия подписи или текста
            if text and text != 'None':
                message_text = f"📝 Текст рекламы 📝\n\n{text}"
            else:
                message_text = ""

            # Если есть медиафайлы, отправляем их вместе с текстом
            if 'files' in advertisement and advertisement['files']:
                media_group = []
                first_file = True
                for file in advertisement['files']:
                    if first_file:
                        caption = message_text
                    else:
                        caption = None
                    if file['type'] == 'photo':
                        media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption))
                    elif file['type'] == 'video':
                        media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption))
                    elif file['type'] == 'document':
                        media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption))
                    elif file['type'] == 'animation':
                        media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption))
                    elif file['type'] == 'sticker':
                        bot.send_sticker(message.chat.id, file['file_id'])
                    elif file['type'] == 'audio':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                    elif file['type'] == 'voice':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                    elif file['type'] == 'video_note':
                        bot.send_video_note(message.chat.id, file['file_id'])
                    first_file = False

                if media_group:
                    bot.send_media_group(message.chat.id, media_group)
            else:
                # Если нет медиафайлов, отправляем только текст
                if message_text:
                    bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

            markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            markup.add('Принять рекламу', 'Отклонить рекламу')
            markup.add('Вернуться в рекламу')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "Выберите действие для рекламы:", reply_markup=markup)
            bot.register_next_step_handler(message, handle_advertisement_request_action, index)
        else:
            bot.send_message(message.chat.id, "Неверный номер запроса! Попробуйте снова")
            bot.register_next_step_handler(message, show_advertisement_request_details)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер запроса!")
        bot.register_next_step_handler(message, show_advertisement_request_details)

#@log_user_actions
def handle_advertisement_request_action(message, index):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    advertisement_id = list(advertisements['advertisements'].keys())[index]
    advertisement = advertisements['advertisements'][advertisement_id]
    current_time = datetime.now()
    end_datetime = datetime.strptime(f"{advertisement['end_date']} {advertisement['end_time']}", "%d.%m.%Y %H:%M")

    if current_time >= end_datetime:
        user_id = advertisement['user_id']
        theme = advertisement['theme']
        del advertisements['advertisements'][advertisement_id]
        save_advertisements()
        bot.send_message(message.chat.id, "Реклама была отклонена, так как срок истек!")
        bot.send_message(user_id, f"Ваша заявка по теме *{theme.lower()}* была отклонена, так как срок истек!", parse_mode="Markdown")
        show_advertisement_menu(message)
        return

    if message.text == 'Принять рекламу':
        advertisements['advertisements'][advertisement_id]['status'] = 'accepted'
        save_advertisements()
        bot.send_message(message.chat.id, "Реклама была принята! Выберите действия для отправки:")
        choose_send_advertisement_action(message, advertisement_id)
    elif message.text == 'Отклонить рекламу':
        user_id = advertisement['user_id']
        theme = advertisement['theme']
        del advertisements['advertisements'][advertisement_id]
        save_advertisements()
        bot.send_message(message.chat.id, "Реклама была отклонена!")
        bot.send_message(user_id, f"Ваша заявка по теме *{theme.lower()}* была отклонена администратором!", parse_mode="Markdown")
        show_advertisement_menu(message)
    else:
        bot.send_message(message.chat.id, "Неверное действие! Попробуйте снова")
        show_advertisement_request_details(message)

#@log_user_actions
def schedule_advertisement(message, advertisement_id):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    advertisement = advertisements['advertisements'][advertisement_id]
    bot.send_message(message.chat.id, f"Тема: {advertisement['theme']}")

    # Формируем текст сообщения в зависимости от наличия подписи или текста
    if advertisement['text'] and advertisement['text'] != 'None':
        message_text = f"📝 Текст рекламы 📝\n\n{advertisement['text']}"
    else:
        message_text = ""

    # Если есть медиафайлы, отправляем их вместе с текстом
    if 'files' in advertisement and advertisement['files']:
        media_group = []
        first_file = True
        for file in advertisement['files']:
            if first_file:
                caption = message_text
            else:
                caption = None
            if file['type'] == 'photo':
                media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption))
            elif file['type'] == 'video':
                media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption))
            elif file['type'] == 'document':
                media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption))
            elif file['type'] == 'animation':
                media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption))
            elif file['type'] == 'sticker':
                bot.send_sticker(message.chat.id, file['file_id'])
            elif file['type'] == 'audio':
                media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
            elif file['type'] == 'voice':
                media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
            elif file['type'] == 'video_note':
                bot.send_video_note(message.chat.id, file['file_id'])
            first_file = False

        if media_group:
            bot.send_media_group(message.chat.id, media_group)
    else:
        # Если нет медиафайлов, отправляем только текст
        if message_text:
            bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Отправить по времени', 'Отправить всем')
    markup.add('Вернуться в рекламу')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите действие для рекламы:", reply_markup=markup)
    bot.register_next_step_handler(message, choose_send_advertisement_action, advertisement_id)


#@log_user_actions
def choose_send_advertisement_action(message, advertisement_id):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Отправить по времени', 'Отправить всем')
    markup.add('Вернуться в рекламу')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите действие для рекламы:", reply_markup=markup)
    bot.register_next_step_handler(message, handle_send_advertisement_action, advertisement_id)

#@log_user_actions
def handle_send_advertisement_action(message, advertisement_id):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    if message.text == 'Отправить по времени':
        schedule_notification(message, advertisement_id)
    elif message.text == 'Отправить всем':
        send_advertisement_to_all(message, advertisement_id)
    else:
        bot.send_message(message.chat.id, "Неверное действие! Попробуйте снова")
        choose_send_advertisement_action(message, advertisement_id)

#@log_user_actions
def schedule_notification(message, advertisement_id):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    advertisement = advertisements['advertisements'][advertisement_id]
    expected_datetime = datetime.strptime(f"{advertisement['expected_date']} {advertisement['expected_time']}", "%d.%m.%Y %H:%M")
    current_time = datetime.now()

    if current_time >= expected_datetime:
        send_advertisement_to_all(message, advertisement_id)
    else:
        delay = (expected_datetime - current_time).total_seconds()
        threading.Timer(delay, send_advertisement_to_all, [message, advertisement_id]).start()
        bot.send_message(message.chat.id, f"Реклама *{advertisement['theme'].lower()}* запланирована на {advertisement['expected_date']} в {advertisement['expected_time']}!", parse_mode="Markdown")
        show_advertisement_menu(message)

#@log_user_actions
def send_advertisement_to_all(message, advertisement_id):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    advertisement = advertisements['advertisements'][advertisement_id]
    users = load_users()
    user_ids = []
    message_ids = []

    for user_id in users.keys():
        media_group = []
        first_file = True
        for file in advertisement['files']:
            if first_file:
                caption = advertisement['text']
            else:
                caption = None
            if file['type'] == 'photo':
                media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption))
            elif file['type'] == 'video':
                media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption))
            elif file['type'] == 'document':
                media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption))
            elif file['type'] == 'animation':
                media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption))
            elif file['type'] == 'sticker':
                bot.send_sticker(user_id, file['file_id'])
            elif file['type'] == 'audio':
                media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
            elif file['type'] == 'voice':
                media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
            elif file['type'] == 'video_note':
                bot.send_video_note(user_id, file['file_id'])
            first_file = False

        if media_group:
            try:
                sent_messages = bot.send_media_group(user_id, media_group)
                for sent_message in sent_messages:
                    message_ids.append(sent_message.message_id)
                    user_ids.append(user_id)
            except Exception:
                pass  # Игнорировать ошибки отправки медиа
        else:
            try:
                sent_message = bot.send_message(user_id, advertisement['text'])
                message_ids.append(sent_message.message_id)
                user_ids.append(user_id)
            except Exception:
                pass  # Игнорировать ошибки отправки текста

    advertisement['user_ids'] = user_ids
    advertisement['message_ids'] = message_ids
    advertisement['status'] = 'accepted'  # Обновляем статус на 'accepted'
    save_advertisements()

    bot.send_message(message.chat.id, "Реклама отправлена всем пользователям!")
    show_advertisement_menu(message)

def check_advertisement_expiration():
    while True:
        now = datetime.now()
        for adv_id, adv in list(advertisements['advertisements'].items()):
            if adv['status'] == 'accepted':
                end_datetime = datetime.strptime(f"{adv['end_date']} {adv['end_time']}", "%d.%m.%Y %H:%M")
                if now >= end_datetime:
                    delete_advertisement_messages(adv_id)
        time.sleep(60)  # Проверяем каждую минуту
threading.Thread(target=check_advertisement_expiration, daemon=True).start()

def check_pending_advertisement_expiration():
    while True:
        current_time = datetime.now()
        for adv_id, adv in list(advertisements['advertisements'].items()):
            if adv['status'] == 'pending':
                end_datetime = datetime.strptime(f"{adv['end_date']} {adv['end_time']}", "%d.%m.%Y %H:%M")
                if current_time >= end_datetime:
                    user_id = adv['user_id']
                    theme = adv['theme']
                    del advertisements['advertisements'][adv_id]
                    save_advertisements()
                    bot.send_message(user_id, f"Ваша заявка по теме *{theme.lower()}* была отклонена, так как срок истек!", parse_mode="Markdown")
        time.sleep(60)  # Проверяем каждую минуту

# Запуск фоновой проверки
threading.Thread(target=check_pending_advertisement_expiration, daemon=True).start()

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить рекламу' and check_admin_access(message))
def delete_advertisement(message):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить рекламу'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if advertisements['advertisements']:
        advertisement_list = [
            f"⭐️ №{i + 1}\n\n"
            f"👤 *Пользователь*: `{adv['user_id']}`\n"
            f"📝 *Тема*: {adv['theme'].lower()}\n"
            f"📅 *Начало*: {adv['expected_date']} в {adv['expected_time']}\n"
            f"⌛️ *Конец*: {adv.get('end_date', 'N/A')} в {adv.get('end_time', 'N/A')}\n"
            for i, adv in enumerate(advertisements['advertisements'].values()) if adv['status'] == 'accepted'
        ]
        if advertisement_list:
            full_message = "*Список опубликованных реклам*:\n\n" + "\n\n".join(advertisement_list)

            # Разбиваем сообщение на части, если оно слишком длинное
            max_length = 4096
            if len(full_message) > max_length:
                parts = [full_message[i:i + max_length] for i in range(0, len(full_message), max_length)]
                for part in parts:
                    bot.send_message(message.chat.id, part, parse_mode="Markdown")
            else:
                bot.send_message(message.chat.id, full_message, parse_mode="Markdown")

            markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('Вернуться в рекламу')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "Введите номер рекламы для удаления:", reply_markup=markup)
            bot.register_next_step_handler(message, process_delete_advertisement)
        else:
            bot.send_message(message.chat.id, "У вас нет *опубликованных* реклам!", parse_mode="Markdown")
    else:
        bot.send_message(message.chat.id, "Нет *опубликованных* реклам!", parse_mode="Markdown")

#@log_user_actions
def process_delete_advertisement(message):
    if message.text == "Вернуться в рекламу":
        show_advertisement_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        index = int(message.text) - 1
        advertisement_list = list(advertisements['advertisements'].values())
        if 0 <= index < len(advertisement_list):
            advertisement_id = list(advertisements['advertisements'].keys())[index]
            delete_advertisement_messages(advertisement_id)
            bot.send_message(message.chat.id, f"Реклама удалена!")
            show_advertisement_menu(message)
        else:
            bot.send_message(message.chat.id, "Неверный номер рекламы! Попробуйте снова")
            delete_advertisement(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер рекламы!")
        delete_advertisement(message)

# Обработчик для кнопки "РЕКЛАМА"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Реклама' and check_admin_access(message))
def show_advertisement_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Реклама'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    handle_admin_advertisement_requests(message)


@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Ваши заявки')
@track_usage('Ваши заявки')  # Добавление отслеживания статистики
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Ваши заявки')
def show_user_advertisement_requests(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, show_user_advertisement_requests)
        return

    user_id = message.chat.id
    user_advertisements = [adv for adv in advertisements['advertisements'].values() if adv['user_id'] == user_id]

    if user_advertisements:
        advertisement_list = [
            f"⭐ №{i + 1}\n\n"
            f"📝 *Тема*: {adv['theme'].lower()}\n"
            f"📅 *Начало*: {adv['expected_date']} в {adv['expected_time']}\n"
            f"⌛ *Конец*: {adv.get('end_date', 'N/A')} в {adv.get('end_time', 'N/A')}\n"
            for i, adv in enumerate(user_advertisements)
        ]
        full_message = "*Ваши заявки на рекламу*:\n\n" + "\n\n".join(advertisement_list)

        # Разбиваем сообщение на части, если оно слишком длинное
        max_length = 4096
        if len(full_message) > max_length:
            parts = [full_message[i:i + max_length] for i in range(0, len(full_message), max_length)]
            for part in parts:
                bot.send_message(message.chat.id, part, parse_mode="Markdown")
        else:
            bot.send_message(message.chat.id, full_message, parse_mode="Markdown")

        markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Вернуться в меню для рекламы')
        markup.add('В главное меню')
        bot.send_message(message.chat.id, "Введите номер заявки для просмотра:", reply_markup=markup)
        bot.register_next_step_handler(message, show_user_advertisement_request_details)
    else:
        bot.send_message(message.chat.id, "*У вас нет активных заявок на рекламу!*", parse_mode="Markdown")

#@log_user_actions
def show_user_advertisement_request_details(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, show_user_advertisement_request_details)
        return

    if message.text == "Вернуться в меню для рекламы":
        view_add_menu(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    try:
        index = int(message.text) - 1
        user_id = message.chat.id
        user_advertisements = [adv for adv in advertisements['advertisements'].values() if adv['user_id'] == user_id]

        if 0 <= index < len(user_advertisements):
            advertisement = user_advertisements[index]
            text = advertisement['text']

            # Основная информация о рекламе
            info_message = (
                f"⭐ *Основная информация о рекламе*:\n\n"
                f"📝 *Тема*: {advertisement['theme'].lower()}\n"
                f"📅 *Начало*: {advertisement['expected_date']} в {advertisement['expected_time']}\n"
                f"⌛ *Конец*: {advertisement.get('end_date', 'N/A')} в {advertisement.get('end_time', 'N/A')}\n\n"
            )

            bot.send_message(message.chat.id, info_message, parse_mode="Markdown")

            # Формируем текст сообщения в зависимости от наличия подписи или текста
            if text and text != 'None':
                message_text = f"📝 Текст рекламы 📝\n\n{text}"
            else:
                message_text = ""

            # Если есть медиафайлы, отправляем их вместе с текстом
            if 'files' in advertisement and advertisement['files']:
                media_group = []
                first_file = True
                for file in advertisement['files']:
                    if first_file:
                        caption = message_text
                    else:
                        caption = None
                    if file['type'] == 'photo':
                        media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption))
                    elif file['type'] == 'video':
                        media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption))
                    elif file['type'] == 'document':
                        media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption))
                    elif file['type'] == 'animation':
                        media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption))
                    elif file['type'] == 'sticker':
                        bot.send_sticker(message.chat.id, file['file_id'])
                    elif file['type'] == 'audio':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                    elif file['type'] == 'voice':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                    elif file['type'] == 'video_note':
                        bot.send_video_note(message.chat.id, file['file_id'])
                    first_file = False

                if media_group:
                    bot.send_media_group(message.chat.id, media_group)
            else:
                # Если нет медиафайлов, отправляем только текст
                if message_text:
                    bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

            markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('Отозвать рекламу')
            markup.add('Вернуться в меню для рекламы')
            markup.add('В главное меню')
            bot.send_message(message.chat.id, "Выберите отозвать заявку или вернуться в меню:", reply_markup=markup)
            bot.register_next_step_handler(message, handle_user_advertisement_request_action, index)
        else:
            bot.send_message(message.chat.id, "Неверный номер заявки! Попробуйте снова")
            show_user_advertisement_requests(message)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер заявки!")
        show_user_advertisement_requests(message)

#@log_user_actions
def handle_user_advertisement_request_action(message, index):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, handle_user_advertisement_request_action, index)
        return

    if message.text == "Вернуться в меню для рекламы":
        view_add_menu(message)
        return

    if message.text == "В главное меню":
        return_to_menu(message)
        return

    user_id = message.chat.id
    user_advertisements = [adv for adv in advertisements['advertisements'].values() if adv['user_id'] == user_id]
    advertisement_id = list(advertisements['advertisements'].keys())[index]
    advertisement = user_advertisements[index]

    if message.text == 'Отозвать рекламу':
        del advertisements['advertisements'][advertisement_id]
        save_advertisements()
        bot.send_message(message.chat.id, "Ваша заявка была успешно отозвана!")

        # Отправка уведомления админам
        with open('data base/admin/admin_sessions.json', 'r', encoding='utf-8') as file:
            admin_data = json.load(file)
            admin_ids = admin_data['admin_sessions']

        for admin_id in admin_ids:
            bot.send_message(admin_id, f"Заявка на рекламу от пользователя `{user_id}` по теме *{advertisement['theme'].lower()}* была отозвана!", parse_mode="Markdown")

        return_to_menu(message)
    else:
        bot.send_message(message.chat.id, "Неверное действие! Попробуйте снова")
        show_user_advertisement_request_details(message)





@bot.message_handler(func=lambda message: message.text == "Для рекламы")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Для рекламы')
@track_usage('Для рекламы')  # Добавление отслеживания статистики
#@log_user_actions
def view_add_menu(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Заявка на рекламу', 'Ваши заявки')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите действие с рекламой:", reply_markup=markup)




@bot.message_handler(func=lambda message: message.text == "Прочее")
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Прочее')
@track_usage('Прочее')  # Добавление отслеживания статистики
#@log_user_actions
def view_reminders(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Чат с админом', 'Для рекламы', 'Новости')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите действие из прочего:", reply_markup=markup)





# Путь к файлу
NEWS_DATABASE_PATH = 'data base/admin/chats/news.json'
ADMIN_SESSIONS_FILE = 'data base/admin/admin_sessions.json'
USER_DATA_PATH = 'data base/admin/users.json'  # Путь к файлу с данными пользователей

# Глобальные переменные
news = {}  # Словарь для хранения новостей
admin_sessions = []

# Загрузка пользователей
def load_users():
    if os.path.exists(USER_DATA_PATH):
        with open(USER_DATA_PATH, 'r') as file:
            return json.load(file)
    return {}

# Загрузка базы данных новостей
def save_news_database():
    for key, value in news.items():
        if 'time' in value and isinstance(value['time'], datetime):
            value['time'] = value['time'].strftime("%d.%m.%Y в %H:%M")
    with open(NEWS_DATABASE_PATH, 'w', encoding='utf-8') as file:
        json.dump(news, file, ensure_ascii=False, indent=4)
    for key, value in news.items():
        if 'time' in value and isinstance(value['time'], str):
            value['time'] = datetime.strptime(value['time'], "%d.%m.%Y в %H:%M")

def load_news_database():
    if os.path.exists(NEWS_DATABASE_PATH):
        with open(NEWS_DATABASE_PATH, 'r', encoding='utf-8') as file:
            data = json.load(file)
            for key, value in data.items():
                if 'time' in value and value['time']:
                    value['time'] = datetime.strptime(value['time'], "%d.%m.%Y в %H:%M")
            return data
    return {}

# Инициализация базы данных при запуске
news = load_news_database()

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    if os.path.exists(ADMIN_SESSIONS_FILE):
        with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
            data = json.load(file)
        return data.get('admin_sessions', [])
    return []

admin_sessions = load_admin_sessions()

#@log_user_actions
def check_admin_access(message):
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "У вас нет прав доступа для выполнения этой операции!")
        return False

# Показ меню новостей для пользователя
@bot.message_handler(func=lambda message: message.text == 'Новости')
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('Новости')
@track_usage('Новости')
#@log_user_actions
def show_news_menu(message):
    markup = telebot.types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True)
    markup.add('3 новости', '5 новостей', '7 новостей')
    markup.add('10 новостей', '15 новостей')
    markup.add('В главное меню')
    bot.send_message(message.chat.id, "Выберите количество новостей:", reply_markup=markup)

# Обработчик для выбора количества новостей
@bot.message_handler(func=lambda message: message.text in ['3 новости', '5 новостей', '7 новостей', '10 новостей', '15 новостей'])
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('3 новости')
@check_function_state_decorator('5 новостей')
@check_function_state_decorator('7 новостей')
@check_function_state_decorator('10 новостей')
@check_function_state_decorator('15 новостей')
@track_usage('3 новости')
@track_usage('5 новостей')
@track_usage('7 новостей')
@track_usage('10 новостей')
@track_usage('15 новостей')
#@log_user_actions
def handle_news_selection(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_news_selection)
        return

    count = int(message.text.split()[0])
    news_list = sorted(news.values(), key=lambda x: x['time'], reverse=True)

    if len(news_list) == 0:
        bot.send_message(message.chat.id, "Новостей нет!")
        return_to_menu(message)
        return

    for i in range(min(count, len(news_list))):
        news_item = news_list[i]
        if 'files' in news_item and news_item['files']:
            media_group = []
            caption = None
            for file in news_item['files']:
                if file.get('caption'):
                    caption = file['caption']
                    break

            first_file = True
            for file in news_item['files']:
                if file['type'] == 'photo':
                    media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption if first_file else None))
                elif file['type'] == 'video':
                    media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption if first_file else None))
                elif file['type'] == 'document':
                    media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption if first_file else None))
                elif file['type'] == 'animation':
                    media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption if first_file else None))
                elif file['type'] == 'sticker':
                    bot.send_sticker(message.chat.id, file['file_id'])
                elif file['type'] == 'audio':
                    media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption if first_file else None))
                elif file['type'] == 'voice':
                    media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption if first_file else None))
                elif file['type'] == 'video_note':
                    bot.send_video_note(message.chat.id, file['file_id'])
                first_file = False
            if media_group:
                bot.send_media_group(message.chat.id, media_group)
        if news_item['text']:
            bot.send_message(message.chat.id, news_item['text'])

    if len(news_list) > count:
        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('Еще новости')
        markup.add('В главное меню')
        bot.send_message(message.chat.id, "Хотите посмотреть Еще новости?", reply_markup=markup)
        bot.register_next_step_handler(message, handle_more_news, count)
    else:
        bot.send_message(message.chat.id, "Новости закончились!")
        return_to_menu(message)

#@log_user_actions
def handle_more_news(message, start_index):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, handle_more_news, start_index)
        return

    if message.text == 'Еще новости':
        news_list = sorted(news.values(), key=lambda x: x['time'], reverse=True)
        if start_index >= len(news_list):
            bot.send_message(message.chat.id, "Новости закончились!")
            return_to_menu(message)
            return

        end_index = min(start_index + 3, len(news_list))
        for i in range(start_index, end_index):
            news_item = news_list[i]
            if 'files' in news_item and news_item['files']:
                media_group = []
                for file in news_item['files']:
                    if file['type'] == 'photo':
                        media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=file.get('caption', "")))
                    elif file['type'] == 'video':
                        media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=file.get('caption', "")))
                    elif file['type'] == 'document':
                        media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=file.get('caption', "")))
                    elif file['type'] == 'animation':
                        media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=file.get('caption', "")))
                    elif file['type'] == 'sticker':
                        bot.send_sticker(message.chat.id, file['file_id'])
                    elif file['type'] == 'audio':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=file.get('caption', "")))
                    elif file['type'] == 'voice':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=file.get('caption', "")))
                    elif file['type'] == 'video_note':
                        bot.send_video_note(message.chat.id, file['file_id'])
                if media_group:
                    bot.send_media_group(message.chat.id, media_group)
            if news_item['text']:
                bot.send_message(message.chat.id, news_item['text'])

        if end_index < len(news_list):
            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('Еще новости')
            markup.add('В главное меню')
            bot.send_message(message.chat.id, "Хотите посмотреть Еще новости?", reply_markup=markup)
            bot.register_next_step_handler(message, handle_more_news, end_index)
        else:
            bot.send_message(message.chat.id, "Новости закончились!")
            return_to_menu(message)
    else:
        return_to_menu(message)

# Показ меню редакции для админа
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Редакция' and check_admin_access(message))
def show_editorial_menu(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Редакция'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Опубликовать новость', 'Отредактировать новость', 'Посмотреть новость', 'Удалить новость')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите действие из редакции:", reply_markup=markup)

# Обработчик для "Опубликовать новость"

# Глобальные переменные для хранения временных данных новости
temp_news = {}

# Обработчик для "Опубликовать новость"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Опубликовать новость' and check_admin_access(message))
def publish_news(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Опубликовать новость'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В редакцию')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите заголовок новости:", reply_markup=markup)
    bot.register_next_step_handler(message, set_news_title)

#@log_user_actions
def set_news_title(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, set_news_title)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    news_title = message.text.capitalize()  # Преобразуем заголовок в правильный регистр
    temp_news['title'] = news_title
    temp_news['files'] = []
    temp_news['chat_id'] = message.chat.id  # Сохраняем chat_id для отправки сообщений

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В редакцию')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите подпись для новости:", reply_markup=markup)
    bot.register_next_step_handler(message, collect_news_caption)

#@log_user_actions
def collect_news_caption(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, collect_news_caption)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    caption = message.text
    temp_news['caption'] = caption

    markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('Пропустить медиафайлы')
    markup.add('В редакцию')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Отправьте мультимедийные файлы или пропустите отправку:", reply_markup=markup)
    bot.register_next_step_handler(message, collect_news_media)

#@log_user_actions
def collect_news_media(message):
    if message.text == "Пропустить медиафайлы":
        temp_news['text'] = temp_news['caption']
        temp_news['caption'] = None
        save_news(message)
        return

    content_type = message.content_type
    file_id = None

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    if file_id:
        temp_news['files'].append({
            'type': content_type,
            'file_id': file_id,
            'caption': temp_news['caption']
        })

        if len(temp_news['files']) >= 10:
            save_news(message)
            return

        markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Добавить еще', 'Завершить отправку')
        markup.add('В редакцию')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Медиафайл добавлен! Хотите добавить еще?", reply_markup=markup)
        bot.register_next_step_handler(message, handle_media_options)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, отправьте мультимедийный файл!")
        bot.register_next_step_handler(message, collect_news_media)

#@log_user_actions
def handle_media_options(message):
    if message.text == "Добавить еще":
        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('В редакцию')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Отправьте следующий мультимедийный файл:", reply_markup=markup)
        bot.register_next_step_handler(message, collect_news_media)
    elif message.text == "Завершить отправку":
        save_news(message)
    elif message.text == "В редакцию":
        show_editorial_menu(message)
    elif message.text == "В меню админ-панели":
        show_admin_panel(message)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, выберите действие!")
        bot.register_next_step_handler(message, handle_media_options)

#@log_user_actions
def save_news(message):
    news_id = str(len(news) + 1)
    news[news_id] = {
        'title': temp_news['title'],
        'text': temp_news.get('text'),
        'time': datetime.now(),
        'files': temp_news['files']
    }
    save_news_database()
    bot.send_message(temp_news['chat_id'], "Новость опубликована!")
    temp_news.clear()
    show_editorial_menu(message)  # Возвращаемся в меню новостей

# Обработчик для "Отредактировать новость"

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Отредактировать новость' and check_admin_access(message))
def edit_news(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Отредактировать новость'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    news_list = [
        f"📰 №{i + 1}. *{n['title']}* - {n['time'].strftime('%d.%m.%Y в %H:%M')}"
        for i, n in enumerate(news.values())
    ]

    if news_list:
        news_text = "\n\n".join(news_list)
        chunks = split_text(news_text)
        for chunk in chunks:
            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('В редакцию')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "📜 *Список новостей*:\n\n\n" + chunk, parse_mode="Markdown", reply_markup=markup)
        bot.send_message(message.chat.id, "Введите номер новости для редактирования:", reply_markup=markup)
        bot.register_next_step_handler(message, choose_news_to_edit)
    else:
        bot.send_message(message.chat.id, "Нет новостей для редактирования!")

#@log_user_actions
def choose_news_to_edit(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_news_to_edit)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        index = int(message.text) - 1
        news_list = list(news.values())
        if 0 <= index < len(news_list):
            news_id = list(news.keys())[index]
            news_item = news[news_id]

            # Извлекаем подпись из первого файла, если есть
            caption = news_item['files'][0]['caption'] if 'files' in news_item and news_item['files'] else None

            # Формируем текст сообщения в зависимости от наличия подписи или текста
            if caption:
                message_text = f"📌 ЗАГОЛОВОК НОВОСТИ 📌\n\n\n{news_item['title']}\n\n\n📢 ПОДПИСЬ НОВОСТИ 📢\n\n\n{caption}"
            elif news_item.get('text'):
                message_text = f"📌 ЗАГОЛОВОК НОВОСТИ 📌\n\n\n{news_item['title']}\n\n\n📝 ТЕКСТ НОВОСТИ 📝\n\n\n{news_item['text']}"
            else:
                message_text = f"📌 ЗАГОЛОВОК НОВОСТИ 📌\n\n\n{news_item['title']}"

            # Если есть медиафайлы, отправляем их вместе с текстом
            if 'files' in news_item and news_item['files']:
                media_group = []
                first_file = True
                for file in news_item['files']:
                    if first_file:
                        caption = message_text
                    else:
                        caption = None
                    if file['type'] == 'photo':
                        media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption))
                    elif file['type'] == 'video':
                        media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption))
                    elif file['type'] == 'document':
                        media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption))
                    elif file['type'] == 'animation':
                        media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption))
                    elif file['type'] == 'sticker':
                        bot.send_sticker(message.chat.id, file['file_id'])
                    elif file['type'] == 'audio':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                    elif file['type'] == 'voice':
                        media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                    elif file['type'] == 'video_note':
                        bot.send_video_note(message.chat.id, file['file_id'])
                    first_file = False

                if media_group:
                    bot.send_media_group(message.chat.id, media_group)
            else:
                # Если нет медиафайлов, отправляем только текст
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

            # Формируем нумерованный список для редактирования
            edit_list = []
            edit_list.append(f"№1. Заголовок новости")
            if news_item['text']:
                edit_list.append(f"№2. Текст новости")
            if 'files' in news_item and news_item['files']:
                edit_list.append(f"№3. Подпись к новости")
                edit_list.append("№4. Медиафайлы")

            edit_text = "\n".join(edit_list)
            bot.send_message(message.chat.id, "Что вы хотите отредактировать?\n\n" + edit_text)

            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('В редакцию')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "Введите номер пункта для редактирования:", reply_markup=markup)
            bot.register_next_step_handler(message, edit_news_item, news_id)
        else:
            bot.send_message(message.chat.id, "Неверный номер новости! Попробуйте снова")
            bot.register_next_step_handler(message, choose_news_to_edit)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер новости!")
        bot.register_next_step_handler(message, choose_news_to_edit)

#@log_user_actions
def edit_news_item(message, news_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, edit_news_item, news_id)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        choice = int(message.text)
        news_item = news[news_id]

        if choice == 1:
            # Редактирование заголовка новости
            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('В редакцию')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "Введите новый заголовок новости:", reply_markup=markup)
            bot.register_next_step_handler(message, edit_news_title, news_id)
        elif choice == 2:
            if 'text' in news_item and news_item['text']:
                # Редактирование текста новости
                markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
                markup.add('В редакцию')
                markup.add('В меню админ-панели')
                bot.send_message(message.chat.id, "Введите новый текст новости:", reply_markup=markup)
                bot.register_next_step_handler(message, edit_news_text, news_id)
            else:
                bot.send_message(message.chat.id, "Текст новости отсутствует! Выберите другой пункт для редактирования:")
                bot.register_next_step_handler(message, edit_news_item, news_id)
        elif choice == 3:
            if 'files' in news_item and news_item['files']:
                # Редактирование подписи к новости
                markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
                markup.add('В редакцию')
                markup.add('В меню админ-панели')
                bot.send_message(message.chat.id, "Введите новую подпись к новости:", reply_markup=markup)
                bot.register_next_step_handler(message, edit_news_caption, news_id)
            else:
                bot.send_message(message.chat.id, "Подпись к новости отсутствует! Выберите другой пункт для редактирования:")
                bot.register_next_step_handler(message, edit_news_item, news_id)
        elif choice == 4:
            if 'files' in news_item and news_item['files']:
                # Редактирование медиафайлов
                markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
                markup.add('В редакцию')
                markup.add('В меню админ-панели')
                bot.send_message(message.chat.id, "Отправьте новые медиафайлы:", reply_markup=markup)
                bot.register_next_step_handler(message, edit_news_media, news_id)
            else:
                bot.send_message(message.chat.id, "Медиафайлы отсутствуют! Выберите другой пункт для редактирования:")
                bot.register_next_step_handler(message, edit_news_item, news_id)
        else:
            bot.send_message(message.chat.id, "Неверный номер пункта! Попробуйте снова")
            bot.register_next_step_handler(message, edit_news_item, news_id)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректный номер пункта!")
        bot.register_next_step_handler(message, edit_news_item, news_id)

#@log_user_actions
def edit_news_title(message, news_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, edit_news_title, news_id)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    news_title = message.text
    news[news_id]['title'] = news_title
    save_news_database()
    bot.send_message(message.chat.id, "Заголовок новости отредактирован!")
    show_editorial_menu(message)

#@log_user_actions
def edit_news_text(message, news_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, edit_news_text, news_id)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    news_text = message.text
    news[news_id]['text'] = news_text
    save_news_database()
    bot.send_message(message.chat.id, "Текст новости отредактирован!")
    show_editorial_menu(message)

#@log_user_actions
def edit_news_caption(message, news_id):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, edit_news_caption, news_id)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    caption = message.text
    for file in news[news_id]['files']:
        file['caption'] = caption
    save_news_database()
    bot.send_message(message.chat.id, "Подпись новости отредактирована!")
    show_editorial_menu(message)

#@log_user_actions
def edit_news_media(message, news_id):
    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    content_type = message.content_type
    file_id = None

    if content_type == 'photo':
        file_id = message.photo[-1].file_id
    elif content_type == 'video':
        file_id = message.video.file_id
    elif content_type == 'document':
        file_id = message.document.file_id
    elif content_type == 'animation':
        file_id = message.animation.file_id
    elif content_type == 'sticker':
        file_id = message.sticker.file_id
    elif content_type == 'audio':
        file_id = message.audio.file_id
    elif content_type == 'voice':
        file_id = message.voice.file_id
    elif content_type == 'video_note':
        file_id = message.video_note.file_id

    if file_id:
        # Сохраняем текущую подпись
        caption = news[news_id]['files'][0]['caption'] if 'files' in news[news_id] and news[news_id]['files'] else None

        # Удаляем старые медиафайлы, если это первый новый файл
        if not news[news_id].get('new_files'):
            news[news_id]['new_files'] = []
            news[news_id]['files'] = []

        # Добавляем новый медиафайл, сохраняя подпись
        news[news_id]['new_files'].append({
            'type': content_type,
            'file_id': file_id,
            'caption': caption
        })

        if len(news[news_id]['new_files']) >= 10:
            # Переносим новые файлы в основной список файлов
            news[news_id]['files'] = news[news_id].get('new_files', [])
            del news[news_id]['new_files']
            save_news_database()
            bot.send_message(message.chat.id, "Медиафайлы новости отредактированы!")
            show_editorial_menu(message)
            return

        markup = telebot.types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Добавить еще', 'Завершить отправку')
        markup.add('В редакцию')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Медиафайл добавлен! Хотите добавить еще?", reply_markup=markup)
        bot.register_next_step_handler(message, handle_edit_media_options, news_id)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, отправьте мультимедийный файл!")
        bot.register_next_step_handler(message, edit_news_media, news_id)

#@log_user_actions
def handle_edit_media_options(message, news_id):
    if message.text == "Добавить еще":
        markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        markup.add('В редакцию')
        markup.add('В меню админ-панели')
        bot.send_message(message.chat.id, "Отправьте следующий мультимедийный файл:", reply_markup=markup)
        bot.register_next_step_handler(message, edit_news_media, news_id)
    elif message.text == "Завершить отправку":
        # Переносим новые файлы в основной список файлов
        news[news_id]['files'] = news[news_id].get('new_files', [])
        del news[news_id]['new_files']
        save_news_database()
        bot.send_message(message.chat.id, "Медиафайлы новости отредактированы!")
        show_editorial_menu(message)
    elif message.text == "В редакцию":
        show_editorial_menu(message)
    elif message.text == "В меню админ-панели":
        show_admin_panel(message)
    else:
        bot.send_message(message.chat.id, "Пожалуйста, выберите действие!")
        bot.register_next_step_handler(message, handle_edit_media_options, news_id)

# Разделение текста

#@log_user_actions
def split_text(text, chunk_size=4096):
    return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

# Обработчик для "Посмотреть новость"

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Посмотреть новость' and check_admin_access(message))
def view_news(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Посмотреть новость'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    news_list = [
        f"📰 №{i + 1}. *{n['title']}* - {n['time'].strftime('%d.%m.%Y в %H:%M')}"
        for i, n in enumerate(news.values())
    ]

    if news_list:
        news_text = "\n\n".join(news_list)
        chunks = split_text(news_text)
        for chunk in chunks:
            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('В редакцию')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "📜 *Список новостей*:\n\n\n" + chunk, parse_mode="Markdown", reply_markup=markup)
        bot.send_message(message.chat.id, "Введите номера новости для просмотра:", reply_markup=markup)
        bot.register_next_step_handler(message, choose_news_to_view)
    else:
        bot.send_message(message.chat.id, "Нет новостей для просмотра!")

#@log_user_actions
def choose_news_to_view(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_news_to_view)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(x.strip()) - 1 for x in message.text.split(',')]
        news_list = list(news.values())
        invalid_indices = []

        # Проверяем все номера новостей
        for index in indices:
            if not (0 <= index < len(news_list)):
                invalid_indices.append(index + 1)

        if invalid_indices:
            bot.send_message(message.chat.id, f"Неверные номера новостей: {', '.join(map(str, invalid_indices))}. Пожалуйста, введите корректные номера новостей")
            bot.register_next_step_handler(message, choose_news_to_view)
        else:
            # Если все номера верны, выводим новости
            for index in indices:
                news_item = news_list[index]

                # Извлекаем подпись из первого файла, если есть
                caption = news_item['files'][0]['caption'] if 'files' in news_item and news_item['files'] else None

                # Формируем текст сообщения в зависимости от наличия подписи или текста
                if caption:
                    message_text = f"📌 ЗАГОЛОВОК НОВОСТИ 📌\n\n\n{news_item['title']}\n\n\n📢 ПОДПИСЬ НОВОСТИ 📢\n\n\n{caption}"
                elif news_item.get('text'):
                    message_text = f"📌 ЗАГОЛОВОК НОВОСТИ 📌\n\n\n{news_item['title']}\n\n\n📝 ТЕКСТ НОВОСТИ 📝\n\n\n{news_item['text']}"
                else:
                    message_text = f"📌 ЗАГОЛОВОК НОВОСТИ 📌\n\n\n{news_item['title']}"

                # Если есть медиафайлы, отправляем их вместе с текстом
                if 'files' in news_item and news_item['files']:
                    media_group = []
                    first_file = True
                    for file in news_item['files']:
                        if first_file:
                            caption = message_text
                        else:
                            caption = None
                        if file['type'] == 'photo':
                            media_group.append(telebot.types.InputMediaPhoto(file['file_id'], caption=caption))
                        elif file['type'] == 'video':
                            media_group.append(telebot.types.InputMediaVideo(file['file_id'], caption=caption))
                        elif file['type'] == 'document':
                            media_group.append(telebot.types.InputMediaDocument(file['file_id'], caption=caption))
                        elif file['type'] == 'animation':
                            media_group.append(telebot.types.InputMediaAnimation(file['file_id'], caption=caption))
                        elif file['type'] == 'sticker':
                            bot.send_sticker(message.chat.id, file['file_id'])
                        elif file['type'] == 'audio':
                            media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                        elif file['type'] == 'voice':
                            media_group.append(telebot.types.InputMediaAudio(file['file_id'], caption=caption))
                        elif file['type'] == 'video_note':
                            bot.send_video_note(message.chat.id, file['file_id'])
                        first_file = False

                    if media_group:
                        bot.send_media_group(message.chat.id, media_group)
                else:
                    # Если нет медиафайлов, отправляем только текст
                    bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

            show_editorial_menu(message)  # Автоматически перенаправляем в меню админ-панели
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректные номера новостей!")
        bot.register_next_step_handler(message, choose_news_to_view)

# Обработчик для "Удалить новость"

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить новость' and check_admin_access(message))
def delete_news(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить новость'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    news_list = [
        f"📰 №{i + 1}. *{n['title']}* - {n['time'].strftime('%d.%m.%Y в %H:%M')}"
        for i, n in enumerate(news.values())
    ]

    if news_list:
        news_text = "\n\n".join(news_list)
        chunks = split_text(news_text)
        for chunk in chunks:
            markup = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
            markup.add('В редакцию')
            markup.add('В меню админ-панели')
            bot.send_message(message.chat.id, "📜 *Список новостей*:\n\n\n" + chunk, parse_mode="Markdown", reply_markup=markup)
        bot.send_message(message.chat.id, "Введите номера новостей:", reply_markup=markup)
        bot.register_next_step_handler(message, choose_news_to_delete)
    else:
        bot.send_message(message.chat.id, "Нет новостей для удаления!")

#@log_user_actions
def choose_news_to_delete(message):
    if message.photo or message.video or message.document or message.animation or message.sticker or message.location or message.audio or message.contact or message.voice or message.video_note:
        bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(message, choose_news_to_delete)
        return

    if message.text == "В редакцию":
        show_editorial_menu(message)
        return

    if message.text == "В меню админ-панели":
        show_admin_panel(message)
        return

    try:
        indices = [int(x.strip()) - 1 for x in message.text.split(',')]
        news_list = list(news.values())
        deleted_news_titles = []
        invalid_indices = []

        for index in indices:
            if 0 <= index < len(news_list):
                news_id = list(news.keys())[index]
                deleted_news = news.pop(news_id)
                deleted_news_titles.append(deleted_news['title'])
            else:
                invalid_indices.append(index + 1)

        if invalid_indices:
            bot.send_message(message.chat.id, f"Неверные номера новостей: {', '.join(map(str, invalid_indices))}. Пожалуйста, введите корректные номера новостей")
            bot.register_next_step_handler(message, choose_news_to_delete)
        else:
            if deleted_news_titles:
                # Перенумеровываем новости после удаления
                new_news = {}
                for i, (key, value) in enumerate(news.items(), start=1):
                    new_news[str(i)] = value
                news.clear()
                news.update(new_news)

                save_news_database()
                deleted_news_titles_lower = [title.lower() for title in deleted_news_titles]
                bot.send_message(message.chat.id, f"Новости (*{', '.join(deleted_news_titles_lower)}*) удалены!", parse_mode="Markdown")
            else:
                bot.send_message(message.chat.id, "Нет новостей для удаления!")

            show_editorial_menu(message)  # Автоматически перенаправляем в меню админ-панели
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите корректные номера новостей!")
        bot.register_next_step_handler(message, choose_news_to_delete)

# (ADMIN n) ------------------------------------------ "ЧАТ АДМИНА И ПОЛЬЗОВАТЕЛЯ ФУУНКЦИЙ ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

# Путь к JSON файлу с админскими сессиями
ADMIN_SESSIONS_FILE = 'data base/admin/admin_sessions.json'

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data['admin_sessions']

# Функция для проверки прав доступа
#@log_user_actions
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return False

# Пути к файлам
USER_DB_PATH = 'data base/admin/users.json'
ADMIN_DB_PATH = 'data base/admin/admin_sessions.json'
ACTIVE_CHATS_PATH = 'data base/admin/chats/active_chats.json'
CHAT_HISTORY_PATH = 'data base/admin/chats/chat_history.json'

# Глобальные переменные
active_chats = {}
user_requests = {}
dialog_states = {}  # Новая структура для хранения состояния диалогов
current_dialogs = {}  # Новая структура для хранения текущих диалогов

# Загрузка активных чатов из файла
def load_active_chats():
    global active_chats, user_requests
    if os.path.exists(ACTIVE_CHATS_PATH):
        with open(ACTIVE_CHATS_PATH, 'r', encoding='utf-8') as file:
            data = json.load(file)
            active_chats = data.get("active_chats", {})
            user_requests = data.get("user_requests", {})

            # Преобразуем строки дат в объекты datetime
            for user_id, requests in user_requests.items():
                user_requests[user_id] = {datetime.strptime(date_str, "%d.%m.%Y").date(): count for date_str, count in requests.items()}
    else:
        active_chats = {}
        user_requests = {}

# Загрузка активных чатов при запуске бота
load_active_chats()

# Сохранение активных чатов в файл
def save_active_chats():
    with open(ACTIVE_CHATS_PATH, 'w', encoding='utf-8') as file:
        data = {
            "active_chats": active_chats,
            "user_requests": {user_id: {date.strftime("%d.%m.%Y"): count for date, count in requests.items()} for user_id, requests in user_requests.items()}
        }
        json.dump(data, file, ensure_ascii=False, indent=4)

#@log_user_actions
def add_user_request(user_id, date, count):
    if user_id not in user_requests:
        user_requests[user_id] = {}
    if date not in user_requests[user_id]:
        user_requests[user_id][date] = 0
    user_requests[user_id][date] += count
    save_active_chats()

# Загрузка истории чата
def load_chat_history():
    if os.path.exists(CHAT_HISTORY_PATH):
        with open(CHAT_HISTORY_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    return {}

# Сохранение сообщения в историю чата
def save_message_to_history(admin_id, user_id, message_content, message_type, caption=None):
    chat_history = load_chat_history()

    # Генерируем ключ для чата
    chat_key = f"{admin_id}_{user_id}"
    if chat_key not in chat_history:
        chat_history[chat_key] = []  # Если чата нет, создаем новый

    # Получаем текущее время и форматируем его
    timestamp = datetime.now().strftime("%d.%m.%Y в %H:%M")

    # Проверяем, существует ли текущий диалог
    if chat_key not in current_dialogs:
        current_dialogs[chat_key] = []  # Создаем новый диалог

    # Добавляем новое сообщение в текущий диалог
    current_dialogs[chat_key].append({
        "type": message_type,  # 'admin' или 'user'
        "content": message_content,
        "timestamp": timestamp,  # Время отправки сообщения
        "caption": caption.lower() if caption else None  # Подпись, если есть
    })

    # Сохраняем обновленную историю в файл
    with open(CHAT_HISTORY_PATH, 'w', encoding='utf-8') as file:
        json.dump(chat_history, file, ensure_ascii=False, indent=4)

# Загрузка базы данных пользователей
def load_users():
    if os.path.exists(USER_DB_PATH):
        with open(USER_DB_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    return {}

# Загрузка базы данных администраторов
def load_admins():
    if os.path.exists(ADMIN_DB_PATH):
        with open(ADMIN_DB_PATH, 'r', encoding='utf-8') as file:
            return json.load(file)
    return []

# Проверка, является ли пользователь администратором
#@log_user_actions
def is_admin(user_id):
    admins = load_admins()
    return user_id in admins

# Экранируем специальные символы Markdown
#@log_user_actions
def escape_markdown(text):
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

# Функция для вывода списка пользователей
#@log_user_actions
def list_users_for_chat(message):
    users_data = load_users()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *заблокирован* 🚫" if data.get('blocked', False) else " - *разблокирован* ✅"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    # Создаем клавиатуру с одной кнопкой "В меню админ-панели"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add(types.KeyboardButton('В меню админ-панели'))

    bot.send_message(
        message.chat.id,
        "Пожалуйста, укажите *один* из следующих параметров для начала чата:\n\n\n"
        "1. *ID* пользователя в формате: `/chat id`\n\n"
        "2. *USERNAME* в формате: `/chat @username`\n\n",
        parse_mode="Markdown",
        reply_markup=markup
    )

# Структуры для хранения активных чатов
active_user_chats = {}  # Хранит, какой администратор общается с пользователем
active_admin_chats = {}  # Хранит, с каким пользователем общается администратор

# Обработчик для меню "Общение"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Общение' and check_admin_access(message))
def show_communication_menu(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Общение'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    markup.add('Чат', 'Запросы')
    markup.add('Оповещения', 'Диалоги')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите действие из общения:", reply_markup=markup)

# Пример использования функции save_last_bot_message

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == "Чат с админом")
@restricted
@track_user_activity
@check_function_state_decorator('Чат с админом')
@track_usage('Чат с админом')
def request_chat_with_admin(message):
    global active_chats
    if active_chats is None:
        active_chats = {}

    user_id = message.from_user.id
    today = datetime.now().date()

    # Проверяем, есть ли уже запрос на чат с статусом "pending"
    if any(chat_data.get("user_id") == user_id and chat_data.get("status") == "pending" for chat_data in active_chats.values()):
        bot.send_message(user_id, "У вас уже есть запрос на чат к администратору! Ожидайте!")
        return

    # Проверяем количество запросов на сегодня
    user_requests_today = user_requests.get(user_id, {}).get(today, 0)
    if user_requests_today >= 3:
        bot.send_message(user_id, "Вы исчерпали лимит запросов на сегодня! Попробуйте завтра...")
        return

    # Запрашиваем тему для общения
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add(types.KeyboardButton('В главное меню'))
    bot.send_message(user_id, "Пожалуйста, укажите тему для общения с администратором!", reply_markup=markup)
    active_chats[user_id] = {"user_id": user_id, "status": "waiting_for_topic", "awaiting_response": False}
    save_active_chats()

#@log_user_actions
@bot.message_handler(func=lambda message: any(chat_data.get("user_id") == message.from_user.id and chat_data.get("status") == "waiting_for_topic" for chat_data in active_chats.values()))
def handle_chat_topic(message):
    user_id = message.from_user.id
    topic = message.text
    today = datetime.now().date()

    # Находим запись с статусом "waiting_for_topic"
    if user_id in active_chats and active_chats[user_id].get("status") == "waiting_for_topic":
        # Обновляем количество запросов
        add_user_request(user_id, today, 1)

        active_chats[user_id] = {
            "user_id": user_id,
            "status": "pending",
            "topic": topic,
            "awaiting_response": False
        }
        save_active_chats()

        bot.send_message(user_id, "Запрос на чат был успешно передан администратору! Ожидаем ответа...")
        return_to_menu(message)  # Перенаправляем пользователя в меню
        return

    bot.send_message(user_id, "У вас уже есть запрос на чат к администратору! Ожидайте!")

# Команда для администратора для связи с пользователем

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Чат' and check_admin_access(message))
@bot.message_handler(commands=['chat'])
def initiate_chat(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Чат'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if not check_admin_access(message):  # Проверяем, является ли пользователь администратором
        return

    command_parts = message.text.split()
    if len(command_parts) == 2:
        user_input = command_parts[1]
        users_data = load_users()

        # Проверка, является ли ввод ID пользователя
        if user_input.isdigit():
            user_id = int(user_input)
            if str(user_id) not in users_data:
                bot.send_message(message.chat.id, f"Пользователь с таким *ID* - `{user_id}` не найден!", parse_mode="Markdown")
                return
            username = users_data[str(user_id)]['username']
        # Проверка, является ли ввод username
        elif user_input.startswith('@'):
            username = user_input
            user_id = None
            for uid, data in users_data.items():
                if data.get('username') == username:
                    user_id = int(uid)
                    break
            if user_id is None:
                bot.send_message(message.chat.id, f"Пользователь с таким *USERNAME* - {escape_markdown(username)} не найден!", parse_mode="Markdown")
                return
        # Проверка, является ли ввод номером из списка
        else:
            try:
                user_number = int(user_input)
                if user_number < 1 or user_number > len(users_data):
                    bot.send_message(message.chat.id, "Неверный номер пользователя!")
                    return
                user_id = list(users_data.keys())[user_number - 1]
                username = users_data[user_id]['username']
            except ValueError:
                bot.send_message(message.chat.id, "Неверный формат номера пользователя!")
                return

        # Проверка, если чат уже активен
        if user_id in active_chats:
            if active_chats[user_id].get("admin_id") is None:
                # Если нет активного админа, регистрируем текущего админа
                active_chats[user_id]["admin_id"] = message.chat.id
                save_active_chats()
            else:
                bot.send_message(message.chat.id, "Этот пользователь уже находится в активном чате с другим администратором!")
                return

        # Отправляем запрос пользователю
        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Принять', 'Отклонить')
        markup.add('В главное меню')
        bot.send_message(user_id, "Администратор хочет связаться с вами!\n\nВыберите *ПРИНЯТЬ* для принятия или *ОТКЛОНИТЬ* для отклонения!", parse_mode="Markdown", reply_markup=markup)
        bot.send_message(message.chat.id, f"Запрос отправлен пользователю {escape_markdown(username)} - `{user_id}`! Ожидаем ответа...", parse_mode="Markdown")

        # Сохраняем запрос в active_chats
        active_chats[user_id] = {"admin_id": message.chat.id, "status": "pending", "awaiting_response": True}
        save_active_chats()

    else:
        list_users_for_chat(message)


# Команда для администратора для просмотра запросов на чат
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Запросы' and check_admin_access(message))
def list_chat_requests(message):
    admin_id = message.from_user.id

    # admin_id = str(message.chat.id)
    # if not check_permission(admin_id, 'Запросы'):
    #     bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
    #     return

    # if not check_admin_access(message):  # Проверяем, является ли пользователь администратором
    #     return

    requests = [(user_id, data.get("topic", "Без темы")) for user_id, data in active_chats.items() if data["status"] == "pending"]

    if not requests:
        bot.send_message(admin_id, "*Нет активных* запросов на чат!", parse_mode="Markdown")
        return

    users_data = load_users()
    request_list = []
    for i, (user_id, topic) in enumerate(requests):
        username = users_data.get(str(user_id), {}).get('username', 'Unknown')
        escaped_username = escape_markdown(username)
        request_list.append(f"🔹 *№{i + 1}.* Запрос от пользователя:\n👤 {escaped_username} - `{user_id}`\n📨 *Тема*: {topic.lower()}")

    request_list_message = "Список *запросов* на чат:\n\n" + "\n\n".join(request_list) + "\n\nВведите номер запроса для начала диалога:"

    # Разбиваем сообщение на части, если оно слишком длинное
    parts = [request_list_message[i:i + MAX_MESSAGE_LENGTH] for i in range(0, len(request_list_message), MAX_MESSAGE_LENGTH)]

    for part in parts:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.add(types.KeyboardButton('Вернуться в общение'))
        markup.add(types.KeyboardButton('В меню админ-панели'))

        bot.send_message(admin_id, part, parse_mode="Markdown", reply_markup=markup)

    active_chats[admin_id] = {"status": "select_request", "requests": requests}
    save_active_chats()

#@log_user_actions
@bot.message_handler(func=lambda message: message.from_user.id in active_chats and active_chats[message.from_user.id]["status"] == "select_request")
def handle_request_selection(message):
    admin_id = message.from_user.id

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    try:
        selected_index = int(message.text) - 1
        requests = active_chats[admin_id]["requests"]
        if selected_index < 0 or selected_index >= len(requests):
            raise IndexError("Номер запроса вне диапазона!")

        selected_user_id, topic = requests[selected_index]
        users_data = load_users()
        username = users_data.get(str(selected_user_id), {}).get('username', 'Unknown')

        if selected_user_id in active_chats:
            if active_chats[selected_user_id].get("admin_id") is None:
                active_chats[selected_user_id]["admin_id"] = admin_id
                save_active_chats()
            else:
                admin_id_in_chat = active_chats[selected_user_id]["admin_id"]
                bot.send_message(admin_id, f"Этот пользователь уже *находится в активном чате* с администратором `{admin_id_in_chat}`!", parse_mode="Markdown")
                show_communication_menu(message)  
                return

        markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        markup.add('Принять', 'Отклонить')
        markup.add('В главное меню')

        bot.send_message(selected_user_id, "Администратор хочет связаться с вами!\n\nВыберите *ПРИНЯТЬ* для принятия или *ОТКЛОНИТЬ* для отклонения!", parse_mode="Markdown", reply_markup=markup)
        bot.send_message(admin_id, f"Запрос отправлен пользователю {escape_markdown(username)} - `{selected_user_id}`! Ожидаем ответа...", parse_mode="Markdown")

        active_chats[selected_user_id] = {"admin_id": admin_id, "status": "pending", "awaiting_response": True}
        save_active_chats()

        del active_chats[admin_id]["requests"][selected_index]
        save_active_chats()

        del active_chats[admin_id]
        save_active_chats()

    except (ValueError, IndexError) as e:
        bot.send_message(admin_id, f"Ошибка: {e}! Пожалуйста, введите корректный номер запроса")

#@log_user_actions
def return_admin_to_menu(admin_id):
    # Перенаправляем администратора в меню админ-панели
    bot.send_message(admin_id, "Чат с пользователем был *завершен*!", parse_mode="Markdown")
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add('Админ', 'Бан', 'Функции')
    markup.add('Общение', 'Реклама', 'Статистика')
    markup.add('Файлы', 'Резервная копия', 'Редакция')
    markup.add('Экстренная остановка')
    markup.add('Выход')
    bot.send_message(admin_id, "Выберите действие из админ-панели:", reply_markup=markup)

# Обработка ответов пользователя на запрос чата

#@log_user_actions
@bot.message_handler(func=lambda message: message.text.lower() in ["принять", "отклонить"] and message.from_user.id in active_chats and active_chats[message.from_user.id]["status"] == "pending")
def handle_chat_response(message):
    user_id = message.from_user.id

    # Проверяем, есть ли активный запрос на чат для этого пользователя
    if user_id in active_chats and active_chats[user_id]["status"] == "pending" and active_chats[user_id]["awaiting_response"]:
        admin_id = active_chats[user_id]["admin_id"]
        users_data = load_users()
        username = users_data[str(user_id)]['username']

        if message.text.lower() == "принять":
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('Стоп'))
            bot.send_message(user_id, "Вы *на связи* с администратором!", parse_mode="Markdown", reply_markup=markup)
            bot.send_message(admin_id, f"Пользователь {escape_markdown(username)} - `{user_id}` *принял* запрос на чат!", parse_mode="Markdown", reply_markup=markup)
            active_chats[user_id]["status"] = "active"
            active_chats[user_id]["awaiting_response"] = False
            active_user_chats[user_id] = admin_id  # Добавляем в активные чаты для пользователя
            active_admin_chats[admin_id] = user_id  # Добавляем в активные чаты для администратора
            save_active_chats()
        else:
            bot.send_message(user_id, "Вы *отклонили* запрос на чат!", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            bot.send_message(admin_id, f"Пользователь {escape_markdown(username)} - `{user_id}` *отклонил* запрос на чат!", parse_mode="Markdown")
            del active_chats[user_id]
            if admin_id in active_chats:
                del active_chats[admin_id]
            save_active_chats()
            return_to_menu(message)  # Перенаправляем пользователя в меню
            return_admin_to_menu(admin_id)  # Перенаправляем администратора в меню
    else:
        # Если нет активного запроса на чат, игнорируем сообщение
        return
        
# Обработчик для возврата в главное меню
@restricted
@track_user_activity
@check_chat_state
@check_function_state_decorator('В главное меню')
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == "В главное меню")
@bot.message_handler(commands=['mainmenu'])
def return_to_menu(message):
    user_id = message.from_user.id
    chat_id = message.chat.id

    # Проверяем, есть ли у пользователя активный запрос на чат
    if user_id in active_chats and active_chats[user_id].get("status") == "waiting_for_topic":
        del active_chats[user_id]  # Сбрасываем статус пользователя
        save_active_chats()

    # Проверяем, есть ли временные данные о поездке
    if user_id in temporary_trip_data:
        temporary_trip_data[user_id] = []

    # Выполняем основную функцию старта
    start(message)

# Функция для отправки сообщения пользователю по его user_id
#@log_user_actions
def send_message_to_user(user_id, text, reply_markup=None):
    bot.send_message(user_id, text, reply_markup=reply_markup)

# Функция для завершения чата

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Стоп')
@bot.message_handler(commands=['stopchat'])
def stop_chat(message):
    user_id = message.from_user.id

    # Проверяем, есть ли у пользователя активный чат
    if user_id in active_user_chats:
        admin_id = active_user_chats[user_id]
        users_data = load_users()
        username = users_data.get(str(user_id), {}).get('username', 'Unknown')
        escaped_username = escape_markdown(username)

        bot.send_message(admin_id, f"Пользователь {escaped_username} - `{user_id}` завершил чат!", parse_mode="Markdown")
        bot.send_message(user_id, "Чат с администратором был *завершен*!", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        del active_user_chats[user_id]
        del active_admin_chats[admin_id]
        if user_id in active_chats:
            del active_chats[user_id]
        save_active_chats()
        start_menu(user_id)  # Перенаправляем пользователя в меню

        # Сохраняем текущий диалог в историю
        chat_key = f"{admin_id}_{user_id}"
        if chat_key in current_dialogs:
            chat_history = load_chat_history()
            chat_history[chat_key].append(current_dialogs[chat_key])
            del current_dialogs[chat_key]
            with open(CHAT_HISTORY_PATH, 'w', encoding='utf-8') as file:
                json.dump(chat_history, file, ensure_ascii=False, indent=4)

        # Перенаправляем администратора в меню админ-панели
        return_admin_to_menu(admin_id)

    elif user_id in active_admin_chats:
        target_user_id = active_admin_chats[user_id]
        bot.send_message(target_user_id, "Администратор *завершил* чат!", parse_mode="Markdown")
        del active_admin_chats[user_id]
        del active_user_chats[target_user_id]
        if target_user_id in active_chats:
            del active_chats[target_user_id]
        save_active_chats()
        start_menu(target_user_id)  # Перенаправляем пользователя в меню

        # Сохраняем текущий диалог в историю
        chat_key = f"{user_id}_{target_user_id}"
        if chat_key in current_dialogs:
            chat_history = load_chat_history()
            chat_history[chat_key].append(current_dialogs[chat_key])
            del current_dialogs[chat_key]
            with open(CHAT_HISTORY_PATH, 'w', encoding='utf-8') as file:
                json.dump(chat_history, file, ensure_ascii=False, indent=4)

        # Перенаправляем администратора в меню админ-панели
        return_admin_to_menu(user_id)

    else:
        bot.send_message(user_id, "Нет активного чата для завершения!")

# Функция для отображения главного меню
#@log_user_actions
def start_menu(user_id):
    # Получаем информацию о пользователе
    user_data = load_user_data()
    username = user_data.get(user_id, {}).get('username', 'Неизвестный пользователь')

    # Создание кнопок меню
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Расход топлива")
    item2 = types.KeyboardButton("Траты и ремонты")
    item3 = types.KeyboardButton("Найти транспорт")
    item4 = types.KeyboardButton("Поиск мест")
    item5 = types.KeyboardButton("Погода")
    item6 = types.KeyboardButton("Код региона")
    item7 = types.KeyboardButton("Цены на топливо")
    item8 = types.KeyboardButton("Анти-радар")
    item9 = types.KeyboardButton("Напоминания")
    item10 = types.KeyboardButton("Коды OBD2")
    item11 = types.KeyboardButton("Прочее")

    markup.add(item1, item2)
    markup.add(item3, item4)
    markup.add(item5, item7)
    markup.add(item6, item8)
    markup.add(item9, item10)
    markup.add(item11)

    welcome_message = f"Добро пожаловать, @{escape_markdown(username)}!\nВыберите действие из меню:"
    send_message_to_user(user_id, welcome_message, parse_mode="Markdown", reply_markup=markup)

#@log_user_actions
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return False



# (ADMIN N) ------------------------------------------ "ФАЙЛЫ ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------
    
# Максимальная длина сообщения в Telegram
TELEGRAM_MESSAGE_LIMIT = 4096

# Определение корневой директории на основе исполняемого файла
EXECUTABLE_FILE = '(93 update ИСПРАВЛЕНИЕ25  ( (  )) CAR MANAGER TG BOT (official) v0924.py'
BASE_DIR = os.path.dirname(os.path.abspath(EXECUTABLE_FILE))

# Пути к директориям и файлам
BACKUP_DIR = os.path.join(BASE_DIR, 'backups')
FILES_PATH = os.path.join(BASE_DIR, 'data base')
ADDITIONAL_FILES_PATH = os.path.join(BASE_DIR, 'files')
ADMIN_SESSIONS_FILE = os.path.join(BASE_DIR, 'data base', 'admin', 'admin_sessions.json')

# Словарь для хранения данных о файлах и директориях
bot_data = {}

# Загрузка админских сессий из JSON файла
def load_admin_sessions():
    with open(ADMIN_SESSIONS_FILE, 'r', encoding='utf-8') as file:
        data = json.load(file)
    return data['admin_sessions']

# Функция для проверки прав доступа
#@log_user_actions
def check_admin_access(message):
    admin_sessions = load_admin_sessions()
    if str(message.chat.id) in admin_sessions:
        return True
    else:
        bot.send_message(message.chat.id, "У вас нет прав доступа для выполнения этой операции.")
        return False

# Общий обработчик для команды "Файлы"
#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Файлы' and check_admin_access(message))
def show_files_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Файлы'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(row_width=3, resize_keyboard=True)
    markup.add('Просмотр файлов', 'Поиск файлов по ID')
    markup.add('Добавить файлы', 'Замена файлов', 'Удалить файлы')
    markup.add('В меню админ-панели')

    bot.send_message(message.chat.id, "Выберите действие с файлами:", reply_markup=markup)

@bot.message_handler(func=lambda message: message.text == 'Просмотр файлов' and check_admin_access(message))
def view_files(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Просмотр файлов'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    files_list = []
    extensions = set()

    # Сбор файлов из корневой директории
    for file_name in os.listdir(BASE_DIR):
        file_path = os.path.join(BASE_DIR, file_name)
        if os.path.isfile(file_path):
            files_list.append(file_path)
            extension = os.path.splitext(file_name)[1]
            extensions.add(extension)

    # Сбор файлов из FILES_PATH
    for root, dirs, files in os.walk(FILES_PATH):
        for file_name in files:
            files_list.append(os.path.join(root, file_name))
            extension = os.path.splitext(file_name)[1]
            extensions.add(extension)

    # Сбор файлов из ADDITIONAL_FILES_PATH
    for root, dirs, files in os.walk(ADDITIONAL_FILES_PATH):
        for file_name in files:
            files_list.append(os.path.join(root, file_name))
            extension = os.path.splitext(file_name)[1]
            extensions.add(extension)

    # Проверка наличия файлов
    if not files_list:
        bot.send_message(message.chat.id, "Файлы не найдены!")
        return

    # Формирование списка расширений
    sorted_extensions = sorted(extensions)
    response = "*Список расширений файлов:*\n\n"
    response += "📁 1. *Отправка всех файлов*\n\n"
    response += "\n".join([f"📄 {i + 2}. *{ext[1:]}*" for i, ext in enumerate(sorted_extensions)])

    # Сохранение данных в bot_data
    bot_data[message.chat.id] = {
        "files_list": files_list,
        "extensions": sorted_extensions
    }

    # Отправка сообщения с нумерованным списком
    for start in range(0, len(response), TELEGRAM_MESSAGE_LIMIT):
        bot.send_message(message.chat.id, response[start:start + TELEGRAM_MESSAGE_LIMIT], parse_mode="Markdown")

    # Подготовка кнопок для дальнейшего взаимодействия
    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню файлы')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер для выбора расширения:", reply_markup=markup)
    bot.register_next_step_handler(message, process_extension_selection)

# @log_user_actions
def process_extension_selection(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_extension_selection)
        return

    try:
        selection = int(message.text.strip())
        if selection == 1:
            files_list = bot_data[message.chat.id]["files_list"]
            selected_extension = None  # Для случая выбора всех файлов
        else:
            extensions = bot_data[message.chat.id]["extensions"]
            if 1 < selection <= len(extensions) + 1:
                selected_extension = extensions[selection - 2]
                files_list = [file for file in bot_data[message.chat.id]["files_list"] if file.endswith(selected_extension)]
            else:
                bot.send_message(message.chat.id, "Некорректный номер!")
                bot.register_next_step_handler(message, process_extension_selection)
                return

        if files_list:
            if selected_extension:  # Для конкретного расширения
                response = f"Показаны файлы с расширением {selected_extension[1:]}:\n\n"
            else:  # Для всех файлов
                response = "Показаны все файлы:\n\n"

            response += "\n".join([f"📄 {i + 1}. {os.path.basename(file_path)}" for i, file_path in enumerate(files_list)])
            for start in range(0, len(response), TELEGRAM_MESSAGE_LIMIT):
                bot.send_message(message.chat.id, response[start:start + TELEGRAM_MESSAGE_LIMIT])
            bot.send_message(message.chat.id, "Введите номера файлов через запятую для отправки:")
            bot.register_next_step_handler(message, process_file_selection, files_list)
        else:
            bot.send_message(message.chat.id, "Файлы с выбранным расширением не найдены!")
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите номер!")
        bot.register_next_step_handler(message, process_extension_selection)

# @log_user_actions
def process_file_selection(message, matched_files):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_file_selection, matched_files)
        return

    try:
        file_numbers = [int(num.strip()) - 1 for num in message.text.split(',')]
        valid_files = [matched_files[num] for num in file_numbers if 0 <= num < len(matched_files)]

        if valid_files:
            for file_path in valid_files:
                with open(file_path, 'rb') as file:
                    bot.send_document(message.chat.id, file)
            show_files_menu(message)
        else:
            bot.send_message(message.chat.id, "Некорректные номера файлов!")
            bot.register_next_step_handler(message, process_file_selection, matched_files)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите номера файлов через запятую!")
        bot.register_next_step_handler(message, process_file_selection, matched_files)

# Поиск файлов пользователя по ID
@bot.message_handler(func=lambda message: message.text == 'Поиск файлов по ID' and check_admin_access(message))
#@log_user_actions
def search_files_by_id(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Поиск файлов по ID'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню файлы')
    markup.add('В меню админ-панели')
    list_users_for_files(message)

#@log_user_actions
def search_id_in_json(data, user_id):
    if isinstance(data, dict):
        for key, value in data.items():
            if key == user_id or (isinstance(value, str) and user_id in value):
                return True
            if search_id_in_json(value, user_id):
                return True
    elif isinstance(data, list):
        for item in data:
            if search_id_in_json(item, user_id):
                return True
    return False

USER_DATA_PATH = 'data base/admin/users.json'

#@log_user_actions
def escape_markdown(text):
    # Экранируем специальные символы Markdown
    return re.sub(r'([_*\[\]()~`>#+\-=|{}.!])', r'\\\1', text)

def load_user_data():
    with open(USER_DATA_PATH, 'r', encoding='utf-8') as file:
        return json.load(file)

#@log_user_actions
def list_users_for_files(message):
    users_data = load_user_data()
    user_list = []
    for user_id, data in users_data.items():
        username = escape_markdown(data['username'])
        status = " - *заблокирован* 🚫" if data.get('blocked', False) else " - *разблокирован* ✅"
        user_list.append(f"№{len(user_list) + 1}. {username} - `{user_id}`{status}")

    response_message = "📋 Список *всех* пользователей:\n\n\n" + "\n\n".join(user_list)
    if len(response_message) > 4096:  # Ограничение Telegram по количеству символов в сообщении
        bot.send_message(message.chat.id, "📜 Список пользователей слишком большой для отправки в одном сообщении!")
    else:
        bot.send_message(message.chat.id, response_message, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню файлы')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Введите номер пользователя, username или ID для поиска файлов:", reply_markup=markup)
    bot.register_next_step_handler(message, process_user_input_for_file_search)

#@log_user_actions
def process_user_input_for_file_search(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_user_input_for_file_search)
        return

    user_input = message.text.strip()
    users_data = load_user_data()

    user_id = None
    username = None

    if user_input.isdigit():
        # Если длина ввода меньше 4 символов, это номер из списка
        if len(user_input) < 4:
            # Проверяем, является ли ввод номером пользователя (индекс из списка)
            user_index = int(user_input) - 1
            if 0 <= user_index < len(users_data):
                user_id = list(users_data.keys())[user_index]
                username = users_data[user_id]['username']  # Получаем username пользователя по ID
            else:
                bot.send_message(message.chat.id, "Некорректный номер пользователя!")
                bot.register_next_step_handler(message, process_user_input_for_file_search)
                return
        else:
            # Если длина ввода >= 4 символов, это ID
            if user_input in users_data:
                user_id = user_input
                username = users_data[user_id]['username']  # Получаем username пользователя по ID
            else:
                bot.send_message(message.chat.id, "Пользователь с таким *ID* не найден!", parse_mode="Markdown")
                bot.register_next_step_handler(message, process_user_input_for_file_search)
                return
    elif user_input.startswith('@'):
        # Проверяем, является ли ввод username
        username = user_input
        user_id = next((user_id for user_id, data in users_data.items() if data['username'].lower() == username.lower()), None)
        if not user_id:
            bot.send_message(message.chat.id, "Пользователь с таким *username* не найден!", parse_mode="Markdown")
            bot.register_next_step_handler(message, process_user_input_for_file_search)
            return
    else:
        bot.send_message(message.chat.id, "Некорректный ввод! Пожалуйста, введите номер, username или ID")
        bot.register_next_step_handler(message, process_user_input_for_file_search)
        return

    bot.send_message(message.chat.id, f"Поиск файлов для пользователя: {escape_markdown(username)} - `{user_id}` ...", parse_mode="Markdown")
    process_file_search(message, user_id)

#@log_user_actions
def process_file_search(message, user_id):
    matched_files = []

    search_paths = [BASE_DIR, BACKUP_DIR, FILES_PATH, ADDITIONAL_FILES_PATH]

    for search_path in search_paths:
        for root, dirs, files in os.walk(search_path):
            for file_name in files:
                file_path = os.path.join(root, file_name)
                if user_id in file_name:
                    matched_files.append(file_path)
                else:
                    if file_name.endswith('.json'):
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                content = json.load(f)
                                if search_id_in_json(content, user_id):
                                    matched_files.append(file_path)
                        except (json.JSONDecodeError, UnicodeDecodeError):
                            pass  # Игнорировать ошибки чтения
                    elif file_name.endswith(('.txt', '.log', '.csv')):
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                content = f.read()
                                if re.search(rf'\b{user_id}\b', content):
                                    matched_files.append(file_path)
                        except UnicodeDecodeError:
                            pass  # Игнорировать ошибки чтения

    if matched_files:
        response = "\n".join([f"📄 {i + 1}. {os.path.basename(path)}" for i, path in enumerate(matched_files)])
        for start in range(0, len(response), TELEGRAM_MESSAGE_LIMIT):
            bot.send_message(message.chat.id, response[start:start + TELEGRAM_MESSAGE_LIMIT])
        bot.send_message(message.chat.id, "Выберите номера файлов через запятую для отправки:")
        bot.register_next_step_handler(message, process_file_selection, matched_files)
    else:
        bot.send_message(message.chat.id, "Файлы с указанным ID не найдены!")

# Замена файла
@bot.message_handler(func=lambda message: message.text == 'Замена файлов' and check_admin_access(message))
#@log_user_actions
def handle_file_replacement(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Замена файлов'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню файлы')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Отправьте новый файл для замены:", reply_markup=markup)
    bot.register_next_step_handler(message, process_file_replacement)

#@log_user_actions
def process_file_replacement(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.document:
        file_name = message.document.file_name
        file_path = None

        # Поиск файла в директориях BACKUP_DIR, FILES_PATH и ADDITIONAL_FILES_PATH
        search_paths = [BASE_DIR, FILES_PATH, ADDITIONAL_FILES_PATH, BACKUP_DIR]

        for search_path in search_paths:
            for root, dirs, files in os.walk(search_path):
                if file_name in files:
                    file_path = os.path.join(root, file_name)
                    break
            if file_path:
                break

        if file_path:
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            with open(file_path, 'wb') as new_file:
                new_file.write(downloaded_file)
            bot.send_message(message.chat.id, "Файл успешно заменен!")
            show_files_menu(message)
        else:
            bot.send_message(message.chat.id, "Файл для замены не найден! Попробуйте отправить другой файл")
            bot.register_next_step_handler(message, process_file_replacement)
    else:
        bot.send_message(message.chat.id, "Неверный формат! Пожалуйста, отправьте файл в формате документа")
        bot.register_next_step_handler(message, process_file_replacement)

@bot.message_handler(func=lambda message: message.text == 'Добавить файлы' and check_admin_access(message))
#@log_user_actions
def add_files(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Добавить файлы'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    directories = [BASE_DIR, FILES_PATH, ADDITIONAL_FILES_PATH, BACKUP_DIR]
    response = "*Список директорий:*\n\n\n"
    response += "\n\n".join([f"📁 {i + 1}. {escape_markdown(dir)}" for i, dir in enumerate(directories)])
    bot.send_message(message.chat.id, response, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню файлы')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите директорию для добавления файла:", reply_markup=markup)
    bot.register_next_step_handler(message, process_add_file_directory_selection)

#@log_user_actions
def process_add_file_directory_selection(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_add_file_directory_selection)
        return

    try:
        selection = int(message.text.strip())
        if 1 <= selection <= 4:  # Убедитесь, что индексы соответствуют количеству директорий
            selected_directory = [BASE_DIR, FILES_PATH, ADDITIONAL_FILES_PATH, BACKUP_DIR][selection - 1]
            bot_data[message.chat.id] = {"selected_directory": selected_directory}
            bot.send_message(message.chat.id, "Отправьте файл для добавления:")
            bot.register_next_step_handler(message, process_add_file)
        else:
            bot.send_message(message.chat.id, "Некорректный номер!")
            bot.register_next_step_handler(message, process_add_file_directory_selection)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите номер!")
        bot.register_next_step_handler(message, process_add_file_directory_selection)

#@log_user_actions
def process_add_file(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.document:
        file_name = message.document.file_name
        selected_directory = bot_data[message.chat.id]["selected_directory"]
        file_path = os.path.join(selected_directory, file_name)

        if os.path.exists(file_path):
            bot.send_message(message.chat.id, "Файл с таким именем уже существует! Пожалуйста, отправьте файл с другим именем")
            bot.register_next_step_handler(message, process_add_file)
        else:
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            with open(file_path, 'wb') as new_file:
                new_file.write(downloaded_file)
            bot.send_message(message.chat.id, "Файл успешно добавлен!")
            show_files_menu(message)
    else:
        bot.send_message(message.chat.id, "Неверный формат! Пожалуйста, отправьте файл в формате документа")
        bot.register_next_step_handler(message, process_add_file)

@bot.message_handler(func=lambda message: message.text == 'Удалить файлы' and check_admin_access(message))
#@log_user_actions
def delete_files(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить файлы'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    directories = [BASE_DIR, FILES_PATH, ADDITIONAL_FILES_PATH, BACKUP_DIR]
    response = "*Список директорий:*\n\n\n"
    response += "\n\n".join([f"📁 {i + 1}. {escape_markdown(dir)}" for i, dir in enumerate(directories)])
    bot.send_message(message.chat.id, response, parse_mode="Markdown")

    markup = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    markup.add('В меню файлы')
    markup.add('В меню админ-панели')
    bot.send_message(message.chat.id, "Выберите директорию для удаления файла:", reply_markup=markup)
    bot.register_next_step_handler(message, process_delete_file_directory_selection)

#@log_user_actions
def process_delete_file_directory_selection(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_delete_file_directory_selection)
        return

    try:
        selection = int(message.text.strip())
        if 1 <= selection <= 4:  # Убедитесь, что индексы соответствуют количеству директорий
            selected_directory = [BASE_DIR, FILES_PATH, ADDITIONAL_FILES_PATH, BACKUP_DIR][selection - 1]
            files_list = [os.path.join(selected_directory, file) for file in os.listdir(selected_directory) if os.path.isfile(os.path.join(selected_directory, file))]
            response = "\n\n".join([f"📄 {i + 1}. {os.path.basename(file_path)}" for i, file_path in enumerate(files_list)])
            bot_data[message.chat.id] = {"files_list": files_list}

            # Разбиваем сообщение на части, если оно слишком длинное
            for start in range(0, len(response), TELEGRAM_MESSAGE_LIMIT):
                bot.send_message(message.chat.id, response[start:start + TELEGRAM_MESSAGE_LIMIT])

            bot.send_message(message.chat.id, "Выберите файл через запятую для удаления:")
            bot.register_next_step_handler(message, process_delete_file_selection)
        else:
            bot.send_message(message.chat.id, "Некорректный номер!")
            bot.register_next_step_handler(message, process_delete_file_directory_selection)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите номер!")
        bot.register_next_step_handler(message, process_delete_file_directory_selection)

#@log_user_actions
def process_delete_file_selection(message):
    if message.text == 'В меню админ-панели':
        show_admin_panel(message)
        return
    if message.text == 'В меню файлы':
        show_files_menu(message)
        return

    if message.photo or message.video or message.document or message.animation or message.sticker or message.audio or message.contact or message.voice or message.video_note:
        sent = bot.send_message(message.chat.id, "Извините, но отправка мультимедийных файлов не разрешена. Пожалуйста, введите текстовое сообщение.")
        bot.register_next_step_handler(sent, process_delete_file_selection)
        return

    try:
        file_numbers = [int(num.strip()) - 1 for num in message.text.split(',')]
        files_list = bot_data[message.chat.id]["files_list"]
        valid_files = [files_list[num] for num in file_numbers if 0 <= num < len(files_list)]

        if valid_files:
            for file_path in valid_files:
                os.remove(file_path)
            bot.send_message(message.chat.id, "Файлы успешно удалены!")
            show_files_menu(message)
        else:
            bot.send_message(message.chat.id, "Некорректные номера файлов!")
            bot.register_next_step_handler(message, process_delete_file_selection)
    except ValueError:
        bot.send_message(message.chat.id, "Пожалуйста, введите номера файлов через запятую!")
        bot.register_next_step_handler(message, process_delete_file_selection)

#----------------------------------------------------------(ДИАЛОГИ)-----------------------------------------------------

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Диалоги' and check_admin_access(message))
def show_dialogs_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Диалоги'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Просмотр диалогов", "Удалить диалоги")
    markup.add("Вернуться в общение") 
    markup.add("В меню админ-панели")

    bot.send_message(message.chat.id, "Выберите действие для диалогов:", reply_markup=markup)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Просмотр диалогов' and check_admin_access(message))
def show_user_dialogs(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Просмотр диалогов'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    chat_history = load_chat_history()
    users = load_users()

    # Формируем список пользователей, с кем были диалоги
    user_ids = [user_id.split('_')[1] for user_id in chat_history.keys()]
    user_ids = list(set(user_ids))  # Убираем повторяющиеся идентификаторы

    user_list = "\n\n".join(
        f"№{i + 1}. {escape_markdown(users.get(user_id, {}).get('username', 'N/A'))} - `{user_id}`"
        for i, user_id in enumerate(user_ids)
    )

    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(KeyboardButton("Вернуться в общение"))
    keyboard.add(KeyboardButton("В меню админ-панели"))

    bot.send_message(
        message.chat.id,
        f"*Список* пользователей *для просмотра* диалогов:\n\n{user_list}\n\nВведите номер пользователя:",
        parse_mode="Markdown",
        reply_markup=keyboard
    )

    dialog_states[message.chat.id] = {"state": "select_user", "user_ids": user_ids}
    save_dialog_states()

#@log_user_actions
@bot.message_handler(func=lambda message: message.chat.id in dialog_states and dialog_states[message.chat.id].get("state") == "select_user")
def handle_user_selection(message):

    if message.text == "В меню админ-панели":
        dialog_states.pop(message.chat.id, None)
        save_dialog_states()
        return


    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return
        
    if message.text == "Удалить диалоги":
        dialog_states.pop(message.chat.id, None)
        save_dialog_states()
        show_delete_dialogs_menu(message)
        return


    user_ids = dialog_states[message.chat.id]["user_ids"]
    users = load_users()

    try:
        selected_index = int(message.text) - 1
        if selected_index < 0 or selected_index >= len(user_ids):
            raise IndexError  # Неверный индекс

        selected_user_id = user_ids[selected_index]
        selected_username = users.get(selected_user_id, {}).get("username", "N/A")

        # Загружаем историю диалогов
        chat_key = f"{message.chat.id}_{selected_user_id}"
        chat_history = load_chat_history().get(chat_key, [])

        if not chat_history:
            bot.send_message(message.chat.id, "*История* переписки с этим пользователем *пуста*!", parse_mode="Markdown")
            return

        # Формируем список с временными интервалами
        dialog_list = []
        for i, dialog in enumerate(chat_history):
            if dialog:
                timestamps = [entry['timestamp'] for entry in dialog]
                start_time = timestamps[0].split(" в ")[1]
                end_time = timestamps[-1].split(" в ")[1]
                date = timestamps[0].split(" в ")[0]
                dialog_list.append(f"№{i + 1}. *{date}* (с {start_time} до {end_time})")
            else:
                dialog_list.append(f"№{i + 1}. (Пустой диалог)")

        dialog_text = "\n".join(dialog_list)
        bot.send_message(
            message.chat.id,
            f"*Выберите диалог* с пользователем {escape_markdown(selected_username)} - `{selected_user_id}`:\n\n{dialog_text}\n\nВведите номер диалога:",
            parse_mode="Markdown"
        )

        dialog_states[message.chat.id] = {
            "state": "select_dialog",
            "selected_user_id": selected_user_id,
            "chat_history": chat_history,
        }
        save_dialog_states()

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный ввод! Пожалуйста, введите номер пользователя")

#@log_user_actions
@bot.message_handler(func=lambda message: message.chat.id in dialog_states and dialog_states[message.chat.id].get("state") == "select_dialog")
def handle_dialog_selection(message):
    selected_user_id = dialog_states[message.chat.id]["selected_user_id"]
    chat_key = f"{message.chat.id}_{selected_user_id}"
    chat_history = load_chat_history().get(chat_key, [])

    if not chat_history:
        bot.send_message(message.chat.id, "*История* переписки с этим пользователем *пуста*!", parse_mode="Markdown")
        show_communication_menu(message)  # Возврат в меню
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    try:
        selected_dialog_index = int(message.text) - 1
        if selected_dialog_index < 0 or selected_dialog_index >= len(chat_history):
            raise IndexError  # Если индекс вне диапазона

        selected_dialog = chat_history[selected_dialog_index]

        if not selected_dialog:
            bot.send_message(message.chat.id, "Выбранный диалог пуст!")
            show_communication_menu(message)  # Возврат в меню
            return

        # Отправка всех сообщений выбранного диалога
        for entry in selected_dialog:
            timestamp = entry['timestamp']
            sender = entry['type']
            content = entry['content']
            caption = entry.get('caption')  # Получаем подпись

            if caption is not None:
                caption = caption.lower()  # Приводим к нижнему регистру, если подпись существует

            if content.startswith("photo:"):
                # Отправляем фото
                photo_id = content.replace("photo:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Фотография]"
                if caption:
                    message_text += f"\n✍ Подпись - {caption}"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_photo(message.chat.id, photo_id)

            elif content.startswith("sticker:"):
                # Отправляем стикер
                sticker_id = content.replace("sticker:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Стикер]"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_sticker(message.chat.id, sticker_id)

            elif content.startswith("voice:"):
                # Отправляем голосовое сообщение
                voice_id = content.replace("voice:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Голосовое сообщение]"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_voice(message.chat.id, voice_id)

            elif content.startswith("video:"):
                # Отправляем видео
                video_id = content.replace("video:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Видео]"
                if caption:
                    message_text += f"\n✍ Подпись - {caption}"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_video(message.chat.id, video_id)

            elif content.startswith("document:"):
                # Отправляем документ
                document_id = content.replace("document:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Документ]"
                if caption:
                    message_text += f"\n✍ Подпись - {caption}"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_document(message.chat.id, document_id)

            elif content.startswith("animation:"):
                # Отправка анимации отдельно
                animation_id = content.replace("animation:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Анимация]"
                if caption:
                    message_text += f"\n✍ Подпись - {escape_markdown(caption)}"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_animation(message.chat.id, animation_id)

            elif content.startswith("audio:"):
                # Отправляем аудиофайл
                audio_id = content.replace("audio:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [Аудио]"
                if caption:
                    message_text += f"\n✍ Подпись - {caption}"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_audio(message.chat.id, audio_id)

            elif content.startswith("location:"):
                # Отправляем локацию
                location_data = content.replace("location:", "").strip()
                lat, lon = map(float, location_data.split(","))
                message_text = f"👤 *{sender.upper()}* - [Локация]"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_location(message.chat.id, latitude=lat, longitude=lon)

            elif content.startswith("contact:"):
                # Отправляем контакт
                contact_data = content.replace("contact:", "").strip()
                phone, first_name, last_name = contact_data.split(",", maxsplit=2)
                message_text = f"👤 *{sender.upper()}* - [Контакт]"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_contact(message.chat.id, phone_number=phone, first_name=first_name, last_name=last_name)

            elif content.startswith("gif:"):
                # Отправляем GIF
                gif_id = content.replace("gif:", "").strip()
                message_text = f"👤 *{sender.upper()}* - [GIF]"
                if caption:
                    message_text += f"\n✍ Подпись - {caption}"
                message_text += f"\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")
                bot.send_document(message.chat.id, gif_id)

            else:
                # Отправляем текстовое сообщение
                message_text = f"👤 *{sender.upper()}* - [Текст]\n📝 Текст - {content}\n📅 *Дата и время*: _{timestamp}_"
                bot.send_message(message.chat.id, message_text, parse_mode="Markdown")

        # Завершаем диалог
        del dialog_states[message.chat.id]
        save_dialog_states()

        # Возврат в меню
        show_communication_menu(message)

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный ввод! Пожалуйста, введите номер диалога")

def save_dialog_states():
    with open('data base/admin/chats/dialog_states.json', 'w', encoding='utf-8') as file:
        json.dump(dialog_states, file, ensure_ascii=False, indent=4)


#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить диалоги' and check_admin_access(message))
def show_delete_dialogs_menu(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить диалоги'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Удалить диалог", "Удалить все диалоги")
    markup.add("Вернуться в общение")
    markup.add("В меню админ-панели")

    bot.send_message(message.chat.id, "Выберите действие для удаления диалогов:", reply_markup=markup)

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить диалог' and check_admin_access(message))
def delete_dialog(message):

    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить диалог'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    chat_history = load_chat_history()
    users = load_users()

    user_ids = [user_id.split('_')[1] for user_id in chat_history.keys()]
    user_ids = list(set(user_ids))

    user_list = "\n\n".join(
        f"№{i + 1}. {escape_markdown(users.get(user_id, {}).get('username', 'N/A'))} - `{user_id}`"
        for i, user_id in enumerate(user_ids)
    )

    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(KeyboardButton("Вернуться в общение"))
    keyboard.add(KeyboardButton("В меню админ-панели"))

    bot.send_message(
        message.chat.id,
        f"*Список* пользователей *для удаления* диалогов:\n\n{user_list}\n\nВведите номер пользователя:",
        parse_mode="Markdown",
        reply_markup=keyboard
    )

    dialog_states[message.chat.id] = {"state": "delete_dialog_select_user", "user_ids": user_ids}
    save_dialog_states()

#@log_user_actions
@bot.message_handler(func=lambda message: message.chat.id in dialog_states and dialog_states[message.chat.id].get("state") == "delete_dialog_select_user")
def handle_delete_dialog_user_selection(message):

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        dialog_states.pop(message.chat.id, None)
        save_dialog_states()
        return show_admin_panel(message)

    user_ids = dialog_states[message.chat.id]["user_ids"]
    users = load_users()

    try:
        selected_index = int(message.text) - 1
        if selected_index < 0 or selected_index >= len(user_ids):
            raise IndexError

        selected_user_id = user_ids[selected_index]
        selected_username = users.get(selected_user_id, {}).get("username", "N/A")

        chat_key = f"{message.chat.id}_{selected_user_id}"
        chat_history = load_chat_history().get(chat_key, [])

        if not chat_history:
            bot.send_message(message.chat.id, "*История* переписки с этим пользователем *пуста*!", parse_mode="Markdown")
            return show_admin_panel(message)

        dialog_list = []
        for i, dialog in enumerate(chat_history):
            if dialog:
                timestamps = [entry['timestamp'] for entry in dialog]
                start_time = timestamps[0].split(" в ")[1]
                end_time = timestamps[-1].split(" в ")[1]
                date = timestamps[0].split(" в ")[0]
                dialog_list.append(f"№{i + 1}. *{date}* (с {start_time} до {end_time})")
            else:
                dialog_list.append(f"№{i + 1}. (Пустой диалог)")

        dialog_text = "\n".join(dialog_list)
        bot.send_message(
            message.chat.id,
            f"*Выберите диалог* для удаления с пользователем {escape_markdown(selected_username)} - `{selected_user_id}`:\n\n{dialog_text}\n\nВведите номер диалога:",
            parse_mode="Markdown"
        )

        dialog_states[message.chat.id] = {
            "state": "delete_dialog_select_dialog",
            "selected_user_id": selected_user_id,
            "chat_history": chat_history,
        }
        save_dialog_states()

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный ввод! Пожалуйста, введите номер пользователя")

#@log_user_actions
@bot.message_handler(func=lambda message: message.chat.id in dialog_states and dialog_states[message.chat.id].get("state") == "delete_dialog_select_dialog")
def handle_delete_dialog_selection(message):
    selected_user_id = dialog_states[message.chat.id]["selected_user_id"]
    chat_key = f"{message.chat.id}_{selected_user_id}"
    chat_history = load_chat_history().get(chat_key, [])

    if not chat_history:
        bot.send_message(message.chat.id, "*История* переписки с этим пользователем *пуста*!", parse_mode="Markdown")
        return show_communication_menu(message)

    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    try:
        selected_dialog_index = int(message.text) - 1
        if selected_dialog_index < 0 or selected_dialog_index >= len(chat_history):
            raise IndexError

        del chat_history[selected_dialog_index]

        chat_history_data = load_chat_history()
        chat_history_data[chat_key] = chat_history

        with open(CHAT_HISTORY_PATH, 'w', encoding='utf-8') as file:
            json.dump(chat_history_data, file, ensure_ascii=False, indent=4)

        bot.send_message(message.chat.id, "Диалог успешно *удален*!", parse_mode="Markdown")
        return show_communication_menu(message)

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный ввод! Пожалуйста, введите номер диалога")

#@log_user_actions
@bot.message_handler(func=lambda message: message.text == 'Удалить все диалоги' and check_admin_access(message))
def delete_all_dialogs(message):
    admin_id = str(message.chat.id)
    if not check_permission(admin_id, 'Удалить все диалоги'):
        bot.send_message(message.chat.id, "⛔️ У вас *нет прав доступа* к этой функции!", parse_mode="Markdown")
        return

    chat_history = load_chat_history()
    users = load_users()

    user_ids = [user_id.split('_')[1] for user_id in chat_history.keys()]
    user_ids = list(set(user_ids))

    user_list = "\n\n".join(
        f"№{i + 1}. {escape_markdown(users.get(user_id, {}).get('username', 'N/A'))} - `{user_id}`"
        for i, user_id in enumerate(user_ids)
    )

    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    keyboard.add(KeyboardButton("Вернуться в общение"))
    keyboard.add(KeyboardButton("В меню админ-панели"))

    bot.send_message(
        message.chat.id,
        f"*Список* пользователей *для удаления* всех диалогов:\n\n{user_list}\n\nВведите номер пользователя:",
        parse_mode="Markdown",
        reply_markup=keyboard
    )

    dialog_states[message.chat.id] = {"state": "delete_all_dialogs_select_user", "user_ids": user_ids}
    save_dialog_states()

#@log_user_actions
@bot.message_handler(func=lambda message: message.chat.id in dialog_states and dialog_states[message.chat.id].get("state") == "delete_all_dialogs_select_user")
def handle_delete_all_dialogs_user_selection(message):
    if message.text == "Вернуться в общение":
        show_communication_menu(message)
        return

    if message.text == "В меню админ-панели":
        dialog_states.pop(message.chat.id, None)
        save_dialog_states()
        return show_admin_panel(message)

    user_ids = dialog_states[message.chat.id]["user_ids"]
    users = load_users()

    try:
        selected_index = int(message.text) - 1
        if selected_index < 0 or selected_index >= len(user_ids):
            raise IndexError

        selected_user_id = user_ids[selected_index]
        selected_username = users.get(selected_user_id, {}).get("username", "N/A")

        bot.send_message(
            message.chat.id,
            f"Вы уверены, что хотите *удалить все диалоги* с пользователем {escape_markdown(selected_username)} - `{selected_user_id}`?\n\nВведите *ДА* для принятия или *НЕТ* для отклонения",
            parse_mode="Markdown"
        )

        dialog_states[message.chat.id] = {
            "state": "confirm_delete_all_dialogs",
            "selected_user_id": selected_user_id,
        }
        save_dialog_states()

    except (ValueError, IndexError):
        bot.send_message(message.chat.id, "Неверный ввод! Пожалуйста, введите номер пользователя")

#@log_user_actions
@bot.message_handler(func=lambda message: message.chat.id in dialog_states and dialog_states[message.chat.id].get("state") == "confirm_delete_all_dialogs")
def handle_confirm_delete_all_dialogs(message):
    if message.text.lower() not in ["да", "нет"]:
        bot.send_message(message.chat.id, "Неверный ввод. Пожалуйста, введите *ДА* для удаления или *НЕТ* для отмены", parse_mode="Markdown")
        return

    if message.text.lower() == "да":
        selected_user_id = dialog_states[message.chat.id]["selected_user_id"]
        chat_key = f"{message.chat.id}_{selected_user_id}"

        chat_history_data = load_chat_history()
        if chat_key in chat_history_data:
            del chat_history_data[chat_key]

        with open(CHAT_HISTORY_PATH, 'w', encoding='utf-8') as file:
            json.dump(chat_history_data, file, ensure_ascii=False, indent=4)

        bot.send_message(message.chat.id, "*Все диалоги* с пользователем успешно *удалены*!", parse_mode="Markdown")
    else:
        bot.send_message(message.chat.id, "Удаление *всех диалогов* с пользователем *отменено*!", parse_mode="Markdown")

    return show_communication_menu(message)

# Изменения в функции handle_chat_messages для сохранения сообщений в историю

#@log_user_actions
@bot.message_handler(content_types=['text', 'photo', 'video', 'document', 'animation', 'sticker', 'audio', 'contact', 'voice', 'video_note', 'gif'])
@check_chat_state
def handle_chat_messages(message):
    user_id = message.from_user.id

    # Игнорируем команды, чтобы они не пересылались в чате
    if message.text and (message.text.startswith('/') or message.text.startswith('Стоп')):
        return

    # Обновляем время последней активности
    if user_id in active_chats:
        active_chats[user_id]["last_activity_time"] = time.time()
        save_active_chats()

    # Если сообщение от администратора
    if user_id in active_admin_chats:
        target_user_id = active_admin_chats[user_id]
        if message.content_type == 'text':
            bot.send_message(target_user_id, f"{message.text}")
            save_message_to_history(user_id, target_user_id, message.text, 'admin')  # Сохранение сообщения
        elif message.content_type == 'photo':
            media_group = []
            media_group.append(types.InputMediaPhoto(message.photo[-1].file_id, caption=message.caption))
            bot.send_media_group(target_user_id, media_group)
            save_message_to_history(user_id, target_user_id, f"photo: {message.photo[-1].file_id}", 'admin', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'video':
            media_group = []
            media_group.append(types.InputMediaVideo(message.video.file_id, caption=message.caption))
            bot.send_media_group(target_user_id, media_group)
            save_message_to_history(user_id, target_user_id, f"video: {message.video.file_id}", 'admin', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'document':
            media_group = []
            media_group.append(types.InputMediaDocument(message.document.file_id, caption=message.caption))
            bot.send_media_group(target_user_id, media_group)
            save_message_to_history(user_id, target_user_id, f"document: {message.document.file_id}", 'admin', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'animation':
            # Отправка анимации отдельно
            bot.send_animation(target_user_id, message.animation.file_id, caption=message.caption)
            save_message_to_history(user_id, target_user_id, f"animation: {message.animation.file_id}", 'admin', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'sticker':
            bot.send_sticker(target_user_id, message.sticker.file_id)
            save_message_to_history(user_id, target_user_id, f"sticker: {message.sticker.file_id}", 'admin')  # Сохранение сообщения
        elif message.content_type == 'audio':
            bot.send_audio(target_user_id, message.audio.file_id, caption=message.caption)
            save_message_to_history(user_id, target_user_id, f"audio: {message.audio.file_id}", 'admin', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'contact':
            bot.send_contact(target_user_id, message.contact.phone_number, message.contact.first_name)
            save_message_to_history(user_id, target_user_id, f"contact: {message.contact.phone_number}", 'admin')  # Сохранение сообщения
        elif message.content_type == 'voice':
            bot.send_voice(target_user_id, message.voice.file_id, caption=message.caption)
            save_message_to_history(user_id, target_user_id, f"voice: {message.voice.file_id}", 'admin', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'video_note':
            bot.send_video_note(target_user_id, message.video_note.file_id)
            save_message_to_history(user_id, target_user_id, f"video_note: {message.video_note.file_id}", 'admin')  # Сохранение сообщения
        elif message.content_type == 'gif':
            bot.send_document(target_user_id, message.document.file_id, caption=message.caption)
            save_message_to_history(user_id, target_user_id, f"gif: {message.document.file_id}", 'admin', caption=message.caption)  # Сохранение сообщения

    # Если сообщение от пользователя
    elif user_id in active_user_chats:
        target_admin_id = active_user_chats[user_id]
        if message.content_type == 'text':
            bot.send_message(target_admin_id, f"{message.text}")
            save_message_to_history(target_admin_id, user_id, message.text, 'user')  # Сохранение сообщения
        elif message.content_type == 'photo':
            media_group = []
            media_group.append(types.InputMediaPhoto(message.photo[-1].file_id, caption=message.caption))
            bot.send_media_group(target_admin_id, media_group)
            save_message_to_history(target_admin_id, user_id, f"photo: {message.photo[-1].file_id}", 'user', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'video':
            media_group = []
            media_group.append(types.InputMediaVideo(message.video.file_id, caption=message.caption))
            bot.send_media_group(target_admin_id, media_group)
            save_message_to_history(target_admin_id, user_id, f"video: {message.video.file_id}", 'user', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'document':
            media_group = []
            media_group.append(types.InputMediaDocument(message.document.file_id, caption=message.caption))
            bot.send_media_group(target_admin_id, media_group)
            save_message_to_history(target_admin_id, user_id, f"document: {message.document.file_id}", 'user', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'animation':
            # Отправка анимации отдельно
            bot.send_animation(target_admin_id, message.animation.file_id, caption=message.caption)
            save_message_to_history(target_admin_id, user_id, f"animation: {message.animation.file_id}", 'user', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'sticker':
            bot.send_sticker(target_admin_id, message.sticker.file_id)
            save_message_to_history(target_admin_id, user_id, f"sticker: {message.sticker.file_id}", 'user')  # Сохранение сообщения
        elif message.content_type == 'audio':
            bot.send_audio(target_admin_id, message.audio.file_id, caption=message.caption)
            save_message_to_history(target_admin_id, user_id, f"audio: {message.audio.file_id}", 'user', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'contact':
            bot.send_contact(target_admin_id, message.contact.phone_number, message.contact.first_name)
            save_message_to_history(target_admin_id, user_id, f"contact: {message.contact.phone_number}", 'user')  # Сохранение сообщения
        elif message.content_type == 'voice':
            bot.send_voice(target_admin_id, message.voice.file_id, caption=message.caption)
            save_message_to_history(target_admin_id, user_id, f"voice: {message.voice.file_id}", 'user', caption=message.caption)  # Сохранение сообщения
        elif message.content_type == 'video_note':
            bot.send_video_note(target_admin_id, message.video_note.file_id)
            save_message_to_history(target_admin_id, user_id, f"video_note: {message.video_note.file_id}", 'user')  # Сохранение сообщения
        elif message.content_type == 'gif':
            bot.send_document(target_admin_id, message.document.file_id, caption=message.caption)
            save_message_to_history(target_admin_id, user_id, f"gif: {message.document.file_id}", 'user', caption=message.caption)  # Сохранение сообщения


def check_chat_activity():
    while True:
        current_time = time.time()
        for user_id, chat_data in list(active_chats.items()):
            if chat_data["status"] == "active":
                last_activity_time = chat_data.get("last_activity_time", current_time)
                if current_time - last_activity_time > 65:  # 5 минут = 300 секунд
                    admin_id = chat_data.get("admin_id")
                    users_data = load_users()
                    username = users_data.get(str(user_id), {}).get('username', 'Unknown')
                    escaped_username = escape_markdown(username)

                    if admin_id:
                        bot.send_message(admin_id, f"Чат с пользователем {escaped_username} - `{user_id}` был автоматически *завершен* из-за неактивности!", parse_mode="Markdown")
                        # Перенаправляем администратора в меню админ-панели
                        return_admin_to_menu(admin_id)

                    bot.send_message(user_id, "Чат с администратором был автоматически *завершен* из-за неактивности!", parse_mode="Markdown")
                    start_menu(user_id)  # Перебросить пользователя в главное меню

                    # Сохраняем текущий диалог в историю
                    chat_key = f"{admin_id}_{user_id}"
                    if chat_key in current_dialogs:
                        chat_history = load_chat_history()
                        chat_history[chat_key].append(current_dialogs[chat_key])
                        del current_dialogs[chat_key]
                        with open(CHAT_HISTORY_PATH, 'w', encoding='utf-8') as file:
                            json.dump(chat_history, file, ensure_ascii=False, indent=4)

                    del active_chats[user_id]
                    save_active_chats()
        time.sleep(60)  # Проверяем каждую минуту

# Запуск фонового потока для проверки активности чатов
threading.Thread(target=check_chat_activity, daemon=True).start()




# (ADMIN 6) ------------------------------------------ "ВЫХОД ДЛЯ АДМИН-ПАНЕЛИ" ---------------------------------------------------

# Функция выхода из админ-панели
@bot.message_handler(func=lambda message: message.text == 'Выход' and message.chat.id in admin_sessions)
#@log_user_actions
def admin_logout(message):
    try:
        bot.send_message(message.chat.id, "Вы вышли из админ-панели!\nБыстрый вход сохранен")
        return_to_menu(message)
    except telebot.apihelper.ApiTelegramException as e:
        if e.result_json['description'] == 'Bad Request: chat not found':
            print(f"Чат не найден по запросу user_id: {message.chat.id}")
        else:
            print(f"Произошла ошибка: {e}")
            
# (16) --------------- КОД ДЛЯ "ФУНКЦИЯ ДЛЯ ОБНОВЛЕНИЙ ОТ TG" ---------------

# Функция для обработки повторных попыток
def start_bot_with_retries(retries=10000000, delay=5):
    attempt = 0
    while attempt < retries:
        try:
            print(f"Запуск попытки #{attempt + 1}")
            bot.polling(none_stop=True, interval=1, timeout=120, long_polling_timeout=120)
        except (ReadTimeout, ConnectionError) as e:
            print(f"Ошибка: {e}. Попытка {attempt + 1} из {retries}")
            attempt += 1
            if attempt < retries:
                print(f"Повторная попытка через {delay} секунд...")
                time.sleep(delay)  # Ожидание перед повторной попыткой
            else:
                print("Превышено количество попыток! Бот не смог запуститься!")
        except Exception as e:
            print(f"Неожиданная ошибка: {e}")
            break

# Запуск бота с повторными попытками
start_bot_with_retries()

@bot.message_handler(func=lambda message: True)
@restricted
@track_user_activity
@check_chat_state
#@log_user_actions
def echo_all(message):
    bot.reply_to(message, message.text)

# (17) --------------- КОД ДЛЯ "ЗАПУСК БОТА" ---------------
if __name__ == '__main__':
    # Запуск бота
    bot_thread = threading.Thread(target=lambda: bot.polling(none_stop=True, interval=1, timeout=120, long_polling_timeout=120), daemon=True)
    bot_thread.start()

    # Основной цикл для проверки расписания
    while True:
        schedule.run_pending()
        time.sleep(60)  # Проверка расписания каждые 60 секунд